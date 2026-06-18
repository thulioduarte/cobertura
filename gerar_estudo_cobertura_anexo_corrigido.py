# -*- coding: utf-8 -*-
"""
Gera estudo de cobertura usando APENAS 2 arquivos (versão com diagnóstico):

1) Sell-in
   - Aba esperada: "Sell-in" ou equivalente, como "Tabela Sell In".
   - Mesmo template analisado anteriormente, com cabeçalho geralmente na linha 14, a partir da coluna B.
   - Aceita UF SM ou UF, EAN ou SKU, Categoria SM ou CATEGORIA SCANN,
     SELL IN (Kg / L) ou SELL IN (QUANTIDADE).

2) Sell-out
   - Arquivo Excel ou CSV no padrão Publicar/VTA, já contendo:
     DATA, UF, NIVEL1, NIVEL2, Fabricante, Marca, SKU,
     Qtd_de_Vendas, Vendas_em_volume, etc.

O que o script faz:
- Permite escolher se a cobertura será por Quantia ou Volume.
- Permite escolher a regra de categoria: Categoria, NIVEL1 ou NIVEL2.
- Quando usa NIVEL1/NIVEL2, mapeia os SKUs do Sell-in para o PROD do Sell-out, como um PROCX, e gera uma aba para cada valor do nível escolhido.
- Quando usa Categoria, prioriza a coluna Categoria do Sell-out; se ela não existir, usa a Categoria do Sell-in por SKU.
- Cria uma aba por categoria ou por PROD do nível escolhido.
- Mantém a cobertura mensal como 12 meses móveis.
- Usa todos os meses em comum entre Sell-in e Sell-out.
- No bloco MAT, prioriza anos fechados. Se o último ano estiver incompleto, usa YTD.
- Calcula a importância dos SKUs no Sell-in e no Sell-out.
- Calcula a cobertura de importância do Sell-out limitada ao fabricante predominante do Sell-in.
  Esse fabricante é identificado pelos SKUs do Sell-in encontrados no Sell-out.

Instalação:
    pip install pandas openpyxl xlsxwriter numpy

Observação:
    Sell-in e Sell-out podem ser Excel ou CSV/TXT/TSV. O CSV é lido com detecção automática
    de separador e encoding.

Uso com janelas de seleção:
    python gerar_estudo_cobertura_sellin_sellout_final.py

Uso via linha de comando:
    python gerar_estudo_cobertura_sellin_sellout_final.py ^
        --sellin "sellin.xlsx" ^
        --sellout "sellout.xlsx" ^
        --metrica volume ^
        --nivel CATEGORIA ^
        --saida "estudo_cobertura.xlsx"

Parâmetros:
    --metrica: volume, quantia ou volume_variavel
        volume          -> usa Vendas_em_volume no Sell-out
        quantia         -> usa Qtd_de_Vendas no Sell-out
        volume_variavel -> converte o Sell-in em volume usando gramatura média ponderada localizada no nome do SKU do Sell-out

    --nivel: CATEGORIA, NIVEL1 ou NIVEL2
        Define a regra de categoria usada no cálculo.
        CATEGORIA -> usa a categoria do Sell-out; se não existir, usa a categoria do Sell-in por SKU.
        NIVEL1    -> usa cada valor de NIVEL1 do Sell-out como uma aba própria.
        NIVEL2    -> usa cada valor de NIVEL2 do Sell-out como uma aba própria.
"""

from __future__ import annotations

import argparse
import math
import queue
import re
import sys
import threading
import tempfile
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


MESES_MAT = 12
MESES_MOVEL = 6


# ============================================================
# Opções de geração do Excel
# ============================================================

OPCOES_SAIDA_PADRAO = {
    # Estudo individual / Cobertura Dash
    "resumo_categorias": True,
    "abas_categorias": True,
    "base_skus": True,
    "base_contribuicao_sellout": False,
    "skus_por_categoria": True,
    "crosschecks": True,
    "parametros": True,
    "descricao_calculos": True,
    "avisos": True,
    "graficos_cobertura": True,
    # Comparação 2.0 x 3.0 / Estudo com 2 Sell-outs
    "abas_auxiliares_comparacao": True,
    "top20_sku_canal_uf": False,
}

ROTULOS_OPCOES_SAIDA = {
    "resumo_categorias": "Resumo Categorias",
    "abas_categorias": "Abas por categoria/PROD",
    "base_skus": "Base SKUs",
    "base_contribuicao_sellout": "Base Contribuição Sell-out",
    "skus_por_categoria": "SKUs por Categoria",
    "crosschecks": "Crosschecks",
    "parametros": "Parâmetros",
    "descricao_calculos": "Descrição Cálculos",
    "avisos": "Avisos",
    "graficos_cobertura": "Gráficos Cobertura",
    "abas_auxiliares_comparacao": "Abas auxiliares da comparação",
    "top20_sku_canal_uf": "TOP 20 SKU por UF/Canal",
}


def normalizar_opcoes_saida(opcoes: Optional[Dict[str, object]] = None) -> Dict[str, bool]:
    """Normaliza as opções de geração, mantendo o comportamento antigo como padrão."""
    saida = dict(OPCOES_SAIDA_PADRAO)
    if opcoes:
        for chave, valor in dict(opcoes).items():
            if chave in saida:
                saida[chave] = bool(valor)

    # Compatibilidade com a chave antiga usada no primeiro ajuste do TOP 20.
    if opcoes and "gerar_top20_sku_canal_uf" in opcoes:
        saida["top20_sku_canal_uf"] = bool(opcoes.get("gerar_top20_sku_canal_uf"))
    return saida


def opcao_saida(opcoes: Optional[Dict[str, object]], chave: str) -> bool:
    return bool(normalizar_opcoes_saida(opcoes).get(chave, True))


def opcoes_saida_para_parametros(opcoes: Optional[Dict[str, object]]) -> str:
    opcoes_norm = normalizar_opcoes_saida(opcoes)
    partes = []
    for chave, rotulo in ROTULOS_OPCOES_SAIDA.items():
        partes.append(f"{rotulo}: {'Sim' if opcoes_norm.get(chave, False) else 'Não'}")
    return "; ".join(partes)


def garantir_aba_info_se_vazio(writer, workbook, titulo: str, opcoes: Optional[Dict[str, object]] = None):
    """Evita gerar um Excel sem nenhuma aba quando o usuário desmarca todas as saídas."""
    if getattr(writer, "sheets", None):
        if len(writer.sheets) > 0:
            return
    ws = workbook.add_worksheet("Info")
    writer.sheets["Info"] = ws
    fmt_titulo = workbook.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white"})
    fmt_texto = workbook.add_format({"text_wrap": True, "valign": "top"})
    ws.write("A1", titulo, fmt_titulo)
    ws.write("A3", "Nenhuma saída foi selecionada para geração. Marque pelo menos uma opção na tela de geração para ver as tabelas no Excel.", fmt_texto)
    if opcoes is not None:
        ws.write("A5", "Opções usadas:", fmt_titulo)
        ws.write("A6", opcoes_saida_para_parametros(opcoes), fmt_texto)
    ws.set_column("A:A", 120)


# ============================================================
# Utilitários gerais
# ============================================================


def normalizar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def normalizar_categoria(valor) -> str:
    return normalizar_texto(valor)


# Ordem final exigida para comparação por UF/agrupamento de estados.
UF_COMPARACAO_ORDEM = [
    "AL - SE", "BA", "CE", "DF", "ES", "GO", "MA - PI", "MG", "MS", "MT",
    "PB", "PE", "PR", "RJ", "RN", "RR - AM - RO - AC", "RS", "SC", "SP", "TO - PA - AP",
]
UF_COMPARACAO_RANK = {uf: i for i, uf in enumerate(UF_COMPARACAO_ORDEM)}
UF_PARA_GRUPO_COMPARACAO = {
    "AL": "AL - SE", "SE": "AL - SE",
    "MA": "MA - PI", "PI": "MA - PI",
    "RR": "RR - AM - RO - AC", "AM": "RR - AM - RO - AC", "RO": "RR - AM - RO - AC", "AC": "RR - AM - RO - AC",
    "TO": "TO - PA - AP", "PA": "TO - PA - AP", "AP": "TO - PA - AP",
    "BA": "BA", "CE": "CE", "DF": "DF", "ES": "ES", "GO": "GO", "MG": "MG", "MS": "MS", "MT": "MT",
    "PB": "PB", "PE": "PE", "PR": "PR", "RJ": "RJ", "RN": "RN", "RS": "RS", "SC": "SC", "SP": "SP",
}


def padronizar_uf_comparacao(valor) -> str:
    """
    Padroniza UF/região para a comparação Sell-in x Sell-out.

    Regras principais:
    - AL e SE viram AL - SE.
    - MA e PI viram MA - PI.
    - RR, AM, RO e AC viram RR - AM - RO - AC.
    - TO, PA e AP viram TO - PA - AP.
    - Qualquer abertura de São Paulo, como INT - SP, MET - SP, SP - INT,
      SP - Interior, SP - MET ou SP-REGMET, vira SP.
    - Demais UFs permanecem como UF individual.
    """
    if pd.isna(valor):
        return "SEM UF"
    original = str(valor).strip()
    if not original or normalizar_texto(original) in {"nan", "none", "sem uf", "sem_uf", "sem-uf"}:
        return "SEM UF"
    if normalizar_texto(original) in {"total", "todos", "geral"}:
        return "TOTAL"

    norm = normalizar_texto(original).upper()
    norm = re.sub(r"\bGRANDE\b", "MET", norm)
    norm = re.sub(r"\bREGIAO\b", "REG", norm)
    norm = re.sub(r"\bMETROPOLITANA\b", "MET", norm)
    norm = re.sub(r"\bINTERIOR\b", "INT", norm)

    # São Paulo pode vir como INT - SP, MET - SP, SP - INT, SP - Interior,
    # SP - MET, SP-REGMET, SP_REGMET etc.
    if re.search(r"(^|\W)SP($|\W)", f" {norm} ") or norm.startswith("SP") or norm.endswith("SP"):
        return "SP"

    tokens = re.findall(r"[A-Z]{2}", norm)
    tokens = [t for t in tokens if t in UF_PARA_GRUPO_COMPARACAO]
    if not tokens:
        return original

    # Grupos específicos do Sell-out. Mesmo que venha apenas AL, por exemplo,
    # ele entra no grupo AL - SE para casar com AL_SE do Sell-out.
    grupos_prioridade = [
        ({"AL", "SE"}, "AL - SE"),
        ({"MA", "PI"}, "MA - PI"),
        ({"RR", "AM", "RO", "AC"}, "RR - AM - RO - AC"),
        ({"TO", "PA", "AP"}, "TO - PA - AP"),
    ]
    token_set = set(tokens)
    for grupo_tokens, grupo_nome in grupos_prioridade:
        if token_set & grupo_tokens:
            return grupo_nome

    # Se tiver mais de uma UF fora dos grupos conhecidos, mantém a primeira UF reconhecida.
    return UF_PARA_GRUPO_COMPARACAO.get(tokens[0], tokens[0])


def adicionar_uf_comparacao(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "uf" not in out.columns:
        out["uf"] = "TOTAL"
    if "uf_original" not in out.columns:
        out["uf_original"] = out["uf"]

    s = out["uf"].astype(str).str.strip().str.upper()
    s = s.replace(
        {"": "SEM UF", "NAN": "SEM UF", "NONE": "SEM UF", "TOTAL": "TOTAL", "TODOS": "TOTAL", "GERAL": "TOTAL"})

    s = s.str.replace(r"\bGRANDE\b", "MET", regex=True)
    s = s.str.replace(r"\bREGIAO\b", "REG", regex=True)
    s = s.str.replace(r"\bMETROPOLITANA\b", "MET", regex=True)
    s = s.str.replace(r"\bINTERIOR\b", "INT", regex=True)

    # Consolidar SP
    mask_sp = s.str.contains(r"(?:^|\W)SP(?:$|\W)", regex=True) | s.str.startswith("SP") | s.str.endswith("SP")
    s = s.mask(mask_sp, "SP")

    mapa_grupos = {
        "AL": "AL - SE", "SE": "AL - SE",
        "MA": "MA - PI", "PI": "MA - PI",
        "RR": "RR - AM - RO - AC", "AM": "RR - AM - RO - AC", "RO": "RR - AM - RO - AC", "AC": "RR - AM - RO - AC",
        "TO": "TO - PA - AP", "PA": "TO - PA - AP", "AP": "TO - PA - AP"
    }

    # Extrai e mapeia em bloco, mantendo o índice original.
    tokens = s.str.extract(r"([A-Z]{2})")[0]
    tokens_mapeados = tokens.map(mapa_grupos).fillna(tokens)

    out["uf_comparacao"] = np.where(s.isin(["SP", "TOTAL", "SEM UF"]), s, tokens_mapeados)
    out["uf_comparacao"] = out["uf_comparacao"].replace("", "TOTAL").fillna("TOTAL")

    return out


def ordenar_por_uf_comparacao(df: pd.DataFrame, coluna_uf: str = "UF", coluna_secundaria: Optional[str] = None) -> pd.DataFrame:
    if df.empty or coluna_uf not in df.columns:
        return df
    out = df.copy()
    out["_ordem_uf"] = out[coluna_uf].map(lambda x: UF_COMPARACAO_RANK.get(str(x), 10_000))
    sort_cols = ["_ordem_uf"]
    asc = [True]
    if coluna_secundaria and coluna_secundaria in out.columns:
        sort_cols.append(coluna_secundaria)
        asc.append(False)
    else:
        sort_cols.append(coluna_uf)
        asc.append(True)
    out = out.sort_values(sort_cols, ascending=asc).drop(columns=["_ordem_uf"], errors="ignore")
    return out


def agrupar_valor_por_uf_comparacao(df: pd.DataFrame, coluna_valor: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["uf", coluna_valor])
    temp = adicionar_uf_comparacao(df)
    return temp.groupby("uf_comparacao", as_index=False)[coluna_valor].sum().rename(columns={"uf_comparacao": "uf"})


def montar_tabela_uf_comparacao(si_base: pd.DataFrame, so_base: pd.DataFrame) -> pd.DataFrame:
    """
    Monta a tabela por UF no padrão de comparação do projeto.

    Regras:
    - Sell-in e Sell-out passam pela mesma consolidação de UF.
    - SP é sempre consolidado como SP, somando INT/MET/Interior/RegMet.
    - Linhas sem UF entram como SEM UF.
    - Inclui todas as UFs/grupos padrão, mesmo quando o valor for zero.
    - Inclui TOTAL ao final.
    """
    si_uf = agrupar_valor_por_uf_comparacao(si_base, "valor_sellin") if si_base is not None and not si_base.empty else pd.DataFrame(columns=["uf", "valor_sellin"])
    so_uf = agrupar_valor_por_uf_comparacao(so_base, "valor_sellout") if so_base is not None and not so_base.empty else pd.DataFrame(columns=["uf", "valor_sellout"])
    merge = si_uf.merge(so_uf, on="uf", how="outer").fillna(0)

    # Remove qualquer TOTAL originado do arquivo antes de recalcular o total final.
    merge = merge[merge["uf"].astype(str).str.upper() != "TOTAL"].copy()

    base_padrao = pd.DataFrame({"uf": UF_COMPARACAO_ORDEM})
    uf = base_padrao.merge(merge, on="uf", how="left").fillna(0)
    extras = merge[~merge["uf"].isin(UF_COMPARACAO_ORDEM)].copy()
    extras = extras.sort_values("uf")
    uf = pd.concat([uf, extras], ignore_index=True)

    total_si = float(pd.to_numeric(uf.get("valor_sellin", 0), errors="coerce").fillna(0).sum())
    total_so = float(pd.to_numeric(uf.get("valor_sellout", 0), errors="coerce").fillna(0).sum())
    uf["Cobertura"] = uf["valor_sellout"] / uf["valor_sellin"].replace(0, np.nan)
    uf["Importância Sell-in"] = uf["valor_sellin"] / total_si if total_si else np.nan
    uf["Importância Sell-out"] = uf["valor_sellout"] / total_so if total_so else np.nan

    total = pd.DataFrame([{
        "uf": "TOTAL",
        "valor_sellin": total_si,
        "valor_sellout": total_so,
        "Cobertura": divisao_segura(total_so, total_si),
        "Importância Sell-in": 1 if total_si else np.nan,
        "Importância Sell-out": 1 if total_so else np.nan,
    }])
    uf = pd.concat([uf, total], ignore_index=True)
    uf = uf.rename(columns={"uf": "UF", "valor_sellin": "Sell-in", "valor_sellout": "Sell-out"})
    return uf


MESES_PT_ABREV = {
    1: "jan.", 2: "fev.", 3: "mar.", 4: "abr.", 5: "mai.", 6: "jun.",
    7: "jul.", 8: "ago.", 9: "set.", 10: "out.", 11: "nov.", 12: "dez.",
}

# Colunas possíveis de mês/data em arquivos Sell-in e Dash.
# Mantém a lista centralizada para não precisar corrigir em vários pontos do código.
COLUNAS_MES_ALTERNATIVAS = [
    "Dia de Selector Data", "Selector Data",
    "ANO MÊS", "ANO MES", "ANO_MÊS", "ANO_MES", "Ano Mês", "Ano Mes", "AnoMes",
    "AAAAMM", "YYYYMM",
    "Mês/Ano", "Mes/Ano", "Mês Ano", "Mes Ano",
    "Competência", "Competencia",
    "Referência", "Referencia",
    "Data", "DATA",
    "Período", "Periodo", "PERIODO",
    "Mês", "Mes",
]

MESES_PT = {
    "jan": 1, "janeiro": 1,
    "fev": 2, "fevereiro": 2,
    "mar": 3, "marco": 3, "março": 3,
    "abr": 4, "abril": 4,
    "mai": 5, "maio": 5,
    "jun": 6, "junho": 6,
    "jul": 7, "julho": 7,
    "ago": 8, "agosto": 8,
    "set": 9, "setembro": 9,
    "out": 10, "outubro": 10,
    "nov": 11, "novembro": 11,
    "dez": 12, "dezembro": 12,
}


def label_mes_pt(mes) -> str:
    if pd.isna(mes):
        return "Sem mês"
    ts = pd.Timestamp(mes)
    return f"{MESES_PT_ABREV.get(ts.month, ts.strftime('%b').lower())} {str(ts.year)[-2:]}"


def ean_texto(valor) -> str:
    """Mantém EAN/SKU como texto numérico, sem .0 e sem notação científica quando possível."""
    if pd.isna(valor):
        return ""
    if isinstance(valor, (int, np.integer)):
        return str(int(valor))
    if isinstance(valor, (float, np.floating)):
        if math.isnan(valor):
            return ""
        if float(valor).is_integer():
            return str(int(valor))
        return re.sub(r"\D", "", str(valor))

    texto = str(valor).strip()
    texto = re.sub(r"\.0$", "", texto)
    # Caso venha como 7.89123E+12, tenta converter.
    if re.search(r"e\+?\d+$", texto, flags=re.IGNORECASE):
        try:
            return str(int(float(texto)))
        except Exception:
            pass
    return re.sub(r"\D", "", texto)


def numero_brasil(valor) -> float:
    """
    Converte números sem dividir por mil.

    Exemplos:
        '173.400'   -> 173400
        '173,400'   -> 173400
        '173,40'    -> 173.40
        '1.234,56'  -> 1234.56
        '1,234.56'  -> 1234.56
    """
    if pd.isna(valor) or valor == "":
        return 0.0
    if isinstance(valor, (int, float, np.integer, np.floating)):
        if pd.isna(valor):
            return 0.0
        return float(valor)

    texto = str(valor).strip()
    if texto == "":
        return 0.0

    texto = texto.replace("R$", "").replace("%", "").replace(" ", "")
    negativo = texto.startswith("(") and texto.endswith(")")
    texto = texto.strip("()")

    tem_ponto = "." in texto
    tem_virgula = "," in texto

    if tem_ponto and tem_virgula:
        # Último separador define decimal; o outro é milhar.
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif tem_ponto:
        partes = texto.split(".")
        # 173.400, 1.234.567 etc. -> milhar
        if all(p.isdigit() for p in partes) and all(len(p) == 3 for p in partes[1:]):
            texto = "".join(partes)
    elif tem_virgula:
        partes = texto.split(",")
        # 173,400 -> milhar
        if all(p.isdigit() for p in partes) and all(len(p) == 3 for p in partes[1:]):
            texto = "".join(partes)
        else:
            texto = texto.replace(",", ".")

    try:
        numero = float(texto)
        return -numero if negativo else numero
    except ValueError:
        return 0.0


def limpar_coluna_numerica_vetorizada(serie: pd.Series) -> pd.Series:
    """
    Converte coluna numérica respeitando formatos brasileiros e internacionais.

    Correções importantes:
    - '340.123' vira 340123, não 340.123.
    - '320.150.850.321' vira 320150850321.
    - '320.150,850321' vira 320150.850321.
    - '1.234,56' e '1,234.56' viram 1234.56.

    Isso evita erro no ajuste de escala entre Sell-in e Sell-out quando uma base
    vem em unidade, milhar, milhão, grama, kg ou tonelada.
    """
    if serie is None:
        return pd.Series(dtype="float64")

    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0.0).astype(float)

    s = serie.astype("string").fillna("").str.strip()
    s = s.str.replace("\u00a0", "", regex=False)
    s = s.str.replace(r"[R\$%\s]", "", regex=True)

    mask_parenteses = s.str.startswith("(") & s.str.endswith(")")
    s = s.str.strip("()")

    mask_sinal_negativo = s.str.startswith("-")
    mask_negativo = mask_parenteses | mask_sinal_negativo
    s = s.str.replace(r"^[+-]", "", regex=True)

    # Mantém somente dígitos e separadores decimais/milhares.
    s = s.str.replace(r"[^0-9,\.]", "", regex=True)

    tem_ponto = s.str.contains(r"\.", regex=True, na=False)
    tem_virgula = s.str.contains(",", regex=False, na=False)

    tratado = s.copy()

    # Quando há ponto e vírgula, o último separador encontrado é tratado como decimal.
    mask_ambos = tem_ponto & tem_virgula
    if mask_ambos.any():
        last_dot = s.str.rfind(".")
        last_comma = s.str.rfind(",")
        mask_decimal_virgula = mask_ambos & (last_comma > last_dot)
        mask_decimal_ponto = mask_ambos & ~mask_decimal_virgula

        tratado.loc[mask_decimal_virgula] = (
            s.loc[mask_decimal_virgula]
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        tratado.loc[mask_decimal_ponto] = s.loc[mask_decimal_ponto].str.replace(",", "", regex=False)

    # Somente ponto: se estiver em blocos de 3 dígitos, é separador de milhar.
    mask_so_ponto = tem_ponto & ~tem_virgula
    if mask_so_ponto.any():
        mask_ponto_milhar = mask_so_ponto & s.str.match(r"^\d{1,3}(?:\.\d{3})+$", na=False)
        tratado.loc[mask_ponto_milhar] = s.loc[mask_ponto_milhar].str.replace(".", "", regex=False)
        # Os demais casos ficam como decimal com ponto.

    # Somente vírgula: se estiver em blocos de 3 dígitos, é separador de milhar;
    # caso contrário, é decimal brasileiro.
    mask_so_virgula = tem_virgula & ~tem_ponto
    if mask_so_virgula.any():
        mask_virgula_milhar = mask_so_virgula & s.str.match(r"^\d{1,3}(?:,\d{3})+$", na=False)
        mask_virgula_decimal = mask_so_virgula & ~mask_virgula_milhar
        tratado.loc[mask_virgula_milhar] = s.loc[mask_virgula_milhar].str.replace(",", "", regex=False)
        tratado.loc[mask_virgula_decimal] = s.loc[mask_virgula_decimal].str.replace(",", ".", regex=False)

    numeros = pd.to_numeric(tratado, errors="coerce").fillna(0.0).astype(float)
    numeros = numeros.mask(mask_negativo, -numeros)
    return numeros


def otimizar_colunas_categoricas_sellout(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reduz uso de RAM convertendo colunas repetitivas de texto do Sell-out para category.

    Mantém a otimização restrita às colunas originais mais repetitivas do Sell-out,
    para não interferir nos groupbys posteriores das bases já tratadas.
    """
    if df is None or df.empty:
        return df

    colunas_categoricas = ["NIVEL1", "NIVEL2", "Fabricante", "Marca", "Categoria", "Canal", "PDV_CANAL"]
    for col in colunas_categoricas:
        if col in df.columns:
            try:
                df[col] = df[col].astype("category")
            except Exception:
                pass
    return df


def converter_mes(valor) -> pd.Timestamp:
    """
    Converte datas/períodos para o primeiro dia do mês.

    Formatos aceitos, entre outros:
    - 01/12/2024
    - dez. 24, dez/24, dez-24, dez24
    - dezembro 2024
    - 202412, 2024-12
    - 12/2024, 12/24
    - serial de data do Excel
    """
    if pd.isna(valor) or valor == "":
        return pd.NaT

    if isinstance(valor, pd.Timestamp):
        return valor.to_period("M").to_timestamp()
    if isinstance(valor, datetime):
        return pd.Timestamp(valor).to_period("M").to_timestamp()

    def ano_4_digitos(ano: int) -> int:
        # Para o contexto do estudo, 24/25 significa 2024/2025.
        if 0 <= ano < 100:
            return 2000 + ano
        return ano

    def timestamp_ano_mes(ano: int, mes: int) -> pd.Timestamp:
        ano = ano_4_digitos(int(ano))
        mes = int(mes)
        if 1900 <= ano <= 2999 and 1 <= mes <= 12:
            return pd.Timestamp(year=ano, month=mes, day=1)
        return pd.NaT

    if isinstance(valor, (int, float, np.integer, np.floating)):
        if pd.isna(valor):
            return pd.NaT
        num = int(valor)

        # Serial de data do Excel.
        if 25000 <= num <= 80000:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(num, unit="D")).to_period("M").to_timestamp()

        # Período YYYYMM, exemplo: 202412.
        if 190001 <= num <= 299912:
            return timestamp_ano_mes(num // 100, num % 100)

        return pd.NaT

    texto = str(valor).strip()
    texto = texto.replace("\ufeff", "").replace("\ufffe", "").replace("\x00", "").strip()
    texto = texto.strip("'\"")
    texto = re.sub(r"\.0$", "", texto)
    texto = re.sub(r"\s+", " ", texto)

    if not texto:
        return pd.NaT

    # 202412
    if re.fullmatch(r"\d{6}", texto):
        return timestamp_ano_mes(int(texto[:4]), int(texto[4:]))

    # 2024-12 ou 2024/12
    m = re.fullmatch(r"((?:19|20)\d{2})\s*[-/]\s*(\d{1,2})", texto)
    if m:
        return timestamp_ano_mes(int(m.group(1)), int(m.group(2)))

    # 12/2024, 12/24, 12-2024 ou 12-24
    m = re.fullmatch(r"(\d{1,2})\s*[-/]\s*(\d{2}|\d{4})", texto)
    if m:
        return timestamp_ano_mes(int(m.group(2)), int(m.group(1)))

    # Datas ISO completas, como 2024-12-01 ou 2024/12/01.
    # Não usar dayfirst=True aqui para evitar inverter ano-mês-dia.
    if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", texto):
        data = pd.to_datetime(texto, errors="coerce", dayfirst=False)
        if pd.isna(data):
            return pd.NaT
        return data.to_period("M").to_timestamp()

    # Meses em português vindos do Dash:
    # dez. 24, dez/24, dez-24, dez24, dezembro 2024, 2024 dez etc.
    texto_norm_mes = normalizar_texto(texto)
    texto_norm_mes = re.sub(r"^([a-z]+)(\d{2}|\d{4})$", r"\1 \2", texto_norm_mes)
    texto_norm_mes = re.sub(r"^(\d{2}|\d{4})([a-z]+)$", r"\1 \2", texto_norm_mes)
    partes_mes = texto_norm_mes.split()

    if len(partes_mes) >= 2:
        # mês ano: dez 24 / dezembro 2024
        mes_txt = partes_mes[0]
        ano_txt = partes_mes[-1]
        if mes_txt in MESES_PT and re.fullmatch(r"\d{2}|\d{4}", ano_txt):
            return timestamp_ano_mes(int(ano_txt), MESES_PT[mes_txt])

        # ano mês: 2024 dez / 24 dez
        ano_txt = partes_mes[0]
        mes_txt = partes_mes[-1]
        if mes_txt in MESES_PT and re.fullmatch(r"\d{2}|\d{4}", ano_txt):
            return timestamp_ano_mes(int(ano_txt), MESES_PT[mes_txt])

    # Fallback para datas comuns, como 01/12/2024.
    data = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if pd.isna(data):
        return pd.NaT
    return data.to_period("M").to_timestamp()


def converter_ano(valor):
    """Converte uma coluna anual para int, sem quebrar se receber data/Timestamp."""
    if pd.isna(valor) or valor == "":
        return np.nan
    if isinstance(valor, (pd.Timestamp, datetime)):
        return pd.Timestamp(valor).year
    if isinstance(valor, (int, float, np.integer, np.floating)):
        if pd.isna(valor):
            return np.nan
        num = int(valor)
        if 1900 <= num <= 2999:
            return num
        if 190001 <= num <= 299912:
            return num // 100
        if 25000 <= num <= 80000:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(num, unit="D")).year
        return np.nan
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return np.nan
    m = re.search(r"(19|20)\d{2}", texto)
    if m:
        return int(m.group(0))
    return np.nan

def meses_entre(inicio: pd.Timestamp, fim: pd.Timestamp) -> List[pd.Timestamp]:
    if pd.isna(inicio) or pd.isna(fim) or inicio > fim:
        return []
    return list(pd.date_range(inicio, fim, freq="MS"))


def divisao_segura(numerador, denominador):
    denominador = float(denominador or 0)
    if denominador == 0:
        return np.nan
    return float(numerador or 0) / denominador


def variacao(atual, anterior):
    if anterior is None or pd.isna(anterior) or float(anterior) == 0:
        return np.nan
    return float(atual or 0) / float(anterior) - 1


def _mes_normalizado(valor):
    if pd.isna(valor):
        return pd.NaT
    return pd.Timestamp(valor).to_period("M").to_timestamp()


def meses_unicos_normalizados(valores) -> List[pd.Timestamp]:
    meses = []
    for m in list(valores or []):
        try:
            ts = _mes_normalizado(m)
            if pd.notna(ts):
                meses.append(ts)
        except Exception:
            continue
    return sorted(set(meses))


def tem_janela_movel_completa(meses, max_mes, qtd_meses: int) -> bool:
    """Confirma se existem todos os meses da janela móvel até max_mes."""
    if pd.isna(max_mes) or qtd_meses <= 0:
        return False
    max_mes = _mes_normalizado(max_mes)
    meses_set = set(meses_unicos_normalizados(meses))
    inicio = (max_mes - pd.DateOffset(months=qtd_meses - 1)).to_period("M").to_timestamp()
    esperados = set(pd.date_range(inicio, max_mes, freq="MS"))
    return esperados.issubset(meses_set)


def filtrar_bases_sku_12m_por_categoria(sellin: pd.DataFrame, sellout: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Limita bases de SKU à janela de 12 meses por categoria/PROD.
    Se não houver mês em algum lado, mantém a base original para não perder dados anuais/totalizados.
    """
    si = sellin.copy() if sellin is not None else pd.DataFrame()
    so = sellout.copy() if sellout is not None else pd.DataFrame()
    if si.empty or so.empty or "categoria_key" not in si.columns or "categoria_key" not in so.columns:
        return si, so
    if "mes" not in si.columns or "mes" not in so.columns:
        return si, so
    if not si["mes"].notna().any() or not so["mes"].notna().any():
        return si, so

    si["mes"] = pd.to_datetime(si["mes"], errors="coerce").dt.to_period("M").dt.to_timestamp()
    so["mes"] = pd.to_datetime(so["mes"], errors="coerce").dt.to_period("M").dt.to_timestamp()
    cats = sorted(set(si["categoria_key"].dropna().astype(str)) | set(so["categoria_key"].dropna().astype(str)))
    si_partes = []
    so_partes = []
    for cat in cats:
        si_cat = si[si["categoria_key"].astype(str) == cat]
        so_cat = so[so["categoria_key"].astype(str) == cat]
        if si_cat.empty and so_cat.empty:
            continue
        max_si = si_cat["mes"].dropna().max() if not si_cat.empty else pd.NaT
        max_so = so_cat["mes"].dropna().max() if not so_cat.empty else pd.NaT
        if pd.notna(max_si) and pd.notna(max_so):
            fim = min(max_si, max_so).to_period("M").to_timestamp()
        else:
            fim = max([m for m in [max_si, max_so] if pd.notna(m)], default=pd.NaT)
        if pd.isna(fim):
            si_partes.append(si_cat)
            so_partes.append(so_cat)
            continue
        inicio = (fim - pd.DateOffset(months=MESES_MAT - 1)).to_period("M").to_timestamp()
        si_partes.append(si_cat[(si_cat["mes"] >= inicio) & (si_cat["mes"] <= fim)])
        so_partes.append(so_cat[(so_cat["mes"] >= inicio) & (so_cat["mes"] <= fim)])
    si_out = pd.concat(si_partes, ignore_index=True) if si_partes else si.iloc[0:0].copy()
    so_out = pd.concat(so_partes, ignore_index=True) if so_partes else so.iloc[0:0].copy()
    return si_out, so_out


def montar_skus_excluidos_em_comum(si: pd.DataFrame, so: pd.DataFrame) -> pd.DataFrame:
    """
    Lista os SKUs que ficaram fora do cálculo de SKU em comum.

    Regras:
    - SKU em comum = existe no Sell-in e no Sell-out.
    - SKU excluído = existe somente no Sell-in ou somente no Sell-out.
    - A tabela é montada por categoria/aba, usando as bases já filtradas da categoria.
    """
    colunas = [
        "SKU", "Status", "Sell-in", "Sell-out",
        "Nome SKU", "Marca", "Fabricante",
    ]

    if si is None:
        si = pd.DataFrame()
    if so is None:
        so = pd.DataFrame()

    if "ean" not in si.columns and "ean" not in so.columns:
        return pd.DataFrame(columns=colunas)

    si, so = filtrar_bases_sku_12m_por_categoria(si, so)
    si_base = si.copy() if not si.empty else pd.DataFrame(columns=["ean", "valor_sellin"])
    so_base = so.copy() if not so.empty else pd.DataFrame(columns=["ean", "valor_sellout"])

    if "ean" not in si_base.columns:
        si_base["ean"] = ""
    if "ean" not in so_base.columns:
        so_base["ean"] = ""

    si_base["SKU"] = si_base["ean"].map(ean_texto)
    so_base["SKU"] = so_base["ean"].map(ean_texto)
    si_base = si_base[si_base["SKU"] != ""].copy()
    so_base = so_base[so_base["SKU"] != ""].copy()

    si_sku = (
        si_base.groupby("SKU", as_index=False)["valor_sellin"].sum()
        .rename(columns={"valor_sellin": "Sell-in"})
        if not si_base.empty and "valor_sellin" in si_base.columns
        else pd.DataFrame(columns=["SKU", "Sell-in"])
    )

    agg_so = {"valor_sellout": "sum"}
    if "nome_sku" in so_base.columns:
        agg_so["nome_sku"] = "first"
    if "marca" in so_base.columns:
        agg_so["marca"] = "first"
    if "fabricante" in so_base.columns:
        agg_so["fabricante"] = "first"

    so_sku = (
        so_base.groupby("SKU", as_index=False).agg(agg_so)
        if not so_base.empty and "valor_sellout" in so_base.columns
        else pd.DataFrame(columns=["SKU", "valor_sellout"])
    )
    so_sku = so_sku.rename(columns={
        "valor_sellout": "Sell-out",
        "nome_sku": "Nome SKU",
        "marca": "Marca",
        "fabricante": "Fabricante",
    })

    base = si_sku.merge(so_sku, on="SKU", how="outer")
    if base.empty:
        return pd.DataFrame(columns=colunas)

    base["Sell-in"] = pd.to_numeric(base.get("Sell-in", 0), errors="coerce").fillna(0)
    base["Sell-out"] = pd.to_numeric(base.get("Sell-out", 0), errors="coerce").fillna(0)

    base["Status"] = np.select(
        [
            (base["Sell-in"] > 0) & (base["Sell-out"] == 0),
            (base["Sell-in"] == 0) & (base["Sell-out"] > 0),
        ],
        ["Somente no Sell-in", "Somente no Sell-out"],
        default="Em comum",
    )
    base = base[base["Status"].isin(["Somente no Sell-in", "Somente no Sell-out"])].copy()

    for c in ["Nome SKU", "Marca", "Fabricante"]:
        if c not in base.columns:
            base[c] = ""
        base[c] = base[c].fillna("")

    base = base[colunas].sort_values(["Status", "Sell-in", "Sell-out", "SKU"], ascending=[True, False, False, True])
    return base.reset_index(drop=True)




def resumo_base_para_diagnostico(nome: str, df: pd.DataFrame, valor_col: str, categoria_col: str = "categoria_key") -> Dict[str, object]:
    """Resume uma base para mensagens de diagnóstico sem expor linhas pesadas."""
    info: Dict[str, object] = {"base": nome}
    try:
        info["linhas"] = int(len(df)) if df is not None else 0
        info["colunas"] = list(df.columns)[:40] if df is not None else []
        if df is None or df.empty:
            info["soma_valor"] = 0.0
            info["categorias"] = 0
            info["skus"] = 0
            info["meses"] = 0
            info["ufs"] = 0
            return info
        if valor_col in df.columns:
            serie_val = pd.to_numeric(df[valor_col], errors="coerce").fillna(0)
            info["linhas_valor_nao_zero"] = int((serie_val != 0).sum())
            info["soma_valor"] = float(serie_val.sum())
        else:
            info["linhas_valor_nao_zero"] = 0
            info["soma_valor"] = 0.0
        if categoria_col in df.columns:
            cats = df[categoria_col].fillna("").astype(str).map(normalizar_categoria)
            info["categorias"] = int(cats[cats != ""].nunique())
            info["exemplos_categorias"] = sorted(df.loc[cats != "", "categoria"].fillna("").astype(str).drop_duplicates().head(8).tolist()) if "categoria" in df.columns else []
        else:
            info["categorias"] = 0
            info["exemplos_categorias"] = []
        if "ean" in df.columns:
            skus = df["ean"].fillna("").astype(str).map(ean_texto)
            info["skus"] = int(skus[skus != ""].nunique())
        else:
            info["skus"] = 0
        if "mes" in df.columns:
            info["meses"] = int(df["mes"].dropna().nunique())
        else:
            info["meses"] = 0
        if "ano" in df.columns:
            info["anos"] = sorted([int(x) for x in pd.to_numeric(df["ano"], errors="coerce").dropna().drop_duplicates().tolist()])[:20]
        else:
            info["anos"] = []
        if "uf" in df.columns:
            info["ufs"] = int(df["uf"].fillna("").astype(str).str.strip().replace("", np.nan).dropna().nunique())
        else:
            info["ufs"] = 0
    except Exception as exc:
        info["erro_diagnostico"] = str(exc)
    return info


def montar_mensagem_diagnostico(
    sellin: pd.DataFrame,
    sellout: pd.DataFrame,
    avisos: List[str],
    metrica: str,
    nivel: str,
    fabricante_filtro: str = "",
) -> str:
    """Monta erro claro quando não há base suficiente para calcular."""
    si = resumo_base_para_diagnostico("Sell-in", sellin, "valor_sellin")
    so = resumo_base_para_diagnostico("Sell-out", sellout, "valor_sellout")

    cats_si = set(sellin.get("categoria_key", pd.Series(dtype=str)).dropna().astype(str).unique()) if sellin is not None and not sellin.empty else set()
    cats_so = set(sellout.get("categoria_key", pd.Series(dtype=str)).dropna().astype(str).unique()) if sellout is not None and not sellout.empty else set()
    comuns = cats_si & cats_so

    partes = [
        "Não encontrei dados suficientes para calcular cobertura.",
        "",
        "Diagnóstico do que o código conseguiu ler:",
        f"- Métrica escolhida: {metrica}",
        f"- Regra escolhida: {nivel}",
        f"- Fabricante digitado: {fabricante_filtro or 'não informado'}",
        "",
        f"Sell-in: {si.get('linhas', 0)} linhas após limpeza; "
        f"{si.get('linhas_valor_nao_zero', 0)} linhas com valor; "
        f"soma={si.get('soma_valor', 0):,.2f}; "
        f"categorias={si.get('categorias', 0)}; SKUs={si.get('skus', 0)}; meses={si.get('meses', 0)}; UFs={si.get('ufs', 0)}.",
        f"Sell-out: {so.get('linhas', 0)} linhas após limpeza/filtros; "
        f"{so.get('linhas_valor_nao_zero', 0)} linhas com valor; "
        f"soma={so.get('soma_valor', 0):,.2f}; "
        f"categorias/PRODs={so.get('categorias', 0)}; SKUs={so.get('skus', 0)}; meses={so.get('meses', 0)}; UFs={so.get('ufs', 0)}.",
        f"Categorias/PRODs em comum: {len(comuns)}.",
    ]

    if si.get("linhas", 0) == 0:
        partes.append("\nProvável causa: o Sell-in ficou vazio depois de converter a coluna de valor. Confira se a métrica escolhida bate com a coluna do template: SELL IN (Kg / L) para volume ou SELL IN (QUANTIDADE) para quantia.")
    if so.get("linhas", 0) == 0:
        partes.append("\nProvável causa: o Sell-out ficou vazio depois de ler a métrica, categoria/PROD ou fabricante. Confira se a coluna usada existe: Vendas_em_volume para volume ou Qtd_de_Vendas para quantia.")
    if fabricante_filtro and so.get("linhas", 0) == 0:
        partes.append("Também confira se o fabricante foi digitado exatamente como aparece no Sell-out. Nesta versão, quando o filtro exato zera a base, o código também tenta buscar por 'contém', mas pode não encontrar se houver acento, abreviação ou CNPJ no nome.")
    if si.get("linhas", 0) > 0 and so.get("linhas", 0) > 0 and not comuns:
        partes.append("\nAs duas bases possuem valores, mas não ficaram com categorias/PRODs em comum. Neste caso, o problema costuma ser o mapeamento por SKU/EAN ou a escolha CATEGORIA/NIVEL1/NIVEL2.")

    if avisos:
        partes.append("\nAvisos gerados até o erro:")
        partes.extend([f"- {a}" for a in avisos[-50:]])

    # Formato BR para milhares no texto.
    return "\n".join(partes).replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_periodo(mes: pd.Timestamp) -> str:
    if pd.isna(mes):
        return ""
    return pd.Timestamp(mes).strftime("%m/%Y")


def nome_aba_seguro(nome: str, usados: set) -> str:
    nome = re.sub(r"[\[\]\:\*\?\/\\]", "-", str(nome))
    nome = re.sub(r"\s+", "_", nome).strip("_")
    if not nome:
        nome = "Categoria"
    nome = nome[:31]
    base = nome
    i = 2
    while nome in usados:
        sufixo = f"_{i}"
        nome = base[:31 - len(sufixo)] + sufixo
        i += 1
    usados.add(nome)
    return nome


def nomes_unicos(headers: List) -> List[str]:
    saida = []
    contagem = {}
    for i, h in enumerate(headers, start=1):
        nome = str(h).strip() if h is not None and str(h).strip() != "" else f"COL_{i}"
        if nome in contagem:
            contagem[nome] += 1
            nome = f"{nome}_{contagem[nome]}"
        else:
            contagem[nome] = 1
        saida.append(nome)
    return saida


def coluna_excel(numero_coluna_1based: int) -> str:
    """Converte 1 -> A, 2 -> B, 27 -> AA."""
    resultado = ""
    n = numero_coluna_1based
    while n:
        n, resto = divmod(n - 1, 26)
        resultado = chr(65 + resto) + resultado
    return resultado


# ============================================================
# Leitura de arquivos e cabeçalhos
# ============================================================



CSV_EXTENSOES = {".csv", ".txt", ".tsv"}
CSV_ABA_FAKE = "__CSV__"


def eh_csv(arquivo: Path) -> bool:
    """Identifica arquivos tabulares em texto que devem ser lidos como CSV."""
    return Path(arquivo).suffix.lower() in CSV_EXTENSOES


def ler_csv_flexivel(arquivo: Path) -> pd.DataFrame:
    """
    Lê CSV/TXT/TSV tentando detectar encoding e separador.

    Suporta, entre outros:
    - separador ; , | tab
    - UTF-8, UTF-8 com BOM, Latin-1, CP1252
    - cabeçalho na primeira linha ou em linhas abaixo
    - campos entre aspas

    Para não ficar pesado em arquivos grandes, primeiro lê uma amostra
    e depois carrega o arquivo completo apenas com a melhor combinação.
    """
    arquivo = Path(arquivo)
    encodings = ["utf-16", "utf-16-le", "utf-16-be", "utf-8-sig", "utf-8", "cp1252", "latin1", "iso-8859-1"]
    if arquivo.suffix.lower() == ".tsv":
        separadores = ["\t", None, ";", ",", "|"]
    else:
        separadores = [None, ";", ",", "\t", "|"]

    melhor = None  # (score, encoding, sep)
    erros = []

    def limpar_bom(df: pd.DataFrame) -> pd.DataFrame:
        return limpar_bom_dataframe(df)

    # 1) Amostra pequena para escolher separador/encoding.
    for enc in encodings:
        for sep in separadores:
            try:
                sample = pd.read_csv(
                    arquivo,
                    sep=sep,
                    engine="python",
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    skip_blank_lines=False,
                    on_bad_lines="skip",
                    encoding=enc,
                    nrows=100,
                )
                if sample is None or sample.empty:
                    continue
                sample = limpar_bom(sample)
                qtd_cols = int(sample.shape[1])
                qtd_linhas = int(sample.shape[0])
                sample_texto = sample.astype(str)
                conteudo = int((sample_texto.apply(lambda col: col.str.strip() != "")).sum().sum())
                texto_amostra = " ".join(sample_texto.head(5).fillna("").astype(str).values.ravel())
                penalidade_garbage = texto_amostra.count("ÿ") + texto_amostra.count("þ") + texto_amostra.count(" ") + texto_amostra.count("\x00")
                score = (qtd_cols, conteudo, qtd_linhas, -penalidade_garbage)
                if melhor is None or score > melhor[0]:
                    melhor = (score, enc, sep)
            except Exception as exc:
                erros.append(f"encoding={enc}, sep={repr(sep)}: {exc}")

    if melhor is None:
        detalhe = "\n".join(erros[-5:]) if erros else "sem detalhes"
        raise ValueError(f"Não foi possível ler o CSV '{arquivo.name}'. Últimos erros:\n{detalhe}")

    _, enc_final, sep_final = melhor

    # 2) Leitura completa usando a melhor configuração encontrada.
    try:
        df = pd.read_csv(
            arquivo,
            sep=sep_final,
            engine="python",
            header=None,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=False,
            on_bad_lines="skip",
            encoding=enc_final,
        )
    except Exception as exc:
        raise ValueError(
            f"Não consegui ler o CSV completo '{arquivo.name}' com encoding={enc_final} e sep={repr(sep_final)}.\n{exc}"
        )

    df = limpar_bom(df)

    # Remove colunas totalmente vazias no fim, mas preserva colunas vazias no começo
    # porque alguns templates exportados para CSV mantêm cabeçalho a partir da coluna B.
    while df.shape[1] > 0:
        ultima = df.iloc[:, -1].astype(str).str.strip()
        if ultima.eq("").all():
            df = df.iloc[:, :-1]
        else:
            break
    return df




def limpar_bom_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove BOM e caracteres nulos de cabeçalhos/células.

    Também limpa o caso em que um CSV UTF-16 foi lido com encoding errado e
    aparece como ``ÿþ`` ou ``þÿ`` no primeiro cabeçalho.
    """
    if df is None:
        return df

    def limpar_texto_csv(x):
        if not isinstance(x, str):
            return x
        return (
            x.replace("\ufeff", "")
             .replace("\ufffe", "")
             .replace("\x00", "")
             .replace("ÿþ", "")
             .replace("þÿ", "")
             .replace("ï»¿", "")
             .strip()
        )

    out = df.copy()
    out.columns = [limpar_texto_csv(str(c)) for c in out.columns]
    return out.apply(lambda col: col.map(limpar_texto_csv))


def remover_linhas_colunas_vazias(df: pd.DataFrame) -> pd.DataFrame:
    """Remove linhas e colunas totalmente vazias, preservando dados de templates/CSVs."""
    if df is None or df.empty:
        return df
    out = df.copy()
    out = out.dropna(how="all")
    if out.empty:
        return out
    mask_linhas = out.astype(str).apply(lambda r: any(str(v).strip() not in {"", "nan", "None"} for v in r), axis=1)
    out = out[mask_linhas]
    if out.empty:
        return out
    mask_cols = out.astype(str).apply(lambda c: any(str(v).strip() not in {"", "nan", "None"} for v in c), axis=0)
    out = out.loc[:, mask_cols]
    return out.reset_index(drop=True)


def detectar_csv_formato(arquivo: Path, nrows: int = 150) -> Tuple[str, Optional[str], pd.DataFrame]:
    """
    Detecta encoding e separador sem carregar o CSV inteiro.

    Retorna: encoding, separador, amostra_raw.
    Prefere separador explícito quando empata com sep=None, porque é mais rápido para leitura em blocos.
    """
    arquivo = Path(arquivo)
    encodings = ["utf-16", "utf-16-le", "utf-16-be", "utf-8-sig", "utf-8", "cp1252", "latin1", "iso-8859-1"]
    if arquivo.suffix.lower() == ".tsv":
        separadores = ["\t", ";", ",", "|", None]
    else:
        separadores = [";", ",", "\t", "|", None]

    melhor = None  # (score, enc, sep, sample)
    erros = []

    for enc in encodings:
        for sep in separadores:
            try:
                sample = pd.read_csv(
                    arquivo,
                    sep=sep,
                    engine="python",
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    skip_blank_lines=False,
                    on_bad_lines="skip",
                    encoding=enc,
                    nrows=nrows,
                )
                if sample is None or sample.empty:
                    continue
                sample = limpar_bom_dataframe(sample)
                qtd_cols = int(sample.shape[1])
                qtd_linhas = int(sample.shape[0])
                sample_texto = sample.astype(str)
                conteudo = int((sample_texto.apply(lambda col: col.str.strip() != "")).sum().sum())
                texto_amostra = " ".join(sample_texto.head(5).fillna("").astype(str).values.ravel())
                penalidade_garbage = texto_amostra.count("ÿ") + texto_amostra.count("þ") + texto_amostra.count(" ") + texto_amostra.count("\x00")
                # Prioriza mais colunas, mais conteúdo, separador explícito e penaliza leitura com encoding errado.
                score = (qtd_cols, conteudo, qtd_linhas, 1 if sep is not None else 0, -penalidade_garbage)
                if melhor is None or score > melhor[0]:
                    melhor = (score, enc, sep, sample)
            except Exception as exc:
                erros.append(f"encoding={enc}, sep={repr(sep)}: {exc}")

    if melhor is None:
        detalhe = "\n".join(erros[-5:]) if erros else "sem detalhes"
        raise ValueError(f"Não foi possível ler a amostra do CSV '{arquivo.name}'. Últimos erros:\n{detalhe}")

    _, enc_final, sep_final, sample_final = melhor
    return enc_final, sep_final, sample_final


def localizar_cabecalho_em_raw_csv(
    raw_csv: pd.DataFrame,
    obrigatorias: List[str],
    max_linhas_scan: int = 150,
    linha_preferencial: Optional[int] = None,
    coluna_inicial_preferencial: Optional[int] = None,
) -> Tuple[int, List[str], int, int]:
    """
    Localiza o cabeçalho dentro de uma amostra raw de CSV.

    Retorna:
        header_row_idx  -> linha 1-based no arquivo
        headers         -> nomes únicos de cabeçalho já recortados
        header_start    -> coluna inicial zero-based no arquivo
        header_end      -> coluna final exclusiva zero-based no arquivo
    """
    header_row_idx = None
    header_values = None
    header_start = 0
    header_end = None

    def valores_linha(idx: int, start_col_1based: int = 1) -> List:
        if idx < 1 or idx > len(raw_csv):
            return []
        start = max(start_col_1based - 1, 0)
        return raw_csv.iloc[idx - 1, start:].tolist()

    def tentar_linha(idx: int, start_col_1based: int = 1) -> bool:
        nonlocal header_row_idx, header_values, header_start, header_end
        valores = valores_linha(idx, start_col_1based)
        if not valores:
            return False
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if linha_contem_cabecalho(valores_recortados, obrigatorias):
            header_row_idx = idx
            header_values = valores_recortados
            header_start = start_col_1based - 1 + inicio_rel
            header_end = start_col_1based - 1 + fim_rel
            return True
        return False

    if linha_preferencial is not None:
        tentar_linha(linha_preferencial, coluna_inicial_preferencial or 1)

    if header_row_idx is None:
        limite = min(max_linhas_scan, len(raw_csv))
        for idx in range(1, limite + 1):
            valores = valores_linha(idx, 1)
            valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
            if linha_contem_cabecalho(valores_recortados, obrigatorias):
                header_row_idx = idx
                header_values = valores_recortados
                header_start = inicio
                header_end = fim
                break

    if header_row_idx is None and linha_preferencial is not None:
        valores = valores_linha(linha_preferencial, coluna_inicial_preferencial or 1)
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if any(v is not None and str(v).strip() != "" for v in valores_recortados):
            header_row_idx = linha_preferencial
            header_values = valores_recortados
            header_start = (coluna_inicial_preferencial or 1) - 1 + inicio_rel
            header_end = (coluna_inicial_preferencial or 1) - 1 + fim_rel

    if header_row_idx is None or header_values is None:
        raise ValueError(f"Não localizei o cabeçalho no CSV. Colunas esperadas: {obrigatorias}.")

    headers = nomes_unicos(header_values)
    return header_row_idx, headers, header_start, header_end if header_end is not None else len(headers)


def preparar_leitura_csv_com_cabecalho(
    arquivo: Path,
    obrigatorias: List[str],
    max_linhas_scan: int = 150,
    linha_preferencial: Optional[int] = None,
    coluna_inicial_preferencial: Optional[int] = None,
) -> Tuple[str, Optional[str], int, List[str], int, int]:
    """Detecta formato e cabeçalho de CSV sem carregar o arquivo completo."""
    enc, sep, sample = detectar_csv_formato(arquivo, nrows=max_linhas_scan)
    try:
        header_row_idx, headers, header_start, header_end = localizar_cabecalho_em_raw_csv(
            sample,
            obrigatorias=obrigatorias,
            max_linhas_scan=max_linhas_scan,
            linha_preferencial=linha_preferencial,
            coluna_inicial_preferencial=coluna_inicial_preferencial,
        )
    except ValueError as exc:
        raise ValueError(f"Não localizei o cabeçalho no CSV '{Path(arquivo).name}'. Colunas esperadas: {obrigatorias}.\n{exc}")
    return enc, sep, header_row_idx, headers, header_start, header_end


def iterar_csv_com_cabecalho_em_blocos(
    arquivo: Path,
    obrigatorias: List[str],
    colunas_desejadas: Optional[List[str]] = None,
    chunksize: int = 200_000,
    max_linhas_scan: int = 150,
    linha_preferencial: Optional[int] = None,
    coluna_inicial_preferencial: Optional[int] = None,
):
    """
    Itera CSV em blocos, lendo somente as colunas desejadas quando possível.

    Evita o erro de memória causado por carregar milhões de linhas e dezenas de colunas de uma vez.
    """
    enc, sep, header_row_idx, headers, header_start, header_end = preparar_leitura_csv_com_cabecalho(
        arquivo=arquivo,
        obrigatorias=obrigatorias,
        max_linhas_scan=max_linhas_scan,
        linha_preferencial=linha_preferencial,
        coluna_inicial_preferencial=coluna_inicial_preferencial,
    )

    header_pos = {nome: header_start + i for i, nome in enumerate(headers)}

    if colunas_desejadas:
        cols_ok = [c for c in colunas_desejadas if c in header_pos]
        # Mantém ordem original do arquivo para evitar comportamento estranho do read_csv com usecols posicional.
        cols_ok = sorted(set(cols_ok), key=lambda c: header_pos[c])
        usecols = [header_pos[c] for c in cols_ok]
        names = cols_ok
    else:
        usecols = list(range(header_start, header_end))
        names = headers

    if not usecols:
        raise ValueError(f"Nenhuma coluna útil encontrada no CSV '{Path(arquivo).name}'.")

    engine = "python" if sep is None else "c"
    read_kwargs = dict(
        filepath_or_buffer=arquivo,
        sep=sep,
        engine=engine,
        header=None,
        skiprows=header_row_idx,
        usecols=usecols,
        names=names,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
        on_bad_lines="skip",
        encoding=enc,
        chunksize=chunksize,
    )
    if engine == "c":
        read_kwargs["low_memory"] = False
    leitor = pd.read_csv(**read_kwargs)

    for chunk in leitor:
        yield limpar_bom_dataframe(chunk), headers, header_row_idx

def ler_csv_com_cabecalho(
    arquivo: Path,
    obrigatorias: List[str],
    max_linhas_scan: int = 150,
    linha_preferencial: Optional[int] = None,
    coluna_inicial_preferencial: Optional[int] = None,
) -> Tuple[pd.DataFrame, int]:
    """Lê CSV procurando a linha de cabeçalho, usando a mesma lógica do Excel."""
    raw_csv = ler_csv_flexivel(arquivo)
    header_row_idx = None
    header_values = None
    header_start = 0
    header_end = None

    def valores_linha(idx: int, start_col_1based: int = 1) -> List:
        if idx < 1 or idx > len(raw_csv):
            return []
        start = max(start_col_1based - 1, 0)
        return raw_csv.iloc[idx - 1, start:].tolist()

    def tentar_linha(idx: int, start_col_1based: int = 1) -> bool:
        nonlocal header_row_idx, header_values, header_start, header_end
        valores = valores_linha(idx, start_col_1based)
        if not valores:
            return False
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if linha_contem_cabecalho(valores_recortados, obrigatorias):
            header_row_idx = idx
            header_values = valores_recortados
            header_start = start_col_1based - 1 + inicio_rel
            header_end = start_col_1based - 1 + fim_rel
            return True
        return False

    # 1) Tentativa preferencial, principalmente para Sell-in exportado do template.
    if linha_preferencial is not None:
        tentar_linha(linha_preferencial, coluna_inicial_preferencial or 1)

    # 2) Scan automático.
    if header_row_idx is None:
        limite = min(max_linhas_scan, len(raw_csv))
        for idx in range(1, limite + 1):
            valores = valores_linha(idx, 1)
            valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
            if linha_contem_cabecalho(valores_recortados, obrigatorias):
                header_row_idx = idx
                header_values = valores_recortados
                header_start = inicio
                header_end = fim
                break

    # 3) Fallback explícito para linha preferencial, se tiver algum conteúdo.
    if header_row_idx is None and linha_preferencial is not None:
        valores = valores_linha(linha_preferencial, coluna_inicial_preferencial or 1)
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if any(v is not None and str(v).strip() != "" for v in valores_recortados):
            header_row_idx = linha_preferencial
            header_values = valores_recortados
            header_start = (coluna_inicial_preferencial or 1) - 1 + inicio_rel
            header_end = (coluna_inicial_preferencial or 1) - 1 + fim_rel

    if header_row_idx is None or header_values is None:
        raise ValueError(
            f"Não localizei o cabeçalho no CSV '{Path(arquivo).name}'. "
            f"Colunas esperadas: {obrigatorias}."
        )

    headers = nomes_unicos(header_values)
    dados = []
    max_col_exclusive = header_end if header_end is not None else raw_csv.shape[1]
    for i in range(header_row_idx, len(raw_csv)):
        linha = raw_csv.iloc[i, header_start:max_col_exclusive].tolist()
        if len(linha) < len(headers):
            linha += [None] * (len(headers) - len(linha))
        linha = linha[:len(headers)]
        if any(v is not None and str(v).strip() != "" for v in linha):
            dados.append(linha)

    return pd.DataFrame(dados, columns=headers), header_row_idx


def escolher_aba(arquivo: Path, candidatas: Iterable[str], usar_primeira_como_fallback: bool = True) -> str:
    """
    Escolhe uma aba pelo nome.

    Para CSV/TXT/TSV, retorna uma aba fake, pois o arquivo não possui abas.
    Por padrão ainda permite fallback para a primeira aba, útil para arquivos de Sell-out
    que costumam ter apenas uma aba. Para o Sell-in, use fallback=False para evitar
    ler abas como "Instrução" por engano.
    """
    if eh_csv(Path(arquivo)):
        return CSV_ABA_FAKE

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    nomes = wb.sheetnames
    wb.close()

    mapa = {normalizar_texto(n): n for n in nomes}

    # 1) Match exato normalizado.
    for c in candidatas:
        chave = normalizar_texto(c)
        if chave in mapa:
            return mapa[chave]

    # 2) Match por contém, para casos como "Base Sell-in", "SELL IN Cliente", etc.
    candidatas_norm = [normalizar_texto(c) for c in candidatas if normalizar_texto(c)]
    for nome_norm, nome_real in mapa.items():
        for cand in candidatas_norm:
            if cand in nome_norm or nome_norm in cand:
                return nome_real

    if usar_primeira_como_fallback and nomes:
        return nomes[0]

    raise ValueError(
        f"Não encontrei nenhuma aba com nome esperado {list(candidatas)} no arquivo '{arquivo.name}'. "
        f"Abas disponíveis: {nomes}"
    )


def aba_tem_cabecalho(ws, obrigatorias: List[str], max_linhas_scan: int = 150) -> bool:
    """Confere se uma aba contém uma linha de cabeçalho com todas as colunas obrigatórias."""
    obrigatorias_norm = [normalizar_texto(x) for x in obrigatorias]
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        valores_norm = [normalizar_texto(v) for v in row]
        achou = all(any(req == col or req in col for col in valores_norm) for req in obrigatorias_norm)
        if achou:
            return True
        if idx >= max_linhas_scan:
            break
    return False


def escolher_aba_sellin(arquivo: Path) -> str:
    """
    Para Sell-in, força a aba do template.

    O objetivo é evitar o erro de pegar a aba "Instrução" quando o arquivo possui
    várias abas. Se o nome não estiver exatamente como esperado, procura a aba que
    realmente contém o cabeçalho UF / ANO MÊS / EAN / Sell-in.
    """
    if eh_csv(Path(arquivo)):
        return CSV_ABA_FAKE

    candidatas = [
        "Sell-in", "Sell in", "SELL-IN", "SELL IN",
        "Sell-in Cliente", "Sell In Cliente", "Base Sell-in", "Base Sell in",
    ]

    try:
        return escolher_aba(arquivo, candidatas, usar_primeira_como_fallback=False)
    except ValueError:
        pass

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        # Evita priorizar abas claramente instrutivas.
        abas_ordenadas = sorted(
            wb.sheetnames,
            key=lambda n: 1 if normalizar_texto(n) in {"instrucao", "instrucoes", "instructions", "instruction"} else 0,
        )
        for aba in abas_ordenadas:
            # Header flexível: EAN/SKU, UF/UF SM e Categoria podem não existir.
            if aba_tem_cabecalho(wb[aba], ["Sell-in"]):
                return aba
    finally:
        wb.close()

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    nomes = wb.sheetnames
    wb.close()
    raise ValueError(
        f"Não encontrei a aba 'Sell-in' nem uma aba com o cabeçalho do template no arquivo '{arquivo.name}'. "
        f"Abas disponíveis: {nomes}"
    )


def localizar_coluna(df: pd.DataFrame, alternativas: List[str], obrigatoria: bool = True) -> Optional[str]:
    mapa = {normalizar_texto(c): c for c in df.columns}
    alternativas_norm = [normalizar_texto(a) for a in alternativas]

    # Match exato normalizado.
    for alt in alternativas_norm:
        if alt in mapa:
            return mapa[alt]

    # Fallback por contém.
    for alt in alternativas_norm:
        if not alt:
            continue
        for col_norm, col_real in mapa.items():
            if alt in col_norm:
                return col_real

    if obrigatoria:
        raise KeyError(f"Coluna não encontrada. Alternativas testadas: {alternativas}")
    return None


EAN_SKU_ALTERNATIVAS_AMPLAS = [
    "EAN", "SKU", "EAN/SKU", "SKU/EAN", "EAN SKU", "SKU EAN",
    "Código Barras", "Codigo Barras", "Código_Barras", "Codigo_Barras",
    "Código de Barras", "Codigo de Barras", "Código_de_Barras", "Codigo_de_Barras",
    "Código Barras SKU", "Codigo Barras SKU", "Código_Barras_SKU", "Codigo_Barras_SKU",
    "Código de Barras SKU", "Codigo de Barras SKU", "Código_de_Barras_SKU", "Codigo_de_Barras_SKU",
    "Código Barra", "Codigo Barra", "Código_Barra", "Codigo_Barra",
    "Código Barra SKU", "Codigo Barra SKU", "Código_Barra_SKU", "Codigo_Barra_SKU",
    "Cód Barras", "Cod Barras", "Cód. Barras", "Cod. Barras",
    "Cód Barras SKU", "Cod Barras SKU", "Cód. Barras SKU", "Cod. Barras SKU",
    "Cód. de Barras", "Cod. de Barras", "Cód. de Barras SKU", "Cod. de Barras SKU",
    "CODIGO_BARRAS", "CODIGO BARRAS", "CODIGO_BARRAS_SKU", "CODIGO BARRAS SKU",
    "COD_BARRAS", "COD BARRAS", "COD_BARRAS_SKU", "COD BARRAS SKU",
    "Codigo de barras del SKU", "Código de barras del SKU",
    "Código Barras Contenido", "Codigo Barras Contenido", "CODIGO_BARRAS_CONTENIDO",
]


def localizar_coluna_ean_sku(df: pd.DataFrame, obrigatoria: bool = False) -> Optional[str]:
    """
    Localiza coluna de EAN/SKU/Código de Barras com mais segurança.

    O localizar_coluna comum aceita fallback por "contém". Para EAN/SKU isso pode ser perigoso,
    porque pode confundir "Nome SKU" com a coluna de código. Esta função prioriza campos de código
    e aceita variações com acento, sem acento, underline, português e espanhol.
    """
    mapa = {normalizar_texto(c): c for c in df.columns}
    alternativas_norm = [normalizar_texto(a) for a in EAN_SKU_ALTERNATIVAS_AMPLAS]

    # 1) Match exato normalizado.
    for alt in alternativas_norm:
        if alt in mapa:
            return mapa[alt]

    # 2) Campos que claramente são código de barras, mesmo com sufixos/prefixos.
    for col_norm, col_real in mapa.items():
        tokens = set(col_norm.split())
        if (
            (("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens))
            or "codigo barras" in col_norm
            or "cod barras" in col_norm
            or "codigo de barras" in col_norm
        ):
            return col_real

    # 3) EAN/SKU puro ou combinado, evitando colunas descritivas como Nome SKU.
    termos_descritivos = {
        "nome", "nombre", "descricao", "descripcion", "descrição", "desc",
        "categoria", "fabricante", "marca", "preco", "precio", "price",
        "valor", "vendas", "volume", "conteudo", "contenido", "qtd", "cant",
        "medida", "analise", "análise", "index",
    }
    for col_norm, col_real in mapa.items():
        tokens = set(col_norm.split())
        tem_codigo_sku = "ean" in tokens or col_norm in {"sku", "ean sku", "sku ean"}
        if tem_codigo_sku and not (tokens & termos_descritivos):
            return col_real

    if obrigatoria:
        raise KeyError(
            "Coluna de EAN/SKU/Código de Barras não encontrada. "
            f"Alternativas testadas: {EAN_SKU_ALTERNATIVAS_AMPLAS}. "
            f"Colunas disponíveis: {resumo_colunas_disponiveis(df.columns)}"
        )
    return None


def resumo_colunas_disponiveis(colunas, limite: int = 80) -> str:
    """Resume colunas disponíveis para mensagens de erro sem ficar gigante."""
    colunas = [str(c) for c in list(colunas)]
    if not colunas:
        return "nenhuma coluna identificada"
    texto = "; ".join(colunas[:limite])
    if len(colunas) > limite:
        texto += f"; ... (+{len(colunas) - limite} colunas)"
    return texto


def localizar_coluna_obrigatoria_com_erro(
    df: pd.DataFrame,
    alternativas: List[str],
    nome_logico: str,
    base: str,
    arquivo: Path | str = "",
) -> str:
    """Localiza coluna obrigatória e informa exatamente o que faltou."""
    try:
        return localizar_coluna(df, alternativas, obrigatoria=True)
    except Exception:
        arq = f" no arquivo '{Path(arquivo).name}'" if arquivo else ""
        raise ValueError(
            f"{base}: não encontrei a coluna obrigatória de {nome_logico}{arq}.\n"
            f"Alternativas testadas: {', '.join(map(str, alternativas))}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(df.columns)}"
        )


def localizar_coluna_valor_sellin(df: pd.DataFrame, metrica: str, arquivo: Path | str = "") -> Tuple[str, List[str]]:
    """
    Localiza a coluna de valor do Sell-in com mais segurança.

    Evita pegar por engano colunas agregadas/resumidas como Total, FY, MAT,
    YTD, média etc. Quando isso acontece, o mesmo valor pode aparecer repetido
    em todos os meses da tabela mensal.
    """
    metrica_norm = "volume_variavel" if metrica_eh_volume_variavel(metrica) else str(metrica or "").lower()
    alternativas = coluna_metrica_sellin(metrica_norm)
    alternativas_norm = [normalizar_texto(a) for a in alternativas]

    termos_ruins = {
        "total", "totais", "geral", "mat", "ytd", "fy", "ano", "anual",
        "acumulado", "acum", "media", "média", "percentual", "porcentagem",
        "var", "variacao", "variação", "cobertura", "share", "ranking",
        "12m", "6m", "24m", "ult", "ultimo", "último", "ant", "anterior",
    }

    def tem_termo_ruim(col_norm: str) -> bool:
        tokens = set(col_norm.split())
        return bool(tokens & termos_ruins) or any(t in col_norm for t in ["12 meses", "6 meses", "ult 12", "ult 24", "fy ", " ytd"])

    candidatos = []
    for col in df.columns:
        col_real = str(col)
        col_norm = normalizar_texto(col_real)
        if not col_norm:
            continue

        match = False
        score = -999.0
        if col_norm in alternativas_norm:
            match = True
            score = 1000.0
        else:
            for alt in alternativas_norm:
                if alt and alt in col_norm:
                    match = True
                    score = max(score, 500.0 + len(alt))

        if not match:
            continue

        if metrica_norm == "volume":
            if any(t in col_norm for t in ["kg", "kl", "lt", "litro", "litros", "volume"]):
                score += 250
            if any(t in col_norm for t in ["quantidade", "qtd", "unidade", "unidades"]):
                score -= 150
        elif metrica_norm in {"quantia", "volume_variavel"}:
            if any(t in col_norm for t in ["quantidade", "qtd", "unidade", "unidades"]):
                score += 250
            if any(t in col_norm for t in ["kg", "kl", "lt", "litro", "litros", "volume"]):
                score -= 150

        if tem_termo_ruim(col_norm):
            score -= 600

        score -= min(len(col_norm), 120) / 1000
        candidatos.append((score, col_real, col_norm))

    if not candidatos:
        arq = f" no arquivo '{Path(arquivo).name}'" if arquivo else ""
        raise ValueError(
            f"Sell-in: não encontrei a coluna obrigatória de valor de Sell-in para a métrica {metrica}{arq}.\n"
            f"Alternativas testadas: {', '.join(map(str, alternativas))}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(df.columns)}"
        )

    candidatos.sort(key=lambda x: x[0], reverse=True)
    escolhido = candidatos[0][1]
    avisos = []

    if tem_termo_ruim(candidatos[0][2]):
        avisos.append(
            f"Atenção: a coluna de Sell-in escolhida foi '{escolhido}', mas o nome parece ser agregado/resumido. "
            "Se o Sell-in mensal aparecer repetido, confira se existe uma coluna mensal bruta de Sell-in no arquivo."
        )

    if len(candidatos) > 1:
        resumo = ", ".join([f"{c[1]}" for c in candidatos[:5]])
        avisos.append(f"Sell-in: coluna de valor escolhida: '{escolhido}'. Candidatas avaliadas: {resumo}.")
    else:
        avisos.append(f"Sell-in: coluna de valor escolhida: '{escolhido}'.")

    return escolhido, avisos




def _colunas_candidatas_por_alternativas(df: pd.DataFrame, alternativas: List[str]) -> List[str]:
    """Retorna todas as colunas que batem com as alternativas, preservando a ordem do arquivo."""
    alts = [normalizar_texto(a) for a in alternativas if normalizar_texto(a)]
    cols = []
    for col in df.columns:
        col_norm = normalizar_texto(col)
        if not col_norm:
            continue
        if any(alt == col_norm or alt in col_norm or col_norm in alt for alt in alts):
            cols.append(col)
    return cols


def _pontuar_coluna_valor_sellin(nome_coluna: str, metrica: str) -> float:
    """Pontuação leve para desempatar colunas de Sell-in pela métrica escolhida."""
    col_norm = normalizar_texto(nome_coluna)
    score = 0.0
    if metrica_eh_volume_variavel(metrica):
        metrica_norm = "volume_variavel"
    else:
        metrica_norm = normalizar_texto(metrica)

    if "sell" in col_norm and "in" in col_norm:
        score += 100
    if metrica_norm.startswith("vol"):
        if any(t in col_norm for t in ["kg", "kl", "lt", "litro", "litros", "volume"]):
            score += 80
        if any(t in col_norm for t in ["quantidade", "qtd", "unidade", "unidades"]):
            score -= 40
    else:
        if any(t in col_norm for t in ["quantidade", "qtd", "unidade", "unidades"]):
            score += 80
        if any(t in col_norm for t in ["kg", "kl", "lt", "litro", "litros", "volume"]):
            score -= 40

    termos_ruins = ["total", "geral", "mat", "ytd", "fy", "ano", "anual", "media", "média", "acumulado", "cobertura", "share", "variacao", "variação"]
    if any(t in col_norm for t in termos_ruins):
        score -= 120
    return score


def _serie_mensal_teste(raw: pd.DataFrame, col_mes: Optional[str], col_valor: Optional[str]) -> pd.Series:
    """Soma valor por mês para testar se a combinação mês/valor é plausível."""
    if not col_mes or not col_valor or col_mes not in raw.columns or col_valor not in raw.columns:
        return pd.Series(dtype=float)
    meses = raw[col_mes].map(converter_mes)
    valores = limpar_coluna_numerica_vetorizada(raw[col_valor])
    tmp = pd.DataFrame({"mes": meses, "valor": valores})
    tmp = tmp[(tmp["mes"].notna()) & (tmp["valor"].fillna(0) != 0)].copy()
    if tmp.empty:
        return pd.Series(dtype=float)
    return tmp.groupby("mes")["valor"].sum().sort_index()


def escolher_colunas_sellin_mes_valor_mais_plausiveis(
    raw: pd.DataFrame,
    c_mes_atual: Optional[str],
    c_val_atual: str,
    metrica: str,
    avisos: List[str],
) -> Tuple[Optional[str], str]:
    """
    Evita um erro comum do Dash/Excel: pegar uma coluna agregada ou um período errado
    e repetir o mesmo Sell-in em todos os meses.

    Só troca a coluna quando a combinação atual gera totais mensais constantes ou sem meses
    e outra combinação gera uma série mensal claramente mais plausível.
    """
    alternativas_mes = COLUNAS_MES_ALTERNATIVAS
    candidatas_mes = _colunas_candidatas_por_alternativas(raw, alternativas_mes)
    if c_mes_atual and c_mes_atual not in candidatas_mes:
        candidatas_mes.insert(0, c_mes_atual)

    candidatas_valor = _colunas_candidatas_por_alternativas(raw, coluna_metrica_sellin(metrica))
    if c_val_atual and c_val_atual not in candidatas_valor:
        candidatas_valor.insert(0, c_val_atual)

    if not candidatas_mes or not candidatas_valor:
        return c_mes_atual, c_val_atual

    serie_atual = _serie_mensal_teste(raw, c_mes_atual, c_val_atual)
    atual_meses = int(len(serie_atual))
    atual_unicos = int(serie_atual.round(8).nunique()) if atual_meses else 0
    atual_soma = float(serie_atual.sum()) if atual_meses else 0.0
    atual_constante = atual_meses >= 3 and atual_unicos <= 1

    melhor = None
    for cm in candidatas_mes:
        for cv in candidatas_valor:
            serie = _serie_mensal_teste(raw, cm, cv)
            meses = int(len(serie))
            if meses == 0:
                continue
            unicos = int(serie.round(8).nunique())
            soma = float(serie.sum())
            # Base do score: ter mais meses e variar entre os meses.
            score = meses * 1000 + unicos * 100 + _pontuar_coluna_valor_sellin(str(cv), metrica)
            # Penaliza série mensal totalmente constante, mas não impede caso seja a única opção.
            if meses >= 3 and unicos <= 1:
                score -= 5000
            # Preferência leve por ANO MÊS quando existir.
            cm_norm = normalizar_texto(cm)
            if "ano mes" in cm_norm:
                score += 250
            if cm == c_mes_atual:
                score += 20
            if cv == c_val_atual:
                score += 20
            cand = (score, meses, unicos, soma, cm, cv)
            if melhor is None or cand > melhor:
                melhor = cand

    if melhor is None:
        return c_mes_atual, c_val_atual

    _, melhor_meses, melhor_unicos, melhor_soma, melhor_mes, melhor_val = melhor

    deve_trocar = False
    if atual_meses == 0 and melhor_meses > 0:
        deve_trocar = True
    elif atual_constante and melhor_unicos > 1:
        deve_trocar = True
    elif atual_meses > 0 and melhor_meses >= atual_meses and melhor_unicos > max(atual_unicos, 1) and melhor_soma > 0 and atual_soma > 0:
        # Só troca por melhoria clara de variabilidade, evitando mudar casos válidos por pouca diferença.
        deve_trocar = atual_unicos <= 1

    if deve_trocar and (melhor_mes != c_mes_atual or melhor_val != c_val_atual):
        avisos.append(
            "Sell-in: ajuste automático de coluna mensal/valor aplicado para evitar Sell-in repetido por mês. "
            f"Antes: mês='{c_mes_atual}', valor='{c_val_atual}', meses={atual_meses}, valores mensais distintos={atual_unicos}. "
            f"Depois: mês='{melhor_mes}', valor='{melhor_val}', meses={melhor_meses}, valores mensais distintos={melhor_unicos}."
        )
        return melhor_mes, melhor_val

    if atual_constante:
        avisos.append(
            "Atenção: o Sell-in mensal ficou com o mesmo valor em todos os meses usando as colunas reconhecidas. "
            "Não encontrei uma combinação alternativa de mês/valor melhor; confira se a coluna do Sell-in no arquivo está mensal mesmo, e não total/agregada."
        )
    return c_mes_atual, c_val_atual

def adicionar_diagnostico_sellin_repetido(avisos: List[str], sellin: pd.DataFrame) -> None:
    """Registra aviso quando o Sell-in mensal ficou idêntico em muitos meses."""
    try:
        if sellin is None or sellin.empty or "mes" not in sellin.columns or "valor_sellin" not in sellin.columns:
            return
        base = sellin[sellin["mes"].notna()].copy()
        if base.empty:
            return
        mensal = base.groupby("mes", as_index=False)["valor_sellin"].sum().sort_values("mes")
        valores = pd.to_numeric(mensal["valor_sellin"], errors="coerce").dropna()
        if len(valores) >= 6 and valores.nunique(dropna=True) == 1:
            avisos.append(
                "Atenção: o total mensal de Sell-in ficou exatamente igual em todos os meses lidos. "
                "Isso geralmente indica que a coluna de valor selecionada é um total/média agregada, "
                "ou que o arquivo de Sell-in realmente contém o mesmo valor para todos os meses."
            )
    except Exception:
        return


def registrar_colunas_reconhecidas(avisos: List[str], base: str, mapa: Dict[str, Optional[str]]) -> None:
    """Inclui aviso objetivo com as colunas reconhecidas, útil para diagnosticar leitura."""
    itens = []
    for nome, coluna in mapa.items():
        itens.append(f"{nome}='{coluna}'" if coluna else f"{nome}=não encontrada")
    avisos.append(f"{base} - colunas reconhecidas: " + "; ".join(itens) + ".")


def registrar_colunas_nao_encontradas(
    avisos: List[str],
    base: str,
    itens: List[Tuple[str, Optional[str], List[str]]],
) -> None:
    """Registra exatamente quais colunas opcionais não foram encontradas."""
    faltantes = []
    for nome, coluna, alternativas in itens:
        if not coluna:
            faltantes.append(f"{nome} (testei: {', '.join(map(str, alternativas[:10]))})")
    if faltantes:
        avisos.append(f"{base} - colunas opcionais não encontradas: " + " | ".join(faltantes) + ".")



def linha_contem_cabecalho(valores: List, obrigatorias: List[str]) -> bool:
    """Valida se uma linha parece ser o cabeçalho esperado, ignorando células vazias."""
    valores_norm = [normalizar_texto(v) for v in valores if v is not None and str(v).strip() != ""]
    obrigatorias_norm = [normalizar_texto(x) for x in obrigatorias]
    return all(any(req == col or req in col or col in req for col in valores_norm) for req in obrigatorias_norm)


def recortar_linha_por_cabecalho(valores: List) -> Tuple[List, int, int]:
    """
    Recorta a linha do cabeçalho removendo colunas totalmente vazias no começo e no fim.
    Retorna: valores_recortados, índice inicial zero-based, índice final exclusivo.
    """
    inicio = 0
    fim = len(valores)

    while inicio < fim and (valores[inicio] is None or str(valores[inicio]).strip() == ""):
        inicio += 1
    while fim > inicio and (valores[fim - 1] is None or str(valores[fim - 1]).strip() == ""):
        fim -= 1

    if inicio >= fim:
        return valores, 0, len(valores)
    return valores[inicio:fim], inicio, fim


def ler_aba_com_cabecalho(
    arquivo: Path,
    aba: str,
    obrigatorias: List[str],
    max_linhas_scan: int = 150,
    linha_preferencial: Optional[int] = None,
    coluna_inicial_preferencial: Optional[int] = None,
) -> Tuple[pd.DataFrame, int]:
    """
    Lê uma aba procurando a linha de cabeçalho que contenha as colunas obrigatórias.

    Para o template de Sell-in, também aceita leitura preferencial na linha 14,
    começando pela coluna B. Isso evita erro quando a aba possui título/instruções
    antes da tabela real.
    """
    if eh_csv(Path(arquivo)):
        return ler_csv_com_cabecalho(
            arquivo=arquivo,
            obrigatorias=obrigatorias,
            max_linhas_scan=max_linhas_scan,
            linha_preferencial=linha_preferencial,
            coluna_inicial_preferencial=coluna_inicial_preferencial,
        )

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba não encontrada no arquivo {arquivo.name}: {aba}")

    ws = wb[aba]
    header_row_idx = None
    header_values = None
    header_start = 0
    header_end = None

    def tentar_linha(idx: int, start_col_1based: int = 1) -> bool:
        nonlocal header_row_idx, header_values, header_start, header_end
        if idx < 1 or idx > ws.max_row:
            return False
        valores = [ws.cell(idx, c).value for c in range(start_col_1based, ws.max_column + 1)]
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if linha_contem_cabecalho(valores_recortados, obrigatorias):
            header_row_idx = idx
            header_values = valores_recortados
            header_start = start_col_1based - 1 + inicio_rel
            header_end = start_col_1based - 1 + fim_rel
            return True
        return False

    # 1) Tentativa preferencial, usada principalmente para Sell-in: linha 14, coluna B.
    if linha_preferencial is not None:
        start_col = coluna_inicial_preferencial or 1
        tentar_linha(linha_preferencial, start_col)

    # 2) Scan automático, caso a posição preferencial não exista ou não bata.
    if header_row_idx is None:
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            valores = list(row)
            valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
            if linha_contem_cabecalho(valores_recortados, obrigatorias):
                header_row_idx = idx
                header_values = valores_recortados
                header_start = inicio
                header_end = fim
                break
            if idx >= max_linhas_scan:
                break

    # 3) Fallback explícito para template: se pediu linha preferencial, lê essa linha mesmo assim
    # e deixa a validação das colunas acusar exatamente o que estiver faltando depois.
    if header_row_idx is None and linha_preferencial is not None:
        start_col = coluna_inicial_preferencial or 1
        valores = [ws.cell(linha_preferencial, c).value for c in range(start_col, ws.max_column + 1)]
        valores_recortados, inicio_rel, fim_rel = recortar_linha_por_cabecalho(valores)
        if any(v is not None and str(v).strip() != "" for v in valores_recortados):
            header_row_idx = linha_preferencial
            header_values = valores_recortados
            header_start = start_col - 1 + inicio_rel
            header_end = start_col - 1 + fim_rel

    if header_row_idx is None or header_values is None:
        wb.close()
        dica_sellin = " Para Sell-in, confirme se o cabeçalho está na linha 14, a partir da coluna B." if linha_preferencial is not None else ""
        raise ValueError(
            f"Não localizei o cabeçalho da aba '{aba}' no arquivo '{arquivo.name}'. "
            f"Colunas esperadas: {obrigatorias}." + dica_sellin
        )

    headers = nomes_unicos(header_values)
    dados = []
    max_col_exclusive = header_end if header_end is not None else ws.max_column
    for row in ws.iter_rows(
        min_row=header_row_idx + 1,
        min_col=header_start + 1,
        max_col=max_col_exclusive,
        values_only=True,
    ):
        linha = list(row)
        if len(linha) < len(headers):
            linha += [None] * (len(headers) - len(linha))
        linha = linha[:len(headers)]
        if any(v is not None and str(v).strip() != "" for v in linha):
            dados.append(linha)

    wb.close()
    return pd.DataFrame(dados, columns=headers), header_row_idx



def _bate_coluna_por_alternativa(nome_coluna, alternativas: List[str]) -> bool:
    """Confere se um nome de coluna bate com alguma alternativa, usando normalização leve."""
    col_norm = normalizar_texto(nome_coluna)
    if not col_norm:
        return False
    for alt in alternativas:
        alt_norm = normalizar_texto(alt)
        if alt_norm and (alt_norm == col_norm or alt_norm in col_norm or col_norm in alt_norm):
            return True
    return False


def pontuar_linha_cabecalho_sellout_padrao(
    valores: List,
    metrica: str,
    nivel_prod: str,
    usar_mapa_categoria_forcado: bool = False,
) -> int:
    """
    Pontua uma possível linha de cabeçalho do Sell-out padrão/VTA.

    A versão anterior exigia a primeira alternativa da métrica, por exemplo
    Qtd_de_Vendas. Isso quebrava quando a aba tinha uma coluna equivalente,
    como Sell-out, Sell out ou Sellout. Aqui a busca aceita qualquer alternativa
    válida da métrica e usa SKU/categoria/UF/mês como reforço.
    """
    valores_norm = [normalizar_texto(v) for v in valores if v is not None and str(v).strip() != ""]
    if not valores_norm:
        return 0

    alts_valor = coluna_metrica_sellout(metrica) + [
        "Sell-out", "Sell out", "Sellout", "Valor Sell-out", "Valor Sellout",
        "Qtd Sell-out", "Quantidade Sell-out", "Volume Sell-out",
    ]
    alts_sku = EAN_SKU_ALTERNATIVAS_AMPLAS + ["SKU", "EAN", "EAN SKU", "SKU EAN"]
    alts_uf = ["UF SM", "UF_SM", "UF Scann", "UF Scanntech", "UF", "Estado"]
    alts_mes = COLUNAS_MES_ALTERNATIVAS + ["Mes", "Mês", "Ano Mes", "Ano Mês"]
    alts_cat = [
        "Categoria SM", "CATEGORIA SM", "CATEGORIA SCANN", "Categoria Scann",
        "Categoria Scanntech", "Categoria", "CATEGORIA", "Categoria Cliente",
        "Categoria Sell-out", "Categoria Sellout", "NIVEL1", "NIVEL2",
    ]
    if nivel_prod in {"NIVEL1", "NIVEL2"} and not usar_mapa_categoria_forcado:
        alts_cat = [nivel_prod]

    def tem(alts: List[str]) -> bool:
        return any(_bate_coluna_por_alternativa(col, alts) for col in valores_norm)

    tem_valor = tem(alts_valor)
    tem_sku = tem(alts_sku)
    tem_cat = tem(alts_cat)
    tem_uf = tem(alts_uf)
    tem_mes = tem(alts_mes)

    # Precisa ter pelo menos valor ou SKU. Com ambos, a chance de ser cabeçalho correto é alta.
    if not (tem_valor or tem_sku):
        return 0

    score = 0
    if tem_valor:
        score += 5
    if tem_sku:
        score += 4
    if tem_cat:
        score += 3
    if tem_uf:
        score += 1
    if tem_mes:
        score += 1

    # Para NIVEL1/NIVEL2, quando não há Congelado, prioriza abas que tenham o nível.
    if nivel_prod in {"NIVEL1", "NIVEL2"} and not usar_mapa_categoria_forcado and not tem_cat:
        score -= 3

    return max(score, 0)


def ler_sellout_excel_com_cabecalho_flexivel(
    arquivo: Path,
    metrica: str,
    nivel_prod: str,
    usar_mapa_categoria_forcado: bool = False,
    max_linhas_scan: int = 250,
) -> Tuple[pd.DataFrame, int, str]:
    """
    Lê um Sell-out Excel procurando automaticamente a melhor aba e o melhor cabeçalho.

    Isso evita pegar por engano uma aba auxiliar e também evita exigir apenas
    Qtd_de_Vendas quando a coluna equivalente se chama Sell-out/Sellout.
    """
    arquivo = Path(arquivo)
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        nomes = wb.sheetnames
        preferidas_exatas = {
            normalizar_texto(x) for x in [
                "Sell-out - SM", "Sell-out", "Sellout", "VTA", "Dados", "Publicar",
                "Base Sell-out", "Base Sellout", "Base", "Planilha1", "Sheet1",
            ]
        }

        melhor = None  # (score_total, score_linha, bonus_nome, ordem_neg, aba, linha, valores, inicio, fim)
        for ordem, aba in enumerate(nomes):
            ws = wb[aba]
            aba_norm = normalizar_texto(aba)
            bonus_nome = 0
            if aba_norm in preferidas_exatas:
                bonus_nome += 8
            elif any(x in aba_norm for x in ["sell out", "sellout", "vta", "publicar", "dados", "base"]):
                bonus_nome += 3
            # Abas de relatório/categoria ainda podem ser úteis como fallback, mas não devem vencer
            # uma aba bruta quando ela existir.
            if any(x in aba_norm for x in ["resumo", "parametro", "parâmetro", "aviso", "crosscheck", "descricao", "descrição"]):
                bonus_nome -= 4

            limite = min(max_linhas_scan, ws.max_row)
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), start=1):
                valores = list(row)
                valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
                score_linha = pontuar_linha_cabecalho_sellout_padrao(
                    valores_recortados,
                    metrica=metrica,
                    nivel_prod=nivel_prod,
                    usar_mapa_categoria_forcado=usar_mapa_categoria_forcado,
                )
                if not score_linha:
                    continue
                # Cabeçalhos de templates costumam aparecer nas primeiras linhas ou perto da 14/15.
                bonus_linha = 2 if idx in (1, 2, 14, 15) else 0
                score_total = score_linha + bonus_nome + bonus_linha
                candidato = (score_total, score_linha, bonus_nome, -ordem, aba, idx, valores_recortados, inicio, fim)
                if melhor is None or candidato[:4] > melhor[:4]:
                    melhor = candidato

        if melhor is None:
            raise ValueError(
                f"Não localizei uma aba de Sell-out válida no arquivo '{arquivo.name}'. "
                "Procurei uma linha de cabeçalho contendo SKU/EAN e/ou uma coluna de valor de Sell-out "
                f"compatível com a métrica '{metrica}'."
            )

        _, _score_linha, _bonus_nome, _ordem, aba, header_row_idx, header_values, header_start, header_end = melhor
        ws = wb[aba]
        headers = nomes_unicos(header_values)
        dados = []
        max_col_exclusive = header_end if header_end is not None else ws.max_column
        for row in ws.iter_rows(
            min_row=header_row_idx + 1,
            min_col=header_start + 1,
            max_col=max_col_exclusive,
            values_only=True,
        ):
            linha = list(row)
            if len(linha) < len(headers):
                linha += [None] * (len(headers) - len(linha))
            linha = linha[:len(headers)]
            if any(v is not None and str(v).strip() != "" for v in linha):
                dados.append(linha)

        return pd.DataFrame(dados, columns=headers), header_row_idx, aba
    finally:
        wb.close()

def metrica_eh_volume_variavel(metrica: str) -> bool:
    """Indica a opção que converte Sell-in em volume por gramatura média ponderada."""
    m = normalizar_texto(metrica)
    return m in {"volume variavel", "volume variável", "volume_variavel", "volumevariavel", "variavel", "variável"}


def alternativas_sellout_quantia() -> List[str]:
    return [
        "Qtd_de_Vendas", "Qtd de Vendas", "Quantidade de Vendas", "Quantidade",
        "Qtd Vendas", "Qtd", "Unidades", "Sell-Out", "Sell out", "Sellout"
    ]


def alternativas_sellout_volume() -> List[str]:
    return [
        "Vendas_em_volume", "Vendas em volume", "Volume", "Vendas Volume",
        "Sell-out Volume", "Sellout Volume"
    ]


def alternativas_sellin_quantia() -> List[str]:
    return [
        # Para Sell-in, evita termos genéricos como "Qtd" ou "Quantidade",
        # porque o template pode ter colunas de validação como "Qtd meses" e "Qtd UF".
        "SELL IN (QUANTIDADE)", "SELL-IN (QUANTIDADE)", "Sell-in (Quantidade)", "Sell in (Quantidade)",
        "SELL IN QUANTIDADE", "SELL-IN QUANTIDADE", "Sell-in Quantidade", "Sell in Quantidade",
        "Sell-in Quantia", "Sell in Quantia",
        "Quantidade Sell-in", "Quantidade Sell in", "Qtd Sell-in", "Qtd Sell in",
        "Sell-in Unidades", "Sell in Unidades", "SELL IN UNIDADES", "SELL-IN UNIDADES",
    ]


def alternativas_sellin_volume() -> List[str]:
    return [
        # Nomes mais específicos primeiro, para priorizar volume quando existirem
        # Sell-in em quantidade e Sell-in em Kg/L na mesma base.
        "SELL IN (Kg / L)", "SELL IN (KG/L)", "SELL IN KG L",
        "Sell-in (Kg / L)", "Sell in (Kg / L)", "Sell-in Kg/L", "Sell in Kg/L",
        "Sell-in Volume", "Sell in Volume", "Volume Sell-in", "Volume Sell in",
        "Vendas_em_volume", "Vendas em volume", "Volume", "Kg/L", "Kg L",
        # Genéricos por último.
        "Sell-in", "Sell in"
    ]


def coluna_metrica_sellout(metrica: str) -> List[str]:
    metrica_norm = normalizar_texto(metrica)
    if metrica_norm.startswith("q"):
        return alternativas_sellout_quantia()
    if metrica_eh_volume_variavel(metrica):
        # Para volume variável, tenta usar volume do Sell-out quando existir.
        # Se não existir, aceita quantidade e o script tenta converter pelo peso do nome do SKU.
        return alternativas_sellout_volume() + alternativas_sellout_quantia()
    return alternativas_sellout_volume()


def coluna_metrica_sellin(metrica: str) -> List[str]:
    metrica_norm = normalizar_texto(metrica)
    if metrica_norm.startswith("q") or metrica_eh_volume_variavel(metrica):
        # Volume variável usa Sell-in em quantidade como base e converte para Kg/L pela gramatura média.
        return alternativas_sellin_quantia() + ["Sell-in", "Sell in"]
    return alternativas_sellin_volume()


def extrair_peso_gramas(texto) -> float:
    """
    Extrai gramatura/peso do nome do SKU.

    Exemplos aceitos:
    - 450G, 450 g, 450gr, 450 gramas
    - 1KG, 1 kg, 1,5kg, 0.5 kilo, 2 kilogramas

    Retorna o peso em gramas. Se não localizar, retorna NaN.
    """
    if pd.isna(texto):
        return np.nan
    s = str(texto).lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace(",", ".")
    # Evita que x em packs atrapalhe: 24x70g ainda encontra 70g.
    padrao = re.compile(
        r"(?<!\d)(\d+(?:\.\d+)?)\s*(kg|kgs|kilo|kilos|quilo|quilos|kilograma|kilogramas|g|gr|grs|grama|gramas)\b",
        flags=re.IGNORECASE,
    )
    encontrados = []
    for valor_txt, unidade in padrao.findall(s):
        try:
            valor = float(valor_txt)
        except Exception:
            continue
        unidade = unidade.lower()
        if unidade in {"kg", "kgs", "kilo", "kilos", "quilo", "quilos", "kilograma", "kilogramas"}:
            valor *= 1000
        if valor > 0:
            encontrados.append(valor)
    if not encontrados:
        return np.nan
    # Se houver mais de uma gramatura, usa a última ocorrência do nome, que costuma ser a gramatura comercial.
    return float(encontrados[-1])


def _media_ponderada_gramas(df: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=keys + ["gramatura_media"])
    tmp = df.copy()
    tmp = tmp[(tmp["peso_gramas_sku"].fillna(0) > 0) & (tmp["qtd_ref_peso"].fillna(0) > 0)].copy()
    if tmp.empty:
        return pd.DataFrame(columns=keys + ["gramatura_media"])
    tmp["_peso_x_qtd"] = tmp["peso_gramas_sku"] * tmp["qtd_ref_peso"]
    agg = tmp.groupby(keys, dropna=False, as_index=False).agg(_peso_x_qtd=("_peso_x_qtd", "sum"), _qtd=("qtd_ref_peso", "sum"))
    agg["gramatura_media"] = agg["_peso_x_qtd"] / agg["_qtd"].replace(0, np.nan)
    return agg[keys + ["gramatura_media"]]


def aplicar_volume_variavel_por_gramatura(
    sellin: pd.DataFrame,
    sellout: pd.DataFrame,
    avisos: List[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, object]]:
    """
    Converte o Sell-in em volume estimado usando a gramatura média ponderada do Sell-out.

    Regra principal:
        Sell-in convertido = Sell-in quantidade * gramatura média ponderada / 1000

    A gramatura média é calculada a partir do nome do SKU do Sell-out e ponderada pela quantidade
    vendida do Sell-out. O script tenta a maior granularidade primeiro:
      1. Categoria/PROD + UF consolidada + mês
      2. Categoria/PROD + mês
      3. Categoria/PROD + UF consolidada
      4. Categoria/PROD
      5. Média global
    """
    si = sellin.copy()
    so = sellout.copy()
    info = {
        "Volume variável aplicado": "Não",
        "SKUs Sell-out com peso localizado": 0,
        "Gramatura média global": "",
        "Linhas Sell-in convertidas": 0,
        "Critério volume variável": "Não aplicado",
    }

    if si.empty or so.empty:
        avisos.append("Volume variável não aplicado: Sell-in ou Sell-out vazio.")
        return si, so, info

    if "nome_sku" not in so.columns:
        so["nome_sku"] = ""
    if "qtd_sellout" not in so.columns:
        # Fallback: usa o próprio valor do Sell-out como peso de ponderação.
        so["qtd_sellout"] = so.get("valor_sellout", 0)

    so["peso_gramas_sku"] = so.get("peso_gramas_sku", pd.Series([np.nan] * len(so), index=so.index))
    so["peso_gramas_sku"] = pd.to_numeric(so["peso_gramas_sku"], errors="coerce")
    faltou_peso = so["peso_gramas_sku"].isna() | (so["peso_gramas_sku"].fillna(0) <= 0)
    if faltou_peso.any():
        so.loc[faltou_peso, "peso_gramas_sku"] = so.loc[faltou_peso, "nome_sku"].map(extrair_peso_gramas)

    so["qtd_ref_peso"] = pd.to_numeric(so["qtd_sellout"], errors="coerce").fillna(0)
    if so["qtd_ref_peso"].sum() == 0:
        so["qtd_ref_peso"] = pd.to_numeric(so.get("valor_sellout", 0), errors="coerce").fillna(0)

    so = adicionar_uf_comparacao(so)
    si = adicionar_uf_comparacao(si)

    base_peso = so[(so["peso_gramas_sku"].fillna(0) > 0) & (so["qtd_ref_peso"].fillna(0) > 0)].copy()
    skus_com_peso = base_peso[base_peso.get("ean", "") != ""]["ean"].nunique() if not base_peso.empty and "ean" in base_peso.columns else 0
    info["SKUs Sell-out com peso localizado"] = int(skus_com_peso)

    if base_peso.empty:
        avisos.append(
            "Volume variável não aplicado: não localizei gramatura no nome dos SKUs do Sell-out. "
            "Procurei padrões como 450G, 150 g, 1KG, 1,5 kg, gramas, kilo e kilogramas."
        )
        return si, so, info

    media_global = float((base_peso["peso_gramas_sku"] * base_peso["qtd_ref_peso"]).sum() / base_peso["qtd_ref_peso"].sum())
    info["Gramatura média global"] = round(media_global, 4)

    # Garante colunas de chave.
    for df in (si, so):
        if "categoria_key" not in df.columns:
            df["categoria_key"] = ""
        if "mes" not in df.columns:
            df["mes"] = pd.NaT
        if "uf_comparacao" not in df.columns:
            df["uf_comparacao"] = df.get("uf", "TOTAL").map(padronizar_uf_comparacao)

    # Merge por granularidades em cascata.
    si["_row_id_volume_variavel"] = np.arange(len(si))
    si["gramatura_media_aplicada"] = np.nan
    si["base_gramatura_aplicada"] = ""

    granularidades = [
        (["categoria_key", "uf_comparacao", "mes"], "Categoria/PROD + UF + mês"),
        (["categoria_key", "mes"], "Categoria/PROD + mês"),
        (["categoria_key", "uf_comparacao"], "Categoria/PROD + UF"),
        (["categoria_key"], "Categoria/PROD"),
    ]

    for keys, label in granularidades:
        media = _media_ponderada_gramas(base_peso, keys)
        if media.empty:
            continue
        temp = si[["_row_id_volume_variavel"] + keys].merge(media, on=keys, how="left")
        mapa = temp.set_index("_row_id_volume_variavel")["gramatura_media"]
        mask = si["gramatura_media_aplicada"].isna() & si["_row_id_volume_variavel"].map(mapa).notna()
        si.loc[mask, "gramatura_media_aplicada"] = si.loc[mask, "_row_id_volume_variavel"].map(mapa)
        si.loc[mask, "base_gramatura_aplicada"] = label

    mask_global = si["gramatura_media_aplicada"].isna()
    si.loc[mask_global, "gramatura_media_aplicada"] = media_global
    si.loc[mask_global, "base_gramatura_aplicada"] = "Média global"

    si["valor_sellin_original_volume_variavel"] = si["valor_sellin"]
    si["valor_sellin"] = si["valor_sellin"] * si["gramatura_media_aplicada"] / 1000
    convertidas = int(si["valor_sellin"].fillna(0).ne(0).sum())
    info["Linhas Sell-in convertidas"] = convertidas
    info["Volume variável aplicado"] = "Sim"
    info["Critério volume variável"] = (
        "Sell-in em quantidade convertido para Kg/L usando gramatura média ponderada dos SKUs do Sell-out. "
        "A ponderação usa Qtd_de_Vendas quando disponível; caso contrário, usa o próprio valor do Sell-out como peso."
    )

    # Se o Sell-out foi lido em quantidade por falta de Vendas_em_volume, tenta convertê-lo também.
    origem_valor = str(so.get("valor_sellout_origem_volume_variavel", pd.Series([""] * len(so), index=so.index)).iloc[0] if len(so) else "")
    if origem_valor == "quantidade":
        so["valor_sellout_original_volume_variavel"] = so["valor_sellout"]
        # usa peso do SKU; quando faltar, usa média da categoria/global por linha.
        so["peso_resolvido_sellout"] = so["peso_gramas_sku"]
        mask_so_sem_peso = so["peso_resolvido_sellout"].isna() | (so["peso_resolvido_sellout"].fillna(0) <= 0)
        if mask_so_sem_peso.any():
            media_cat = _media_ponderada_gramas(base_peso, ["categoria_key"])
            so = so.merge(media_cat.rename(columns={"gramatura_media": "gramatura_media_categoria"}), on="categoria_key", how="left")
            so.loc[mask_so_sem_peso, "peso_resolvido_sellout"] = so.loc[mask_so_sem_peso, "gramatura_media_categoria"].fillna(media_global)
        so["valor_sellout"] = pd.to_numeric(so["qtd_sellout"], errors="coerce").fillna(0) * so["peso_resolvido_sellout"] / 1000
        avisos.append("Volume variável: Sell-out também foi convertido de quantidade para Kg/L por falta de coluna de volume.")

    si = si.drop(columns=["_row_id_volume_variavel"], errors="ignore")
    avisos.append(
        f"Volume variável aplicado: {convertidas} linhas do Sell-in convertidas por gramatura média. "
        f"Gramatura média global ponderada: {media_global:.4f} g. SKUs do Sell-out com peso localizado: {skus_com_peso}."
    )
    return si, so, info





def pontuar_linha_cabecalho_sellin(valores: List, metrica: str) -> int:
    """
    Dá uma pontuação para identificar o cabeçalho REAL do template de Sell-in.

    Isso evita confundir títulos/instruções como "Tabela de Sell In" ou a linha explicativa
    "SELL IN - Kilograma ou litro" com o cabeçalho verdadeiro.
    """
    valores_norm = [normalizar_texto(v) for v in valores if v is not None and str(v).strip() != ""]
    if not valores_norm:
        return 0

    grupos = {
        "uf": ["uf sm", "uf", "estado"],
        "mes": ["ano mes", "mês", "mes", "data", "periodo"],
        "canal": ["canal", "pdv canal"],
        "categoria": ["categoria sm", "categoria scann", "categoria", "prod", "nivel1", "nivel2"],
        "sku": ["ean", "sku", "codigo de barras", "código de barras", "cod barras"],
        # Para identificar a linha do cabeçalho, aceita qualquer coluna de Sell-in,
        # mesmo que a métrica escolhida depois não exista. Assim o erro final fica
        # mais claro: coluna de valor não encontrada, em vez de cabeçalho não encontrado.
        "valor": coluna_metrica_sellin(metrica) + [
            "SELL IN (Kg / L)", "SELL IN (QUANTIDADE)", "Sell-in", "Sell in"
        ],
    }

    def bate(alt: str, col: str) -> bool:
        alt_n = normalizar_texto(alt)
        if not alt_n or not col:
            return False
        return alt_n == col or alt_n in col or col in alt_n

    score = 0
    tem_valor = False
    for grupo, alternativas in grupos.items():
        encontrou = any(bate(alt, col) for alt in alternativas for col in valores_norm)
        if encontrou:
            score += 1
            if grupo == "valor":
                tem_valor = True

    # Linha de título/instrução pode conter "Sell In", mas geralmente não tem UF/Mês/EAN/Categoria.
    if not tem_valor:
        return 0
    if score < 2:
        return 0
    return score


def ordem_linhas_cabecalho_sellin(limite: int, linha_padrao: int = 15) -> List[int]:
    """
    Define a ordem de procura do cabeçalho do Sell-in.

    Regra:
    - procura primeiro de baixo para cima, da linha 15 até a linha 1;
    - assim cobre arquivos com cabeçalho na primeira linha, sem perder o padrão do template;
    - se não encontrar nas primeiras 15 linhas, continua procurando da linha 16 até o limite atual.
    """
    try:
        limite = int(limite or 0)
    except Exception:
        limite = 0
    if limite <= 0:
        return []

    linha_padrao = min(int(linha_padrao), limite)
    ordem = list(range(linha_padrao, 0, -1))

    # Mantém fallback para arquivos fora do padrão, sem atrapalhar o template.
    if limite > linha_padrao:
        ordem.extend(range(linha_padrao + 1, limite + 1))

    return ordem


def ler_sellin_com_cabecalho_flexivel(arquivo: Path, aba: str, metrica: str) -> Tuple[pd.DataFrame, int]:
    """
    Lê Sell-in com busca de cabeçalho por pontuação.

    Regras:
    - procura primeiro de baixo para cima: linha 15, 14, 13, 12... até a linha 1;
    - cobre arquivos com cabeçalho na primeira linha;
    - mantém fallback até o limite atual de leitura, caso o cabeçalho esteja depois da linha 15;
    - evita pegar títulos/instruções acima da tabela;
    - aceita colunas opcionais, desde que exista alguma coluna de valor de Sell-in.
    """
    arquivo = Path(arquivo)

    if eh_csv(arquivo):
        enc, sep, sample = detectar_csv_formato(arquivo, nrows=250)
        melhor = None
        limite = min(250, len(sample))

        for idx in ordem_linhas_cabecalho_sellin(limite, linha_padrao=15):
            valores = sample.iloc[idx - 1, :].tolist()
            valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
            score = pontuar_linha_cabecalho_sellin(valores_recortados, metrica)
            if score:
                # Como a ordem já começa em 15 -> 1, o primeiro cabeçalho válido
                # tende a ser a tabela real, não títulos/instruções acima dela.
                melhor = (idx, valores_recortados, inicio, fim)
                break

        if melhor is None:
            raise ValueError(
                f"Não localizei o cabeçalho do Sell-in no CSV '{arquivo.name}'. "
                "Procurei primeiro da linha 15 até a linha 1 e, se necessário, até o limite atual de leitura. "
                "A linha precisa ter coluna de Sell-in e pelo menos mais uma coluna estrutural, como UF, Mês, EAN/SKU ou Categoria."
            )

        header_row_idx, header_values, header_start, header_end = melhor
        headers = nomes_unicos(header_values)
        usecols = list(range(header_start, header_end))
        engine = "python" if sep is None else "c"
        kwargs = dict(
            filepath_or_buffer=arquivo,
            sep=sep,
            engine=engine,
            header=None,
            skiprows=header_row_idx,
            usecols=usecols,
            names=headers,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=False,
            on_bad_lines="skip",
            encoding=enc,
        )
        if engine == "c":
            kwargs["low_memory"] = False
        df = pd.read_csv(**kwargs)
        df = limpar_bom_dataframe(df)
        return remover_linhas_colunas_vazias(df), header_row_idx

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba não encontrada no arquivo {arquivo.name}: {aba}")
    ws = wb[aba]

    melhor = None
    limite = min(250, ws.max_row)

    # Lê as primeiras linhas uma vez só. Depois avalia na ordem desejada:
    # 15, 14, 13... 1. Isso mantém performance melhor em modo read_only.
    linhas_amostra = list(ws.iter_rows(min_row=1, max_row=limite, values_only=True))

    for idx in ordem_linhas_cabecalho_sellin(limite, linha_padrao=15):
        valores = list(linhas_amostra[idx - 1])
        valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
        score = pontuar_linha_cabecalho_sellin(valores_recortados, metrica)
        if score:
            # Como a ordem já começa em 15 -> 1, o primeiro cabeçalho válido
            # tende a ser a tabela real, não títulos/instruções acima dela.
            melhor = (idx, valores_recortados, inicio, fim)
            break

    if melhor is None:
        wb.close()
        raise ValueError(
            f"Não localizei o cabeçalho do Sell-in na aba '{aba}' do arquivo '{arquivo.name}'. "
            "Procurei primeiro da linha 15 até a linha 1 e, se necessário, até o limite atual de leitura. "
            "A linha precisa ter coluna de Sell-in e pelo menos mais uma coluna estrutural, como UF, Mês, EAN/SKU ou Categoria."
        )

    header_row_idx, header_values, header_start, header_end = melhor
    headers = nomes_unicos(header_values)
    dados = []
    max_col_exclusive = header_end if header_end is not None else ws.max_column
    for row in ws.iter_rows(
        min_row=header_row_idx + 1,
        min_col=header_start + 1,
        max_col=max_col_exclusive,
        values_only=True,
    ):
        linha = list(row)
        if len(linha) < len(headers):
            linha += [None] * (len(headers) - len(linha))
        linha = linha[:len(headers)]
        if any(v is not None and str(v).strip() != "" for v in linha):
            dados.append(linha)

    wb.close()
    return pd.DataFrame(dados, columns=headers), header_row_idx

def ler_sellin(arquivo: Path, metrica: str) -> Tuple[pd.DataFrame, List[str]]:
    """
    Lê o Sell-in do template, mas sem travar quando algumas colunas não existirem.

    Regras:
    - Valor de Sell-in é obrigatório.
    - UF, mês, ano, SKU/EAN e categoria são opcionais.
    - Se não houver mês, tenta usar ANO.
    - Se não houver SKU/EAN, o código segue com ean vazio e permite uso de fabricante escolhido no Sell-out.
    """
    avisos = []
    aba = escolher_aba_sellin(arquivo)

    # Busca de cabeçalho por pontuação, para não confundir título/instruções com a tabela real.
    raw, linha = ler_sellin_com_cabecalho_flexivel(arquivo, aba, metrica)

    # Preferir UF SM quando existir; se não existir, usa UF normal.
    c_uf = localizar_coluna(raw, ["UF SM", "UF_SM", "UF Scann", "UF Scanntech", "UF", "Estado"], obrigatoria=False)
    c_mes = localizar_coluna(raw, COLUNAS_MES_ALTERNATIVAS, obrigatoria=False)
    c_ano = localizar_coluna(raw, ["Ano", "ANO", "Year", "YEAR"], obrigatoria=False)
    if c_ano == c_mes:
        c_ano = None
    c_ean = localizar_coluna(
        raw,
        [
            "EAN", "SKU", "EAN/SKU", "SKU/EAN",
            "Código Barras", "Codigo Barras", "Código de Barras", "Codigo de Barras",
            "Cod Barras", "Cód Barras", "CODIGO_BARRAS", "COD_BARRAS"
        ],
        obrigatoria=False,
    )
    c_cat = localizar_coluna(
        raw,
        [
            "Categoria SM", "CATEGORIA SM",
            "CATEGORIA SCANN", "Categoria Scann", "Categoria Scanntech",
            "Categoria", "CATEGORIA", "PROD", "NIVEL1", "NIVEL2"
        ],
        obrigatoria=False,
    )
    c_fab_si = localizar_coluna(raw, ["Fabricante", "FABRICANTE", "Fornecedor", "Indústria", "Industria"], obrigatoria=False)
    c_val, avisos_valor_sellin = localizar_coluna_valor_sellin(raw, metrica, arquivo)
    avisos.extend(avisos_valor_sellin)

    # Validação cruzada mês x valor: evita repetir o mesmo Sell-in em todos os meses
    # quando uma coluna agregada ou um período errado é escolhido automaticamente.
    c_mes, c_val = escolher_colunas_sellin_mes_valor_mais_plausiveis(
        raw=raw,
        c_mes_atual=c_mes,
        c_val_atual=c_val,
        metrica=metrica,
        avisos=avisos,
    )

    registrar_colunas_reconhecidas(avisos, "Sell-in", {
        "UF": c_uf,
        "Mês/Data": c_mes,
        "Ano": c_ano,
        "SKU/EAN": c_ean,
        "Categoria": c_cat,
        "Fabricante": c_fab_si,
        "Valor Sell-in": c_val,
    })
    registrar_colunas_nao_encontradas(avisos, "Sell-in", [
        ("UF", c_uf, ["UF SM", "UF", "Estado"]),
        ("Mês/Data", c_mes, ["ANO MÊS", "ANO_MES", "AAAAMM", "YYYYMM", "Mês/Ano", "Competência", "Referência", "Data", "Período"]),
        ("Ano", c_ano, ["Ano", "ANO", "Year"]),
        ("SKU/EAN", c_ean, ["EAN", "SKU", "EAN/SKU", "Código de Barras"]),
        ("Categoria", c_cat, ["Categoria SM", "CATEGORIA SCANN", "Categoria", "PROD"]),
        ("Fabricante", c_fab_si, ["Fabricante", "Fornecedor", "Indústria"]),
    ])

    n = len(raw)
    uf = raw[c_uf].astype(str).str.strip() if c_uf else pd.Series(["TOTAL"] * n, index=raw.index)
    mes = raw[c_mes].map(converter_mes) if c_mes else pd.Series([pd.NaT] * n, index=raw.index)

    if c_ano:
        ano = raw[c_ano].map(converter_ano)
    else:
        ano = mes.map(lambda x: x.year if pd.notna(x) else np.nan)

    ean = raw[c_ean].map(ean_texto) if c_ean else pd.Series([""] * n, index=raw.index)
    categoria = raw[c_cat].astype(str).str.strip() if c_cat else pd.Series([""] * n, index=raw.index)
    fabricante_si = raw[c_fab_si].astype(str).str.strip() if c_fab_si else pd.Series([""] * n, index=raw.index)

    df = pd.DataFrame({
        "uf": uf.replace({"nan": "", "None": ""}).fillna(""),
        "mes": mes,
        "ano": ano,
        "ean": ean,
        "categoria_original_sellin": categoria.replace({"nan": "", "None": ""}).fillna(""),
        "fabricante_sellin": fabricante_si.replace({"nan": "", "None": ""}).fillna(""),
        "valor_sellin": limpar_coluna_numerica_vetorizada(raw[c_val]),
    })

    # Se não houver UF, mantém TOTAL; se houver linhas vazias, também usa TOTAL.
    df["uf"] = df["uf"].astype(str).str.strip().replace("", "TOTAL")

    # Se não houver mês, mas houver ano, mantém a informação anual para cálculo anual.
    df["ano"] = pd.to_numeric(df["ano"], errors="coerce")
    df.loc[df["mes"].notna(), "ano"] = df.loc[df["mes"].notna(), "mes"].dt.year

    # Não filtra por SKU, UF ou mês. Só remove linhas sem valor de Sell-in.
    linhas_antes = len(df)
    df = df[df["valor_sellin"].fillna(0) != 0].copy()

    if not c_uf:
        avisos.append("Sell-in sem coluna UF. A análise por UF usará TOTAL para o Sell-in.")
    if not c_mes:
        if c_ano:
            avisos.append("Sell-in sem mês. Será usada análise anual quando possível, com base na coluna de ano.")
        else:
            avisos.append("Sell-in sem mês e sem ano. As análises temporais serão limitadas ao total disponível.")
    if not c_ean:
        avisos.append("Sell-in sem SKU/EAN. O mapeamento por SKU será ignorado; use o fabricante do Sell-out na interface.")
    if not c_cat:
        avisos.append("Sell-in sem categoria. Se não houver SKU para mapear, será criada uma visão total pelo fabricante selecionado.")
    if linhas_antes and df.empty:
        avisos.append("Sell-in foi lido, mas não sobrou nenhuma linha com valor numérico de Sell-in.")

    adicionar_diagnostico_sellin_repetido(avisos, df)
    avisos.append(f"Sell-in lido da aba '{aba}' a partir da linha de cabeçalho {linha}, usando a coluna '{c_val}' como métrica.")
    return df, avisos

def mapa_categoria_sellin_por_sku(sellin: pd.DataFrame) -> pd.DataFrame:
    """
    Cria um mapa SKU -> Categoria do Sell-in.
    Se o mesmo SKU aparecer em mais de uma categoria, fica a categoria de maior Sell-in.
    """
    if sellin.empty or "categoria_original_sellin" not in sellin.columns:
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])

    base = sellin.copy()
    base["categoria_original_sellin"] = base["categoria_original_sellin"].fillna("").astype(str).str.strip()
    base["categoria_key_original_sellin"] = base["categoria_original_sellin"].map(normalizar_categoria)
    base = base[(base["ean"] != "") & (base["categoria_key_original_sellin"] != "")].copy()

    if base.empty:
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])

    mapa = (
        base.groupby(["ean", "categoria_original_sellin", "categoria_key_original_sellin"], as_index=False)["valor_sellin"]
        .sum()
        .sort_values(["ean", "valor_sellin"], ascending=[True, False])
        .drop_duplicates("ean", keep="first")
        [["ean", "categoria_original_sellin", "categoria_key_original_sellin"]]
        .rename(columns={
            "categoria_original_sellin": "categoria_map_prod",
            "categoria_key_original_sellin": "categoria_key_map_prod",
        })
    )
    return mapa



def _alternativas_categoria_congelado(nivel_prod: str) -> Tuple[List[str], str]:
    """Retorna as colunas de categoria/PROD esperadas no Congelado conforme o agrupamento escolhido."""
    nivel = normalizar_nivel_prod(nivel_prod)
    if nivel == "NIVEL1":
        return [
            "NIVEL1", "Nível 1", "Nivel 1", "N1",
            "Est Mer 6 (Categoria)", "Est Mer 6 Categoria", "EST_MER_6",
            "Categoría congelada ScannMarket", "Categoria congelada ScannMarket",
            "Categoria atual Data Excellence", "Categoria Scanntech", "Categoria", "CATEGORIA",
        ], "NIVEL1/Categoria do Congelado"
    if nivel == "NIVEL2":
        return [
            "NIVEL2", "Nível 2", "Nivel 2", "N2",
            "Est Mer 7 (Subcategoria)", "Est Mer 7 Subcategoria", "Est Mer 7", "EST_MER_7",
            "Subcategoria", "Subcategoria ScannMarket",
            "Categoría congelada ScannMarket", "Categoria congelada ScannMarket",
            "Categoria atual Data Excellence", "Categoria", "CATEGORIA",
        ], "NIVEL2/Subcategoria do Congelado"
    if nivel == "ESTMER7":
        return [
            "Est Mer 7 (Subcategoria)", "Est Mer 7 Subcategoria", "Est Mer 7", "EST_MER_7",
            "Subcategoria", "Subcategoria ScannMarket", "NIVEL2", "Nível 2", "Nivel 2", "N2",
        ], "Est Mer 7 do Congelado"
    return [
        "Categoría congelada ScannMarket", "Categoria congelada ScannMarket",
        "Categoria atual Data Excellence", "Est Mer 6 (Categoria)", "Est Mer 6 Categoria",
        "Categoria Scanntech", "Categoria SM", "CATEGORIA SCANN", "Categoria", "CATEGORIA", "PROD",
    ], "Categoria do Congelado"



def localizar_colunas_codigo_congelado(raw: pd.DataFrame) -> List[str]:
    """
    Localiza todas as colunas de código possíveis do Congelado, priorizando
    Código Barras SKU. Evita confundir Nome SKU, Marca SKU e Fabricante SKU.
    """
    if raw is None or raw.empty:
        return []

    prioridade_exata = {
        normalizar_texto("Código Barras SKU"): 1000,
        normalizar_texto("Codigo Barras SKU"): 1000,
        normalizar_texto("Código_Barras_SKU"): 1000,
        normalizar_texto("Codigo_Barras_SKU"): 1000,
        normalizar_texto("Código de Barras SKU"): 980,
        normalizar_texto("Codigo de Barras SKU"): 980,
        normalizar_texto("CODIGO_BARRAS_SKU"): 980,
        normalizar_texto("CODIGO BARRAS SKU"): 980,
        normalizar_texto("COD_BARRAS_SKU"): 970,
        normalizar_texto("COD BARRAS SKU"): 970,
        normalizar_texto("CODIGO_BARRAS_CONTENIDO"): 950,
        normalizar_texto("CODIGO BARRAS CONTENIDO"): 950,
        normalizar_texto("Código Barras Contenido"): 950,
        normalizar_texto("Codigo Barras Contenido"): 950,
        normalizar_texto("Código Barras Conteúdo"): 950,
        normalizar_texto("Codigo Barras Conteudo"): 950,
        normalizar_texto("EAN"): 900,
        normalizar_texto("SKU/EAN"): 890,
        normalizar_texto("EAN/SKU"): 890,
        normalizar_texto("SKU EAN"): 880,
        normalizar_texto("EAN SKU"): 880,
        normalizar_texto("SKU"): 600,
    }
    termos_descritivos = {
        "nome", "nombre", "descricao", "descrição", "descripcion", "desc",
        "categoria", "fabricante", "marca", "fornecedor", "proveedor",
        "preco", "precio", "price", "valor", "vendas", "volume",
        "conteudo", "contenido", "qtd", "cant", "medida", "analise", "análise",
    }

    candidatos = []
    for col in raw.columns:
        col_norm = normalizar_texto(col)
        if not col_norm:
            continue
        tokens = set(col_norm.split())
        score = None

        if col_norm in prioridade_exata:
            score = prioridade_exata[col_norm]
        elif (("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens) and "sku" in tokens):
            score = 920
        elif (("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens)):
            score = 850
        elif "ean" in tokens and not (tokens & termos_descritivos):
            score = 800
        elif col_norm == "sku":
            score = 600
        elif "sku" in tokens and not (tokens & termos_descritivos):
            score = 520

        if score is None:
            continue
        if tokens & termos_descritivos:
            score -= 500

        amostra = raw[col].head(5000).map(ean_texto)
        qtd_validos = int((amostra != "").sum())
        if qtd_validos <= 0:
            continue
        score += min(qtd_validos, 1000) / 1000
        candidatos.append((score, str(col)))

    candidatos.sort(key=lambda x: x[0], reverse=True)
    saida = []
    for _, col in candidatos:
        if col not in saida:
            saida.append(col)
    return saida


def localizar_coluna_categoria_congelado(raw: pd.DataFrame, alternativas: List[str]) -> Optional[str]:
    """Localiza a coluna de categoria do Congelado, incluindo o espanhol 'Categoría'."""
    if raw is None or raw.empty:
        return None
    prioridade_exata = [
        "Categoría congelada ScannMarket",
        "Categoria congelada ScannMarket",
        "Categoria atual Data Excellence",
        "Est Mer 6 (Categoria)",
        "Est Mer 6 Categoria",
        "Categoria Scanntech",
        "Categoria SM",
        "CATEGORIA SCANN",
        "Categoria",
        "CATEGORIA",
        "PROD",
    ]
    alts = prioridade_exata + list(alternativas or [])
    mapa = {normalizar_texto(c): c for c in raw.columns}
    for alt in alts:
        alt_norm = normalizar_texto(alt)
        if alt_norm in mapa:
            return mapa[alt_norm]

    candidatos = []
    for col in raw.columns:
        col_norm = normalizar_texto(col)
        if not col_norm:
            continue
        score = 0
        for alt in alts:
            alt_norm = normalizar_texto(alt)
            if alt_norm and (alt_norm in col_norm or col_norm in alt_norm):
                score = max(score, 500 + len(alt_norm))
        if "categoria" in col_norm:
            score += 200
        if "congelada" in col_norm or "congelado" in col_norm:
            score += 150
        if "scannmarket" in col_norm or "scann" in col_norm:
            score += 100
        if "data excellence" in col_norm:
            score += 100
        if score > 0:
            qtd_validas = int(raw[col].astype(str).str.strip().replace({"": np.nan, "nan": np.nan, "None": np.nan}).dropna().head(5000).shape[0])
            if qtd_validas > 0:
                score += min(qtd_validas, 1000) / 1000
                candidatos.append((score, str(col)))
    if not candidatos:
        return None
    candidatos.sort(key=lambda x: x[0], reverse=True)
    return candidatos[0][1]



def localizar_coluna_fabricante_congelado(raw: pd.DataFrame) -> Optional[str]:
    """Localiza a coluna de Fabricante no Congelado, priorizando Fabricante SKU."""
    if raw is None or raw.empty:
        return None

    prioridade = [
        "Fabricante SKU", "Fabricante", "Fabricante do SKU", "Fabricante del SKU",
        "Proveedor SKU", "Proveedor", "Fornecedor SKU", "Fornecedor",
        "Fabricante Marca", "Nome Fabricante", "Nombre Fabricante",
    ]
    mapa = {normalizar_texto(c): c for c in raw.columns}

    # 1) Match exato normalizado.
    for alt in prioridade:
        alt_norm = normalizar_texto(alt)
        if alt_norm in mapa:
            return mapa[alt_norm]

    # 2) Fallback por pontuação, evitando colunas que não são fabricante.
    termos_ruins = {
        "categoria", "congelada", "scannmarket", "ean", "sku", "codigo", "barras",
        "marca", "nome", "nombre", "descricao", "descripcion", "volume", "valor",
        "venda", "vendas", "qtd", "quantidade", "canal", "uf", "data", "mes",
    }
    candidatos = []
    for col in raw.columns:
        col_norm = normalizar_texto(col)
        if not col_norm:
            continue
        tokens = set(col_norm.split())
        score = 0
        if "fabricante" in tokens or "fabricante" in col_norm:
            score += 500
        if "proveedor" in tokens or "fornecedor" in tokens:
            score += 450
        if "sku" in tokens:
            score += 50
        if tokens & termos_ruins:
            score -= 200
        if score <= 0:
            continue
        qtd_validas = int(raw[col].astype(str).str.strip().replace({"": np.nan, "nan": np.nan, "None": np.nan}).dropna().head(5000).shape[0])
        if qtd_validas > 0:
            candidatos.append((score + min(qtd_validas, 1000) / 1000, str(col)))
    if not candidatos:
        return None
    candidatos.sort(key=lambda x: x[0], reverse=True)
    return candidatos[0][1]


def filtrar_congelado_por_fabricante(
    raw: pd.DataFrame,
    fabricante_filtro: str = "",
    avisos: Optional[List[str]] = None,
    contexto: str = "Congelado",
    obrigatorio: bool = True,
) -> pd.DataFrame:
    """
    Filtra o Congelado pelo fabricante informado, usando a mesma lógica do Sell-out:
    primeiro match exato normalizado e, como fallback, contém o texto digitado.

    Esta função deve ser aplicada antes de criar qualquer mapa SKU/EAN -> Categoria,
    para evitar misturar categorias de outros fabricantes e reduzir o volume processado.
    """
    if avisos is None:
        avisos = []
    fabricante_filtro = str(fabricante_filtro or "").strip()
    if raw is None or raw.empty or not fabricante_filtro:
        return raw

    c_fab = localizar_coluna_fabricante_congelado(raw)
    if not c_fab:
        msg = (
            f"{contexto}: foi informado o fabricante '{fabricante_filtro}', mas não encontrei coluna de Fabricante no Congelado.\n"
            "Colunas esperadas: Fabricante SKU, Fabricante, Proveedor ou Fornecedor.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(raw.columns)}"
        )
        if obrigatorio:
            raise ValueError(msg)
        avisos.append(msg)
        return raw

    antes = len(raw)
    fab_key = normalizar_texto(fabricante_filtro)
    serie_fab = raw[c_fab].fillna("").astype(str).str.strip()
    serie_key = serie_fab.map(normalizar_texto)
    mask_exata = serie_key == fab_key
    mask_contem = serie_key.str.contains(re.escape(fab_key), na=False) if fab_key else mask_exata
    filtrado = raw[mask_exata | mask_contem].copy()

    avisos.append(
        f"{contexto}: Congelado filtrado por fabricante '{fabricante_filtro}' usando a coluna '{c_fab}'. "
        f"Exato: {int(mask_exata.sum())}; contém: {int(mask_contem.sum())}; linhas mantidas: {len(filtrado)} de {antes}.".replace(",", ".")
    )

    if filtrado.empty:
        exemplos = []
        vistos = set()
        for v in serie_fab.head(20000).tolist():
            v = str(v).strip()
            k = normalizar_texto(v)
            if v and k and k not in vistos:
                vistos.add(k)
                exemplos.append(v)
            if len(exemplos) >= 12:
                break
        detalhe = f" Exemplos encontrados no Congelado: {'; '.join(exemplos)}" if exemplos else ""
        msg = (
            f"{contexto}: nenhum registro do Congelado ficou após filtrar pelo fabricante '{fabricante_filtro}'."
            f"{detalhe}"
        )
        if obrigatorio:
            raise ValueError(msg)
        avisos.append(msg)
    return filtrado

def _score_congelado_raw(raw: pd.DataFrame, alternativas_cat: List[str]) -> Tuple[int, List[str], Optional[str]]:
    if raw is None or raw.empty:
        return 0, [], None
    raw = remover_linhas_colunas_vazias(limpar_bom_dataframe(raw))
    if raw is None or raw.empty:
        return 0, [], None
    cols_codigo = localizar_colunas_codigo_congelado(raw)
    col_cat = localizar_coluna_categoria_congelado(raw, alternativas_cat)
    if not cols_codigo or not col_cat:
        return 0, cols_codigo, col_cat
    mask_codigo = pd.Series(False, index=raw.index)
    for c in cols_codigo:
        mask_codigo = mask_codigo | raw[c].map(ean_texto).ne("")
    mask_cat = raw[col_cat].astype(str).str.strip().map(normalizar_categoria).ne("")
    score = int((mask_codigo & mask_cat).sum())
    return score, cols_codigo, col_cat


def _df_com_header_em_linha(raw_sem_header: pd.DataFrame, header_idx: int) -> pd.DataFrame:
    valores = raw_sem_header.iloc[header_idx].tolist()
    ultimo = len(valores)
    while ultimo > 0 and str(valores[ultimo - 1]).strip() in {"", "nan", "None"}:
        ultimo -= 1
    valores = valores[:ultimo]
    if not valores:
        return pd.DataFrame()
    headers = nomes_unicos(valores)
    dados = raw_sem_header.iloc[header_idx + 1:, :ultimo].copy()
    dados.columns = headers
    return remover_linhas_colunas_vazias(limpar_bom_dataframe(dados))


def ler_congelado_flexivel(caminho: Path, alternativas_cat: List[str], avisos: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Lê o Congelado com cabeçalho flexível.

    Aceita cabeçalho na primeira linha ou em linhas abaixo, abas com nomes variados,
    e o layout Código Barras SKU + Categoría congelada ScannMarket.
    """
    if avisos is None:
        avisos = []
    caminho = Path(caminho)
    candidatos = []

    def avaliar(df: pd.DataFrame, descricao: str):
        try:
            df_limpo = remover_linhas_colunas_vazias(limpar_bom_dataframe(df))
            score, cols_codigo, col_cat = _score_congelado_raw(df_limpo, alternativas_cat)
            if score > 0:
                candidatos.append((score, descricao, df_limpo, cols_codigo, col_cat))
        except Exception:
            pass

    if eh_csv(caminho):
        try:
            raw_primeira = ler_tabela_primeira_linha(caminho, ["Base Congelada", "Congelado", "Fabricante", "Dados", "Planilha1", "Sheet1"])
            avaliar(raw_primeira, "CSV com cabeçalho na primeira linha")
        except Exception:
            pass
        raw_sem_header = ler_csv_flexivel(caminho)
        for idx in range(min(150, len(raw_sem_header))):
            avaliar(_df_com_header_em_linha(raw_sem_header, idx), f"CSV com cabeçalho na linha {idx + 1}")
    else:
        wb = load_workbook(caminho, read_only=True, data_only=True)
        nomes_abas = wb.sheetnames
        wb.close()
        preferidas_norm = [normalizar_texto(x) for x in ["Base Congelada", "Congelado", "Fabricante", "Dados", "Planilha1", "Sheet1"]]
        abas_ordenadas = sorted(nomes_abas, key=lambda n: 0 if any(p and p in normalizar_texto(n) for p in preferidas_norm) else 1)
        for aba in abas_ordenadas:
            try:
                avaliar(pd.read_excel(caminho, sheet_name=aba, dtype=str, engine="openpyxl"), f"aba '{aba}' com cabeçalho na primeira linha")
            except Exception:
                pass
            try:
                raw_sem_header = pd.read_excel(caminho, sheet_name=aba, header=None, dtype=str, engine="openpyxl")
                raw_sem_header = limpar_bom_dataframe(raw_sem_header)
                for idx in range(min(150, len(raw_sem_header))):
                    avaliar(_df_com_header_em_linha(raw_sem_header, idx), f"aba '{aba}' com cabeçalho na linha {idx + 1}")
            except Exception:
                pass

    if not candidatos:
        try:
            raw = ler_tabela_primeira_linha(caminho, ["Base Congelada", "Congelado", "Fabricante", "Dados", "Planilha1", "Sheet1"])
            return remover_linhas_colunas_vazias(limpar_bom_dataframe(raw))
        except Exception:
            return pd.DataFrame()

    candidatos.sort(key=lambda x: x[0], reverse=True)
    score, descricao, df_escolhido, cols_codigo, col_cat = candidatos[0]
    avisos.append(
        f"Congelado lido por detecção flexível: {descricao}; linhas úteis com código + categoria: {score}; "
        f"coluna(s) de código: {', '.join(cols_codigo)}; coluna de categoria: {col_cat}."
    )
    return df_escolhido

def ler_mapa_congelado_categoria(
    congelado_path: str | Path,
    nivel_prod: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Lê o Congelado opcional do Estudo de Cobertura e cria um mapa SKU/EAN -> Categoria/PROD.

    Quando este mapa é informado no Estudo de Cobertura, ele passa a ser a fonte oficial
    de categoria dos SKUs. O Sell-in e o Sell-out deixam de depender da categoria original
    de cada base e passam a procurar o SKU no Congelado.
    """
    if avisos is None:
        avisos = []

    caminho = Path(str(congelado_path or "").strip())
    if not str(caminho):
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {caminho}")

    alternativas_cat, origem = _alternativas_categoria_congelado(nivel_prod)
    raw = ler_congelado_flexivel(caminho, alternativas_cat, avisos)
    if raw is None or raw.empty:
        avisos.append(f"Congelado opcional '{caminho.name}' foi lido, mas está vazio ou sem cabeçalho útil.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    # Quando houver fabricante selecionado, o Congelado também deve ser filtrado
    # antes de montar o mapa SKU/EAN -> Categoria. Isso evita misturar SKUs de
    # outros fabricantes e reduz bastante o volume processado.
    raw = filtrar_congelado_por_fabricante(
        raw,
        fabricante_filtro=fabricante_filtro,
        avisos=avisos,
        contexto="Congelado opcional do Estudo de Cobertura",
        obrigatorio=bool(str(fabricante_filtro or "").strip()),
    )

    if raw is None or raw.empty:
        avisos.append(
            f"Congelado opcional '{caminho.name}' ficou vazio após o filtro de fabricante."
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    # Localização reforçada para o layout informado:
    # Código Barras SKU + Categoría congelada ScannMarket.
    colunas_codigo = localizar_colunas_codigo_congelado(raw)
    c_cat = localizar_coluna_categoria_congelado(raw, alternativas_cat)
    c_vol = localizar_coluna(raw, ["Cant Vta", "Quantidade Venda", "Volume", "Vendas em volume", "Qtd_de_Vendas", "Qtd de Vendas"], obrigatoria=False)
    c_val = localizar_coluna(raw, ["Imp Vta (Ult.24 Meses)", "Imp Vta Ult 24 Meses", "Imp Vta", "Valor", "Vendas em valor"], obrigatoria=False)

    if not colunas_codigo:
        raise ValueError(
            "No Congelado opcional, não encontrei coluna de SKU/EAN/Código de Barras preenchida.\n"
            "Para o seu layout, a coluna esperada é 'Código Barras SKU'.\n"
            f"Alternativas aceitas: {', '.join(EAN_SKU_ALTERNATIVAS_AMPLAS)}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(raw.columns)}"
        )
    if not c_cat:
        raise ValueError(
            f"No Congelado opcional, não encontrei a coluna de categoria para a regra {normalizar_nivel_prod(nivel_prod)}.\n"
            "Para o seu layout, a coluna esperada é 'Categoría congelada ScannMarket'.\n"
            f"Alternativas testadas: {', '.join(alternativas_cat)}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(raw.columns)}"
        )

    # O Congelado pode ter mais de uma coluna de código.
    # O mapa é criado com TODAS as colunas de código encontradas,
    # priorizando Código Barras SKU quando existir.
    partes_codigo = []
    for c_cod in colunas_codigo:
        temp = pd.DataFrame({
            "ean": raw[c_cod].map(ean_texto),
            "categoria_map_prod": raw[c_cat].astype(str).str.strip(),
            "_rank_congelado": (
                (limpar_coluna_numerica_vetorizada(raw[c_vol]) if c_vol else 0.0)
                + (limpar_coluna_numerica_vetorizada(raw[c_val]) if c_val else 0.0) / 1_000_000_000
            ),
            "_coluna_codigo_congelado": c_cod,
        })
        partes_codigo.append(temp)

    base = pd.concat(partes_codigo, ignore_index=True) if partes_codigo else pd.DataFrame(columns=["ean", "categoria_map_prod", "_rank_congelado", "_coluna_codigo_congelado"])
    base["categoria_key_map_prod"] = base["categoria_map_prod"].map(normalizar_categoria)
    base = base[(base["ean"] != "") & (base["categoria_key_map_prod"] != "")].copy()

    if base.empty:
        avisos.append(
            f"Congelado opcional '{caminho.name}' foi lido, mas nenhum SKU ficou com categoria válida em '{c_cat}'."
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    mapa = (
        base.sort_values(["ean", "_rank_congelado"], ascending=[True, False])
        .drop_duplicates("ean", keep="first")
        [["ean", "categoria_map_prod", "categoria_key_map_prod"]]
    )

    colunas_codigo_txt = ", ".join(colunas_codigo) if colunas_codigo else "não identificada"
    avisos.append(
        f"Congelado opcional aplicado no Estudo de Cobertura: {len(mapa):,} códigos de SKU/EAN mapeados usando {origem} "
        f"pela coluna de categoria '{c_cat}' e coluna(s) de código {colunas_codigo_txt}.".replace(",", ".")
    )
    return mapa, avisos




# ============================================================
# Modo Cobertura Dash
# ============================================================


def alternativas_sellout_dash_quantia() -> List[str]:
    return [
        "Vendas Medida.", "Vendas Medida", "VENDAS MEDIDA", "Medida", "Vendas",
        "SELL OUT (QUANTIDADE)", "SELL-OUT (QUANTIDADE)", "Sell-out (Quantidade)", "Sell out (Quantidade)",
        "SELL OUT QUANTIDADE", "SELL-OUT QUANTIDADE", "Sell-out Quantidade", "Sell out Quantidade",
        "Sell-out Quantia", "Sell out Quantia", "Quantidade Sell-out", "Quantidade Sell out",
        "Qtd Sell-out", "Qtd Sell out", "Qtd_de_Vendas", "Qtd de Vendas", "Quantidade de Vendas",
        "Sell-out", "Sell out", "Sellout",
    ]


def alternativas_sellout_dash_volume() -> List[str]:
    return [
        "Vendas Medida.", "Vendas Medida", "VENDAS MEDIDA", "Medida", "Vendas",
        "SELL OUT (Kg / L)", "SELL-OUT (Kg / L)", "SELL OUT (KG/L)", "SELL-OUT (KG/L)",
        "SELL OUT KG L", "SELL-OUT KG L", "Sell-out (Kg / L)", "Sell out (Kg / L)",
        "Sell-out Kg/L", "Sell out Kg/L", "Sell-out Volume", "Sell out Volume",
        "Volume Sell-out", "Volume Sell out", "Vendas_em_volume", "Vendas em volume",
        "Volume", "Kg/L", "Kg L", "Sell-out", "Sell out", "Sellout",
    ]


def coluna_metrica_sellout_dash(metrica: str) -> List[str]:
    metrica_norm = normalizar_texto(metrica)
    if metrica_norm.startswith("q"):
        return alternativas_sellout_dash_quantia()
    if metrica_eh_volume_variavel(metrica):
        # Volume variável precisa partir de quantidade quando possível; se só houver volume, usa volume.
        return alternativas_sellout_dash_volume() + alternativas_sellout_dash_quantia()
    return alternativas_sellout_dash_volume()


def ler_tabela_primeira_linha(arquivo: Path, abas_preferidas: Optional[List[str]] = None) -> pd.DataFrame:
    """Lê uma tabela simples cujo cabeçalho está na primeira linha, aceitando Excel e CSV."""
    arquivo = Path(arquivo)
    if eh_csv(arquivo):
        enc, sep, _sample = detectar_csv_formato(arquivo, nrows=80)
        engine = "python" if sep is None else "c"
        kwargs = dict(
            filepath_or_buffer=arquivo,
            sep=sep,
            engine=engine,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=False,
            on_bad_lines="skip",
            encoding=enc,
        )
        if engine == "c":
            kwargs["low_memory"] = False
        return limpar_bom_dataframe(remover_linhas_colunas_vazias(pd.read_csv(**kwargs)))

    aba = escolher_aba(arquivo, abas_preferidas or ["Dados", "SKU", "Fabricante", "Base", "Planilha1", "Sheet1"])
    df = pd.read_excel(arquivo, sheet_name=aba, dtype=str, engine="openpyxl")
    return limpar_bom_dataframe(remover_linhas_colunas_vazias(df))




def arquivo_excel_zip_valido(arquivo: Path) -> bool:
    """Retorna True quando o arquivo parece um XLSX/XLSM real, e não um CSV renomeado."""
    arquivo = Path(arquivo)
    try:
        with open(arquivo, "rb") as f:
            inicio = f.read(8)
    except Exception:
        return False
    # XLSX/XLSM são arquivos zip e começam com PK.
    return inicio.startswith(b"PK")


def arquivo_excel_ole_valido(arquivo: Path) -> bool:
    """Retorna True para XLS antigo real. Esse caso usa pandas/openpyxl fallback quando possível."""
    arquivo = Path(arquivo)
    try:
        with open(arquivo, "rb") as f:
            inicio = f.read(8)
    except Exception:
        return False
    return inicio.startswith(b"\xd0\xcf\x11\xe0")


def detectar_encoding_texto_por_bom(arquivo: Path) -> Optional[str]:
    """Detecta encoding pelo BOM antes de tentar heurística geral."""
    try:
        with open(arquivo, "rb") as f:
            inicio = f.read(4)
    except Exception:
        return None
    if inicio.startswith(b"\xff\xfe") or inicio.startswith(b"\xfe\xff"):
        return "utf-16"
    if inicio.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    return None


def detectar_texto_sep_rapido(arquivo: Path) -> Tuple[str, Optional[str]]:
    """
    Detecta encoding e separador de arquivo texto, incluindo CSV/TXT/TSV UTF-16.

    Essa função é usada principalmente para Congelado do Dash, porque alguns
    arquivos chegam como texto/CSV mesmo quando a extensão não deixa isso claro.
    """
    arquivo = Path(arquivo)
    enc_preferido = detectar_encoding_texto_por_bom(arquivo)
    encodings = [enc_preferido] if enc_preferido else []
    encodings += [e for e in ["utf-8-sig", "utf-8", "cp1252", "latin1", "iso-8859-1", "utf-16"] if e not in encodings]

    melhor = None
    for enc in encodings:
        try:
            with open(arquivo, "r", encoding=enc, errors="replace", newline="") as f:
                amostra = f.read(20000)
            if not amostra:
                continue
            penalidade = amostra.count(" ") + amostra.count("\x00") + amostra.count("ÿþ") + amostra.count("þÿ")
            linhas = [l for l in amostra.splitlines()[:30] if l.strip()]
            if not linhas:
                continue
            contagens = {sep: sum(l.count(sep) for l in linhas[:10]) for sep in ["\t", ";", ",", "|"]}
            sep, qtd_sep = max(contagens.items(), key=lambda x: x[1])
            # Pontuação: separadores e texto legível valem mais que encoding com lixo.
            score = (qtd_sep, -penalidade, 1 if enc_preferido and enc == enc_preferido else 0)
            if melhor is None or score > melhor[0]:
                melhor = (score, enc, sep if qtd_sep > 0 else None)
        except Exception:
            continue

    if melhor is None:
        return "utf-8-sig", None
    return melhor[1], melhor[2]


def _normalizar_header_congelado(valor) -> str:
    return normalizar_texto(str(valor).replace("\ufeff", "").replace("\ufffe", "").replace("\x00", "").replace("ÿþ", "").replace("þÿ", ""))


def _indice_por_prioridade(headers: List[str], alternativas: List[str]) -> Optional[int]:
    """Localiza uma coluna por nome, com prioridade e sem olhar os dados."""
    headers_norm = [_normalizar_header_congelado(h) for h in headers]
    alts_norm = [_normalizar_header_congelado(a) for a in alternativas if str(a).strip()]

    for alt in alts_norm:
        for i, h in enumerate(headers_norm):
            if h == alt:
                return i

    for alt in alts_norm:
        if not alt:
            continue
        for i, h in enumerate(headers_norm):
            if alt in h or h in alt:
                return i
    return None


def _indices_codigo_congelado_por_headers(headers: List[str]) -> List[int]:
    """
    Localiza as colunas de código do Congelado sem criar DataFrame.

    Prioridade principal:
    1) Código Barras SKU
    2) CODIGO_BARRAS_CONTENIDO
    3) outros EAN/SKU compatíveis
    """
    prioridades = [
        ("Código Barras SKU", 1000),
        ("Codigo Barras SKU", 1000),
        ("Código_Barras_SKU", 1000),
        ("Codigo_Barras_SKU", 1000),
        ("CODIGO_BARRAS_SKU", 990),
        ("CODIGO BARRAS SKU", 990),
        ("Código de Barras SKU", 980),
        ("Codigo de Barras SKU", 980),
        ("CODIGO_BARRAS_CONTENIDO", 950),
        ("CODIGO BARRAS CONTENIDO", 950),
        ("Código Barras Contenido", 950),
        ("Codigo Barras Contenido", 950),
        ("EAN", 850),
        ("SKU/EAN", 840),
        ("EAN/SKU", 840),
        ("EAN SKU", 830),
        ("SKU EAN", 830),
        ("SKU", 500),
    ]
    headers_norm = [_normalizar_header_congelado(h) for h in headers]
    candidatos = []
    termos_ruins = {"nome", "nombre", "marca", "fabricante", "categoria", "descricao", "descripcion", "conteudo", "contenido", "qtd", "cant", "valor", "volume", "vta"}

    for i, h in enumerate(headers_norm):
        if not h:
            continue
        score = None
        for alt, prioridade in prioridades:
            alt_norm = _normalizar_header_congelado(alt)
            if h == alt_norm:
                score = prioridade if score is None else max(score, prioridade)
            elif alt_norm in h or h in alt_norm:
                score = max(score or 0, prioridade - 80)
        tokens = set(h.split())
        if score is None:
            if (("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens) and "sku" in tokens):
                score = 920
            elif (("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens)):
                score = 850
            elif "ean" in tokens and not (tokens & termos_ruins):
                score = 760
            elif h == "sku":
                score = 500
        if score is None:
            continue
        if tokens & termos_ruins and not any(x in h for x in ["codigo barras sku", "codigo de barras sku", "codigo barras contenido"]):
            score -= 400
        if score > 0:
            candidatos.append((score, i))

    candidatos.sort(key=lambda x: x[0], reverse=True)
    saida = []
    for _, i in candidatos:
        if i not in saida:
            saida.append(i)
    return saida


def _selecionar_indices_congelado_dash(headers: List[str]) -> Dict[str, object]:
    """Seleciona somente as colunas que o modo Dash realmente usa do Congelado."""
    idxs_codigo = _indices_codigo_congelado_por_headers(headers)
    idx_cat = _indice_por_prioridade(headers, [
        "Categoría congelada ScannMarket",
        "Categoria congelada ScannMarket",
        "Categoria atual Data Excellence",
        "Est Mer 6 (Categoria)",
        "Est Mer 6 Categoria",
        "Categoria",
    ])
    idx_estmer7 = _indice_por_prioridade(headers, [
        "Est Mer 7 (Subcategoria)",
        "Est Mer 7 Subcategoria",
        "Est Mer 7",
        "Subcategoria",
        "Subcategoria ScannMarket",
    ])
    idx_fab = _indice_por_prioridade(headers, [
        "Fabricante SKU", "Fabricante", "Fabricante do SKU", "Fabricante del SKU", "Proveedor", "Fornecedor",
    ])
    idx_nome = _indice_por_prioridade(headers, ["Nome SKU", "PROD_NOMBRE_ORIGINAL", "NOMBRE_SKU", "Descrição", "Descricao"])
    idx_marca = _indice_por_prioridade(headers, ["Marca SKU", "Marca"])
    idx_qtd = _indice_por_prioridade(headers, ["Qtd Conteúdo SKU", "Qtd Conteudo SKU", "UNIDADES_CONTENIDO"])
    idx_vol = _indice_por_prioridade(headers, ["Cant Vta", "Quantidade Venda", "Volume", "Vendas em volume"])
    idx_val = _indice_por_prioridade(headers, ["Imp Vta (Ult.24 Meses)", "Imp Vta Ult 24 Meses", "Imp Vta", "Valor"])

    return {
        "codigo": idxs_codigo,
        "categoria": idx_cat,
        "estmer7": idx_estmer7,
        "fabricante": idx_fab,
        "nome": idx_nome,
        "marca": idx_marca,
        "qtd": idx_qtd,
        "volume": idx_vol,
        "valor": idx_val,
    }


def _headers_congelado_validos(headers: List[str]) -> bool:
    sel = _selecionar_indices_congelado_dash(headers)
    return bool(sel.get("codigo")) and sel.get("categoria") is not None


def _mensagem_erro_colunas_congelado_dash(headers: List[str]) -> str:
    comentario_colunas_esperadas = (
        "Colunas conhecidas do Congelado: Nome SKU; Código Barras SKU; Qtd Conteúdo SKU; "
        "Est Mer Codigo; Marca SKU; Fabricante SKU; Categoría congelada ScannMarket; "
        "Categoria atual Data Excellence; Est Mer 6 (Categoria); Est Mer 7 (Subcategoria); "
        "CODIGO_BARRAS_CONTENIDO; UNIDADES_CONTENIDO; Imp Vta (Ult.24 Meses); Cant Vta."
    )
    return (
        "No arquivo Congelado, não encontrei as colunas mínimas para o modo Dash.\n"
        "Mínimo necessário: 'Código Barras SKU' ou 'CODIGO_BARRAS_CONTENIDO' + "
        "'Categoría congelada ScannMarket'.\n"
        f"{comentario_colunas_esperadas}\n"
        f"Colunas disponíveis identificadas: {resumo_colunas_disponiveis(headers)}"
    )


def _linha_valor(row, idx: Optional[int], default: str = ""):
    if idx is None:
        return default
    try:
        valor = row[idx]
    except Exception:
        return default
    if valor is None:
        return default
    return str(valor).strip()


def _fabricante_linha_aceito(valor: str, fabricante_filtro: str) -> bool:
    filtro = normalizar_texto(fabricante_filtro)
    if not filtro:
        return True
    atual = normalizar_texto(valor)
    return atual == filtro or filtro in atual or atual in filtro


def _montar_fab_dataframe_congelado_dash(linhas: List[Dict[str, object]], avisos: List[str], origem: str) -> pd.DataFrame:
    colunas = ["ean", "nome_sku_fab", "marca_fab", "fabricante_fab", "categoria_fab", "estmer7_fab", "volume_fab", "valor_fab"]
    if not linhas:
        avisos.append(f"Cobertura Dash: Congelado lido por {origem}, mas nenhum EAN válido ficou após filtros.")
        return pd.DataFrame(columns=colunas)
    fab = pd.DataFrame(linhas)
    for c in colunas:
        if c not in fab.columns:
            fab[c] = 0.0 if c in {"volume_fab", "valor_fab"} else ""
    fab = fab[colunas]
    fab = fab[fab["ean"].map(ean_texto) != ""].copy()
    fab["ean"] = fab["ean"].map(ean_texto)
    avisos.append(f"Cobertura Dash: Congelado lido por {origem}; linhas úteis com EAN: {len(fab):,}.".replace(",", "."))
    return fab


def ler_congelado_dash_otimizado(
    fabricante_path: str | Path,
    avisos: List[str],
    fabricante_filtro: str = "",
) -> pd.DataFrame:
    """
    Lê o Congelado do modo Cobertura Dash sem carregar a planilha inteira.

    Regras:
    - aceita XLSX real e CSV/TXT/TSV, inclusive UTF-16;
    - também trata arquivo texto renomeado como XLS/XLSX;
    - procura somente as colunas usadas no Dash;
    - filtra Fabricante durante a leitura;
    - não faz fallback para leitura pesada do arquivo inteiro.
    """
    arquivo = Path(fabricante_path)
    if not arquivo.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {arquivo}")

    filtro_norm = normalizar_texto(fabricante_filtro)

    # XLSX real: leitura em streaming com openpyxl.
    if arquivo_excel_zip_valido(arquivo):
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        try:
            abas_preferidas = ["Fabricante", "Base Congelada", "Congelado", "Dados", "Planilha1", "Sheet1"]
            abas = sorted(
                wb.sheetnames,
                key=lambda n: 0 if normalizar_texto(n) in {normalizar_texto(a) for a in abas_preferidas} else 1,
            )
            melhor = None
            for aba in abas:
                ws = wb[aba]
                for linha_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    headers = [str(v).strip() if v is not None else "" for v in row]
                    if _headers_congelado_validos(headers):
                        melhor = (aba, linha_idx, headers, _selecionar_indices_congelado_dash(headers))
                        break
                    if linha_idx >= 80:
                        break
                if melhor:
                    break
            if not melhor:
                primeira = wb[wb.sheetnames[0]]
                headers_amostra = []
                for row in primeira.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers_amostra = [str(v).strip() if v is not None else "" for v in row]
                    break
                raise ValueError(_mensagem_erro_colunas_congelado_dash(headers_amostra))

            aba, header_row, headers, sel = melhor
            idxs_codigo = list(sel.get("codigo") or [])
            idx_fab = sel.get("fabricante")
            if filtro_norm and idx_fab is None:
                raise ValueError(
                    f"Congelado auxiliar do modo Dash: foi informado o fabricante '{fabricante_filtro}', "
                    "mas não encontrei a coluna 'Fabricante SKU' no Congelado."
                )

            indices_uteis = set(idxs_codigo)
            for chave in ["categoria", "estmer7", "fabricante", "nome", "marca", "qtd", "volume", "valor"]:
                idx = sel.get(chave)
                if idx is not None:
                    indices_uteis.add(idx)
            max_idx = max(indices_uteis) if indices_uteis else 0

            linhas = []
            vazias_consecutivas = 0
            ws = wb[aba]
            for row in ws.iter_rows(min_row=header_row + 1, max_col=max_idx + 1, values_only=True):
                codigos = [ean_texto(_linha_valor(row, idx)) for idx in idxs_codigo]
                ean = next((c for c in codigos if c), "")
                categoria = _linha_valor(row, sel.get("categoria"))
                fabricante_linha = _linha_valor(row, idx_fab)

                if not ean and not categoria and not fabricante_linha:
                    vazias_consecutivas += 1
                    if vazias_consecutivas >= 200:
                        break
                    continue
                vazias_consecutivas = 0

                if filtro_norm and not _fabricante_linha_aceito(fabricante_linha, fabricante_filtro):
                    continue
                if not ean:
                    continue

                linhas.append({
                    "ean": ean,
                    "nome_sku_fab": _linha_valor(row, sel.get("nome")),
                    "marca_fab": _linha_valor(row, sel.get("marca")),
                    "fabricante_fab": fabricante_linha,
                    "categoria_fab": categoria,
                    "estmer7_fab": _linha_valor(row, sel.get("estmer7")),
                    "volume_fab": numero_brasil(_linha_valor(row, sel.get("volume"))),
                    "valor_fab": numero_brasil(_linha_valor(row, sel.get("valor"))),
                })
        finally:
            wb.close()
        return _montar_fab_dataframe_congelado_dash(linhas, avisos, f"streaming XLSX/aba '{aba}'")

    # Arquivo texto: CSV/TXT/TSV ou export UTF-16 renomeado.
    enc, sep = detectar_texto_sep_rapido(arquivo)
    try:
        sample = pd.read_csv(
            arquivo,
            sep=sep,
            engine="python",
            header=None,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=False,
            on_bad_lines="skip",
            encoding=enc,
            nrows=100,
        )
        sample = limpar_bom_dataframe(sample)
    except Exception as exc:
        raise ValueError(f"Não consegui ler a amostra do Congelado como texto/CSV. Encoding={enc}, sep={repr(sep)}.\n{exc}")

    header_idx = None
    headers = None
    for i in range(len(sample)):
        linha = [str(v).strip() if v is not None else "" for v in sample.iloc[i].tolist()]
        if _headers_congelado_validos(linha):
            header_idx = i
            headers = linha
            break
    if header_idx is None or headers is None:
        primeira = [str(v).strip() if v is not None else "" for v in sample.iloc[0].tolist()] if len(sample) else []
        raise ValueError(_mensagem_erro_colunas_congelado_dash(primeira))

    sel = _selecionar_indices_congelado_dash(headers)
    idxs_codigo = list(sel.get("codigo") or [])
    idx_fab = sel.get("fabricante")
    if filtro_norm and idx_fab is None:
        raise ValueError(
            f"Congelado auxiliar do modo Dash: foi informado o fabricante '{fabricante_filtro}', "
            "mas não encontrei a coluna 'Fabricante SKU' no Congelado."
        )

    indices_uteis = set(idxs_codigo)
    for chave in ["categoria", "estmer7", "fabricante", "nome", "marca", "qtd", "volume", "valor"]:
        idx = sel.get(chave)
        if idx is not None:
            indices_uteis.add(idx)
    usecols = sorted(indices_uteis)
    nomes_usados = [f"COL_{i}" for i in usecols]
    idx_para_nome = {idx: f"COL_{idx}" for idx in usecols}

    linhas = []
    leitor = pd.read_csv(
        arquivo,
        sep=sep,
        engine="python",
        header=None,
        skiprows=header_idx + 1,
        usecols=usecols,
        names=nomes_usados,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
        on_bad_lines="skip",
        encoding=enc,
        chunksize=100_000,
    )

    for chunk in leitor:
        chunk = limpar_bom_dataframe(chunk)
        for col in chunk.columns:
            chunk[col] = chunk[col].astype(str).str.strip()

        codigo_series = []
        for idx in idxs_codigo:
            col = idx_para_nome.get(idx)
            if col in chunk.columns:
                codigo_series.append(chunk[col].map(ean_texto))
        if not codigo_series:
            continue
        ean_final = codigo_series[0]
        for serie in codigo_series[1:]:
            ean_final = ean_final.where(ean_final != "", serie)

        fab_col = idx_para_nome.get(idx_fab) if idx_fab is not None else None
        if filtro_norm and fab_col in chunk.columns:
            mask_fab = chunk[fab_col].map(lambda x: _fabricante_linha_aceito(x, fabricante_filtro))
        else:
            mask_fab = pd.Series(True, index=chunk.index)

        cat_col = idx_para_nome.get(sel.get("categoria")) if sel.get("categoria") is not None else None
        est_col = idx_para_nome.get(sel.get("estmer7")) if sel.get("estmer7") is not None else None
        nome_col = idx_para_nome.get(sel.get("nome")) if sel.get("nome") is not None else None
        marca_col = idx_para_nome.get(sel.get("marca")) if sel.get("marca") is not None else None
        vol_col = idx_para_nome.get(sel.get("volume")) if sel.get("volume") is not None else None
        val_col = idx_para_nome.get(sel.get("valor")) if sel.get("valor") is not None else None

        temp = pd.DataFrame({
            "ean": ean_final,
            "nome_sku_fab": chunk[nome_col] if nome_col in chunk.columns else "",
            "marca_fab": chunk[marca_col] if marca_col in chunk.columns else "",
            "fabricante_fab": chunk[fab_col] if fab_col in chunk.columns else "",
            "categoria_fab": chunk[cat_col] if cat_col in chunk.columns else "",
            "estmer7_fab": chunk[est_col] if est_col in chunk.columns else "",
            "volume_fab": limpar_coluna_numerica_vetorizada(chunk[vol_col]) if vol_col in chunk.columns else 0.0,
            "valor_fab": limpar_coluna_numerica_vetorizada(chunk[val_col]) if val_col in chunk.columns else 0.0,
        })
        temp = temp[(temp["ean"].map(ean_texto) != "") & mask_fab].copy()
        if not temp.empty:
            linhas.extend(temp.to_dict("records"))

    return _montar_fab_dataframe_congelado_dash(linhas, avisos, f"CSV/texto otimizado encoding={enc}, sep={repr(sep)}")

def preparar_auxiliares_dash(
    sku_path: str | Path,
    fabricante_path: str | Path | None = None,
    avisos: List[str] | None = None,
    fabricante_filtro: str = "",
) -> pd.DataFrame:
    """
    Lê os arquivos auxiliares do modo Cobertura Dash.

    SKU: traz Código Barras, nome do SKU, Categoria Scanntech, NIVEL1, NIVEL2, vendas em volume/valor.
    Congelado opcional: traz Código Barras SKU/CODIGO_BARRAS_CONTENIDO, Marca SKU, Fabricante SKU e categorias congeladas.
    """
    if avisos is None:
        avisos = []

    sku_path = Path(sku_path)
    fabricante_path = Path(fabricante_path) if fabricante_path else None
    if not sku_path.exists():
        raise FileNotFoundError(f"Arquivo SKU não encontrado: {sku_path}")

    sku_raw = ler_tabela_primeira_linha(sku_path, ["SKU", "Dados", "Planilha1", "Sheet1"])
    # O Congelado do Dash é opcional. Quando informado, ele complementa Marca,
    # Fabricante, categoria congelada e Est Mer 7. Quando não informado, o Dash
    # segue com o arquivo SKU como fonte principal de categoria/PROD.
    if fabricante_path and fabricante_path.exists():
        fab = ler_congelado_dash_otimizado(fabricante_path, avisos, fabricante_filtro=fabricante_filtro)
    else:
        fab = pd.DataFrame(columns=[
            "ean", "nome_sku_fab", "marca_fab", "fabricante_fab",
            "categoria_fab", "estmer7_fab", "volume_fab", "valor_fab",
        ])
        avisos.append(
            "Cobertura Dash: Congelado não informado. "
            "O mapeamento seguirá pelo arquivo SKU; Marca, Fabricante e Est Mer 7 podem ficar vazios."
        )

    c_sku_ean = localizar_coluna_ean_sku(sku_raw, obrigatoria=False)
    c_sku_nome = localizar_coluna(sku_raw, ["PROD_NOMBRE_ORIGINAL", "Nome SKU", "NOMBRE_SKU", "Descrição", "Descricao"], obrigatoria=False)
    c_sku_cat = localizar_coluna(sku_raw, ["Categoria Scanntech", "Categoria SM", "CATEGORIA SCANN", "Categoria", "PROD"], obrigatoria=False)
    c_sku_n1 = localizar_coluna(sku_raw, ["NIVEL1", "Nível 1", "Nivel 1", "N1"], obrigatoria=False)
    c_sku_n2 = localizar_coluna(sku_raw, ["NIVEL2", "Nível 2", "Nivel 2", "N2"], obrigatoria=False)
    c_sku_qtd = localizar_coluna(sku_raw, ["Qtd Conteúdo SKU", "Qtd Conteudo SKU", "Qtd_Conteudo", "Qtd Conteúdo", "Qtd Conteudo"], obrigatoria=False)
    c_sku_vlr = localizar_coluna(sku_raw, ["Vendas em valor R$", "Vendas em valor", "Valor"], obrigatoria=False)
    c_sku_vol = localizar_coluna(sku_raw, ["Vendas em volume", "Vendas_em_volume", "Volume"], obrigatoria=False)

    if not c_sku_ean:
        raise ValueError(
            "No arquivo SKU, não encontrei coluna de Código Barras/EAN/SKU.\n"
            f"Alternativas aceitas: {', '.join(EAN_SKU_ALTERNATIVAS_AMPLAS)}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(sku_raw.columns)}"
        )

    sku = pd.DataFrame({
        "ean": sku_raw[c_sku_ean].map(ean_texto),
        "nome_sku_aux": sku_raw[c_sku_nome].astype(str).str.strip() if c_sku_nome else "",
        "categoria_aux": sku_raw[c_sku_cat].astype(str).str.strip() if c_sku_cat else "",
        "nivel1_aux": sku_raw[c_sku_n1].astype(str).str.strip() if c_sku_n1 else "",
        "nivel2_aux": sku_raw[c_sku_n2].astype(str).str.strip() if c_sku_n2 else "",
        "qtd_conteudo_aux": limpar_coluna_numerica_vetorizada(sku_raw[c_sku_qtd]) if c_sku_qtd else 0.0,
        "valor_aux": limpar_coluna_numerica_vetorizada(sku_raw[c_sku_vlr]) if c_sku_vlr else 0.0,
        "volume_aux": limpar_coluna_numerica_vetorizada(sku_raw[c_sku_vol]) if c_sku_vol else 0.0,
    })
    sku = sku[sku["ean"] != ""].copy()
    if sku.empty:
        avisos.append("Arquivo SKU foi lido, mas não possui EAN/Código Barras válido.")

    if not sku.empty:
        sku_rank = sku.copy()
        sku_rank["_rank_aux"] = pd.to_numeric(sku_rank["volume_aux"], errors="coerce").fillna(0) + pd.to_numeric(sku_rank["valor_aux"], errors="coerce").fillna(0) / 1_000_000_000
        sku_rank = sku_rank.sort_values(["ean", "_rank_aux"], ascending=[True, False]).drop_duplicates("ean", keep="first")
    else:
        sku_rank = pd.DataFrame(columns=["ean", "nome_sku_aux", "categoria_aux", "nivel1_aux", "nivel2_aux", "qtd_conteudo_aux", "valor_aux", "volume_aux"])

    if not fab.empty:
        fab_rank = fab.copy()
        fab_rank["_rank_fab"] = pd.to_numeric(fab_rank["volume_fab"], errors="coerce").fillna(0) + pd.to_numeric(fab_rank["valor_fab"], errors="coerce").fillna(0) / 1_000_000_000
        fab_rank = fab_rank.sort_values(["ean", "_rank_fab"], ascending=[True, False]).drop_duplicates("ean", keep="first")
    else:
        fab_rank = pd.DataFrame(columns=["ean", "nome_sku_fab", "marca_fab", "fabricante_fab", "categoria_fab", "estmer7_fab", "volume_fab", "valor_fab"])

    aux = sku_rank.merge(fab_rank, on="ean", how="outer")
    for col in ["nome_sku_aux", "categoria_aux", "nivel1_aux", "nivel2_aux", "nome_sku_fab", "marca_fab", "fabricante_fab", "categoria_fab", "estmer7_fab"]:
        if col not in aux.columns:
            aux[col] = ""
        aux[col] = aux[col].fillna("").astype(str).str.strip()

    aux["nome_sku"] = aux["nome_sku_aux"].where(aux["nome_sku_aux"].map(normalizar_texto) != "", aux["nome_sku_fab"])

    # Cobertura Dash - de/para oficial:
    # - A categoria operacional deve vir primeiro do arquivo SKU (Categoria Scanntech),
    #   pois é ela que deve classificar os EANs do Sell-in e do Vendas SKU da mesma forma.
    # - A categoria do Congelado fica como double check/fallback.
    # - NIVEL1/NIVEL2 vêm do SKU.
    # - Est Mer 7 vem do Congelado como opção adicional de abertura.
    aux["categoria_congelado"] = aux["categoria_fab"].fillna("").astype(str).str.strip()
    aux["categoria_sku"] = aux["categoria_aux"].fillna("").astype(str).str.strip()
    aux["categoria_aux_final"] = aux["categoria_sku"].where(
        aux["categoria_sku"].map(normalizar_texto) != "",
        aux["categoria_congelado"],
    )
    aux["categoria_aux_origem"] = np.where(
        aux["categoria_sku"].map(normalizar_texto) != "",
        "SKU - Categoria Scanntech",
        np.where(aux["categoria_congelado"].map(normalizar_texto) != "", "Congelado - categoria", "Não encontrada"),
    )
    aux["estmer7_aux"] = aux["estmer7_fab"].fillna("").astype(str).str.strip()
    aux["marca"] = aux["marca_fab"]
    aux["fabricante"] = aux["fabricante_fab"]
    aux["peso_gramas_aux"] = aux["nome_sku"].map(extrair_peso_gramas)
    if "qtd_conteudo_aux" in aux.columns:
        aux["peso_gramas_aux"] = aux["peso_gramas_aux"].where(aux["peso_gramas_aux"].fillna(0) > 0, pd.to_numeric(aux["qtd_conteudo_aux"], errors="coerce"))

    qtd_cat_congelado = int(aux["categoria_congelado"].map(normalizar_texto).ne("").sum()) if "categoria_congelado" in aux.columns else 0
    qtd_cat_sku = int(aux["categoria_sku"].map(normalizar_texto).ne("").sum()) if "categoria_sku" in aux.columns else 0
    avisos.append(
        f"Cobertura Dash: auxiliares lidos com cabeçalho na primeira linha. "
        f"SKU: {len(sku):,} linhas válidas; Congelado: {len(fab):,} linhas válidas; EANs auxiliares únicos: {aux['ean'].nunique():,}; "
        f"categorias encontradas no Congelado: {qtd_cat_congelado:,}; categorias encontradas no SKU: {qtd_cat_sku:,}.".replace(",", ".")
    )
    return aux


def montar_mapa_categoria_aux_dash(aux_dash: pd.DataFrame, nivel_prod: str, avisos: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Monta mapa EAN -> categoria/PROD a partir dos arquivos auxiliares do modo Dash.

    Esse mapa é usado principalmente para classificar o Sell-in do mesmo jeito
    que o Dash, sem depender de o EAN aparecer em Vendas SKU.
    """
    if avisos is None:
        avisos = []
    if aux_dash is None or aux_dash.empty or "ean" not in aux_dash.columns:
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])

    nivel = normalizar_nivel_prod(nivel_prod)
    base = aux_dash.copy()
    for col in ["ean", "categoria_aux_final", "categoria_sku", "categoria_congelado", "nivel1_aux", "nivel2_aux", "estmer7_aux", "volume_fab", "volume_aux", "valor_fab", "valor_aux"]:
        if col not in base.columns:
            base[col] = 0.0 if col in {"volume_fab", "volume_aux", "valor_fab", "valor_aux"} else ""

    if nivel == "NIVEL1":
        cat = base["nivel1_aux"]
        origem = "NIVEL1 do arquivo SKU"
    elif nivel == "NIVEL2":
        cat = base["nivel2_aux"]
        origem = "NIVEL2 do arquivo SKU"
    elif nivel == "ESTMER7":
        cat = base["estmer7_aux"]
        origem = "Est Mer 7 do Congelado"
    else:
        cat = base["categoria_aux_final"]
        origem = "Categoria Scanntech do arquivo SKU, com fallback no Congelado"

    base["categoria_map_prod"] = cat.fillna("").astype(str).str.strip()
    base["categoria_key_map_prod"] = base["categoria_map_prod"].map(normalizar_categoria)
    base["_rank_aux_dash"] = (
        pd.to_numeric(base.get("volume_fab", 0), errors="coerce").fillna(0)
        + pd.to_numeric(base.get("volume_aux", 0), errors="coerce").fillna(0)
        + pd.to_numeric(base.get("valor_fab", 0), errors="coerce").fillna(0) / 1_000_000_000
        + pd.to_numeric(base.get("valor_aux", 0), errors="coerce").fillna(0) / 1_000_000_000
    )
    mapa = (
        base[(base["ean"].map(ean_texto) != "") & (base["categoria_key_map_prod"] != "")]
        .assign(ean=lambda d: d["ean"].map(ean_texto))
        .sort_values(["ean", "_rank_aux_dash"], ascending=[True, False])
        .drop_duplicates("ean", keep="first")
        [["ean", "categoria_map_prod", "categoria_key_map_prod"]]
    )
    avisos.append(f"Cobertura Dash: mapa auxiliar SKU -> categoria/PROD criado por {origem}; EANs mapeados: {len(mapa):,}.".replace(",", "."))
    return mapa


def extrair_ean_nome_sku_misto(valor) -> Tuple[str, str]:
    """
    Lê textos no padrão "123456789 - NOME DO SKU 100g".
    Retorna o EAN/código antes do hífen quando existir e o nome do SKU sem o código.
    """
    if pd.isna(valor):
        return "", ""
    texto = str(valor).strip()
    if not texto:
        return "", ""

    m = re.match(r"^\s*(\d{6,14})\s*[-–—]\s*(.+?)\s*$", texto)
    if m:
        return ean_texto(m.group(1)), m.group(2).strip()

    m = re.search(r"(?<!\d)(\d{6,14})(?!\d)", texto)
    if m:
        ean = ean_texto(m.group(1))
        nome = (texto[:m.start()] + " " + texto[m.end():]).strip(" -–—\t")
        return ean, nome.strip()

    return ean_texto(texto), texto


def coluna_parece_ean_nome_sku(nome_coluna: Optional[str]) -> bool:
    n = normalizar_texto(nome_coluna or "")
    return bool(("ean" in n and "nome" in n) or ("ean" in n and "sku" in n and "+" in str(nome_coluna or "")))


def distribuir_vendas_uf_por_vendas_sku(vendas_uf: pd.DataFrame, vendas_sku: pd.DataFrame, avisos: List[str]) -> pd.DataFrame:
    """
    Modo Cobertura Dash: quando existem dois exports do Dash:
    - Vendas UF: mês/UF/valor, sem SKU/categoria;
    - Vendas SKU: mês/EAN+nome/valor, com categoria/PROD via arquivo SKU/Congelado.

    Para permitir cobertura por categoria/PROD e também por UF, distribui o total UF/mês
    conforme a participação mensal de cada categoria/PROD observada em Vendas SKU.
    """
    if vendas_uf is None or vendas_uf.empty:
        avisos.append("Cobertura Dash: Vendas UF não possui dados válidos; usando Vendas SKU como fonte principal.")
        return vendas_sku.copy()
    if vendas_sku is None or vendas_sku.empty:
        avisos.append("Cobertura Dash: Vendas SKU não foi informado ou não possui dados válidos; usando Vendas UF como total disponível.")
        return vendas_uf.copy()

    sku = vendas_sku.copy()
    uf = vendas_uf.copy()
    if "categoria_key" not in sku.columns or sku["categoria_key"].map(normalizar_categoria).eq("").all():
        avisos.append("Cobertura Dash: Vendas SKU não trouxe categoria/PROD; usando Vendas UF sem distribuição por SKU.")
        return uf

    # Se Vendas UF já vier com categoria/PROD real, não redistribui.
    if "categoria_key" in uf.columns:
        cats_uf = {normalizar_categoria(x) for x in uf["categoria_key"].dropna().astype(str).unique() if normalizar_categoria(x)}
        cats_total = {normalizar_categoria("Total Dash"), normalizar_categoria("Total Sell-out"), normalizar_categoria("Total disponível")}
        if len(cats_uf - cats_total) > 0:
            avisos.append("Cobertura Dash: Vendas UF já possui categoria/PROD; distribuição por Vendas SKU não foi necessária.")
            return uf

    def add_periodo_key(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        if "mes" in out.columns and out["mes"].notna().any():
            out["_periodo_dash"] = out["mes"].map(lambda x: pd.Timestamp(x).strftime("%Y-%m") if pd.notna(x) else "SEM_MES")
        elif "ano" in out.columns and pd.to_numeric(out["ano"], errors="coerce").notna().any():
            out["_periodo_dash"] = pd.to_numeric(out["ano"], errors="coerce").fillna(0).astype(int).astype(str)
        else:
            out["_periodo_dash"] = "TOTAL"
        return out

    sku = add_periodo_key(sku)
    uf = add_periodo_key(uf)

    texto_cols = ["categoria", "categoria_key", "fabricante", "fabricante_key", "marca", "canal"]
    for c in texto_cols:
        if c not in sku.columns:
            sku[c] = ""
        sku[c] = sku[c].fillna("").astype(str).str.strip()
    if "valor_sellout" not in sku.columns or "valor_sellout" not in uf.columns:
        return uf

    cat = (
        sku.groupby(["_periodo_dash", "categoria_key", "categoria"], dropna=False, as_index=False)
        .agg(
            valor_categoria_sku=("valor_sellout", "sum"),
            fabricante=("fabricante", lambda s: next((str(x).strip() for x in s if str(x).strip()), "")),
            fabricante_key=("fabricante_key", lambda s: next((str(x).strip() for x in s if str(x).strip()), "")),
            marca=("marca", lambda s: next((str(x).strip() for x in s if str(x).strip()), "")),
            canal=("canal", lambda s: next((str(x).strip() for x in s if str(x).strip()), "")),
        )
    )
    total_periodo = cat.groupby("_periodo_dash", as_index=False)["valor_categoria_sku"].sum().rename(columns={"valor_categoria_sku": "total_sku_periodo"})
    cat = cat.merge(total_periodo, on="_periodo_dash", how="left")
    cat["participacao_categoria"] = cat["valor_categoria_sku"] / cat["total_sku_periodo"].replace(0, np.nan)
    cat = cat[cat["participacao_categoria"].fillna(0) > 0].copy()

    if cat.empty:
        avisos.append("Cobertura Dash: não consegui calcular participação de categoria em Vendas SKU; usando Vendas UF sem distribuição.")
        return uf

    if "uf" not in uf.columns:
        uf["uf"] = "TOTAL"
    uf_base = (
        uf.groupby(["_periodo_dash", "uf"], dropna=False, as_index=False)
        .agg(
            valor_uf=("valor_sellout", "sum"),
            mes=("mes", "first") if "mes" in uf.columns else ("_periodo_dash", "first"),
            ano=("ano", "first") if "ano" in uf.columns else ("_periodo_dash", "first"),
        )
    )
    dist = uf_base.merge(cat, on="_periodo_dash", how="left")
    dist["valor_sellout"] = pd.to_numeric(dist["valor_uf"], errors="coerce").fillna(0) * pd.to_numeric(dist["participacao_categoria"], errors="coerce").fillna(0)
    dist["ean"] = ""
    dist["nome_sku"] = ""
    dist = dist[["uf", "mes", "ano", "ean", "categoria", "categoria_key", "fabricante", "fabricante_key", "marca", "nome_sku", "canal", "valor_sellout"]].copy()
    dist = dist[(dist["categoria_key"].map(normalizar_categoria) != "") & (dist["valor_sellout"].fillna(0) != 0)].copy()

    avisos.append(
        "Cobertura Dash: Vendas UF foi distribuído por categoria/PROD usando a participação mensal de Vendas SKU. "
        "Use essa leitura como estimativa quando Vendas UF não trouxer SKU/categoria diretamente."
    )
    return dist


def pontuar_linha_cabecalho_sellout_dash(valores: List, metrica: str) -> int:
    valores_norm = [normalizar_texto(v) for v in valores if v is not None and str(v).strip() != ""]
    if not valores_norm:
        return 0
    grupos = {
        "uf": ["uf sm", "uf", "estado"],
        "mes": ["ano mes", "mês", "mes", "data", "periodo", "dia de selector data", "selector data"],
        "categoria": ["categoria sm", "categoria scann", "categoria", "prod", "nivel1", "nivel2"],
        "sku": ["ean", "sku", "ean nome sku", "ean nome", "ean sku", "codigo de barras", "código de barras", "cod barras"],
        "valor": coluna_metrica_sellout_dash(metrica) + ["SELL OUT", "Sell-out", "Sell out", "Sellout"],
    }

    def bate(alt: str, col: str) -> bool:
        alt_n = normalizar_texto(alt)
        return bool(alt_n and col and (alt_n == col or alt_n in col or col in alt_n))

    score = 0
    tem_valor = False
    for grupo, alternativas in grupos.items():
        encontrou = any(bate(alt, col) for alt in alternativas for col in valores_norm)
        if encontrou:
            score += 1
            if grupo == "valor":
                tem_valor = True
    if not tem_valor or score < 2:
        return 0
    return score


def ler_sellout_dash_com_cabecalho_flexivel(arquivo: Path, metrica: str) -> Tuple[pd.DataFrame, int, str]:
    arquivo = Path(arquivo)
    if eh_csv(arquivo):
        enc, sep, sample = detectar_csv_formato(arquivo, nrows=250)
        melhor = None
        for idx in range(1, min(250, len(sample)) + 1):
            valores = sample.iloc[idx - 1, :].tolist()
            valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
            score = pontuar_linha_cabecalho_sellout_dash(valores_recortados, metrica)
            if score:
                bonus = 2 if idx in (14, 15) else 0
                candidato = (score + bonus, idx, valores_recortados, inicio, fim)
                if melhor is None or candidato[0] > melhor[0]:
                    melhor = candidato
        if melhor is None:
            raise ValueError(
                f"Não localizei o cabeçalho do Sell-out Dash no CSV '{arquivo.name}'. "
                "Procurei uma linha com coluna de Sell-out e pelo menos UF, Mês, EAN/SKU ou Categoria."
            )
        _, header_row_idx, header_values, header_start, header_end = melhor
        headers = nomes_unicos(header_values)
        usecols = list(range(header_start, header_end))
        engine = "python" if sep is None else "c"
        kwargs = dict(
            filepath_or_buffer=arquivo,
            sep=sep,
            engine=engine,
            header=None,
            skiprows=header_row_idx,
            usecols=usecols,
            names=headers,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=False,
            on_bad_lines="skip",
            encoding=enc,
        )
        if engine == "c":
            kwargs["low_memory"] = False
        return remover_linhas_colunas_vazias(limpar_bom_dataframe(pd.read_csv(**kwargs))), header_row_idx, "CSV"

    aba = escolher_aba(arquivo, ["Sell-out", "Sellout", "Tabela Sell Out", "Tabela Sell-out", "Tabela Sell In", "Tabela Sell-in", "Sell-in", "Dash", "Dados", "Planilha1", "Sheet1"])
    wb = load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb[aba]
    melhor = None
    limite = min(250, ws.max_row)
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), start=1):
        valores = list(row)
        valores_recortados, inicio, fim = recortar_linha_por_cabecalho(valores)
        score = pontuar_linha_cabecalho_sellout_dash(valores_recortados, metrica)
        if score:
            bonus = 3 if idx in (14, 15) else 0
            candidato = (score + bonus, idx, valores_recortados, inicio, fim)
            if melhor is None or candidato[0] > melhor[0]:
                melhor = candidato
    if melhor is None:
        wb.close()
        raise ValueError(
            f"Não localizei o cabeçalho do Sell-out Dash na aba '{aba}' do arquivo '{arquivo.name}'. "
            "Procurei uma linha com coluna de Sell-out e pelo menos UF, Mês, EAN/SKU ou Categoria."
        )
    _, header_row_idx, header_values, header_start, header_end = melhor
    headers = nomes_unicos(header_values)
    dados = []
    max_col_exclusive = header_end if header_end is not None else ws.max_column
    for row in ws.iter_rows(
        min_row=header_row_idx + 1,
        min_col=header_start + 1,
        max_col=max_col_exclusive,
        values_only=True,
    ):
        linha = list(row)
        if len(linha) < len(headers):
            linha += [None] * (len(headers) - len(linha))
        linha = linha[:len(headers)]
        if any(v is not None and str(v).strip() != "" for v in linha):
            dados.append(linha)
    wb.close()
    return pd.DataFrame(dados, columns=headers), header_row_idx, aba


def ler_sellout_dash(
    arquivo: Path,
    metrica: str,
    nivel_prod: str,
    aux_dash: pd.DataFrame,
    sellin_para_categoria: Optional[pd.DataFrame] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Lê o Sell-out no formato Dash/template e enriquece com arquivos SKU/Fabricante."""
    avisos: List[str] = []
    nivel_prod = normalizar_nivel_prod(nivel_prod)
    raw, linha, aba = ler_sellout_dash_com_cabecalho_flexivel(Path(arquivo), metrica)
    raw = otimizar_colunas_categoricas_sellout(raw)

    c_uf = localizar_coluna(raw, ["UF SM", "UF_SM", "UF Scann", "UF Scanntech", "UF", "Estado"], obrigatoria=False)
    c_mes = localizar_coluna(raw, COLUNAS_MES_ALTERNATIVAS, obrigatoria=False)
    c_ano = localizar_coluna(raw, ["Ano", "ANO", "Year", "YEAR"], obrigatoria=False)
    if c_ano == c_mes:
        c_ano = None
    c_ean = localizar_coluna(raw, ["EAN + Nome Sku", "EAN + Nome SKU", "EAN Nome Sku", "EAN Nome SKU", "EAN", "SKU", "EAN/SKU", "SKU/EAN", "Código Barras", "Codigo Barras", "Código de Barras", "Codigo de Barras", "Cod Barras", "Cód Barras", "CODIGO_BARRAS", "COD_BARRAS"], obrigatoria=False)
    c_cat = localizar_coluna(raw, ["Categoria SM", "CATEGORIA SM", "CATEGORIA SCANN", "Categoria Scann", "Categoria Scanntech", "Categoria", "CATEGORIA", "PROD"], obrigatoria=False)
    c_fab_dash = localizar_coluna(raw, ["Fabricante", "Fabricante SKU", "FABRICANTE"], obrigatoria=False)
    c_marca_dash = localizar_coluna(raw, ["Marca", "Marca SKU"], obrigatoria=False)
    c_nome_dash = localizar_coluna(raw, ["NOMBRE_SKU", "Nome SKU", "PROD_NOMBRE_ORIGINAL", "Descrição", "Descricao"], obrigatoria=False)
    c_canal = localizar_coluna(raw, ["CANAL", "Canal", "PDV_CANAL", "PDV Canal"], obrigatoria=False)
    c_val = localizar_coluna_obrigatoria_com_erro(
        raw,
        coluna_metrica_sellout_dash(metrica),
        f"valor de Sell-out Dash para a métrica {metrica}",
        "Sell-out Dash",
        arquivo,
    )

    c_qtd_var = localizar_coluna(raw, alternativas_sellout_dash_quantia(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None
    c_volume_var = localizar_coluna(raw, alternativas_sellout_dash_volume(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None

    registrar_colunas_reconhecidas(avisos, "Sell-out Dash", {
        "UF": c_uf,
        "Mês/Data": c_mes,
        "Ano": c_ano,
        "SKU/EAN": c_ean,
        "Categoria": c_cat,
        "Fabricante no Dash": c_fab_dash,
        "Marca no Dash": c_marca_dash,
        "Nome SKU no Dash": c_nome_dash,
        "Canal": c_canal,
        "Valor Sell-out Dash": c_val,
    })

    n = len(raw)
    if c_ean:
        if coluna_parece_ean_nome_sku(c_ean):
            pares_ean_nome = raw[c_ean].map(extrair_ean_nome_sku_misto)
            ean = pares_ean_nome.map(lambda x: x[0])
            nome_misto_ean = pares_ean_nome.map(lambda x: x[1])
        else:
            ean = raw[c_ean].map(ean_texto)
            nome_misto_ean = pd.Series([""] * n, index=raw.index)
    else:
        ean = pd.Series([""] * n, index=raw.index)
        nome_misto_ean = pd.Series([""] * n, index=raw.index)
    mes = raw[c_mes].map(converter_mes) if c_mes else pd.Series([pd.NaT] * n, index=raw.index)
    ano = raw[c_ano].map(converter_ano) if c_ano else mes.map(lambda x: x.year if pd.notna(x) else np.nan)

    df = pd.DataFrame({
        "uf": raw[c_uf].astype(str).str.strip() if c_uf else pd.Series(["TOTAL"] * n, index=raw.index),
        "mes": mes,
        "ano": ano,
        "ean": ean,
        "categoria_dash": raw[c_cat].astype(str).str.strip() if c_cat else pd.Series([""] * n, index=raw.index),
        "valor_sellout": limpar_coluna_numerica_vetorizada(raw[c_val]),
        "fabricante_dash": raw[c_fab_dash].astype(str).str.strip() if c_fab_dash else pd.Series([""] * n, index=raw.index),
        "marca_dash": raw[c_marca_dash].astype(str).str.strip() if c_marca_dash else pd.Series([""] * n, index=raw.index),
        "nome_sku_dash": raw[c_nome_dash].astype(str).str.strip() if c_nome_dash else nome_misto_ean,
        "canal": raw[c_canal].astype(str).str.strip() if c_canal else pd.Series([""] * n, index=raw.index),
    })
    df["uf"] = df["uf"].astype(str).str.strip().replace("", "TOTAL")
    df["ano"] = pd.to_numeric(df["ano"], errors="coerce")
    df.loc[df["mes"].notna(), "ano"] = df.loc[df["mes"].notna(), "mes"].dt.year

    aux_cols = [
        "ean", "nome_sku", "categoria_aux_final", "categoria_congelado", "categoria_sku", "categoria_aux_origem",
        "nivel1_aux", "nivel2_aux", "estmer7_aux", "marca", "fabricante", "peso_gramas_aux", "volume_aux", "valor_aux"
    ]
    for col in aux_cols:
        if col not in aux_dash.columns:
            aux_dash[col] = "" if col not in {"peso_gramas_aux", "volume_aux", "valor_aux"} else 0.0
    df = df.merge(aux_dash[aux_cols], on="ean", how="left")

    for col in ["nome_sku", "categoria_aux_final", "categoria_congelado", "categoria_sku", "categoria_aux_origem", "nivel1_aux", "nivel2_aux", "estmer7_aux", "marca", "fabricante"]:
        df[col] = df[col].fillna("").astype(str).str.strip()
    df["fabricante"] = df["fabricante"].where(df["fabricante"].map(normalizar_texto) != "", df["fabricante_dash"].fillna(""))
    df["marca"] = df["marca"].where(df["marca"].map(normalizar_texto) != "", df["marca_dash"].fillna(""))
    df["nome_sku"] = df["nome_sku"].where(df["nome_sku"].map(normalizar_texto) != "", df["nome_sku_dash"].fillna(""))

    if nivel_prod == "NIVEL1":
        categoria = df["nivel1_aux"]
        origem_categoria = "NIVEL1 do arquivo SKU"
    elif nivel_prod == "NIVEL2":
        categoria = df["nivel2_aux"]
        origem_categoria = "NIVEL2 do arquivo SKU"
    elif nivel_prod == "ESTMER7":
        categoria = df["estmer7_aux"]
        origem_categoria = "Est Mer 7 (Subcategoria) do arquivo Congelado"
    else:
        # Para CATEGORIA, a prioridade no modo Dash agora é:
        # 1) Categoria Scanntech do arquivo SKU por EAN/SKU;
        # 2) Categoria do Congelado por EAN/SKU como double check/fallback;
        # 3) Categoria vinda no próprio Dash, se existir;
        # 4) Total Dash.
        categoria = df["categoria_aux_final"].where(
            df["categoria_aux_final"].map(normalizar_categoria) != "",
            df["categoria_dash"],
        )
        origem_categoria = "Categoria Scanntech do arquivo SKU por EAN/SKU, com fallback no Congelado e depois no Dash"

    # Fallbacks para não zerar a base.
    categoria = categoria.fillna("").astype(str).str.strip()
    categoria = categoria.where(categoria.map(normalizar_categoria) != "", df["categoria_aux_final"])
    categoria = categoria.where(categoria.map(normalizar_categoria) != "", df["categoria_dash"])
    if categoria.map(normalizar_categoria).eq("").all():
        if fabricante_filtro:
            categoria = pd.Series([f"Total Fabricante - {fabricante_filtro}"] * len(df), index=df.index)
        else:
            categoria = pd.Series(["Total Dash"] * len(df), index=df.index)
        avisos.append("Cobertura Dash: não foi possível mapear categoria/PROD por SKU; criada visão Total Dash.")

    df["categoria"] = categoria
    df["categoria_key"] = df["categoria"].map(normalizar_categoria)
    df["fabricante"] = df["fabricante"].fillna("").astype(str).str.strip()
    df["fabricante_key"] = df["fabricante"].map(normalizar_texto)

    fabricante_filtro = str(fabricante_filtro or "").strip()
    if fabricante_filtro:
        fab_key = normalizar_texto(fabricante_filtro)
        antes = len(df)
        tem_fabricante_na_base = df["fabricante_key"].fillna("").map(normalizar_texto).ne("").any()
        if tem_fabricante_na_base:
            mask_exata = df["fabricante_key"] == fab_key
            mask_contem = df["fabricante_key"].str.contains(re.escape(fab_key), na=False) if fab_key else mask_exata
            df = df[mask_exata | mask_contem].copy()
            avisos.append(
                f"Cobertura Dash: base filtrada por fabricante '{fabricante_filtro}'. Linhas mantidas: {len(df)} de {antes}."
            )
        else:
            avisos.append(
                f"Cobertura Dash: fabricante '{fabricante_filtro}' foi informado, mas esta base não possui Fabricante por linha; "
                "o filtro não foi aplicado nesta leitura para não zerar os dados. O Fabricante será registrado nos parâmetros."
            )

    if metrica_eh_volume_variavel(metrica):
        df["qtd_sellout"] = limpar_coluna_numerica_vetorizada(raw[c_qtd_var]) if c_qtd_var and c_qtd_var in raw.columns else df["valor_sellout"]
        df["peso_gramas_sku"] = pd.to_numeric(df["peso_gramas_aux"], errors="coerce")
        faltou_peso = df["peso_gramas_sku"].isna() | (df["peso_gramas_sku"].fillna(0) <= 0)
        df.loc[faltou_peso, "peso_gramas_sku"] = df.loc[faltou_peso, "nome_sku"].map(extrair_peso_gramas)
        val_norm = normalizar_texto(c_val)
        origem = "volume" if any(normalizar_texto(x) == val_norm for x in alternativas_sellout_dash_volume()) else "quantidade"
        df["valor_sellout_origem_volume_variavel"] = origem

    linhas_antes = len(df)
    df = df[(df["categoria_key"] != "") & (df["valor_sellout"].fillna(0) != 0)].copy()
    if linhas_antes and df.empty:
        avisos.append("Cobertura Dash: Sell-out Dash ficou sem linhas válidas após filtros de categoria/valor/fabricante.")

    group_cols = ["uf", "mes", "ano", "ean", "categoria", "categoria_key", "fabricante", "fabricante_key", "marca", "nome_sku", "canal"]
    if metrica_eh_volume_variavel(metrica):
        df["_peso_x_qtd_volume_variavel"] = pd.to_numeric(df.get("peso_gramas_sku", np.nan), errors="coerce") * pd.to_numeric(df.get("qtd_sellout", 0), errors="coerce").fillna(0)
        df = (
            df.groupby(group_cols, dropna=False, as_index=False)
            .agg(
                valor_sellout=("valor_sellout", "sum"),
                qtd_sellout=("qtd_sellout", "sum"),
                _peso_x_qtd_volume_variavel=("_peso_x_qtd_volume_variavel", "sum"),
            )
        )
        df["peso_gramas_sku"] = df["_peso_x_qtd_volume_variavel"] / df["qtd_sellout"].replace(0, np.nan)
        if "valor_sellout_origem_volume_variavel" not in df.columns:
            df["valor_sellout_origem_volume_variavel"] = "volume"
    else:
        df = df.groupby(group_cols, dropna=False, as_index=False)["valor_sellout"].sum()

    if not df.empty and df["ean"].map(normalizar_texto).ne("").any():
        mapa = (
            df[df["ean"] != ""].groupby(["ean", "categoria_key", "categoria"], as_index=False)["valor_sellout"]
            .sum()
            .sort_values(["ean", "valor_sellout"], ascending=[True, False])
            .drop_duplicates("ean", keep="first")
            [["ean", "categoria", "categoria_key"]]
            .rename(columns={"categoria": "categoria_map_prod", "categoria_key": "categoria_key_map_prod"})
        )
    else:
        mapa = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
        avisos.append("Cobertura Dash: não foi possível criar mapa SKU -> categoria/PROD pelo Sell-out Dash.")

    avisos.append(
        f"Sell-out Dash lido da aba '{aba}', cabeçalho na linha {linha}, usando {origem_categoria} e '{c_val}' como métrica. "
        f"Linhas válidas agregadas: {len(df):,}.".replace(",", ".")
    )
    return df, mapa, avisos


def converter_vendas_dash_para_referencia(df: pd.DataFrame, nome_referencia: str = "Vendas UF") -> pd.DataFrame:
    """
    Converte uma base do modo Dash, lida como valor_sellout, para o lado de referência
    usado nos cálculos antigos como Sell-in.

    No modo Cobertura Dash sem Sell-in, a referência passa a ser o total do Dash
    por UF/mês/categoria, normalmente vindo de Vendas UF.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "uf", "mes", "ano", "ean", "categoria", "categoria_key",
            "fabricante", "fabricante_key", "marca", "nome_sku", "canal",
            "valor_sellin", "categoria_original_sellin", "fabricante_sellin"
        ])
    ref = df.copy()
    if "valor_sellin" not in ref.columns:
        ref["valor_sellin"] = pd.to_numeric(ref.get("valor_sellout", 0), errors="coerce").fillna(0)
    for col in ["uf", "mes", "ano", "ean", "categoria", "categoria_key", "fabricante", "fabricante_key", "marca", "nome_sku", "canal"]:
        if col not in ref.columns:
            ref[col] = ""
    ref["categoria_original_sellin"] = ref.get("categoria", "").fillna("").astype(str).str.strip()
    ref["fabricante_sellin"] = ref.get("fabricante", "").fillna("").astype(str).str.strip()
    ref["origem_referencia_dash"] = nome_referencia
    ref = ref[ref["valor_sellin"].fillna(0) != 0].copy()
    return ref


def normalizar_nivel_prod(nivel_prod: str) -> str:
    nivel = normalizar_texto(nivel_prod).upper().strip()
    if nivel in {"CATEGORIA", "CATEGORIA SM", "CAT"}:
        return "CATEGORIA"
    if nivel in {"NIVEL 1", "NIVEL1", "N1"}:
        return "NIVEL1"
    if nivel in {"NIVEL 2", "NIVEL2", "N2"}:
        return "NIVEL2"
    if nivel in {"EST MER 7", "ESTMER7", "EST_MER_7", "SUBCATEGORIA", "EST MER 7 SUBCATEGORIA"}:
        return "ESTMER7"
    raise ValueError("Regra de categoria inválida. Use CATEGORIA, NIVEL1, NIVEL2 ou ESTMER7.")


def colunas_existentes_por_alternativa(headers: List[str], alternativas: List[str], obrigatoria: bool = False) -> Optional[str]:
    """Versão leve de localizar_coluna usando só lista de headers."""
    return localizar_coluna(pd.DataFrame(columns=headers), alternativas, obrigatoria=obrigatoria)


def ler_sellout_csv_em_blocos(
    arquivo: Path,
    metrica: str,
    nivel_prod: str,
    sellin_para_categoria: Optional[pd.DataFrame] = None,
    fabricante_filtro: str = "",
    progress_callback=None,
    mapa_categoria_forcado: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Lê Sell-out CSV em blocos, usando apenas as colunas necessárias.

    Essa função evita carregar CSVs grandes inteiros na memória. É especialmente útil para VTA
    com milhões de linhas, como arquivos Publicar mensais/acumulados.
    """
    avisos: List[str] = []
    arquivo = Path(arquivo)
    nivel_prod = normalizar_nivel_prod(nivel_prod)
    usar_mapa_categoria_forcado = mapa_categoria_forcado is not None and not mapa_categoria_forcado.empty

    def progresso(msg: str, pct: Optional[float] = None):
        if progress_callback:
            try:
                progress_callback(msg, pct)
            except TypeError:
                progress_callback(msg)

    obrigatorias = ["SKU"] if metrica_eh_volume_variavel(metrica) else coluna_metrica_sellout(metrica)[:1]
    if nivel_prod in {"NIVEL1", "NIVEL2"} and not usar_mapa_categoria_forcado:
        obrigatorias.append(nivel_prod)

    enc, sep, header_row_idx, headers, header_start, header_end = preparar_leitura_csv_com_cabecalho(
        arquivo=arquivo,
        obrigatorias=obrigatorias,
        max_linhas_scan=200,
    )

    c_uf = colunas_existentes_por_alternativa(headers, ["UF", "Estado"], obrigatoria=False)
    c_mes = colunas_existentes_por_alternativa(headers, ["DATA", "Data", "ANO MÊS", "ANO MES", "Mês", "Mes", "PERIODO", "Periodo"], obrigatoria=False)
    c_ano = colunas_existentes_por_alternativa(headers, ["Ano", "ANO", "Year", "YEAR"], obrigatoria=False)
    if c_ano == c_mes:
        c_ano = None
    c_sku = colunas_existentes_por_alternativa(
        headers,
        [
            "SKU", "EAN", "EAN/SKU", "SKU/EAN",
            "Código Barras", "Codigo Barras", "Código de Barras", "Codigo de Barras",
            "Cod Barras", "Cód Barras", "CODIGO_BARRAS", "COD_BARRAS"
        ],
        obrigatoria=False,
    )
    c_qtd_var = colunas_existentes_por_alternativa(headers, alternativas_sellout_quantia(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None
    c_volume_var = colunas_existentes_por_alternativa(headers, alternativas_sellout_volume(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None
    alternativas_valor_sellout = coluna_metrica_sellout(metrica)
    if metrica_eh_volume_variavel(metrica) and c_volume_var:
        c_val = c_volume_var
        origem_valor_volume_variavel = "volume"
    elif metrica_eh_volume_variavel(metrica) and c_qtd_var:
        c_val = c_qtd_var
        origem_valor_volume_variavel = "quantidade"
    else:
        c_val = localizar_coluna_obrigatoria_com_erro(
            pd.DataFrame(columns=headers),
            alternativas_valor_sellout,
            f"valor de Sell-out para a métrica {metrica}",
            "Sell-out CSV",
            arquivo,
        )
        origem_valor_volume_variavel = "volume" if normalizar_texto(c_val) in {normalizar_texto(x) for x in alternativas_sellout_volume()} else "quantidade"
    c_fab = colunas_existentes_por_alternativa(headers, ["Fabricante", "Fabricante SKU", "FABRICANTE"], obrigatoria=False)
    if str(fabricante_filtro or "").strip() and not c_fab:
        raise ValueError("Você informou um fabricante para filtrar, mas não encontrei a coluna Fabricante no Sell-out.")
    c_marca = colunas_existentes_por_alternativa(headers, ["Marca", "Marca SKU"], obrigatoria=False)
    c_nome = colunas_existentes_por_alternativa(headers, ["NOMBRE_SKU", "Nome SKU", "Descrição", "Descricao", "PROD_NOMBRE_ORIGINAL"], obrigatoria=False)
    c_canal = colunas_existentes_por_alternativa(headers, ["PDV_CANAL", "Canal", "PDV Canal"], obrigatoria=False)

    mapa_si = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
    if nivel_prod == "CATEGORIA" and not usar_mapa_categoria_forcado:
        mapa_si = mapa_categoria_sellin_por_sku(sellin_para_categoria if sellin_para_categoria is not None else pd.DataFrame())
    origem_categoria = ""
    c_cat = None
    usar_total_categoria = False
    dict_cat_forcado = {}
    if usar_mapa_categoria_forcado and c_sku:
        mapa_forcado = mapa_categoria_forcado.copy()
        mapa_forcado["ean"] = mapa_forcado["ean"].map(ean_texto)
        mapa_forcado = mapa_forcado[(mapa_forcado["ean"] != "") & (mapa_forcado["categoria_key_map_prod"].map(normalizar_categoria) != "")].copy()
        dict_cat_forcado = mapa_forcado.drop_duplicates("ean", keep="first").set_index("ean")["categoria_map_prod"].to_dict()
        origem_categoria = "Congelado opcional mapeado por SKU"
    elif usar_mapa_categoria_forcado and not c_sku:
        avisos.append("Congelado opcional informado, mas o Sell-out CSV não possui SKU/EAN. A categoria do Congelado não pôde ser aplicada ao Sell-out.")

    if dict_cat_forcado:
        c_cat = None
    elif nivel_prod == "CATEGORIA":
        c_cat = colunas_existentes_por_alternativa(
            headers,
            [
                "Categoria SM", "CATEGORIA SM", "CATEGORIA SCANN", "Categoria Scann",
                "Categoria Scanntech", "Categoria", "CATEGORIA",
                "Categoria Cliente", "Categoria Sell-out", "Categoria Sellout",
            ],
            obrigatoria=False,
        )
        if c_cat:
            origem_categoria = f"coluna '{c_cat}' do Sell-out"
            if not mapa_si.empty and c_sku:
                origem_categoria += " com fallback para Categoria do Sell-in por SKU"
        else:
            if not mapa_si.empty and c_sku:
                origem_categoria = "Categoria do Sell-in mapeada por SKU"
            else:
                usar_total_categoria = True
                origem_categoria = "total disponível"
                avisos.append("Sell-out sem coluna Categoria e sem mapeamento por SKU. Será usada visão total quando necessário.")
    else:
        c_cat = colunas_existentes_por_alternativa(headers, [nivel_prod], obrigatoria=False)
        if c_cat:
            origem_categoria = f"coluna '{nivel_prod}' do Sell-out"
        else:
            usar_total_categoria = True
            origem_categoria = f"{nivel_prod} não encontrado / total disponível"
            avisos.append(f"Sell-out sem coluna {nivel_prod}. Será usada visão total quando necessário.")

    registrar_colunas_reconhecidas(avisos, "Sell-out CSV", {
        "UF": c_uf,
        "Mês/Data": c_mes,
        "Ano": c_ano,
        "SKU/EAN": c_sku,
        "Categoria/PROD": c_cat,
        "Fabricante": c_fab,
        "Marca": c_marca,
        "Nome SKU": c_nome,
        "Canal": c_canal,
        "Valor Sell-out": c_val,
        "Qtd Sell-out para ponderação": c_qtd_var if metrica_eh_volume_variavel(metrica) else None,
        "Volume Sell-out para comparação": c_volume_var if metrica_eh_volume_variavel(metrica) else None,
    })
    registrar_colunas_nao_encontradas(avisos, "Sell-out CSV", [
        ("UF", c_uf, ["UF", "Estado"]),
        ("Mês/Data", c_mes, ["DATA", "ANO MÊS", "ANO_MES", "AAAAMM", "YYYYMM", "Mês/Ano", "Competência", "Referência", "Período"]),
        ("Ano", c_ano, ["Ano", "ANO", "Year"]),
        ("SKU/EAN", c_sku, ["SKU", "EAN", "EAN/SKU", "Código de Barras"]),
        ("Categoria/PROD", c_cat, ["Categoria", "CATEGORIA SCANN", "NIVEL1", "NIVEL2"]),
        ("Fabricante", c_fab, ["Fabricante", "Fabricante SKU"]),
        ("Marca", c_marca, ["Marca", "Marca SKU"]),
        ("Nome SKU", c_nome, ["NOMBRE_SKU", "Nome SKU", "Descrição"]),
        ("Canal", c_canal, ["PDV_CANAL", "Canal"]),
    ])

    colunas_necessarias = [
        c_uf, c_mes, c_ano, c_sku, c_val, c_qtd_var, c_volume_var, c_fab, c_marca, c_nome, c_canal, c_cat
    ]
    colunas_necessarias = [c for c in colunas_necessarias if c]
    # Remove duplicadas preservando ordem.
    colunas_necessarias = list(dict.fromkeys(colunas_necessarias))

    # Dicionários usados para mapear Categoria do Sell-in por SKU, sem merge pesado.
    dict_cat = {}
    if not mapa_si.empty:
        dict_cat = mapa_si.set_index("ean")["categoria_map_prod"].to_dict()

    fabricante_filtro = str(fabricante_filtro or "").strip()
    fab_key = normalizar_texto(fabricante_filtro) if fabricante_filtro else ""

    header_pos = {nome: header_start + i for i, nome in enumerate(headers)}
    pares_colunas = [(header_pos[c], c) for c in colunas_necessarias if c in header_pos]
    pares_colunas = sorted(set(pares_colunas), key=lambda x: x[0])
    usecols = [pos for pos, _ in pares_colunas]
    names = [nome for _, nome in pares_colunas]
    if not usecols or c_val not in names:
        raise ValueError(f"Não foi possível identificar as colunas necessárias no CSV '{arquivo.name}'.")

    engine = "python" if sep is None else "c"
    read_kwargs = dict(
        filepath_or_buffer=arquivo,
        sep=sep,
        engine=engine,
        header=None,
        skiprows=header_row_idx,
        usecols=usecols,
        names=names,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
        on_bad_lines="skip",
        encoding=enc,
        chunksize=200_000,
    )
    if engine == "c":
        read_kwargs["low_memory"] = False

    partes: List[pd.DataFrame] = []
    linhas_lidas = 0
    linhas_validas = 0
    linhas_apos_filtro_fab_exato = 0
    linhas_apos_filtro_fab_contem = 0
    amostras_fabricantes: List[str] = []
    vistos_fabricantes = set()

    progresso("Lendo Sell-out em blocos...", 20)
    for idx_chunk, chunk in enumerate(pd.read_csv(**read_kwargs), start=1):
        chunk = limpar_bom_dataframe(chunk)
        chunk = otimizar_colunas_categoricas_sellout(chunk)
        n = len(chunk)
        linhas_lidas += n

        def serie_col(col: Optional[str], default: str = "") -> pd.Series:
            if col and col in chunk.columns:
                return chunk[col]
            return pd.Series([default] * n, index=chunk.index)

        sku_serie = serie_col(c_sku).map(ean_texto) if c_sku else pd.Series([""] * n, index=chunk.index)

        if dict_cat_forcado:
            categoria_serie = sku_serie.map(dict_cat_forcado).fillna("")
        elif usar_total_categoria:
            if fabricante_filtro:
                categoria_serie = pd.Series([f"Total Fabricante - {fabricante_filtro}"] * n, index=chunk.index)
            else:
                categoria_serie = pd.Series(["Total Sell-out"] * n, index=chunk.index)
        elif nivel_prod == "CATEGORIA" and not c_cat and dict_cat:
            categoria_serie = sku_serie.map(dict_cat).fillna("")
        elif c_cat:
            categoria_serie = serie_col(c_cat).astype(str).str.strip()
            if nivel_prod == "CATEGORIA" and dict_cat:
                categoria_sellin = sku_serie.map(dict_cat).fillna("")
                categoria_serie = categoria_serie.where(categoria_serie.map(normalizar_categoria) != "", categoria_sellin)
        else:
            categoria_serie = pd.Series([""] * n, index=chunk.index)

        mes = serie_col(c_mes).map(converter_mes) if c_mes else pd.Series([pd.NaT] * n, index=chunk.index)
        if c_ano:
            ano = serie_col(c_ano).map(converter_ano)
        else:
            ano = mes.map(lambda x: x.year if pd.notna(x) else np.nan)

        out = pd.DataFrame({
            "uf": serie_col(c_uf, "TOTAL").astype(str).str.strip() if c_uf else pd.Series(["TOTAL"] * n, index=chunk.index),
            "mes": mes,
            "ano": ano,
            "ean": sku_serie,
            "categoria": categoria_serie.replace({"nan": "", "None": ""}).fillna(""),
            "valor_sellout": limpar_coluna_numerica_vetorizada(serie_col(c_val)),
            "fabricante": serie_col(c_fab).astype(str).str.strip() if c_fab else pd.Series([""] * n, index=chunk.index),
            "marca": serie_col(c_marca).astype(str).str.strip() if c_marca else pd.Series([""] * n, index=chunk.index),
            "nome_sku": serie_col(c_nome).astype(str).str.strip() if c_nome else pd.Series([""] * n, index=chunk.index),
            "canal": serie_col(c_canal).astype(str).str.strip() if c_canal else pd.Series([""] * n, index=chunk.index),
        })
        if metrica_eh_volume_variavel(metrica):
            out["qtd_sellout"] = limpar_coluna_numerica_vetorizada(serie_col(c_qtd_var)) if c_qtd_var else out["valor_sellout"]
            out["peso_gramas_sku"] = out["nome_sku"].map(extrair_peso_gramas)
            out["valor_sellout_origem_volume_variavel"] = origem_valor_volume_variavel

        out["uf"] = out["uf"].astype(str).str.strip().replace("", "TOTAL")
        out["ano"] = pd.to_numeric(out["ano"], errors="coerce")
        out.loc[out["mes"].notna(), "ano"] = out.loc[out["mes"].notna(), "mes"].dt.year
        out["fabricante"] = out["fabricante"].fillna("").astype(str).str.strip()
        out["fabricante_key"] = out["fabricante"].map(normalizar_texto)

        if c_fab:
            vals_fab = out["fabricante"].dropna().astype(str).str.strip().unique().tolist()
            for v in vals_fab:
                k = normalizar_texto(v)
                if v and k and k not in vistos_fabricantes and len(amostras_fabricantes) < 12:
                    vistos_fabricantes.add(k)
                    amostras_fabricantes.append(v)

        if fabricante_filtro:
            mask_exata = out["fabricante_key"] == fab_key
            # Se o usuário digitar só parte do nome, permite conter como fallback sem carregar lista completa.
            mask_contem = out["fabricante_key"].str.contains(re.escape(fab_key), na=False) if fab_key else mask_exata
            linhas_apos_filtro_fab_exato += int(mask_exata.sum())
            linhas_apos_filtro_fab_contem += int(mask_contem.sum())
            out = out[mask_exata | mask_contem].copy()

        out["categoria"] = out["categoria"].astype(str).str.strip()
        # Quando CATEGORIA vem por mapeamento de SKU do Sell-in, não pode zerar a base inteira
        # só porque parte dos SKUs não encontrou categoria. Nesses casos cria total de fallback.
        if (out["categoria"].map(normalizar_categoria) == "").all():
            if dict_cat_forcado:
                # Congelado selecionado para Categoria: não criar Total Fabricante.
                # Linhas sem SKU encontrado no Congelado ficam sem categoria e serão removidas.
                pass
            elif fabricante_filtro:
                out["categoria"] = f"Total Fabricante - {fabricante_filtro}"
            elif usar_total_categoria:
                out["categoria"] = "Total Sell-out"
        else:
            if fabricante_filtro and not dict_cat_forcado:
                out.loc[out["categoria"].map(normalizar_categoria) == "", "categoria"] = f"Total Fabricante - {fabricante_filtro}"
        out["categoria_key"] = out["categoria"].map(normalizar_categoria)
        out = out[(out["categoria_key"] != "") & (out["valor_sellout"].fillna(0) != 0)].copy()
        linhas_validas += len(out)

        if not out.empty:
            group_cols = [
                "uf", "mes", "ano", "ean", "categoria", "categoria_key",
                "fabricante", "fabricante_key", "marca", "nome_sku", "canal"
            ]
            if metrica_eh_volume_variavel(metrica):
                out["_peso_x_qtd_volume_variavel"] = out["peso_gramas_sku"] * out["qtd_sellout"]
                agrupado = (
                    out.groupby(group_cols, dropna=False, as_index=False)
                    .agg(
                        valor_sellout=("valor_sellout", "sum"),
                        qtd_sellout=("qtd_sellout", "sum"),
                        _peso_x_qtd_volume_variavel=("_peso_x_qtd_volume_variavel", "sum"),
                    )
                )
                agrupado["peso_gramas_sku"] = agrupado["_peso_x_qtd_volume_variavel"] / agrupado["qtd_sellout"].replace(0, np.nan)
                agrupado["valor_sellout_origem_volume_variavel"] = origem_valor_volume_variavel
            else:
                agrupado = out.groupby(group_cols, dropna=False, as_index=False)["valor_sellout"].sum()
            partes.append(agrupado)

        # Atualiza por bloco, sem calcular total exato de linhas para não deixar mais lento.
        if idx_chunk % 3 == 0:
            pct_aprox = min(34, 20 + idx_chunk)
            progresso(f"Lendo Sell-out em blocos... {linhas_lidas:,} linhas lidas".replace(",", "."), pct_aprox)

    if partes:
        df = pd.concat(partes, ignore_index=True)
        group_cols_final = [
            "uf", "mes", "ano", "ean", "categoria", "categoria_key",
            "fabricante", "fabricante_key", "marca", "nome_sku", "canal"
        ]
        if metrica_eh_volume_variavel(metrica) and "qtd_sellout" in df.columns:
            df = (
                df.groupby(group_cols_final, dropna=False, as_index=False)
                .agg(
                    valor_sellout=("valor_sellout", "sum"),
                    qtd_sellout=("qtd_sellout", "sum"),
                    _peso_x_qtd_volume_variavel=("_peso_x_qtd_volume_variavel", "sum"),
                )
            )
            df["peso_gramas_sku"] = df["_peso_x_qtd_volume_variavel"] / df["qtd_sellout"].replace(0, np.nan)
            df["valor_sellout_origem_volume_variavel"] = origem_valor_volume_variavel
        else:
            df = df.groupby(group_cols_final, dropna=False, as_index=False)["valor_sellout"].sum()
    else:
        df = pd.DataFrame(columns=[
            "uf", "mes", "ano", "ean", "categoria", "categoria_key", "valor_sellout",
            "fabricante", "fabricante_key", "marca", "nome_sku", "canal",
            "qtd_sellout", "peso_gramas_sku", "valor_sellout_origem_volume_variavel",
        ])

    if fabricante_filtro:
        avisos.append(
            f"Sell-out filtrado pelo fabricante selecionado: {fabricante_filtro}. "
            f"Linhas com fabricante exato: {linhas_apos_filtro_fab_exato:,}; "
            f"linhas com fabricante contendo o texto: {linhas_apos_filtro_fab_contem:,}; "
            f"linhas válidas após filtros finais: {linhas_validas:,}.".replace(",", ".")
        )
        if linhas_validas == 0 and amostras_fabricantes:
            avisos.append("Exemplos de fabricantes encontrados no Sell-out: " + "; ".join(amostras_fabricantes[:12]))
    if not c_uf:
        avisos.append("Sell-out sem UF. A análise por UF usará TOTAL para o Sell-out.")
    if not c_mes:
        if c_ano:
            avisos.append("Sell-out sem mês. Será usada análise anual quando possível, com base na coluna de ano.")
        else:
            avisos.append("Sell-out sem mês e sem ano. As análises temporais serão limitadas ao total disponível.")
    if not c_sku:
        avisos.append("Sell-out sem SKU/EAN. A base de SKUs será limitada ao total disponível.")
    if c_fab is None:
        avisos.append("Sell-out sem coluna Fabricante. Não será possível escolher/validar fabricante no Sell-out.")

    if not df.empty and df["ean"].map(normalizar_texto).ne("").any():
        mapa = (
            df[df["ean"] != ""].groupby(["ean", "categoria_key", "categoria"], as_index=False)["valor_sellout"]
            .sum()
            .sort_values(["ean", "valor_sellout"], ascending=[True, False])
            .drop_duplicates("ean", keep="first")
            [["ean", "categoria", "categoria_key"]]
            .rename(columns={"categoria": "categoria_map_prod", "categoria_key": "categoria_key_map_prod"})
        )
    else:
        mapa = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
        avisos.append("Não foi possível criar mapa SKU -> categoria pelo Sell-out.")

    if df.empty:
        avisos.append("Sell-out ficou sem linhas válidas após filtros de categoria/valor/fabricante.")

    avisos.append(
        f"Sell-out CSV lido em blocos, usando {origem_categoria} como categoria/PROD e '{c_val}' como métrica. "
        f"Linhas lidas: {linhas_lidas:,}; linhas válidas agregadas: {len(df):,}.".replace(",", ".")
    )
    return df, mapa, avisos


def ler_sellout(
    arquivo: Path,
    metrica: str,
    nivel_prod: str,
    sellin_para_categoria: Optional[pd.DataFrame] = None,
    fabricante_filtro: str = "",
    progress_callback=None,
    mapa_categoria_forcado: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    avisos = []
    nivel_prod = normalizar_nivel_prod(nivel_prod)
    usar_mapa_categoria_forcado = mapa_categoria_forcado is not None and not mapa_categoria_forcado.empty

    if eh_csv(Path(arquivo)):
        return ler_sellout_csv_em_blocos(
            arquivo=arquivo,
            metrica=metrica,
            nivel_prod=nivel_prod,
            sellin_para_categoria=sellin_para_categoria,
            fabricante_filtro=fabricante_filtro,
            progress_callback=progress_callback,
            mapa_categoria_forcado=mapa_categoria_forcado,
        )

    raw, linha, aba = ler_sellout_excel_com_cabecalho_flexivel(
        arquivo=arquivo,
        metrica=metrica,
        nivel_prod=nivel_prod,
        usar_mapa_categoria_forcado=usar_mapa_categoria_forcado,
    )
    raw = otimizar_colunas_categoricas_sellout(raw)

    # Preferir UF SM quando existir; se não existir, usa UF normal.
    c_uf = localizar_coluna(raw, ["UF SM", "UF_SM", "UF Scann", "UF Scanntech", "UF", "Estado"], obrigatoria=False)
    c_mes = localizar_coluna(raw, COLUNAS_MES_ALTERNATIVAS, obrigatoria=False)
    c_ano = localizar_coluna(raw, ["Ano", "ANO", "Year", "YEAR"], obrigatoria=False)
    if c_ano == c_mes:
        c_ano = None
    c_sku = localizar_coluna(
        raw,
        [
            "SKU", "EAN", "EAN/SKU", "SKU/EAN",
            "Código Barras", "Codigo Barras", "Código de Barras", "Codigo de Barras",
            "Cod Barras", "Cód Barras", "CODIGO_BARRAS", "COD_BARRAS"
        ],
        obrigatoria=False,
    )
    c_qtd_var = localizar_coluna(raw, alternativas_sellout_quantia(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None
    c_volume_var = localizar_coluna(raw, alternativas_sellout_volume(), obrigatoria=False) if metrica_eh_volume_variavel(metrica) else None
    alternativas_valor_sellout = coluna_metrica_sellout(metrica)
    if metrica_eh_volume_variavel(metrica) and c_volume_var:
        c_val = c_volume_var
        origem_valor_volume_variavel = "volume"
    elif metrica_eh_volume_variavel(metrica) and c_qtd_var:
        c_val = c_qtd_var
        origem_valor_volume_variavel = "quantidade"
    else:
        c_val = localizar_coluna_obrigatoria_com_erro(
            raw,
            alternativas_valor_sellout,
            f"valor de Sell-out para a métrica {metrica}",
            "Sell-out",
            arquivo,
        )
        origem_valor_volume_variavel = "volume" if normalizar_texto(c_val) in {normalizar_texto(x) for x in alternativas_sellout_volume()} else "quantidade"
    c_fab = localizar_coluna(raw, ["Fabricante", "Fabricante SKU", "FABRICANTE"], obrigatoria=False)
    if str(fabricante_filtro or "").strip() and not c_fab:
        raise ValueError("Você informou um fabricante para filtrar, mas não encontrei a coluna Fabricante no Sell-out.")
    c_marca = localizar_coluna(raw, ["Marca", "Marca SKU"], obrigatoria=False)
    c_nome = localizar_coluna(raw, ["NOMBRE_SKU", "Nome SKU", "Descrição", "Descricao", "PROD_NOMBRE_ORIGINAL"], obrigatoria=False)
    c_canal = localizar_coluna(raw, ["PDV_CANAL", "Canal", "PDV Canal"], obrigatoria=False)

    n = len(raw)
    sku_serie = raw[c_sku].map(ean_texto) if c_sku else pd.Series([""] * n, index=raw.index)
    dict_cat_forcado = {}
    if usar_mapa_categoria_forcado and c_sku:
        mapa_forcado = mapa_categoria_forcado.copy()
        mapa_forcado["ean"] = mapa_forcado["ean"].map(ean_texto)
        mapa_forcado = mapa_forcado[(mapa_forcado["ean"] != "") & (mapa_forcado["categoria_key_map_prod"].map(normalizar_categoria) != "")].copy()
        dict_cat_forcado = mapa_forcado.drop_duplicates("ean", keep="first").set_index("ean")["categoria_map_prod"].to_dict()
    elif usar_mapa_categoria_forcado and not c_sku:
        avisos.append("Congelado opcional informado, mas o Sell-out não possui SKU/EAN. A categoria do Congelado não pôde ser aplicada ao Sell-out.")

    if dict_cat_forcado:
        c_cat = None
        categoria_serie = sku_serie.map(dict_cat_forcado).fillna("")
        origem_categoria = "Congelado opcional mapeado por SKU"
    elif nivel_prod == "CATEGORIA":
        c_cat = localizar_coluna(
            raw,
            [
                "Categoria SM", "CATEGORIA SM", "CATEGORIA SCANN", "Categoria Scann",
                "Categoria Scanntech", "Categoria", "CATEGORIA",
                "Categoria Cliente", "Categoria Sell-out", "Categoria Sellout"
            ],
            obrigatoria=False,
        )

        mapa_si = mapa_categoria_sellin_por_sku(sellin_para_categoria if sellin_para_categoria is not None else pd.DataFrame())
        dict_cat_si = mapa_si.set_index("ean")["categoria_map_prod"].to_dict() if not mapa_si.empty and c_sku else {}

        if c_cat:
            categoria_serie = raw[c_cat].astype(str).str.strip()
            if dict_cat_si:
                categoria_sellin = sku_serie.map(dict_cat_si).fillna("")
                categoria_serie = categoria_serie.where(categoria_serie.map(normalizar_categoria) != "", categoria_sellin)
                origem_categoria = f"coluna '{c_cat}' do Sell-out com fallback para Categoria do Sell-in por SKU"
            else:
                origem_categoria = f"coluna '{c_cat}' do Sell-out"
        else:
            if dict_cat_si:
                categoria_serie = sku_serie.map(dict_cat_si).fillna("")
                origem_categoria = "Categoria do Sell-in mapeada por SKU"
            else:
                categoria_serie = pd.Series([""] * n, index=raw.index)
                origem_categoria = "sem categoria disponível"
                avisos.append("Sell-out sem coluna Categoria e sem mapeamento por SKU. Será usada visão total quando necessário.")
    else:
        c_cat = localizar_coluna(raw, [nivel_prod], obrigatoria=False)
        if c_cat:
            categoria_serie = raw[c_cat].astype(str).str.strip()
            origem_categoria = f"coluna '{nivel_prod}' do Sell-out"
        else:
            categoria_serie = pd.Series([""] * n, index=raw.index)
            origem_categoria = f"{nivel_prod} não encontrado"
            avisos.append(f"Sell-out sem coluna {nivel_prod}. Será usada visão total quando necessário.")

    registrar_colunas_reconhecidas(avisos, "Sell-out", {
        "UF": c_uf,
        "Mês/Data": c_mes,
        "Ano": c_ano,
        "SKU/EAN": c_sku,
        "Categoria/PROD": c_cat,
        "Fabricante": c_fab,
        "Marca": c_marca,
        "Nome SKU": c_nome,
        "Canal": c_canal,
        "Valor Sell-out": c_val,
        "Qtd Sell-out para ponderação": c_qtd_var if metrica_eh_volume_variavel(metrica) else None,
        "Volume Sell-out para comparação": c_volume_var if metrica_eh_volume_variavel(metrica) else None,
    })
    registrar_colunas_nao_encontradas(avisos, "Sell-out", [
        ("UF", c_uf, ["UF", "Estado"]),
        ("Mês/Data", c_mes, ["DATA", "ANO MÊS", "ANO_MES", "AAAAMM", "YYYYMM", "Mês/Ano", "Competência", "Referência", "Período"]),
        ("Ano", c_ano, ["Ano", "ANO", "Year"]),
        ("SKU/EAN", c_sku, ["SKU", "EAN", "EAN/SKU", "Código de Barras"]),
        ("Categoria/PROD", c_cat, ["Categoria", "CATEGORIA SCANN", "NIVEL1", "NIVEL2"]),
        ("Fabricante", c_fab, ["Fabricante", "Fabricante SKU"]),
        ("Marca", c_marca, ["Marca", "Marca SKU"]),
        ("Nome SKU", c_nome, ["NOMBRE_SKU", "Nome SKU", "Descrição"]),
        ("Canal", c_canal, ["PDV_CANAL", "Canal"]),
    ])

    mes = raw[c_mes].map(converter_mes) if c_mes else pd.Series([pd.NaT] * n, index=raw.index)
    if c_ano:
        ano = raw[c_ano].map(converter_ano)
    else:
        ano = mes.map(lambda x: x.year if pd.notna(x) else np.nan)

    df = pd.DataFrame({
        "uf": raw[c_uf].astype(str).str.strip() if c_uf else pd.Series(["TOTAL"] * n, index=raw.index),
        "mes": mes,
        "ano": ano,
        "ean": sku_serie,
        "categoria": categoria_serie.replace({"nan": "", "None": ""}).fillna(""),
        "valor_sellout": limpar_coluna_numerica_vetorizada(raw[c_val]),
        "fabricante": raw[c_fab].astype(str).str.strip() if c_fab else pd.Series([""] * n, index=raw.index),
        "marca": raw[c_marca].astype(str).str.strip() if c_marca else pd.Series([""] * n, index=raw.index),
        "nome_sku": raw[c_nome].astype(str).str.strip() if c_nome else pd.Series([""] * n, index=raw.index),
        "canal": raw[c_canal].astype(str).str.strip() if c_canal else pd.Series([""] * n, index=raw.index),
    })
    if metrica_eh_volume_variavel(metrica):
        df["qtd_sellout"] = limpar_coluna_numerica_vetorizada(raw[c_qtd_var]) if c_qtd_var else df["valor_sellout"]
        df["peso_gramas_sku"] = df["nome_sku"].map(extrair_peso_gramas)
        df["valor_sellout_origem_volume_variavel"] = origem_valor_volume_variavel
    df["uf"] = df["uf"].astype(str).str.strip().replace("", "TOTAL")
    df["ano"] = pd.to_numeric(df["ano"], errors="coerce")
    df.loc[df["mes"].notna(), "ano"] = df.loc[df["mes"].notna(), "mes"].dt.year
    df["fabricante"] = df["fabricante"].fillna("").astype(str).str.strip()
    df["fabricante_key"] = df["fabricante"].map(normalizar_texto)

    fabricante_filtro = str(fabricante_filtro or "").strip()
    if fabricante_filtro:
        fab_key = normalizar_texto(fabricante_filtro)
        antes = len(df)
        mask_exata = df["fabricante_key"] == fab_key
        mask_contem = df["fabricante_key"].str.contains(re.escape(fab_key), na=False) if fab_key else mask_exata
        df = df[mask_exata | mask_contem].copy()
        avisos.append(
            f"Sell-out filtrado pelo fabricante selecionado: {fabricante_filtro}. "
            f"Exato: {int(mask_exata.sum())}; contém: {int(mask_contem.sum())}; linhas mantidas: {len(df)} de {antes}."
        )

    # Se não existir categoria/PROD, cria uma visão total do fabricante selecionado ou total geral.
    df["categoria"] = df["categoria"].astype(str).str.strip()
    if (df["categoria"].map(normalizar_categoria) == "").all():
        if dict_cat_forcado:
            # Congelado selecionado para Categoria: não criar Total Fabricante.
            # Linhas sem SKU encontrado no Congelado ficam sem categoria e serão removidas.
            origem_categoria += " / nenhum SKU encontrado no Congelado"
            avisos.append(
                "Congelado opcional aplicado no Sell-out, mas nenhum SKU do Sell-out encontrou categoria no Congelado. "
                "Não foi criada visão Total Fabricante, pois a Categoria deve vir do Congelado."
            )
        elif fabricante_filtro:
            df["categoria"] = f"Total Fabricante - {fabricante_filtro}"
            origem_categoria += " / total por fabricante selecionado"
        else:
            df["categoria"] = "Total Sell-out"
            origem_categoria += " / total Sell-out"
    else:
        # Não descarta linhas úteis apenas por falta de categoria mapeada no fluxo padrão.
        if fabricante_filtro and not dict_cat_forcado:
            df.loc[df["categoria"].map(normalizar_categoria) == "", "categoria"] = f"Total Fabricante - {fabricante_filtro}"

    df["categoria_key"] = df["categoria"].map(normalizar_categoria)

    linhas_antes = len(df)
    df = df[
        (df["categoria_key"] != "")
        & (df["valor_sellout"].fillna(0) != 0)
    ].copy()

    if not c_uf:
        avisos.append("Sell-out sem UF. A análise por UF usará TOTAL para o Sell-out.")
    if not c_mes:
        if c_ano:
            avisos.append("Sell-out sem mês. Será usada análise anual quando possível, com base na coluna de ano.")
        else:
            avisos.append("Sell-out sem mês e sem ano. As análises temporais serão limitadas ao total disponível.")
    if not c_sku:
        avisos.append("Sell-out sem SKU/EAN. A base de SKUs será limitada ao total disponível.")
    if c_fab is None:
        avisos.append("Sell-out sem coluna Fabricante. Não será possível escolher/validar fabricante no Sell-out.")

    if not df.empty and df["ean"].map(normalizar_texto).ne("").any():
        mapa = (
            df[df["ean"] != ""].groupby(["ean", "categoria_key", "categoria"], as_index=False)["valor_sellout"]
            .sum()
            .sort_values(["ean", "valor_sellout"], ascending=[True, False])
            .drop_duplicates("ean", keep="first")
            [["ean", "categoria", "categoria_key"]]
            .rename(columns={"categoria": "categoria_map_prod", "categoria_key": "categoria_key_map_prod"})
        )
    else:
        mapa = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
        avisos.append("Não foi possível criar mapa SKU -> categoria pelo Sell-out.")

    if df.empty and linhas_antes:
        avisos.append("Sell-out ficou sem linhas válidas após filtros de categoria/valor.")

    avisos.append(f"Sell-out lido da aba '{aba}', usando {origem_categoria} como categoria/PROD e '{c_val}' como métrica.")
    return df, mapa, avisos


def aplicar_mapeamento_prod_no_sellin(
    sellin: pd.DataFrame,
    mapa_sellout: pd.DataFrame,
    fabricante_filtro: str = "",
    forcar_mapa_sem_fallback: bool = False,
    origem_mapa: str = "Sell-out",
) -> Tuple[pd.DataFrame, List[str]]:
    avisos = []
    si = sellin.copy()

    # Garante colunas esperadas.
    for col in ["ean", "categoria_original_sellin", "fabricante_sellin"]:
        if col not in si.columns:
            si[col] = ""

    tem_sku = si["ean"].map(normalizar_texto).ne("").any()
    if tem_sku and not mapa_sellout.empty:
        si = si.merge(mapa_sellout, on="ean", how="left")

        if forcar_mapa_sem_fallback:
            # Quando o Congelado é usado no modo CATEGORIA, ele é a fonte oficial.
            # SKU não encontrado no Congelado não deve cair para a categoria original,
            # porque isso mistura critérios e pode gerar abas como Total Fabricante.
            si["categoria"] = si["categoria_map_prod"].fillna("").astype(str).str.strip()
        else:
            si["categoria"] = si["categoria_map_prod"].where(
                si["categoria_map_prod"].fillna("").map(normalizar_categoria) != "",
                si["categoria_original_sellin"],
            )

        total_skus = si[si["ean"] != ""]["ean"].nunique()
        skus_mapeados = si[si.get("categoria_map_prod", "").fillna("").map(normalizar_categoria) != ""]["ean"].nunique()
        faltantes = total_skus - skus_mapeados
        avisos.append(f"Mapeamento de categoria no Sell-in via {origem_mapa}: {skus_mapeados} de {total_skus} SKUs do Sell-in encontrados.")
        if faltantes:
            if forcar_mapa_sem_fallback:
                avisos.append(
                    f"{faltantes} SKUs do Sell-in não foram encontrados no {origem_mapa}. "
                    "Como o Congelado foi selecionado para Categoria, esses SKUs ficaram fora da categoria mapeada, sem fallback para a categoria original."
                )
            else:
                avisos.append(
                    f"{faltantes} SKUs do Sell-in não foram encontrados no Sell-out para mapear PROD. "
                    "Nesses casos, foi usada a categoria original do Sell-in como fallback, quando existente."
                )
    else:
        si["categoria"] = si["categoria_original_sellin"].fillna("").astype(str).str.strip()
        if not tem_sku:
            avisos.append("Sell-in sem SKU/EAN: mapeamento PROD por SKU não foi aplicado.")
        elif mapa_sellout.empty:
            avisos.append("Mapa SKU -> categoria do Sell-out vazio: foi usada a categoria original do Sell-in quando existente.")

    # Se ainda não tiver categoria, usa visão total somente no fluxo padrão.
    # Quando o Congelado é a fonte oficial, não criamos Total Fabricante, pois isso esconde falha de mapeamento.
    if si["categoria"].fillna("").map(normalizar_categoria).eq("").all():
        fabricante_filtro = str(fabricante_filtro or "").strip()
        if forcar_mapa_sem_fallback:
            avisos.append(
                f"Sell-in sem categoria após mapeamento via {origem_mapa}. "
                "Nenhuma visão Total Fabricante foi criada, porque o Congelado deve definir a Categoria."
            )
        elif fabricante_filtro:
            si["categoria"] = f"Total Fabricante - {fabricante_filtro}"
            avisos.append("Sell-in sem categoria: criada visão total pelo fabricante selecionado no Sell-out.")
        elif si["fabricante_sellin"].fillna("").map(normalizar_texto).ne("").any():
            # Usa o fabricante de maior Sell-in como visão total.
            fab_ref = (
                si[si["fabricante_sellin"].fillna("").map(normalizar_texto) != ""]
                .groupby("fabricante_sellin", as_index=False)["valor_sellin"].sum()
                .sort_values("valor_sellin", ascending=False)
                .iloc[0]["fabricante_sellin"]
            )
            si["categoria"] = f"Total Fabricante - {fab_ref}"
            avisos.append(f"Sell-in sem categoria: criada visão total pelo fabricante do Sell-in '{fab_ref}'.")
        else:
            si["categoria"] = "Total Sell-in"
            avisos.append("Sell-in sem categoria e sem fabricante: criada visão total Sell-in.")

    si["categoria_key"] = si["categoria"].map(normalizar_categoria)
    si = si[(si["categoria_key"] != "") & (si["valor_sellin"].fillna(0) != 0)].copy()
    return si, avisos



def reagrupar_por_categoria_original_e_nivel(
    sellin: pd.DataFrame,
    sellout: pd.DataFrame,
    nivel: str,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Quando o usuário escolhe NIVEL1 ou NIVEL2, cada valor do nível escolhido vira
    uma categoria/aba própria.

    Exemplo:
        NIVEL1 no Sell-out = PÃO DE FORMA TRADICIONAL
        Resultado: aba PÃO_DE_FORMA_TRADICIONAL

    O Sell-in é classificado por SKU/EAN usando o mapa do Sell-out. Linhas do
    Sell-in sem SKU mapeado para o nível escolhido são mantidas fora da análise
    por nível, porque não existe como saber a qual PROD elas pertencem.
    """
    avisos: List[str] = []
    nivel = normalizar_nivel_prod(nivel)
    if nivel == "CATEGORIA":
        return sellin, sellout, avisos

    si = sellin.copy()
    so = sellout.copy()

    for col in ["ean", "categoria", "categoria_key", "categoria_map_prod", "categoria_key_map_prod"]:
        if col not in si.columns:
            si[col] = ""

    # O aplicar_mapeamento_prod_no_sellin já trouxe o PROD do Sell-out para
    # categoria_map_prod / categoria_key_map_prod. Nesta etapa, para NIVEL1/NIVEL2,
    # a categoria final passa a ser o próprio valor do nível escolhido.
    prod_nome = si["categoria_map_prod"].fillna("").astype(str).str.strip()
    prod_key = si["categoria_key_map_prod"].fillna("").astype(str).str.strip()
    prod_key = prod_key.where(prod_key.map(normalizar_categoria) != "", prod_nome.map(normalizar_categoria))

    total_skus = si[si["ean"].astype(str).str.strip() != ""]["ean"].nunique() if "ean" in si.columns else 0
    skus_mapeados = si[(si["ean"].astype(str).str.strip() != "") & (prod_key.map(normalizar_categoria) != "")]["ean"].nunique() if "ean" in si.columns else 0

    si["categoria"] = prod_nome
    si["categoria_key"] = prod_key
    si["prod_sellout_mapeado"] = prod_nome
    si["prod_sellout_mapeado_key"] = prod_key

    antes = len(si)
    si = si[(si["categoria_key"].map(normalizar_categoria) != "") & (si["valor_sellin"].fillna(0) != 0)].copy()
    removidas = antes - len(si)

    if skus_mapeados:
        avisos.append(
            f"{nivel}: {skus_mapeados} de {total_skus} SKUs do Sell-in foram mapeados para valores do {nivel}. "
            f"Cada valor do {nivel} será gerado em uma aba própria."
        )
    else:
        avisos.append(
            f"{nivel}: nenhum SKU do Sell-in foi mapeado para valores do {nivel}. "
            "Sem esse mapeamento, não é possível separar o Sell-in por PROD."
        )

    if removidas:
        avisos.append(
            f"{nivel}: {removidas} linhas do Sell-in ficaram fora da análise por nível por não terem SKU/EAN mapeado no Sell-out."
        )

    # Filtra o Sell-out para os PRODs efetivamente encontrados no Sell-in.
    # Isso evita gerar abas de níveis sem Sell-in correspondente.
    prods_validos = set(si["categoria_key"].dropna().astype(str).map(normalizar_categoria))
    if prods_validos and "categoria_key" in so.columns:
        antes_so = len(so)
        so = so[so["categoria_key"].astype(str).map(normalizar_categoria).isin(prods_validos)].copy()
        removidas_so = antes_so - len(so)
        if removidas_so:
            avisos.append(
                f"{nivel}: {removidas_so} linhas do Sell-out foram ignoradas por pertencerem a valores do nível sem SKU no Sell-in."
            )

    avisos.append(
        f"{nivel}: as abas serão geradas pelo próprio valor do nível escolhido, "
        "por exemplo PÃO_DE_FORMA_TRADICIONAL, PÃO_DE_FORMA_INTEGRAL etc."
    )

    return si, so, avisos

def determinar_periodos_mat_ou_ytd(meses: List[pd.Timestamp], max_mes: pd.Timestamp) -> Dict[str, object]:
    meses_validos = [pd.Timestamp(m).to_period("M").to_timestamp() for m in meses if not pd.isna(m)]
    meses_set = set(meses_validos)

    anos = sorted({m.year for m in meses_validos})
    anos_fechados = []
    for ano in anos:
        meses_ano = {pd.Timestamp(year=ano, month=mes, day=1) for mes in range(1, 13)}
        if meses_ano.issubset(meses_set):
            anos_fechados.append(ano)

    # Prioriza o último par de anos fechados consecutivos.
    for ano_atual in sorted(anos_fechados, reverse=True):
        ano_anterior = ano_atual - 1
        if ano_anterior in anos_fechados:
            return {
                "tipo": "Ano fechado",
                "label_anterior": f"FY {ano_anterior}",
                "label_atual": f"FY {ano_atual}",
                "inicio_anterior": pd.Timestamp(year=ano_anterior, month=1, day=1),
                "fim_anterior": pd.Timestamp(year=ano_anterior, month=12, day=1),
                "inicio_atual": pd.Timestamp(year=ano_atual, month=1, day=1),
                "fim_atual": pd.Timestamp(year=ano_atual, month=12, day=1),
            }

    # Se não tiver par de anos fechados, usa YTD do último ano disponível contra o mesmo período do ano anterior.
    ano_atual = pd.Timestamp(max_mes).year
    mes_final = pd.Timestamp(max_mes).month
    ano_anterior = ano_atual - 1
    return {
        "tipo": "YTD",
        "label_anterior": f"YTD {ano_anterior}",
        "label_atual": f"YTD {ano_atual}",
        "inicio_anterior": pd.Timestamp(year=ano_anterior, month=1, day=1),
        "fim_anterior": pd.Timestamp(year=ano_anterior, month=mes_final, day=1),
        "inicio_atual": pd.Timestamp(year=ano_atual, month=1, day=1),
        "fim_atual": pd.Timestamp(year=ano_atual, month=mes_final, day=1),
    }





def maior_sequencia_meses_consecutivos(df: pd.DataFrame, col_mes: str, col_valor: str) -> int:
    """
    Retorna a maior sequência de meses consecutivos com valor positivo.
    Usado para validar se o Sell-in possui ao menos 12 meses seguidos para cálculo de cobertura 12M.
    """
    if df is None or df.empty or col_mes not in df.columns or col_valor not in df.columns:
        return 0

    base = df[[col_mes, col_valor]].copy()
    base[col_mes] = pd.to_datetime(base[col_mes], errors="coerce").dt.to_period("M").dt.to_timestamp()
    base[col_valor] = pd.to_numeric(base[col_valor], errors="coerce").fillna(0)
    meses = sorted(base.loc[(base[col_mes].notna()) & (base[col_valor] > 0), col_mes].drop_duplicates().tolist())
    if not meses:
        return 0

    maior = 1
    atual = 1
    anterior = meses[0]
    for mes in meses[1:]:
        esperado = (pd.Timestamp(anterior) + pd.DateOffset(months=1)).to_period("M").to_timestamp()
        if pd.Timestamp(mes).to_period("M").to_timestamp() == esperado:
            atual += 1
        else:
            atual = 1
        maior = max(maior, atual)
        anterior = mes
    return int(maior)



def skus_comuns_sellin_sellout(si: pd.DataFrame, so: pd.DataFrame) -> set:
    """Retorna SKUs/EANs que possuem volume positivo tanto no Sell-in quanto no Sell-out."""
    if si is None or so is None or si.empty or so.empty:
        return set()
    if "ean" not in si.columns or "ean" not in so.columns:
        return set()
    si_tmp = si.copy()
    so_tmp = so.copy()
    si_tmp["ean"] = si_tmp["ean"].map(ean_texto)
    so_tmp["ean"] = so_tmp["ean"].map(ean_texto)
    si_skus = set(si_tmp.loc[(si_tmp["ean"] != "") & (pd.to_numeric(si_tmp.get("valor_sellin", 0), errors="coerce").fillna(0) > 0), "ean"])
    so_skus = set(so_tmp.loc[(so_tmp["ean"] != "") & (pd.to_numeric(so_tmp.get("valor_sellout", 0), errors="coerce").fillna(0) > 0), "ean"])
    return si_skus & so_skus


def calcular_periodos_mat_movel(meses: List[pd.Timestamp], max_mes: pd.Timestamp) -> Dict[str, object]:
    """
    Calcula período MAT móvel real: últimos 12 meses encerrando no último mês disponível
    contra os 12 meses imediatamente anteriores.
    """
    if pd.isna(max_mes):
        return {
            "tipo": "MAT móvel",
            "label_anterior": "MAT-1",
            "label_atual": "MAT",
            "inicio_anterior": pd.NaT,
            "fim_anterior": pd.NaT,
            "inicio_atual": pd.NaT,
            "fim_atual": pd.NaT,
        }
    fim_atual = pd.Timestamp(max_mes).to_period("M").to_timestamp()
    inicio_atual = (fim_atual - pd.DateOffset(months=MESES_MAT - 1)).to_period("M").to_timestamp()
    fim_anterior = (inicio_atual - pd.DateOffset(months=1)).to_period("M").to_timestamp()
    inicio_anterior = (fim_anterior - pd.DateOffset(months=MESES_MAT - 1)).to_period("M").to_timestamp()
    return {
        "tipo": "MAT móvel",
        "label_anterior": "MAT-1",
        "label_atual": "MAT",
        "inicio_anterior": inicio_anterior,
        "fim_anterior": fim_anterior,
        "inicio_atual": inicio_atual,
        "fim_atual": fim_atual,
    }


def somar_periodo_dataframe(df: pd.DataFrame, coluna_valor: str, inicio, fim, coluna_mes: str = "mes") -> float:
    if df is None or df.empty or coluna_valor not in df.columns or coluna_mes not in df.columns:
        return 0.0
    if pd.isna(inicio) or pd.isna(fim):
        return 0.0
    temp = df.copy()
    temp[coluna_mes] = pd.to_datetime(temp[coluna_mes], errors="coerce").dt.to_period("M").dt.to_timestamp()
    mask = (temp[coluna_mes] >= inicio) & (temp[coluna_mes] <= fim)
    return float(pd.to_numeric(temp.loc[mask, coluna_valor], errors="coerce").fillna(0).sum())

def filtrar_periodo_calculo(df: pd.DataFrame, inicio, fim, modo_periodo: str) -> pd.DataFrame:
    """Filtra uma base pelo mesmo período usado no cálculo superior da aba.

    Em modo mensal usa a coluna mes. Em modo anual usa a coluna ano.
    Se o período estiver indisponível, devolve a base completa para manter o fallback total.
    """
    if df is None:
        return pd.DataFrame()
    if df.empty:
        return df.copy()
    if pd.isna(inicio) or pd.isna(fim):
        return df.copy()

    out = df.copy()
    modo = str(modo_periodo or "").lower()
    if modo == "mensal" and "mes" in out.columns:
        out["_mes_periodo_calc"] = pd.to_datetime(out["mes"], errors="coerce").dt.to_period("M").dt.to_timestamp()
        mask = (out["_mes_periodo_calc"] >= pd.Timestamp(inicio).to_period("M").to_timestamp()) & (out["_mes_periodo_calc"] <= pd.Timestamp(fim).to_period("M").to_timestamp())
        return out.loc[mask].drop(columns=["_mes_periodo_calc"], errors="ignore").copy()

    if modo == "anual" and "ano" in out.columns:
        ano_ini = int(pd.Timestamp(inicio).year)
        ano_fim = int(pd.Timestamp(fim).year)
        anos = pd.to_numeric(out["ano"], errors="coerce")
        return out.loc[(anos >= ano_ini) & (anos <= ano_fim)].copy()

    return out.copy()


def somar_sku_comum_no_periodo(si: pd.DataFrame, so: pd.DataFrame, inicio, fim, modo_periodo: str) -> Tuple[float, float, int]:
    """Soma Sell-in/Sell-out apenas dos SKUs em comum no mesmo período informado."""
    si_p = filtrar_periodo_calculo(si, inicio, fim, modo_periodo)
    so_p = filtrar_periodo_calculo(so, inicio, fim, modo_periodo)
    skus = skus_comuns_sellin_sellout(si_p, so_p)
    if not skus:
        return 0.0, 0.0, 0
    si_val = float(pd.to_numeric(si_p.loc[si_p["ean"].map(ean_texto).isin(skus), "valor_sellin"], errors="coerce").fillna(0).sum()) if "ean" in si_p.columns else 0.0
    so_val = float(pd.to_numeric(so_p.loc[so_p["ean"].map(ean_texto).isin(skus), "valor_sellout"], errors="coerce").fillna(0).sum()) if "ean" in so_p.columns else 0.0
    return si_val, so_val, len(skus)


def montar_resumo_sku_comum_mesmo_periodo(si: pd.DataFrame, so: pd.DataFrame, periodos: Dict[str, object], modo_periodo: str) -> pd.DataFrame:
    """Replica o bloco superior, mas calculando somente SKUs em comum.

    A regra é intencionalmente igual à tabela de cima da aba: usa o mesmo período
    anterior e o mesmo período atual definidos em `periodos`; a única diferença é
    restringir as somas aos SKUs/EANs existentes nos dois lados dentro de cada período.
    """
    label_ant = periodos.get("label_anterior", "Período anterior")
    label_atual = periodos.get("label_atual", "Período atual")
    ia, fa = periodos.get("inicio_anterior"), periodos.get("fim_anterior")
    it, ft = periodos.get("inicio_atual"), periodos.get("fim_atual")

    si_ant, so_ant, skus_ant = somar_sku_comum_no_periodo(si, so, ia, fa, modo_periodo)
    si_atual, so_atual, skus_atual = somar_sku_comum_no_periodo(si, so, it, ft, modo_periodo)
    cov_ant = divisao_segura(so_ant, si_ant)
    cov_atual = divisao_segura(so_atual, si_atual)

    out = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out", "Cobertura", "Qtd SKUs em comum"],
        label_ant: [si_ant, so_ant, cov_ant, skus_ant],
        label_atual: [si_atual, so_atual, cov_atual, skus_atual],
        "Tendência %": [
            variacao(si_atual, si_ant),
            variacao(so_atual, so_ant),
            variacao(cov_atual, cov_ant),
            np.nan,
        ],
    })
    return out


def skus_comuns_do_periodo_atual(si: pd.DataFrame, so: pd.DataFrame, periodos: Dict[str, object], modo_periodo: str) -> set:
    """Retorna os SKUs comuns do período atual do cálculo superior."""
    si_p = filtrar_periodo_calculo(si, periodos.get("inicio_atual"), periodos.get("fim_atual"), modo_periodo)
    so_p = filtrar_periodo_calculo(so, periodos.get("inicio_atual"), periodos.get("fim_atual"), modo_periodo)
    return skus_comuns_sellin_sellout(si_p, so_p)


def montar_resumo_mat_movel_individual(mensal: pd.DataFrame, max_mes: pd.Timestamp) -> Tuple[Dict[str, object], pd.DataFrame]:
    """Monta a tabela MAT móvel. Só calcula MAT quando existem 24 meses completos."""
    meses = mensal["mes"].dropna().tolist() if mensal is not None and "mes" in mensal.columns else []
    periodos_mat = calcular_periodos_mat_movel(meses, max_mes)
    la = periodos_mat["label_anterior"]
    lat = periodos_mat["label_atual"]

    if not tem_janela_movel_completa(meses, max_mes, MESES_MAT * 2):
        periodos_mat["tipo"] = "MAT móvel indisponível - menos de 24 meses"
        df = pd.DataFrame({
            "Indicador": ["Sell-in", "Sell-out"],
            la: [np.nan, np.nan],
            lat: [np.nan, np.nan],
            "Tendência %": [np.nan, np.nan],
        })
        return periodos_mat, df

    ia, fa = periodos_mat["inicio_anterior"], periodos_mat["fim_anterior"]
    it, ft = periodos_mat["inicio_atual"], periodos_mat["fim_atual"]
    si_ant = somar_periodo_dataframe(mensal, "valor_sellin", ia, fa, "mes")
    si_atual = somar_periodo_dataframe(mensal, "valor_sellin", it, ft, "mes")
    so_ant = somar_periodo_dataframe(mensal, "valor_sellout", ia, fa, "mes")
    so_atual = somar_periodo_dataframe(mensal, "valor_sellout", it, ft, "mes")
    df = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out"],
        la: [si_ant, so_ant],
        lat: [si_atual, so_atual],
        "Tendência %": [variacao(si_atual, si_ant), variacao(so_atual, so_ant)],
    })
    return periodos_mat, df

def calcular_cobertura_total_disponivel(
    categoria_key: str,
    categoria_nome: str,
    si: pd.DataFrame,
    so: pd.DataFrame,
    motivo: str = "Total disponível",
) -> Dict[str, object]:
    """Fallback para quando não há mês/ano comum, mas existem valores nos dois lados."""
    si_atual = si["valor_sellin"].sum() if not si.empty else 0
    so_atual = so["valor_sellout"].sum() if not so.empty else 0
    skus_comuns = skus_comuns_sellin_sellout(si, so)
    si_comum_atual = si.loc[si["ean"].map(ean_texto).isin(skus_comuns), "valor_sellin"].sum() if skus_comuns and "ean" in si.columns else 0
    so_comum_atual = so.loc[so["ean"].map(ean_texto).isin(skus_comuns), "valor_sellout"].sum() if skus_comuns and "ean" in so.columns else 0

    meses = []
    if not si.empty and "mes" in si.columns:
        meses += [m for m in si["mes"].dropna().tolist()]
    if not so.empty and "mes" in so.columns:
        meses += [m for m in so["mes"].dropna().tolist()]
    inicio = min(meses) if meses else pd.NaT
    fim = max(meses) if meses else pd.NaT

    periodos = {
        "tipo": motivo,
        "label_anterior": "Anterior indisponível",
        "label_atual": "Total disponível",
        "inicio_anterior": pd.NaT,
        "fim_anterior": pd.NaT,
        "inicio_atual": pd.NaT,
        "fim_atual": pd.NaT,
    }

    resumo_periodo = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
        periodos["label_anterior"]: [np.nan, np.nan, np.nan],
        periodos["label_atual"]: [si_atual, so_atual, divisao_segura(so_atual, si_atual)],
        "Tendência %": [np.nan, np.nan, np.nan],
    })

    resumo_6m = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
        "Período anterior": [np.nan, np.nan, np.nan],
        "Período atual": [np.nan, np.nan, np.nan],
        "Tendência %": [np.nan, np.nan, np.nan],
    })

    mensal_saida = pd.DataFrame({
        "Mês": ["Total"],
        "Sell-in": [si_atual],
        "Sell-out": [so_atual],
        "Cobertura": [divisao_segura(so_atual, si_atual)],
        "Sell-in SKU em Comum": [si_comum_atual],
        "Sell-out SKU em Comum": [so_comum_atual],
        "Cobertura SKU em Comum": [divisao_segura(so_comum_atual, si_comum_atual)],
    })

    resumo_sku_comum = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out", "Cobertura", "Qtd SKUs em comum"],
        periodos["label_anterior"]: [np.nan, np.nan, np.nan, np.nan],
        periodos["label_atual"]: [si_comum_atual, so_comum_atual, divisao_segura(so_comum_atual, si_comum_atual), len(skus_comuns)],
        "Tendência %": [np.nan, np.nan, np.nan, np.nan],
    })

    if "uf" in si.columns and "uf" in so.columns:
        uf = montar_tabela_uf_comparacao(si, so)
    else:
        uf = pd.DataFrame(columns=["UF", "Sell-in", "Sell-out", "Cobertura", "Importância Sell-in", "Importância Sell-out"])

    return {
        "categoria": categoria_nome,
        "categoria_key": categoria_key,
        "inicio_comum": inicio,
        "fim_comum": fim,
        "periodos": periodos,
        "resumo_periodo": resumo_periodo,
        "resumo_6m": resumo_6m,
        "resumo_mat_movel": resumo_periodo.copy(),
        "periodos_mat_movel": periodos.copy(),
        "resumo_sku_comum": resumo_sku_comum,
        "mensal": mensal_saida,
        "uf": uf,
        "skus_excluidos": montar_skus_excluidos_em_comum(si, so),
        "si_atual": si_atual,
        "so_atual": so_atual,
        "cobertura_atual": divisao_segura(so_atual, si_atual),
        "tipo_comparacao": motivo,
        "label_anterior": periodos["label_anterior"],
        "label_atual": periodos["label_atual"],
        "modo_periodo": "total",
        "periodo_uf_inicio": pd.NaT,
        "periodo_uf_fim": pd.NaT,
        "periodo_uf_tipo": "Total disponível",
        "status_sellin": "Inválido",
        "meses_consecutivos_sellin": 0,
    }

def calcular_cobertura_categoria(
    categoria_key: str,
    categoria_nome: str,
    sellin: pd.DataFrame,
    sellout: pd.DataFrame,
) -> Dict[str, object]:
    si = sellin[sellin["categoria_key"] == categoria_key].copy()
    so = sellout[sellout["categoria_key"] == categoria_key].copy()

    if si.empty and so.empty:
        return {}

    # Período usado na tabela por UF.
    # Quando há dados mensais, a tabela por UF deve seguir a mesma lógica da cobertura:
    # últimos 12 meses móveis encerrando no último mês comum entre Sell-in e Sell-out.
    periodo_uf_inicio = pd.NaT
    periodo_uf_fim = pd.NaT
    periodo_uf_tipo = "Indisponível"

    # Modo preferencial: mensal, quando os dois lados possuem mês.
    tem_mes_si = (not si.empty) and si["mes"].notna().any()
    tem_mes_so = (not so.empty) and so["mes"].notna().any()

    if tem_mes_si and tem_mes_so:
        min_si = si["mes"].dropna().min()
        min_so = so["mes"].dropna().min()
        max_si = si["mes"].dropna().max()
        max_so = so["mes"].dropna().max()

        inicio_comum = max(min_si, min_so).to_period("M").to_timestamp()
        max_mes = min(max_si, max_so).to_period("M").to_timestamp()
        if inicio_comum > max_mes:
            return calcular_cobertura_total_disponivel(
                categoria_key,
                categoria_nome,
                si,
                so,
                motivo="Total disponível - sem mês em comum",
            )

        meses = meses_entre(inicio_comum, max_mes)
        if not meses:
            return calcular_cobertura_total_disponivel(
                categoria_key,
                categoria_nome,
                si,
                so,
                motivo="Total disponível - sem mês em comum",
            )

        periodo_uf_fim = max_mes
        periodo_uf_inicio = (periodo_uf_fim - pd.DateOffset(months=MESES_MAT - 1)).to_period("M").to_timestamp()
        if periodo_uf_inicio < inicio_comum:
            periodo_uf_inicio = inicio_comum
        periodo_uf_tipo = "Últimos 12 meses móveis"

        periodos = determinar_periodos_mat_ou_ytd(meses, max_mes)

        # SKU em comum segue o mesmo período definido na tabela superior.
        # Ou seja: usa o período anterior/atual de `periodos` e muda apenas o filtro para SKUs presentes nos dois lados.
        resumo_sku_comum = montar_resumo_sku_comum_mesmo_periodo(si, so, periodos, "mensal")
        skus_comuns = skus_comuns_do_periodo_atual(si, so, periodos, "mensal")
        si_periodo_atual_comum = filtrar_periodo_calculo(si, periodos.get("inicio_atual"), periodos.get("fim_atual"), "mensal")
        so_periodo_atual_comum = filtrar_periodo_calculo(so, periodos.get("inicio_atual"), periodos.get("fim_atual"), "mensal")
        si_comum = si_periodo_atual_comum[si_periodo_atual_comum["ean"].map(ean_texto).isin(skus_comuns)].copy() if skus_comuns and "ean" in si_periodo_atual_comum.columns else si.iloc[0:0].copy()
        so_comum = so_periodo_atual_comum[so_periodo_atual_comum["ean"].map(ean_texto).isin(skus_comuns)].copy() if skus_comuns and "ean" in so_periodo_atual_comum.columns else so.iloc[0:0].copy()

        si_mensal = si.groupby("mes", as_index=False)["valor_sellin"].sum()
        so_mensal = so.groupby("mes", as_index=False)["valor_sellout"].sum()
        si_mensal_comum = si_comum.groupby("mes", as_index=False)["valor_sellin"].sum().rename(columns={"valor_sellin": "valor_sellin_comum"}) if not si_comum.empty else pd.DataFrame(columns=["mes", "valor_sellin_comum"])
        so_mensal_comum = so_comum.groupby("mes", as_index=False)["valor_sellout"].sum().rename(columns={"valor_sellout": "valor_sellout_comum"}) if not so_comum.empty else pd.DataFrame(columns=["mes", "valor_sellout_comum"])

        mensal = pd.DataFrame({"mes": meses})
        mensal = (mensal
            .merge(si_mensal, on="mes", how="left")
            .merge(so_mensal, on="mes", how="left")
            .merge(si_mensal_comum, on="mes", how="left")
            .merge(so_mensal_comum, on="mes", how="left")
        )
        for col in ["valor_sellin", "valor_sellout", "valor_sellin_comum", "valor_sellout_comum"]:
            if col not in mensal.columns:
                mensal[col] = 0.0
            mensal[col] = pd.to_numeric(mensal[col], errors="coerce").fillna(0)
        mensal["periodo"] = mensal["mes"].map(formatar_periodo)

        sellin_12m = mensal["valor_sellin"].rolling(window=MESES_MAT, min_periods=MESES_MAT).sum()
        sellout_12m = mensal["valor_sellout"].rolling(window=MESES_MAT, min_periods=MESES_MAT).sum()
        mensal["cobertura"] = sellout_12m / sellin_12m.replace(0, np.nan)
        mensal["valor_sellin_comum_periodo"] = mensal["valor_sellin_comum"]
        mensal["valor_sellout_comum_periodo"] = mensal["valor_sellout_comum"]
        mensal["cobertura_sku_comum"] = mensal["valor_sellout_comum"] / mensal["valor_sellin_comum"].replace(0, np.nan)

        periodos_mat_movel, resumo_mat_movel = montar_resumo_mat_movel_individual(mensal, max_mes)

        meses_consecutivos_sellin = maior_sequencia_meses_consecutivos(mensal, "mes", "valor_sellin")
        status_sellin = "Válido" if meses_consecutivos_sellin >= MESES_MAT else "Inválido"

        inicio_ant = periodos["inicio_anterior"]
        fim_ant = periodos["fim_anterior"]
        inicio_atual = periodos["inicio_atual"]
        fim_atual = periodos["fim_atual"]

        mask_ant = (mensal["mes"] >= inicio_ant) & (mensal["mes"] <= fim_ant)
        mask_atual = (mensal["mes"] >= inicio_atual) & (mensal["mes"] <= fim_atual)

        si_ant = mensal.loc[mask_ant, "valor_sellin"].sum()
        si_atual = mensal.loc[mask_atual, "valor_sellin"].sum()
        so_ant = mensal.loc[mask_ant, "valor_sellout"].sum()
        so_atual = mensal.loc[mask_atual, "valor_sellout"].sum()

        inicio_6_atual = (max_mes - pd.DateOffset(months=MESES_MOVEL - 1)).to_period("M").to_timestamp()
        fim_6_anterior = (inicio_6_atual - pd.DateOffset(months=1)).to_period("M").to_timestamp()
        inicio_6_anterior = (fim_6_anterior - pd.DateOffset(months=MESES_MOVEL - 1)).to_period("M").to_timestamp()

        mask_6_ant = (mensal["mes"] >= inicio_6_anterior) & (mensal["mes"] <= fim_6_anterior)
        mask_6_atual = (mensal["mes"] >= inicio_6_atual) & (mensal["mes"] <= max_mes)

        si_6_ant = mensal.loc[mask_6_ant, "valor_sellin"].sum()
        si_6_atual = mensal.loc[mask_6_atual, "valor_sellin"].sum()
        so_6_ant = mensal.loc[mask_6_ant, "valor_sellout"].sum()
        so_6_atual = mensal.loc[mask_6_atual, "valor_sellout"].sum()

        label_6_ant = f"{formatar_periodo(inicio_6_anterior)} a {formatar_periodo(fim_6_anterior)}"
        label_6_atual = f"{formatar_periodo(inicio_6_atual)} a {formatar_periodo(max_mes)}"

        if tem_janela_movel_completa(meses, max_mes, MESES_MOVEL * 2):
            resumo_6m = pd.DataFrame({
                "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
                label_6_ant: [si_6_ant, so_6_ant, divisao_segura(so_6_ant, si_6_ant)],
                label_6_atual: [si_6_atual, so_6_atual, divisao_segura(so_6_atual, si_6_atual)],
                "Tendência %": [variacao(si_6_atual, si_6_ant), variacao(so_6_atual, so_6_ant), variacao(divisao_segura(so_6_atual, si_6_atual), divisao_segura(so_6_ant, si_6_ant))],
            })
        else:
            resumo_6m = pd.DataFrame({
                "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
                label_6_ant: [np.nan, np.nan, np.nan],
                label_6_atual: [np.nan, np.nan, np.nan],
                "Tendência %": [np.nan, np.nan, np.nan],
            })

        mensal_saida = mensal[["mes", "valor_sellin", "valor_sellout", "cobertura", "valor_sellin_comum_periodo", "valor_sellout_comum_periodo", "cobertura_sku_comum"]].rename(columns={
            "mes": "Mês",
            "valor_sellin": "Sell-in",
            "valor_sellout": "Sell-out",
            "cobertura": "Cobertura 12M Móvel",
            "valor_sellin_comum_periodo": "Sell-in SKU em Comum",
            "valor_sellout_comum_periodo": "Sell-out SKU em Comum",
            "cobertura_sku_comum": "Cobertura SKU em Comum",
        })

        inicio_comum_saida = inicio_comum
        fim_comum_saida = max_mes
        modo_periodo = "mensal"

    else:
        # Modo flexível: anual ou total. Usado quando Sell-in não tem mês.
        si2 = si.copy()
        so2 = so.copy()
        si2["ano_calc"] = pd.to_numeric(si2.get("ano", np.nan), errors="coerce")
        so2["ano_calc"] = pd.to_numeric(so2.get("ano", np.nan), errors="coerce")

        anos_si = set(si2["ano_calc"].dropna().astype(int).tolist())
        anos_so = set(so2["ano_calc"].dropna().astype(int).tolist())
        anos_comuns = sorted(anos_si & anos_so)

        if anos_comuns:
            anual_si = si2[si2["ano_calc"].isin(anos_comuns)].groupby("ano_calc", as_index=False)["valor_sellin"].sum()
            anual_so = so2[so2["ano_calc"].isin(anos_comuns)].groupby("ano_calc", as_index=False)["valor_sellout"].sum()
            base_periodo = pd.DataFrame({"ano_calc": anos_comuns})
            mensal = base_periodo.merge(anual_si, on="ano_calc", how="left").merge(anual_so, on="ano_calc", how="left")
            mensal[["valor_sellin", "valor_sellout"]] = mensal[["valor_sellin", "valor_sellout"]].fillna(0)
            mensal["periodo"] = mensal["ano_calc"].astype(int).astype(str)
            mensal["cobertura"] = mensal["valor_sellout"] / mensal["valor_sellin"].replace(0, np.nan)

            ano_atual = max(anos_comuns)
            ano_anterior = ano_atual - 1 if (ano_atual - 1) in anos_comuns else (anos_comuns[-2] if len(anos_comuns) >= 2 else ano_atual)
            label_ant = f"FY {ano_anterior}"
            label_atual = f"FY {ano_atual}"
            si_ant = mensal.loc[mensal["ano_calc"] == ano_anterior, "valor_sellin"].sum()
            si_atual = mensal.loc[mensal["ano_calc"] == ano_atual, "valor_sellin"].sum()
            so_ant = mensal.loc[mensal["ano_calc"] == ano_anterior, "valor_sellout"].sum()
            so_atual = mensal.loc[mensal["ano_calc"] == ano_atual, "valor_sellout"].sum()
            periodos = {
                "tipo": "Ano fechado/anual disponível",
                "label_anterior": label_ant,
                "label_atual": label_atual,
                "inicio_anterior": pd.Timestamp(year=int(ano_anterior), month=1, day=1),
                "fim_anterior": pd.Timestamp(year=int(ano_anterior), month=12, day=1),
                "inicio_atual": pd.Timestamp(year=int(ano_atual), month=1, day=1),
                "fim_atual": pd.Timestamp(year=int(ano_atual), month=12, day=1),
            }
            inicio_comum_saida = pd.Timestamp(year=int(min(anos_comuns)), month=1, day=1)
            fim_comum_saida = pd.Timestamp(year=int(max(anos_comuns)), month=12, day=1)
            modo_periodo = "anual"
        else:
            # Sem mês e sem ano em comum: total geral.
            si_ant = 0
            so_ant = 0
            si_atual = si2["valor_sellin"].sum()
            so_atual = so2["valor_sellout"].sum()
            periodos = {
                "tipo": "Total disponível",
                "label_anterior": "Anterior indisponível",
                "label_atual": "Total disponível",
                "inicio_anterior": pd.NaT,
                "fim_anterior": pd.NaT,
                "inicio_atual": pd.NaT,
                "fim_atual": pd.NaT,
            }
            mensal = pd.DataFrame({
                "periodo": ["Total"],
                "valor_sellin": [si_atual],
                "valor_sellout": [so_atual],
                "cobertura": [divisao_segura(so_atual, si_atual)],
            })
            inicio_comum_saida = pd.NaT
            fim_comum_saida = pd.NaT
            modo_periodo = "total"

        resumo_6m = pd.DataFrame({
            "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
            "Período anterior": [np.nan, np.nan, np.nan],
            "Período atual": [np.nan, np.nan, np.nan],
            "Tendência %": [np.nan, np.nan, np.nan],
        })
        mensal_saida = mensal[["periodo", "valor_sellin", "valor_sellout", "cobertura"]].rename(columns={
            "periodo": "Mês",
            "valor_sellin": "Sell-in",
            "valor_sellout": "Sell-out",
            "cobertura": "Cobertura",
        })
        meses_consecutivos_sellin = 0
        status_sellin = "Inválido"

    tendencia_disponivel = False
    if modo_periodo == "mensal":
        tendencia_disponivel = tem_janela_movel_completa(meses, max_mes, MESES_MAT * 2)
    elif modo_periodo == "anual":
        tendencia_disponivel = len(anos_comuns) >= 2

    resumo_periodo = pd.DataFrame({
        "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
        periodos["label_anterior"]: [si_ant, so_ant, divisao_segura(so_ant, si_ant)],
        periodos["label_atual"]: [si_atual, so_atual, divisao_segura(so_atual, si_atual)],
        "Tendência %": [
            variacao(si_atual, si_ant) if tendencia_disponivel else np.nan,
            variacao(so_atual, so_ant) if tendencia_disponivel else np.nan,
            variacao(divisao_segura(so_atual, si_atual), divisao_segura(so_ant, si_ant)) if tendencia_disponivel else np.nan,
        ],
    })

    if "resumo_mat_movel" not in locals():
        resumo_mat_movel = resumo_periodo[resumo_periodo["Indicador"].isin(["Sell-in", "Sell-out"])].copy()
        periodos_mat_movel = periodos.copy()

    for col in ["Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"]:
        if col not in mensal_saida.columns:
            mensal_saida[col] = np.nan

    if "resumo_sku_comum" not in locals():
        resumo_sku_comum = montar_resumo_sku_comum_mesmo_periodo(si, so, periodos, modo_periodo)

    # Tabela por UF: só usa o que existir. Se não houver UF, ambos terão TOTAL.
    # Em modo mensal, usa sempre a janela de 12 meses móveis, alinhada ao cálculo de cobertura.
    inicio_atual = periodos.get("inicio_atual")
    fim_atual = periodos.get("fim_atual")
    if modo_periodo == "mensal" and pd.notna(periodo_uf_inicio) and pd.notna(periodo_uf_fim):
        si_uf_base = si[(si["mes"] >= periodo_uf_inicio) & (si["mes"] <= periodo_uf_fim)]
        so_uf_base = so[(so["mes"] >= periodo_uf_inicio) & (so["mes"] <= periodo_uf_fim)]
    elif modo_periodo == "anual" and pd.notna(inicio_atual):
        ano_atual = int(pd.Timestamp(inicio_atual).year)
        si_uf_base = si[pd.to_numeric(si.get("ano", np.nan), errors="coerce") == ano_atual]
        so_uf_base = so[pd.to_numeric(so.get("ano", np.nan), errors="coerce") == ano_atual]
        periodo_uf_inicio = pd.Timestamp(year=int(ano_atual), month=1, day=1)
        periodo_uf_fim = pd.Timestamp(year=int(ano_atual), month=12, day=1)
        periodo_uf_tipo = "Ano atual disponível"
    else:
        si_uf_base = si
        so_uf_base = so
        periodo_uf_tipo = "Total disponível"

    if "uf" in si_uf_base.columns and "uf" in so_uf_base.columns:
        uf = montar_tabela_uf_comparacao(si_uf_base, so_uf_base)
    else:
        uf = pd.DataFrame(columns=["UF", "Sell-in", "Sell-out", "Cobertura", "Importância Sell-in", "Importância Sell-out"])

    return {
        "categoria": categoria_nome,
        "categoria_key": categoria_key,
        "inicio_comum": inicio_comum_saida,
        "fim_comum": fim_comum_saida,
        "periodos": periodos,
        "resumo_periodo": resumo_periodo,
        "resumo_6m": resumo_6m,
        "resumo_mat_movel": resumo_mat_movel,
        "periodos_mat_movel": periodos_mat_movel,
        "resumo_sku_comum": resumo_sku_comum,
        "mensal": mensal_saida,
        "uf": uf,
        "skus_excluidos": montar_skus_excluidos_em_comum(si, so),
        "si_atual": si_atual,
        "so_atual": so_atual,
        "cobertura_atual": divisao_segura(so_atual, si_atual),
        "tipo_comparacao": periodos["tipo"],
        "label_anterior": periodos["label_anterior"],
        "label_atual": periodos["label_atual"],
        "modo_periodo": modo_periodo,
        "periodo_uf_inicio": periodo_uf_inicio,
        "periodo_uf_fim": periodo_uf_fim,
        "periodo_uf_tipo": periodo_uf_tipo,
        "status_sellin": status_sellin,
        "meses_consecutivos_sellin": meses_consecutivos_sellin,
    }

def preparar_skus(sellin: pd.DataFrame, sellout: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Cria detalhe de SKUs e resumo por categoria usando janela de 12 meses por categoria/PROD."""
    sellin, sellout = filtrar_bases_sku_12m_por_categoria(sellin, sellout)
    si_sku = (
        sellin.groupby(["categoria_key", "categoria", "ean"], as_index=False)
        .agg(Volume_Sellin_SKU=("valor_sellin", "sum"))
    )

    so_sku = (
        sellout.groupby(["categoria_key", "categoria", "ean", "fabricante", "marca"], as_index=False)
        .agg(Volume_Sellout_SKU=("valor_sellout", "sum"), Nome_SKU=("nome_sku", "first"))
    )

    # Para cada categoria+EAN, pega o fabricante/marca de maior Sell-out.
    so_sku_principal = (
        so_sku.sort_values(["categoria_key", "ean", "Volume_Sellout_SKU"], ascending=[True, True, False])
        .drop_duplicates(["categoria_key", "ean"], keep="first")
        [["categoria_key", "ean", "fabricante", "marca", "Nome_SKU"]]
    )

    # Fabricante referência do Sell-in por categoria:
    # cruza os SKUs do Sell-in com o fabricante encontrado no Sell-out e escolhe o fabricante com maior peso no Sell-in.
    si_com_fab = si_sku.merge(so_sku_principal, on=["categoria_key", "ean"], how="left")
    si_com_fab["fabricante"] = si_com_fab["fabricante"].fillna("").astype(str).str.strip()
    si_com_fab = si_com_fab[si_com_fab["fabricante"].map(normalizar_texto) != ""].copy()

    if not si_com_fab.empty:
        fab_ref = (
            si_com_fab.groupby(["categoria_key", "fabricante"], as_index=False)
            .agg(
                Volume_Sellin_Encontrado_Fabricante=("Volume_Sellin_SKU", "sum"),
                SKUs_Sellin_Encontrados_Fabricante=("ean", "nunique"),
            )
            .sort_values(
                ["categoria_key", "Volume_Sellin_Encontrado_Fabricante", "SKUs_Sellin_Encontrados_Fabricante"],
                ascending=[True, False, False],
            )
            .drop_duplicates("categoria_key", keep="first")
            .rename(columns={"fabricante": "Fabricante referência Sell-in"})
        )
    else:
        fab_ref = pd.DataFrame(columns=[
            "categoria_key", "Fabricante referência Sell-in",
            "Volume_Sellin_Encontrado_Fabricante", "SKUs_Sellin_Encontrados_Fabricante"
        ])

    # União dos SKUs Sell-in x Sell-out.
    chaves = pd.concat([
        si_sku[["categoria_key", "categoria", "ean"]],
        so_sku[["categoria_key", "categoria", "ean"]],
    ], ignore_index=True).drop_duplicates()

    detalhe = chaves.merge(si_sku, on=["categoria_key", "categoria", "ean"], how="left")

    so_total_sku = (
        so_sku.groupby(["categoria_key", "categoria", "ean"], as_index=False)
        .agg(Volume_Sellout_SKU=("Volume_Sellout_SKU", "sum"))
    )
    detalhe = detalhe.merge(so_total_sku, on=["categoria_key", "categoria", "ean"], how="left")
    detalhe = detalhe.merge(so_sku_principal, on=["categoria_key", "ean"], how="left")
    detalhe = detalhe.merge(fab_ref, on="categoria_key", how="left")

    detalhe["Volume_Sellin_SKU"] = pd.to_numeric(detalhe["Volume_Sellin_SKU"], errors="coerce").fillna(0)
    detalhe["Volume_Sellout_SKU"] = pd.to_numeric(detalhe["Volume_Sellout_SKU"], errors="coerce").fillna(0)
    detalhe["fabricante"] = detalhe["fabricante"].fillna("")
    detalhe["marca"] = detalhe["marca"].fillna("")
    detalhe["Nome_SKU"] = detalhe["Nome_SKU"].fillna("")
    detalhe["Fabricante referência Sell-in"] = detalhe["Fabricante referência Sell-in"].fillna("")

    detalhe["Está no Sell-in Cliente"] = detalhe["Volume_Sellin_SKU"] > 0
    detalhe["Está no nosso Sell-out"] = detalhe["Volume_Sellout_SKU"] > 0
    detalhe["SKU em comum"] = detalhe["Está no Sell-in Cliente"] & detalhe["Está no nosso Sell-out"]

    detalhe["Dentro do fabricante Sell-in"] = (
        detalhe["fabricante"].map(normalizar_texto) == detalhe["Fabricante referência Sell-in"].map(normalizar_texto)
    ) & (detalhe["Fabricante referência Sell-in"].map(normalizar_texto) != "")

    detalhe["SKU em comum no fabricante Sell-in"] = detalhe["SKU em comum"] & detalhe["Dentro do fabricante Sell-in"]

    def status_sku(row):
        if row["Está no Sell-in Cliente"] and row["Está no nosso Sell-out"]:
            return "Em comum"
        if row["Está no Sell-in Cliente"] and not row["Está no nosso Sell-out"]:
            return "Só no Sell-in Cliente"
        if not row["Está no Sell-in Cliente"] and row["Está no nosso Sell-out"]:
            return "Só no nosso Sell-out"
        return "Sem volume identificado"

    detalhe["Status SKU"] = detalhe.apply(status_sku, axis=1)

    total_si_cat = detalhe.groupby("categoria_key")["Volume_Sellin_SKU"].transform("sum")
    total_so_cat = detalhe.groupby("categoria_key")["Volume_Sellout_SKU"].transform("sum")

    detalhe["Importância Sell-in SKU"] = detalhe["Volume_Sellin_SKU"] / total_si_cat.replace(0, np.nan)
    detalhe["Importância Sell-out SKU"] = detalhe["Volume_Sellout_SKU"] / total_so_cat.replace(0, np.nan)

    detalhe["Volume Sell-out Fab. Sell-in SKU"] = np.where(
        detalhe["Dentro do fabricante Sell-in"], detalhe["Volume_Sellout_SKU"], 0.0
    )
    total_so_fab = detalhe.groupby("categoria_key")["Volume Sell-out Fab. Sell-in SKU"].transform("sum")
    detalhe["Importância Sell-out SKU Fab. Sell-in"] = detalhe["Volume Sell-out Fab. Sell-in SKU"] / total_so_fab.replace(0, np.nan)

    detalhe = detalhe.rename(columns={
        "categoria": "Categoria",
        "ean": "SKU",
        "fabricante": "Fabricante",
        "marca": "Marca",
        "Volume_Sellin_SKU": "Volume Sell-in SKU",
        "Volume_Sellout_SKU": "Volume Sell-out SKU",
    })

    # Compatibilidade com a lógica do template antigo:
    # antes o "nosso Sell-out" podia vir de SKU SM + Base Congelada.
    # No novo modelo de 2 arquivos, essas informações já vêm dentro do próprio Sell-out/Publicar.
    detalhe["Volume Sell-out Referência SKU"] = detalhe["Volume Sell-out SKU"]
    detalhe["Está no SKU SM/Publicar"] = detalhe["Está no nosso Sell-out"]

    resumo = detalhe.groupby("categoria_key", as_index=False).agg(
        Categoria=("Categoria", "first"),
        Fabricante_referencia_Sellin=("Fabricante referência Sell-in", "first"),
        SKUs_Sellin=("Está no Sell-in Cliente", "sum"),
        SKUs_Sellout=("Está no nosso Sell-out", "sum"),
        SKUs_Em_Comum=("SKU em comum", "sum"),
        SKUs_Em_Comum_Fabricante_Sellin=("SKU em comum no fabricante Sell-in", "sum"),
        Volume_Sellin=("Volume Sell-in SKU", "sum"),
        Volume_Sellout=("Volume Sell-out SKU", "sum"),
        Volume_Sellout_Fabricante_Sellin=("Volume Sell-out Fab. Sell-in SKU", "sum"),
        Volume_Sellin_Encontrado_Fabricante=("Volume_Sellin_Encontrado_Fabricante", "first"),
        SKUs_Sellin_Encontrados_Fabricante=("SKUs_Sellin_Encontrados_Fabricante", "first"),
    )

    resumo["SKUs_Nosso_Sellout"] = resumo["SKUs_Sellout"]
    resumo["Volume_Sellout_Referencia"] = resumo["Volume_Sellout"]

    comum_si = detalhe[detalhe["SKU em comum"]].groupby("categoria_key")["Volume Sell-in SKU"].sum()
    comum_so = detalhe[detalhe["SKU em comum"]].groupby("categoria_key")["Volume Sell-out SKU"].sum()
    comum_so_fab = detalhe[detalhe["SKU em comum no fabricante Sell-in"]].groupby("categoria_key")["Volume Sell-out SKU"].sum()
    total_so_fab_res = detalhe[detalhe["Dentro do fabricante Sell-in"]].groupby("categoria_key")["Volume Sell-out SKU"].sum()

    resumo = resumo.merge(comum_si.rename("Volume_Sellin_Em_Comum"), on="categoria_key", how="left")
    resumo = resumo.merge(comum_so.rename("Volume_Sellout_Em_Comum"), on="categoria_key", how="left")
    resumo = resumo.merge(comum_so_fab.rename("Volume_Sellout_Em_Comum_Fabricante_Sellin"), on="categoria_key", how="left")
    resumo = resumo.merge(total_so_fab_res.rename("Total_Sellout_Fabricante_Sellin"), on="categoria_key", how="left")

    for col in [
        "Volume_Sellin_Em_Comum", "Volume_Sellout_Em_Comum",
        "Volume_Sellout_Em_Comum_Fabricante_Sellin", "Total_Sellout_Fabricante_Sellin",
        "Volume_Sellin_Encontrado_Fabricante", "SKUs_Sellin_Encontrados_Fabricante",
    ]:
        resumo[col] = pd.to_numeric(resumo[col], errors="coerce").fillna(0)

    resumo["Cobertura importância Sell-in"] = resumo["Volume_Sellin_Em_Comum"] / resumo["Volume_Sellin"].replace(0, np.nan)
    resumo["Cobertura importância Sell-out"] = resumo["Volume_Sellout_Em_Comum"] / resumo["Volume_Sellout"].replace(0, np.nan)
    resumo["Cobertura importância Sell-out Fab. Sell-in"] = (
        resumo["Volume_Sellout_Em_Comum_Fabricante_Sellin"] / resumo["Total_Sellout_Fabricante_Sellin"].replace(0, np.nan)
    )
    resumo["% Sell-in usado para detectar fabricante"] = (
        resumo["Volume_Sellin_Encontrado_Fabricante"] / resumo["Volume_Sellin"].replace(0, np.nan)
    )

    ordem_detalhe = [
        "Categoria", "SKU", "Nome_SKU", "Marca", "Fabricante", "Fabricante referência Sell-in",
        "Está no Sell-in Cliente", "Está no nosso Sell-out", "Está no SKU SM/Publicar", "SKU em comum",
        "Dentro do fabricante Sell-in", "SKU em comum no fabricante Sell-in", "Status SKU",
        "Volume Sell-in SKU", "Volume Sell-out SKU", "Volume Sell-out Referência SKU", "Volume Sell-out Fab. Sell-in SKU",
        "Importância Sell-in SKU", "Importância Sell-out SKU", "Importância Sell-out SKU Fab. Sell-in",
        "categoria_key",
    ]
    existentes = [c for c in ordem_detalhe if c in detalhe.columns]
    outros = [c for c in detalhe.columns if c not in existentes]
    detalhe = detalhe[existentes + outros].sort_values(["Categoria", "Volume Sell-in SKU", "Volume Sell-out SKU"], ascending=[True, False, False])

    return detalhe, resumo


# ============================================================
# Escrita Excel
# ============================================================


def aplicar_formatos_basicos(writer, sheet_name, df, startrow=0, startcol=0, percent_cols=None, number_cols=None, date_cols=None):
    percent_cols = set(percent_cols or [])
    number_cols = set(number_cols or [])
    date_cols = set(date_cols or [])

    workbook = writer.book
    ws = writer.sheets[sheet_name]
    fmt_header = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "align": "center", "valign": "vcenter"})
    fmt_num = workbook.add_format({"num_format": "#,##0", "border": 1})
    fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
    fmt_date = workbook.add_format({"num_format": "mmm/yy", "border": 1})
    fmt_text = workbook.add_format({"border": 1})

    for j, col in enumerate(df.columns):
        ws.write(startrow, startcol + j, col, fmt_header)
        serie = df[col] if len(df) else pd.Series([], dtype=object)
        max_len = max([len(str(col))] + [len(str(v)) for v in serie.head(250).fillna("")])
        width = min(max(max_len + 2, 10), 42)
        fmt = fmt_text
        if col in percent_cols:
            fmt = fmt_pct
            width = min(max(width, 12), 18)
        elif col in number_cols:
            fmt = fmt_num
            width = min(max(width, 14), 20)
        elif col in date_cols:
            fmt = fmt_date
            width = min(max(width, 12), 16)
        ws.set_column(startcol + j, startcol + j, width, fmt)

    if len(df) > 0:
        ws.autofilter(startrow, startcol, startrow + len(df), startcol + len(df.columns) - 1)




def valor_finito(valor) -> Optional[float]:
    """Retorna float somente quando o valor é numérico e finito; caso contrário, None."""
    try:
        if pd.isna(valor):
            return None
        numero = float(valor)
        if not math.isfinite(numero):
            return None
        return numero
    except Exception:
        return None


def write_number_ou_branco(ws, row: int, col: int, valor, fmt=None) -> None:
    """Evita erro do xlsxwriter com NaN/INF, escrevendo célula vazia quando necessário."""
    numero = valor_finito(valor)
    if numero is None:
        ws.write_blank(row, col, None, fmt)
    else:
        ws.write_number(row, col, numero, fmt)

def escrever_categoria(writer, nome_aba: str, resultado: Dict[str, object]):
    """
    Escreve a aba da categoria no padrão visual do template "Calculo Cobertura".
    """
    workbook = writer.book
    ws = workbook.add_worksheet(nome_aba)
    writer.sheets[nome_aba] = ws

    azul_escuro = "#1F4E78"
    azul_claro = "#D9EAF7"
    azul_linha = "#156082"
    laranja_linha = "#E97132"
    cinza_texto = "#595959"

    fmt_titulo = workbook.add_format({
        "bold": True, "font_size": 16, "font_color": "white",
        "bg_color": azul_escuro, "align": "center", "valign": "vcenter",
        "border": 1,
    })
    fmt_top_label = workbook.add_format({
        "bold": True, "font_color": "white", "bg_color": azul_escuro,
        "align": "center", "valign": "vcenter", "border": 1,
    })
    fmt_top_value = workbook.add_format({
        "bold": True, "font_color": azul_escuro, "bg_color": "#EAF3F8",
        "align": "center", "valign": "vcenter", "border": 1,
    })
    fmt_secao = workbook.add_format({
        "bold": True, "bg_color": azul_claro, "border": 1,
        "align": "center", "valign": "vcenter",
    })
    fmt_header = workbook.add_format({
        "bold": True, "bg_color": azul_claro, "border": 1,
        "align": "center", "valign": "vcenter",
    })
    fmt_text = workbook.add_format({"border": 1, "align": "left", "valign": "vcenter"})
    fmt_text_center = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
    fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
    fmt_date = workbook.add_format({"num_format": "mmm/yy", "border": 1, "align": "center"})
    fmt_blank = workbook.add_format({"border": 1})
    fmt_link = workbook.add_format({"font_color": "#0563C1", "underline": 1, "border": 1, "align": "center", "valign": "vcenter"})

    # Larguras iguais/próximas ao template.
    ws.set_column("A:A", 2.7)
    ws.set_column("B:G", 11.7)
    ws.set_column("H:H", 15.9)
    ws.set_column("I:K", 11.7)
    ws.set_column("L:M", 18)
    ws.set_column("N:Q", 11)
    ws.set_default_row(18)

    categoria_nome = str(resultado.get("categoria", ""))
    fabricante_exibido = str(resultado.get("fabricante_exibido", "") or "").strip()
    if not fabricante_exibido:
        fabricante_exibido = "Não informado"
    periodos = resultado.get("periodos", {}) or {}
    resumo_periodo = limpar_dataframe_excel(resultado["resumo_periodo"])
    resumo_6m = limpar_dataframe_excel(resultado["resumo_6m"])
    mensal = limpar_dataframe_excel(resultado["mensal"])
    uf = limpar_dataframe_excel(resultado["uf"])
    skus_excluidos = limpar_dataframe_excel(resultado.get("skus_excluidos", pd.DataFrame()))

    label_ant = resultado.get("label_anterior", periodos.get("label_anterior", "MAT-1"))
    label_atual = resultado.get("label_atual", periodos.get("label_atual", "MAT"))
    tipo_comp = str(resultado.get("tipo_comparacao", ""))

    # Cabeçalho superior.
    ws.write_url("L1", "internal:'Resumo Categorias'!A1", fmt_link, string="Voltar ao Resumo")
    ws.merge_range("B1:D2", "Estudo de Cobertura", fmt_titulo)
    ws.write("E1", "Categoria", fmt_top_label)
    ws.write("E2", categoria_nome, fmt_top_value)
    ws.write("G1", "Fabricante", fmt_top_label)
    ws.write("G2", fabricante_exibido, fmt_top_value)

    # Bloco superior esquerdo: MAT/YTD.
    periodo_atual_txt = periodo_texto(periodos.get("inicio_atual"), periodos.get("fim_atual"))
    ws.write("B6", "Tipo de Cálculo", fmt_secao)
    ws.merge_range("C6:F6", f"{tipo_comp} - {periodo_atual_txt}".strip(" -"), fmt_secao)

    ws.write("B7", "12 meses", fmt_header)
    ws.write_blank("C7", None, fmt_blank)
    ws.write("D7", f"Volume {label_ant}", fmt_header)
    ws.write("E7", f"Volume {label_atual}", fmt_header)
    ws.write("F7", "Tendência", fmt_header)

    def valor_resumo(indicador, coluna):
        return extrair_valor_resumo(resumo_periodo, indicador, coluna)

    indicadores = ["Sell-in", "Sell-out"]
    for i, indicador in enumerate(indicadores, start=8):
        ws.write(i - 1, 2, indicador.replace("-", ""), fmt_text_center)  # C
        write_number_ou_branco(ws, i - 1, 3, valor_resumo(indicador, label_ant), fmt_num)  # D
        write_number_ou_branco(ws, i - 1, 4, valor_resumo(indicador, label_atual), fmt_num)  # E
        write_number_ou_branco(ws, i - 1, 5, valor_resumo(indicador, "Tendência %"), fmt_pct)  # F

    # Bloco superior direito: 6 meses móveis.
    cols_6m = [c for c in resumo_6m.columns if c not in {"Indicador", "Tendência %"}]
    label_6_ant = cols_6m[0] if len(cols_6m) >= 1 else "6M Ant."
    label_6_atual = cols_6m[1] if len(cols_6m) >= 2 else "6M Atual"
    ws.merge_range("H6:K6", f"6 meses móveis - {label_6_atual}", fmt_secao)
    ws.write_blank("H7", None, fmt_blank)
    ws.write("I7", "Volume 6M Ant.", fmt_header)
    ws.write("J7", "Volume 6M Atual", fmt_header)
    ws.write("K7", "Tendência", fmt_header)

    for i, indicador in enumerate(indicadores, start=8):
        ws.write(i - 1, 7, indicador.replace("-", ""), fmt_text_center)  # H
        write_number_ou_branco(ws, i - 1, 8, extrair_valor_resumo(resumo_6m, indicador, label_6_ant), fmt_num)
        write_number_ou_branco(ws, i - 1, 9, extrair_valor_resumo(resumo_6m, indicador, label_6_atual), fmt_num)
        write_number_ou_branco(ws, i - 1, 10, extrair_valor_resumo(resumo_6m, indicador, "Tendência %"), fmt_pct)

    # Períodos.
    ws.merge_range("B12:D12", "Período", fmt_secao)
    ws.write("C13", "Dt Inic", fmt_header)
    ws.write("D13", "Dt fim", fmt_header)
    ws.write("B14", "Sellin", fmt_text_center)
    ws.write("B15", "Sellout", fmt_text_center)
    for row in (14, 15):
        for col, key in [(3, "inicio_atual"), (4, "fim_atual")]:
            dt = periodos.get(key)
            if pd.notna(dt):
                ws.write_datetime(row - 1, col - 1, pd.Timestamp(dt).to_pydatetime(), fmt_date)
            else:
                ws.write_blank(row - 1, col - 1, None, fmt_date)

    # Bloco MAT móvel real: últimos 12 meses contra os 12 meses anteriores.
    periodos_mat_movel = resultado.get("periodos_mat_movel", {}) or {}
    resumo_mat_movel = limpar_dataframe_excel(resultado.get("resumo_mat_movel", resumo_periodo))
    label_mat_ant = periodos_mat_movel.get("label_anterior", "MAT-1")
    label_mat_atual = periodos_mat_movel.get("label_atual", "MAT")

    ws.merge_range("H12:K12", "MAT", fmt_secao)
    ws.write("I13", f"Volume {label_mat_ant}", fmt_header)
    ws.write("J13", f"Volume {label_mat_atual}", fmt_header)
    ws.write("K13", "Tendência", fmt_header)
    for i, row in resumo_mat_movel.reset_index(drop=True).iterrows():
        if i >= 3:
            break
        r = 13 + i
        indicador_mat = row.get("Indicador", "")
        ws.write(r, 7, str(indicador_mat).replace("-", ""), fmt_text_center)
        write_number_ou_branco(ws, r, 8, row.get(label_mat_ant, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 9, row.get(label_mat_atual, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 10, row.get("Tendência %", row.get("Tendência", np.nan)), fmt_pct)

    # Tabela mensal principal em B:E, no mesmo padrão do layout de comparação.
    start_mensal = 17  # linha 18 no Excel
    for j, header in enumerate(["Data", "Sell-in", "Sell-out", "Cobertura"], start=1):
        ws.write(start_mensal, j, header, fmt_header)

    primeira_dados_mensal = start_mensal + 1
    for i in range(len(mensal)):
        row = primeira_dados_mensal + i
        mes_val = mensal.iloc[i].get("Mês", "")
        if resultado.get("modo_periodo") == "mensal" and pd.notna(mes_val):
            try:
                ws.write_datetime(row, 1, pd.Timestamp(mes_val).to_pydatetime(), fmt_date)
            except Exception:
                ws.write(row, 1, str(mes_val), fmt_text)
        else:
            ws.write(row, 1, str(mes_val), fmt_text)

        write_number_ou_branco(ws, row, 2, mensal.iloc[i].get("Sell-in", np.nan), fmt_num)
        write_number_ou_branco(ws, row, 3, mensal.iloc[i].get("Sell-out", np.nan), fmt_num)

        if resultado.get("modo_periodo") == "mensal":
            if i < MESES_MAT - 1:
                ws.write_blank(row, 4, None, fmt_pct)
            else:
                linha_excel = row + 1
                linha_inicio = linha_excel - MESES_MAT + 1
                formula = f'=IFERROR(SUM(D{linha_inicio}:D{linha_excel})/SUM(C{linha_inicio}:C{linha_excel}),"")'
                valor_calc = mensal.iloc[i].get("Cobertura 12M Móvel", np.nan)
                ws.write_formula(row, 4, formula, fmt_pct, float(valor_calc) if pd.notna(valor_calc) else "")
        else:
            cobertura_col = "Cobertura 12M Móvel" if "Cobertura 12M Móvel" in mensal.columns else "Cobertura"
            write_number_ou_branco(ws, row, 4, mensal.iloc[i].get(cobertura_col, np.nan), fmt_pct)

    # Tabela por UF ao lado da tabela mensal, em I:N.
    start_uf = start_mensal
    uf_headers = ["UF", "Sell-in 12M", "Sell-out 12M", "Cobertura", "Importância Sell-in", "Importância Sell-out"]
    for j, h in enumerate(uf_headers, start=8):
        ws.write(start_uf, j, h, fmt_header)

    for i in range(len(uf)):
        row = start_uf + 1 + i
        uf_nome = str(uf.iloc[i].get("UF", ""))
        ws.write(row, 8, uf_nome, fmt_text)
        if normalizar_texto(uf_nome) == "total":
            first_data_excel = start_uf + 2
            total_row_excel = row + 1
            last_sum_excel = total_row_excel - 1
            ws.write_formula(row, 9, f'=SUM(J{first_data_excel}:J{last_sum_excel})', fmt_num, valor_finito(uf.iloc[i].get("Sell-in", np.nan)) or 0)
            ws.write_formula(row, 10, f'=SUM(K{first_data_excel}:K{last_sum_excel})', fmt_num, valor_finito(uf.iloc[i].get("Sell-out", np.nan)) or 0)
            ws.write_formula(row, 11, f'=IFERROR(K{total_row_excel}/J{total_row_excel},"")', fmt_pct, valor_finito(uf.iloc[i].get("Cobertura", np.nan)) or "")
            ws.write_number(row, 12, 1, fmt_pct)
            ws.write_number(row, 13, 1, fmt_pct)
        else:
            write_number_ou_branco(ws, row, 9, uf.iloc[i].get("Sell-in", np.nan), fmt_num)
            write_number_ou_branco(ws, row, 10, uf.iloc[i].get("Sell-out", np.nan), fmt_num)
            write_number_ou_branco(ws, row, 11, uf.iloc[i].get("Cobertura", np.nan), fmt_pct)
            write_number_ou_branco(ws, row, 12, uf.iloc[i].get("Importância Sell-in", np.nan), fmt_pct)
            write_number_ou_branco(ws, row, 13, uf.iloc[i].get("Importância Sell-out", np.nan), fmt_pct)

    # Gráfico principal abaixo das tabelas superiores.
    chart_row = max(primeira_dados_mensal + len(mensal) + 1, start_uf + len(uf) + 2)
    if len(mensal) > 0:
        last_row = primeira_dados_mensal + len(mensal) - 1

        chart = workbook.add_chart({"type": "line"})
        chart.add_series({
            "name": "Sell-in",
            "categories": [nome_aba, primeira_dados_mensal, 1, last_row, 1],
            "values": [nome_aba, primeira_dados_mensal, 2, last_row, 2],
            "line": {"color": azul_linha, "width": 2.25},
        })
        chart.add_series({
            "name": "Sell-out",
            "categories": [nome_aba, primeira_dados_mensal, 1, last_row, 1],
            "values": [nome_aba, primeira_dados_mensal, 3, last_row, 3],
            "line": {"color": laranja_linha, "width": 2.25},
        })
        chart.set_title({"name": "Sell-in x Sell-out por mês", "name_font": {"color": "#000000", "size": 16, "bold": True}})
        chart.set_x_axis({"name": "Mês", "num_format": "mmm/yy", "label_position": "low", "num_font": {"rotation": -90}})
        chart.set_y_axis({"name": "Volume", "num_format": "#,##0"})
        chart.set_legend({"position": "bottom"})
        chart.set_size({"width": 720, "height": 340})
        chart.set_plotarea({"border": {"none": True}})
        chart.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_row, 1, chart)

    # Tabela separada de SKU em comum, usando o mesmo período da tabela superior.
    sku_title_row = chart_row + 18
    resumo_sku_comum = limpar_dataframe_excel(resultado.get("resumo_sku_comum", pd.DataFrame()))
    if resumo_sku_comum.empty:
        resumo_sku_comum = pd.DataFrame({
            "Indicador": ["Sell-in", "Sell-out", "Cobertura"],
            str(label_ant): [np.nan, np.nan, np.nan],
            str(label_atual): [np.nan, np.nan, np.nan],
            "Tendência %": [np.nan, np.nan, np.nan],
        })

    cols_sku_resumo = list(resumo_sku_comum.columns)
    largura_sku_resumo = max(4, len(cols_sku_resumo))
    ws.merge_range(sku_title_row, 1, sku_title_row, largura_sku_resumo, "SKU em Comum - mesmo período do cálculo superior", fmt_secao)
    sku_header_row = sku_title_row + 1
    for j, header in enumerate(cols_sku_resumo, start=1):
        ws.write(sku_header_row, j, header, fmt_header)

    primeira_sku_row = sku_header_row + 1
    for i, (_, linha_sku) in enumerate(resumo_sku_comum.reset_index(drop=True).iterrows()):
        row = primeira_sku_row + i
        indicador_sku = str(linha_sku.get("Indicador", ""))
        for j, header in enumerate(cols_sku_resumo, start=1):
            val = linha_sku.get(header, np.nan)
            if header == "Indicador":
                ws.write(row, j, indicador_sku, fmt_text_center)
            elif "Tendência" in str(header) or indicador_sku == "Cobertura":
                write_number_ou_branco(ws, row, j, val, fmt_pct)
            elif "Qtd" in indicador_sku:
                write_number_ou_branco(ws, row, j, val, fmt_num)
            else:
                write_number_ou_branco(ws, row, j, val, fmt_num)

    # Detalhe mensal de apoio: mostra os meses do período atual do cálculo superior,
    # mas sem recalcular 12M. A cobertura mensal é Sell-out comum / Sell-in comum.
    detalhe_sku_title_row = primeira_sku_row + max(len(resumo_sku_comum), 1) + 2
    ws.merge_range(detalhe_sku_title_row, 1, detalhe_sku_title_row, 4, "Detalhe mensal do SKU em Comum - período atual", fmt_secao)
    detalhe_sku_header_row = detalhe_sku_title_row + 1
    for j, header in enumerate(["Data", "Sell-in", "Sell-out", "Cobertura"], start=1):
        ws.write(detalhe_sku_header_row, j, header, fmt_header)

    inicio_sku_atual = periodos.get("inicio_atual")
    fim_sku_atual = periodos.get("fim_atual")
    mensal_sku_detalhe = mensal.copy()
    if resultado.get("modo_periodo") == "mensal" and pd.notna(inicio_sku_atual) and pd.notna(fim_sku_atual):
        mensal_sku_detalhe = mensal_sku_detalhe[(pd.to_datetime(mensal_sku_detalhe["Mês"], errors="coerce") >= pd.Timestamp(inicio_sku_atual)) & (pd.to_datetime(mensal_sku_detalhe["Mês"], errors="coerce") <= pd.Timestamp(fim_sku_atual))].copy()

    primeira_sku_detalhe_row = detalhe_sku_header_row + 1
    for i in range(len(mensal_sku_detalhe)):
        row = primeira_sku_detalhe_row + i
        mes_val = mensal_sku_detalhe.iloc[i].get("Mês", "")
        if resultado.get("modo_periodo") == "mensal" and pd.notna(mes_val):
            try:
                ws.write_datetime(row, 1, pd.Timestamp(mes_val).to_pydatetime(), fmt_date)
            except Exception:
                ws.write(row, 1, str(mes_val), fmt_text)
        else:
            ws.write(row, 1, str(mes_val), fmt_text)
        write_number_ou_branco(ws, row, 2, mensal_sku_detalhe.iloc[i].get("Sell-in SKU em Comum", np.nan), fmt_num)
        write_number_ou_branco(ws, row, 3, mensal_sku_detalhe.iloc[i].get("Sell-out SKU em Comum", np.nan), fmt_num)
        linha_excel = row + 1
        valor_comum = mensal_sku_detalhe.iloc[i].get("Cobertura SKU em Comum", np.nan)
        ws.write_formula(row, 4, f'=IFERROR(D{linha_excel}/C{linha_excel},"")', fmt_pct, float(valor_comum) if pd.notna(valor_comum) else "")

    # Tabela dos SKUs excluídos do cálculo de SKU em comum.
    # Estes são os SKUs que aparecem em apenas uma das bases e, por isso,
    # não entram no bloco "SKU em Comum".
    excluidos_title_row = primeira_sku_detalhe_row + len(mensal_sku_detalhe) + 2
    ws.merge_range(excluidos_title_row, 1, excluidos_title_row, 7, "SKUs excluídos do cálculo de SKU em Comum", fmt_secao)
    excluidos_header_row = excluidos_title_row + 1
    excluidos_cols = ["SKU", "Status", "Sell-in", "Sell-out", "Nome SKU", "Marca", "Fabricante"]
    for j, header in enumerate(excluidos_cols, start=1):
        ws.write(excluidos_header_row, j, header, fmt_header)

    if skus_excluidos.empty:
        ws.merge_range(excluidos_header_row + 1, 1, excluidos_header_row + 1, 7, "Nenhum SKU excluído. Todos os SKUs relevantes estão em comum entre Sell-in e Sell-out.", fmt_text_center)
        excluidos_fim_row = excluidos_header_row + 1
    else:
        for i, (_, sku_row) in enumerate(skus_excluidos.reset_index(drop=True).iterrows()):
            row = excluidos_header_row + 1 + i
            ws.write(row, 1, str(sku_row.get("SKU", "")), fmt_text)
            ws.write(row, 2, str(sku_row.get("Status", "")), fmt_text_center)
            write_number_ou_branco(ws, row, 3, sku_row.get("Sell-in", np.nan), fmt_num)
            write_number_ou_branco(ws, row, 4, sku_row.get("Sell-out", np.nan), fmt_num)
            ws.write(row, 5, str(sku_row.get("Nome SKU", "")), fmt_text)
            ws.write(row, 6, str(sku_row.get("Marca", "")), fmt_text)
            ws.write(row, 7, str(sku_row.get("Fabricante", "")), fmt_text)
        excluidos_fim_row = excluidos_header_row + len(skus_excluidos)

    # Gráfico do SKU em comum abaixo da tabela de SKU em comum e da lista de excluídos.
    if len(mensal_sku_detalhe) > 0:
        last_sku_row = primeira_sku_detalhe_row + len(mensal_sku_detalhe) - 1
        chart_comum_row = excluidos_fim_row + 2
        chart_comum = workbook.add_chart({"type": "line"})
        chart_comum.add_series({
            "name": "Sell-in SKU em comum",
            "categories": [nome_aba, primeira_sku_detalhe_row, 1, last_sku_row, 1],
            "values": [nome_aba, primeira_sku_detalhe_row, 2, last_sku_row, 2],
            "line": {"color": azul_linha, "width": 2.25},
        })
        chart_comum.add_series({
            "name": "Sell-out SKU em comum",
            "categories": [nome_aba, primeira_sku_detalhe_row, 1, last_sku_row, 1],
            "values": [nome_aba, primeira_sku_detalhe_row, 3, last_sku_row, 3],
            "line": {"color": laranja_linha, "width": 2.25},
        })
        chart_comum.set_title({"name": "Sell-in SKU em Comum x Sell-out SKU em Comum", "name_font": {"color": "#000000", "size": 14, "bold": True}})
        chart_comum.set_x_axis({"name": "Mês", "num_format": "mmm/yy", "label_position": "low", "num_font": {"rotation": -90}})
        chart_comum.set_y_axis({"name": "Volume", "num_format": "#,##0"})
        chart_comum.set_legend({"position": "bottom"})
        chart_comum.set_size({"width": 720, "height": 340})
        chart_comum.set_plotarea({"border": {"none": True}})
        chart_comum.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_comum_row, 1, chart_comum)


def gerar_descricao_calculos() -> pd.DataFrame:
    """Cria a documentação das colunas calculadas do arquivo final."""
    linhas = [
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Cobertura atual",
            "Explicação": "Mostra quanto do Sell-in do período atual está coberto pelo Sell-out correspondente.",
            "Como foi calculado": "Sell-out atual / Sell-in atual. O período atual é FY quando há anos fechados consecutivos; caso contrário, é YTD.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Status Sell-in",
            "Explicação": "Indica se a categoria/PROD tem base mensal suficiente no Sell-in para cobertura 12 meses móveis.",
            "Como foi calculado": "Válido quando existe sequência de pelo menos 12 meses consecutivos com Sell-in maior que zero; caso contrário, Inválido.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Qtd SKUs Sell-in / Qtd SKUs Sell-out / Qtd SKUs em comum",
            "Explicação": "Mostra quantos SKUs aparecem no Sell-in, no Sell-out e nas duas bases ao mesmo tempo.",
            "Como foi calculado": "Contagem distinta de SKU/EAN por categoria/PROD em cada base e na intersecção entre Sell-in e Sell-out.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Importância Sell-in",
            "Explicação": "Mostra o peso da categoria/PROD dentro do Sell-in total do resumo, usando o mesmo período atual do cálculo principal.",
            "Como foi calculado": "Sell-in atual da categoria/PROD / soma do Sell-in atual de todas as categorias/PRODs. A aba Resumo Categorias é ordenada por essa coluna, do maior para o menor.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Tendência Sell-in / Tendência Sell-out / Diferença Tendência",
            "Explicação": "Compara a tendência de crescimento/queda do Sell-in e do Sell-out na categoria/PROD e mostra a diferença entre elas.",
            "Como foi calculado": "Tendência Sell-in = (Sell-in do período atual / Sell-in do período anterior) - 1. Tendência Sell-out = (Sell-out do período atual / Sell-out do período anterior) - 1. Diferença Tendência = Tendência Sell-out - Tendência Sell-in. O período é ano fechado quando houver; caso contrário, YTD.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Sell-in MAT-1 / Sell-in MAT / Tendência Sell-in",
            "Explicação": "Reproduz no resumo principal a comparação do bloco MAT/YTD calculada nas abas de categoria.",
            "Como foi calculado": "Soma do Sell-in no período anterior e no período atual. A variação é (Sell-in atual / Sell-in anterior) - 1.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Sell-out MAT-1 / Sell-out MAT / Tendência Sell-out",
            "Explicação": "Reproduz no resumo principal a comparação do bloco MAT/YTD calculada nas abas de categoria.",
            "Como foi calculado": "Soma do Sell-out no período anterior e no período atual. A variação é (Sell-out atual / Sell-out anterior) - 1.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Sell-in 6M / Sell-out 6M / Cobertura 6M Atual",
            "Explicação": "Resume a comparação dos últimos 6 meses contra os 6 meses anteriores.",
            "Como foi calculado": "Soma dos últimos 6 meses disponíveis e dos 6 meses imediatamente anteriores. Cobertura 6M Atual = Sell-out 6M Atual / Sell-in 6M Atual.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Tipo comparação",
            "Explicação": "Indica se a comparação usou ano fechado ou YTD.",
            "Como foi calculado": "Ano fechado quando existem dois anos completos consecutivos. Se o último ano estiver incompleto, usa YTD até o último mês disponível contra o mesmo período do ano anterior.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Cobertura importância Sell-in",
            "Explicação": "Percentual da importância do Sell-in do cliente que também aparece no nosso Sell-out, considerando os SKUs em comum.",
            "Como foi calculado": "Volume_Sellin_Em_Comum / Volume_Sellin.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Cobertura importância Sell-out",
            "Explicação": "Percentual da importância do nosso Sell-out coberto pelos SKUs em comum com o Sell-in.",
            "Como foi calculado": "Volume_Sellout_Em_Comum / Volume_Sellout.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Cobertura importância Sell-out Fab. Sell-in",
            "Explicação": "Percentual da importância do nosso Sell-out coberto pelos SKUs em comum, limitado ao fabricante de referência do Sell-in.",
            "Como foi calculado": "Volume_Sellout_Em_Comum_Fabricante_Sellin / Total_Sellout_Fabricante_Sellin.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "% Sell-in usado para detectar fabricante",
            "Explicação": "Quanto do Sell-in conseguiu ser usado para identificar o fabricante de referência pelo cruzamento com o Sell-out.",
            "Como foi calculado": "Volume_Sellin_Encontrado_Fabricante / Volume_Sellin.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Cobertura",
            "Explicação": "Na tabela mensal, é a cobertura em 12 meses móveis, não a cobertura mensal simples.",
            "Como foi calculado": "SOMA(Sell-out dos últimos 12 meses) / SOMA(Sell-in dos últimos 12 meses). Exemplo: D40 = SOMA(C29:C40) / SOMA(B29:B40).",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Sell-in SKU em Comum / Sell-out SKU em Comum / Cobertura SKU em Comum",
            "Explicação": "Mostra a cobertura considerando apenas os SKUs que existem simultaneamente no Sell-in e no Sell-out.",
            "Como foi calculado": "O bloco SKU em Comum usa o mesmo período definido na tabela superior da aba, tanto no período anterior quanto no atual. A única diferença é restringir as somas aos SKUs/EANs em comum dentro de cada período; depois calcula Sell-out em comum / Sell-in em comum.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "MAT",
            "Explicação": "A tabela MAT da área I:L usa tendência de 12 meses móveis, separada da tendência FY/YTD da tabela superior.",
            "Como foi calculado": "MAT = últimos 12 meses encerrando no último mês comum disponível. MAT-1 = 12 meses imediatamente anteriores. Tendência = MAT / MAT-1 - 1.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Tendência %",
            "Explicação": "Mostra a variação percentual entre o período anterior e o período atual para Sell-in, Sell-out ou Cobertura.",
            "Como foi calculado": "(Valor atual / Valor anterior) - 1. Quando o valor anterior é zero, o resultado fica vazio.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Resumo MAT/YTD",
            "Explicação": "Resumo comparativo entre os períodos de referência usados para a categoria.",
            "Como foi calculado": "Soma Sell-in e Sell-out no período anterior e no atual. A cobertura do bloco é Sell-out / Sell-in em cada período.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Resumo 6 meses móveis",
            "Explicação": "Compara os últimos 6 meses disponíveis contra os 6 meses anteriores.",
            "Como foi calculado": "Soma os 6 meses mais recentes e compara com a soma dos 6 meses imediatamente anteriores. Cobertura = Sell-out / Sell-in.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Importância Sell-in",
            "Explicação": "Na tabela por UF, mostra o peso daquela UF dentro do Sell-in da categoria na mesma janela usada para a cobertura 12 meses móveis.",
            "Como foi calculado": "Sell-in da UF nos últimos 12 meses móveis / Sell-in total da categoria nos mesmos últimos 12 meses móveis.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Importância Sell-out",
            "Explicação": "Na tabela por UF, mostra o peso daquela UF dentro do Sell-out da categoria na mesma janela usada para a cobertura 12 meses móveis.",
            "Como foi calculado": "Sell-out da UF nos últimos 12 meses móveis / Sell-out total da categoria nos mesmos últimos 12 meses móveis.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "SKU em comum",
            "Explicação": "Indica se o SKU existe tanto no Sell-in quanto no nosso Sell-out.",
            "Como foi calculado": "VERDADEIRO quando o SKU tem volume de Sell-in maior que zero e volume de Sell-out maior que zero.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "Fabricante referência Sell-in",
            "Explicação": "Fabricante usado como referência para limitar a visão de importância do nosso Sell-out.",
            "Como foi calculado": "Identificado pelo fabricante predominante dos SKUs do Sell-in encontrados no Sell-out, considerando o maior volume de Sell-in encontrado.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "Dentro do fabricante Sell-in",
            "Explicação": "Indica se o SKU pertence ao fabricante de referência do Sell-in.",
            "Como foi calculado": "Compara o fabricante do SKU no Sell-out com o Fabricante referência Sell-in da categoria.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "Importância Sell-in SKU",
            "Explicação": "Peso do SKU dentro do Sell-in da categoria.",
            "Como foi calculado": "Volume Sell-in SKU / Volume total de Sell-in da categoria.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "Importância Sell-out SKU",
            "Explicação": "Peso do SKU dentro do nosso Sell-out da categoria.",
            "Como foi calculado": "Volume Sell-out SKU / Volume total de Sell-out da categoria.",
        },
        {
            "Nome da aba": "Base SKUs",
            "Nome da coluna calculada": "Importância Sell-out SKU Fab. Sell-in",
            "Explicação": "Peso do SKU dentro do Sell-out apenas do fabricante de referência do Sell-in.",
            "Como foi calculado": "Volume Sell-out Fab. Sell-in SKU / Volume total do Sell-out do fabricante de referência na categoria.",
        },
        {
            "Nome da aba": "SKUs por Categoria",
            "Nome da coluna calculada": "Cobertura importância Sell-in",
            "Explicação": "Resumo por categoria do percentual do Sell-in coberto pelos SKUs que também estão no Sell-out.",
            "Como foi calculado": "Volume_Sellin_Em_Comum / Volume_Sellin.",
        },
        {
            "Nome da aba": "SKUs por Categoria",
            "Nome da coluna calculada": "Cobertura importância Sell-out",
            "Explicação": "Resumo por categoria do percentual do Sell-out coberto pelos SKUs que também estão no Sell-in.",
            "Como foi calculado": "Volume_Sellout_Em_Comum / Volume_Sellout.",
        },
        {
            "Nome da aba": "SKUs por Categoria",
            "Nome da coluna calculada": "Cobertura importância Sell-out Fab. Sell-in",
            "Explicação": "Resumo por categoria do percentual do Sell-out do fabricante de referência coberto pelos SKUs em comum.",
            "Como foi calculado": "Volume_Sellout_Em_Comum_Fabricante_Sellin / Total_Sellout_Fabricante_Sellin.",
        },
        {
            "Nome da aba": "Resumo Categorias",
            "Nome da coluna calculada": "Abrir aba",
            "Explicação": "Link interno para abrir diretamente a aba da categoria/PROD correspondente.",
            "Como foi calculado": "O código grava um hyperlink interno para a célula A1 da aba gerada para cada categoria/PROD.",
        },
        {
            "Nome da aba": "Parâmetros",
            "Nome da coluna calculada": "Volume variável aplicado / Gramatura média global",
            "Explicação": "Documenta se a métrica Volume Variável foi usada para converter o Sell-in em volume estimado por peso unitário médio.",
            "Como foi calculado": "Quando a opção volume_variavel é escolhida, o Sell-in em quantidade é multiplicado pela gramatura média ponderada localizada no nome dos SKUs do Sell-out e dividido por 1.000.",
        },
        {
            "Nome da aba": "Parâmetros",
            "Nome da coluna calculada": "Ajuste volumetria aplicado",
            "Explicação": "Indica se o código precisou colocar Sell-in e Sell-out na mesma escala antes de calcular cobertura.",
            "Como foi calculado": "Compara a mediana da razão Sell-out/Sell-in por categoria. Se a diferença sugerir escala de centenas, milhares, milhões etc., divide o lado maior pelo fator detectado.",
        },
        {
            "Nome da aba": "Parâmetros",
            "Nome da coluna calculada": "Coluna ajustada por volumetria / Divisor aplicado",
            "Explicação": "Mostra qual lado foi alterado e por qual fator para deixar Sell-in e Sell-out comparáveis.",
            "Como foi calculado": "Exemplo: se Sell-out aparenta estar em unidades e Sell-in em milhares, o Sell-out é dividido por 1.000.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Tabela por UF",
            "Explicação": "Compara Sell-in e Sell-out usando a mesma abertura de UF/grupo nos dois lados.",
            "Como foi calculado": "Antes de somar, as UFs são padronizadas: AL e SE viram AL - SE; MA e PI viram MA - PI; RR/AM/RO/AC viram RR - AM - RO - AC; TO/PA/AP viram TO - PA - AP; aberturas de São Paulo como INT - SP, MET - SP, SP - INT, SP - Interior ou SP-REGMET viram SP. Em modo mensal, a tabela usa os últimos 12 meses móveis encerrando no último mês comum entre Sell-in e Sell-out.",
        },
        {
            "Nome da aba": "Abas de categoria/PROD",
            "Nome da coluna calculada": "Gráfico Sell-in x Sell-out por mês",
            "Explicação": "Gráfico de linhas com Sell-in e Sell-out por mês, no padrão visual solicitado.",
            "Como foi calculado": "Usa os valores mensais já ajustados de volumetria. Sell-in é linha azul e Sell-out é linha laranja. A cobertura permanece na tabela, mas não entra no gráfico.",
        },
        {
            "Nome da aba": "Crosschecks",
            "Nome da coluna calculada": "Sell-out / Dash / Check",
            "Explicação": "Para cada categoria/PROD processado, mostra o Sell-out por UF e mês, deixa uma tabela Dash em branco para preenchimento manual e cria uma tabela Check para validar se Dash e Sell-out estão iguais.",
            "Como foi calculado": "Sell-out é agrupado por categoria/PROD, UF padronizada e mês. Dash fica em branco. Check usa fórmulas do Excel no padrão =IF(célula Dash=célula Sell-out,'OK','NOK').",
        },
        {
            "Nome da aba": "Avisos",
            "Nome da coluna calculada": "Aviso",
            "Explicação": "Registra pontos de atenção da leitura, como colunas não encontradas, filtros aplicados e fallbacks usados.",
            "Como foi calculado": "Texto gerado automaticamente durante a leitura e validação dos arquivos.",
        },
    ]
    return pd.DataFrame(linhas)





def extrair_valor_resumo(df: pd.DataFrame, indicador: str, coluna: str):
    """Busca um valor em um dataframe de resumo no formato Indicador x Período."""
    if df is None or df.empty or "Indicador" not in df.columns or coluna not in df.columns:
        return np.nan
    mask = df["Indicador"].astype(str).map(normalizar_texto) == normalizar_texto(indicador)
    if not mask.any():
        return np.nan
    return df.loc[mask, coluna].iloc[0]


def periodo_texto(inicio, fim) -> str:
    if pd.isna(inicio) or pd.isna(fim):
        return ""
    return f"{formatar_periodo(inicio)} a {formatar_periodo(fim)}"


def montar_linha_resumo_categoria(resultado: Dict[str, object]) -> Dict[str, object]:
    """
    Monta o Resumo Categorias com as mesmas métricas do código antigo,
    mas usando as novas referências: Sell-in + Sell-out/Publicar.
    """
    periodo = resultado.get("periodos", {}) or {}
    resumo_periodo = resultado.get("resumo_periodo", pd.DataFrame())
    resumo_6m = resultado.get("resumo_6m", pd.DataFrame())

    label_ant = resultado.get("label_anterior") or periodo.get("label_anterior", "Período anterior")
    label_atual = resultado.get("label_atual") or periodo.get("label_atual", "Período atual")

    # As colunas 6M têm nomes dinâmicos. Pega a primeira e a segunda coluna de período.
    cols_6m = [c for c in resumo_6m.columns if c not in {"Indicador", "Tendência %"}]
    col_6m_ant = cols_6m[0] if len(cols_6m) >= 1 else "Período anterior"
    col_6m_atual = cols_6m[1] if len(cols_6m) >= 2 else "Período atual"

    linha = {
        "Categoria": resultado.get("categoria", ""),
        "Status Sell-in": resultado.get("status_sellin", "Inválido"),
        "Meses consecutivos Sell-in": resultado.get("meses_consecutivos_sellin", 0),
        "Período exibido": periodo_texto(resultado.get("inicio_comum"), resultado.get("fim_comum")),
        "Tipo comparação MAT": resultado.get("tipo_comparacao", ""),
        "Período MAT-1": periodo_texto(periodo.get("inicio_anterior"), periodo.get("fim_anterior")),
        "Período MAT": periodo_texto(periodo.get("inicio_atual"), periodo.get("fim_atual")),
        "Label MAT-1": label_ant,
        "Label MAT": label_atual,
        "Sell-in MAT-1": extrair_valor_resumo(resumo_periodo, "Sell-in", label_ant),
        "Sell-in MAT": extrair_valor_resumo(resumo_periodo, "Sell-in", label_atual),
        "Tendência Sell-in": extrair_valor_resumo(resumo_periodo, "Sell-in", "Tendência %"),
        "Sell-out MAT-1": extrair_valor_resumo(resumo_periodo, "Sell-out", label_ant),
        "Sell-out MAT": extrair_valor_resumo(resumo_periodo, "Sell-out", label_atual),
        "Tendência Sell-out": extrair_valor_resumo(resumo_periodo, "Sell-out", "Tendência %"),
        "Cobertura MAT": extrair_valor_resumo(resumo_periodo, "Cobertura", label_atual),
        "Sell-in 6M Anterior": extrair_valor_resumo(resumo_6m, "Sell-in", col_6m_ant),
        "Sell-in 6M Atual": extrair_valor_resumo(resumo_6m, "Sell-in", col_6m_atual),
        "Tendência Sell-in 6M": extrair_valor_resumo(resumo_6m, "Sell-in", "Tendência %"),
        "Sell-out 6M Anterior": extrair_valor_resumo(resumo_6m, "Sell-out", col_6m_ant),
        "Sell-out 6M Atual": extrair_valor_resumo(resumo_6m, "Sell-out", col_6m_atual),
        "Tendência Sell-out 6M": extrair_valor_resumo(resumo_6m, "Sell-out", "Tendência %"),
        "Cobertura 6M Atual": extrair_valor_resumo(resumo_6m, "Cobertura", col_6m_atual),
    }

    # Mantém também os nomes curtos criados na versão nova para facilitar leitura/filtro.
    linha["Período comum início"] = resultado.get("inicio_comum")
    linha["Período comum fim"] = resultado.get("fim_comum")
    linha["Tipo comparação"] = resultado.get("tipo_comparacao", "")
    linha["Período anterior"] = label_ant
    linha["Período atual"] = label_atual
    linha["Sell-in atual"] = linha["Sell-in MAT"]
    linha["Sell-out atual"] = linha["Sell-out MAT"]
    linha["Cobertura atual"] = linha["Cobertura MAT"]

    # Tendência principal: usa o mesmo período MAT/YTD definido para a categoria/PROD.
    # Ano fechado quando houver base completa; caso contrário, YTD contra o mesmo intervalo anterior.
    linha["Tendência Sell-in"] = linha["Tendência Sell-in"]
    linha["Tendência Sell-out"] = linha["Tendência Sell-out"]
    try:
        linha["Diferença Tendência"] = float(linha["Tendência Sell-out"]) - float(linha["Tendência Sell-in"])
    except Exception:
        linha["Diferença Tendência"] = np.nan
    return linha

def limpar_dataframe_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Remove NaN/INF que podem quebrar a geração do Excel."""
    if df is None:
        return pd.DataFrame()
    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    return out


def remover_colunas_duplicadas(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas com nomes duplicados, mantendo sempre a primeira ocorrência."""
    if df is None:
        return pd.DataFrame()
    if df.empty and len(df.columns) == 0:
        return df
    return df.loc[:, ~pd.Index(df.columns).duplicated(keep="first")].copy()


COLUNAS_REMOVER_RESUMO_CATEGORIAS = {
    "Período MAT-1", "Período MAT", "Label MAT-1", "Label MAT",
    "Sell-in MAT-1", "Sell-in MAT",
    "Sell-out MAT-1", "Sell-out MAT",
    "Cobertura 6M Atual", "Período comum início", "Período comum fim",
    "Tipo comparação", "Tipo comparação MAT", "Período anterior", "Período atual",
    "Sell-in atual", "Sell-out atual", "Fabricante_referencia_Sellin",
    "SKUs_Em_Comum_Fabricante_Sellin", "Volume_Sellin", "Volume_Sellout",
    "Volume_Sellout_Fabricante_Sellin", "Volume_Sellin_Encontrado_Fabricante",
    "SKUs_Sellin_Encontrados_Fabricante", "SKUs_Nosso_Sellout",
    "Volume_Sellout_Referencia", "Volume_Sellin_Em_Comum", "Volume_Sellout_Em_Comum",
    "Volume_Sellout_Em_Comum_Fabricante_Sellin", "Total_Sellout_Fabricante_Sellin",
    "Cobertura importância Sell-out Fab. Sell-in", "% Sell-in usado para detectar fabricante",
    "Variação Sell-in", "Variação Sell-out", "Variação Sell-in 6M", "Variação Sell-out 6M",
}


def limpar_colunas_resumo_categorias(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas técnicas excedentes do Resumo Categorias sem apagar a Cobertura atual."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "Cobertura atual" not in out.columns and "Cobertura MAT" in out.columns:
        out["Cobertura atual"] = out["Cobertura MAT"]
    out = out.drop(columns=[c for c in COLUNAS_REMOVER_RESUMO_CATEGORIAS if c in out.columns], errors="ignore")
    return out


def nome_arquivo_curto(caminho) -> str:
    """Mostra só o nome do arquivo, sem o caminho completo do diretório."""
    if caminho is None or str(caminho).strip() == "":
        return ""
    try:
        return Path(caminho).name
    except Exception:
        return str(caminho)

def _pivot_crosscheck_base(df: pd.DataFrame, coluna_valor: str, periodos_ordenados: List[pd.Timestamp]) -> pd.DataFrame:
    """
    Monta pivot UF x mês para crosscheck, usando apenas Sell-out e a UF já padronizada.
    A saída sempre traz a ordem oficial de comparação por UF para facilitar preenchimento/conferência no Dash.
    """
    if periodos_ordenados:
        cols = ["UF"] + [label_mes_pt(m) for m in periodos_ordenados]
    else:
        cols = ["UF", "Total"]

    if df is None or df.empty:
        vazio = pd.DataFrame({"UF": UF_COMPARACAO_ORDEM})
        for c in cols[1:]:
            vazio[c] = 0
        return vazio[cols]

    temp = adicionar_uf_comparacao(df)
    temp = temp[temp[coluna_valor].fillna(0) != 0].copy()

    if temp.empty:
        vazio = pd.DataFrame({"UF": UF_COMPARACAO_ORDEM})
        for c in cols[1:]:
            vazio[c] = 0
        return vazio[cols]

    if periodos_ordenados and "mes" in temp.columns and temp["mes"].notna().any():
        temp = temp[temp["mes"].notna()].copy()
        temp["periodo_crosscheck"] = temp["mes"].map(lambda x: pd.Timestamp(x).to_period("M").to_timestamp())
        pivot = temp.pivot_table(
            index="uf_comparacao",
            columns="periodo_crosscheck",
            values=coluna_valor,
            aggfunc="sum",
            fill_value=0,
        )
        pivot = pivot.reindex(index=UF_COMPARACAO_ORDEM, columns=periodos_ordenados, fill_value=0)
        pivot.columns = [label_mes_pt(c) for c in pivot.columns]
    else:
        temp["periodo_crosscheck"] = "Total"
        pivot = temp.pivot_table(
            index="uf_comparacao",
            columns="periodo_crosscheck",
            values=coluna_valor,
            aggfunc="sum",
            fill_value=0,
        )
        if "Total" not in pivot.columns:
            pivot["Total"] = 0
        pivot = pivot.reindex(index=UF_COMPARACAO_ORDEM, fill_value=0)[["Total"]]

    pivot = pivot.reset_index().rename(columns={"uf_comparacao": "UF"})
    return pivot


def montar_crosschecks_fabricante(sellout: pd.DataFrame, resultados: List[Dict[str, object]]) -> List[Dict[str, object]]:
    """
    Cria os blocos da aba Crosschecks por categoria/PROD processado.

    Cada bloco possui apenas a visão de Sell-out por UF/mês, pois a parte Dash será preenchida manualmente
    pelo usuário e a parte Check terá fórmulas comparando Dash x Sell-out.
    """
    if sellout is not None and not sellout.empty and "mes" in sellout.columns:
        periodos = sorted({pd.Timestamp(m).to_period("M").to_timestamp() for m in sellout["mes"].dropna().tolist()})
    else:
        periodos = []

    blocos = []
    vistos = set()
    for r in resultados or []:
        cat_key = str(r.get("categoria_key", ""))
        cat_nome = str(r.get("categoria", cat_key or "Categoria"))
        if not cat_key or cat_key in vistos:
            continue
        vistos.add(cat_key)
        so_cat = sellout[sellout["categoria_key"] == cat_key].copy() if sellout is not None and not sellout.empty else pd.DataFrame()
        pivot = _pivot_crosscheck_base(so_cat, "valor_sellout", periodos)
        blocos.append({"categoria_key": cat_key, "categoria": cat_nome, "sellout": pivot})

    return blocos


def escrever_crosschecks(writer, crosschecks: List[Dict[str, object]]):
    workbook = writer.book
    sheet = "Crosschecks"
    ws = workbook.add_worksheet(sheet)
    writer.sheets[sheet] = ws

    fmt_titulo = workbook.add_format({
        "bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white",
        "align": "left", "valign": "vcenter",
    })
    fmt_secao = workbook.add_format({
        "bold": True, "font_size": 12, "bg_color": "#D9EAF7", "border": 1,
        "align": "left", "valign": "vcenter",
    })
    fmt_header = workbook.add_format({
        "bold": True, "bg_color": "#1F4E78", "font_color": "white", "border": 1,
        "align": "center", "valign": "vcenter",
    })
    fmt_text = workbook.add_format({"border": 1, "align": "left", "valign": "vcenter"})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
    fmt_blank = workbook.add_format({"border": 1})
    fmt_check = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})
    fmt_ok = workbook.add_format({"bg_color": "#E2F0D9", "font_color": "#375623", "border": 1, "align": "center"})
    fmt_nok = workbook.add_format({"bg_color": "#FCE4D6", "font_color": "#9C0006", "border": 1, "align": "center"})

    def col_excel(idx_zero_based: int) -> str:
        n = idx_zero_based + 1
        letras = ""
        while n:
            n, rem = divmod(n - 1, 26)
            letras = chr(65 + rem) + letras
        return letras

    def escrever_tabela_valores(df: pd.DataFrame, row_title: int, titulo: str, categoria: str, valores_em_branco: bool = False):
        """Escreve Categoria/Título/Header/Dados no layout fixo do crosscheck."""
        df = limpar_dataframe_excel(df)
        if df.empty:
            df = pd.DataFrame({"UF": UF_COMPARACAO_ORDEM})

        # Linha da categoria fica imediatamente acima da primeira tabela do bloco.
        if categoria:
            ws.write(row_title - 1, 0, categoria, fmt_secao)

        ws.write(row_title, 0, titulo, fmt_secao)
        for col_idx, col in enumerate(df.columns):
            ws.write(row_title + 1, col_idx, col, fmt_header)

        for i in range(len(df)):
            row = row_title + 2 + i
            for col_idx, col in enumerate(df.columns):
                valor = df.iloc[i][col]
                if col_idx == 0:
                    ws.write(row, col_idx, str(valor), fmt_text)
                elif valores_em_branco:
                    ws.write_blank(row, col_idx, None, fmt_blank)
                else:
                    write_number_ou_branco(ws, row, col_idx, valor, fmt_num)

    def escrever_tabela_check(df: pd.DataFrame, row_title: int, sellout_data_start: int, dash_data_start: int):
        df = limpar_dataframe_excel(df)
        if df.empty:
            df = pd.DataFrame({"UF": UF_COMPARACAO_ORDEM})

        ws.write(row_title, 0, "Check", fmt_secao)
        for col_idx, col in enumerate(df.columns):
            ws.write(row_title + 1, col_idx, col, fmt_header)

        last_data_row = row_title + 1
        last_data_col = max(len(df.columns) - 1, 0)
        for i in range(len(df)):
            row = row_title + 2 + i
            last_data_row = row
            for col_idx, col in enumerate(df.columns):
                if col_idx == 0:
                    ws.write(row, col_idx, str(df.iloc[i].get("UF", "")), fmt_text)
                else:
                    letra = col_excel(col_idx)
                    formula = f'=IF({letra}{dash_data_start + i + 1}={letra}{sellout_data_start + i + 1},"OK","NOK")'
                    ws.write_formula(row, col_idx, formula, fmt_check)

        if last_data_col >= 1 and last_data_row >= row_title + 2:
            first_col = col_excel(1)
            last_col = col_excel(last_data_col)
            first_row_excel = row_title + 3
            last_row_excel = last_data_row + 1
            ws.conditional_format(f"{first_col}{first_row_excel}:{last_col}{last_row_excel}", {
                "type": "text", "criteria": "containing", "value": "OK", "format": fmt_ok,
            })
            ws.conditional_format(f"{first_col}{first_row_excel}:{last_col}{last_row_excel}", {
                "type": "text", "criteria": "containing", "value": "NOK", "format": fmt_nok,
            })

    ws.merge_range("A1:H1", "Crosschecks - Sell-out x Dash por UF e mês", fmt_titulo)
    ws.write("A3", "Regra de UF", fmt_secao)
    ws.write(
        "B3",
        "UFs padronizadas para comparação: AL - SE; MA - PI; RR - AM - RO - AC; TO - PA - AP; SP consolidado com INT/MET/Interior/RegMet. O usuário preenche a tabela Dash; a tabela Check compara Dash x Sell-out.",
        fmt_text,
    )
    ws.set_column("A:A", 24)
    ws.set_column("B:AZ", 14)

    if not crosschecks:
        ws.write("A5", "Sem dados de Sell-out para Crosscheck", fmt_secao)
        return

    bloco_altura = 80
    inicio_bloco = 4  # Excel linha 5.
    for idx, bloco in enumerate(crosschecks):
        base = inicio_bloco + idx * bloco_altura
        categoria = str(bloco.get("categoria", "Categoria"))
        df_so = limpar_dataframe_excel(bloco.get("sellout", pd.DataFrame()))

        # Linhas do primeiro bloco em Excel:
        # Categoria A5, Sell-out A6, header A7, dados A8:...
        sellout_title = base + 1
        sellout_data_start = base + 3
        dash_title = base + 27       # Excel linha 32 no primeiro bloco.
        dash_data_start = base + 29  # Excel linha 34 no primeiro bloco.
        check_title = base + 52      # Excel linha 57 no primeiro bloco.

        escrever_tabela_valores(df_so, sellout_title, "Sell-out por UF", categoria, valores_em_branco=False)
        escrever_tabela_valores(df_so, dash_title, "Dash", "", valores_em_branco=True)
        escrever_tabela_check(df_so, check_title, sellout_data_start=sellout_data_start, dash_data_start=dash_data_start)



# ============================================================
# Base auxiliar para contribuição SKU 2.0 x 3.0
# ============================================================

def preparar_base_contribuicao_sellout(sellout: pd.DataFrame) -> pd.DataFrame:
    """
    Monta uma base compacta do Sell-out para análises de contribuição por UF, Canal e EAN.

    Essa base é gravada nos estudos individuais e depois é lida pelo modo
    "Comparação de Estudo de Cobertura". Assim, a comparação consegue calcular
    a contribuição sem precisar reler os Sell-outs originais.
    """
    colunas = [
        "Aba Categoria", "Categoria", "categoria_key", "UF", "Canal", "EAN",
        "Data", "Ano", "Volume Sell-out",
    ]
    if sellout is None or sellout.empty:
        return pd.DataFrame(columns=colunas)

    base = sellout.copy()
    for c in ["categoria", "categoria_key", "uf", "canal", "ean", "valor_sellout"]:
        if c not in base.columns:
            base[c] = "" if c != "valor_sellout" else 0

    base["EAN"] = base["ean"].map(ean_texto)
    base = base[base["EAN"] != ""].copy()
    if base.empty:
        return pd.DataFrame(columns=colunas)

    if "mes" in base.columns:
        base["Data"] = base["mes"].map(converter_mes)
    else:
        base["Data"] = pd.NaT

    if "ano" in base.columns:
        base["Ano"] = pd.to_numeric(base["ano"], errors="coerce")
    else:
        base["Ano"] = np.nan
    base.loc[base["Data"].notna(), "Ano"] = base.loc[base["Data"].notna(), "Data"].dt.year

    base["Categoria"] = base["categoria"].fillna("").astype(str).str.strip()
    base["categoria_key"] = base["categoria_key"].fillna(base["Categoria"].map(normalizar_categoria)).astype(str)
    base.loc[base["categoria_key"].map(normalizar_categoria) == "", "categoria_key"] = base.loc[
        base["categoria_key"].map(normalizar_categoria) == "", "Categoria"
    ].map(normalizar_categoria)

    base["UF"] = base["uf"].fillna("TOTAL").astype(str).str.strip().replace("", "TOTAL")
    base["Canal"] = base["canal"].fillna("TOTAL").astype(str).str.strip().replace("", "TOTAL")
    base["Volume Sell-out"] = pd.to_numeric(base["valor_sellout"], errors="coerce").fillna(0)
    base = base[(base["categoria_key"].map(normalizar_categoria) != "") & (base["Volume Sell-out"] != 0)].copy()

    if base.empty:
        return pd.DataFrame(columns=colunas)

    group_cols = ["Categoria", "categoria_key", "UF", "Canal", "EAN", "Data", "Ano"]
    out = base.groupby(group_cols, dropna=False, as_index=False)["Volume Sell-out"].sum()
    out["Aba Categoria"] = ""
    out = out[["Aba Categoria", "Categoria", "categoria_key", "UF", "Canal", "EAN", "Data", "Ano", "Volume Sell-out"]]
    return out.reset_index(drop=True)


def aplicar_abas_categoria_base_contribuicao(base_contribuicao: pd.DataFrame, resultados: List[Dict[str, object]]) -> pd.DataFrame:
    """Preenche a coluna Aba Categoria usando os nomes reais das abas geradas no estudo."""
    if base_contribuicao is None or base_contribuicao.empty:
        return pd.DataFrame(columns=[
            "Aba Categoria", "Categoria", "categoria_key", "UF", "Canal", "EAN",
            "Data", "Ano", "Volume Sell-out",
        ])
    out = base_contribuicao.copy()
    mapa_aba = {}
    mapa_nome = {}
    for r in resultados or []:
        key = normalizar_categoria(r.get("categoria_key", ""))
        if not key:
            key = normalizar_categoria(r.get("categoria", ""))
        if key:
            mapa_aba[key] = str(r.get("_nome_aba_excel") or r.get("categoria") or "")
            mapa_nome[key] = str(r.get("categoria") or "")
    out["categoria_key"] = out["categoria_key"].map(normalizar_categoria)
    out["Aba Categoria"] = out["categoria_key"].map(mapa_aba).fillna("")
    out.loc[out["Aba Categoria"].eq(""), "Aba Categoria"] = out.loc[out["Aba Categoria"].eq(""), "Categoria"].astype(str).str.strip()
    out["Categoria"] = out["categoria_key"].map(mapa_nome).fillna(out["Categoria"])
    return out

def _tipo_grafico_legivel(tipo: str) -> str:
    tipo = str(tipo or "Categoria").upper()
    mapa = {"CATEGORIA": "Categoria", "NIVEL1": "Nível 1", "NIVEL2": "Nível 2", "ESTMER7": "Est Mer 7"}
    return mapa.get(tipo, str(tipo).title())


def _valor_para_data_grafico(valor):
    if pd.isna(valor):
        return pd.NaT
    if isinstance(valor, (pd.Timestamp, datetime)):
        return pd.Timestamp(valor).to_period("M").to_timestamp()
    return converter_mes(valor)


def _preparar_tabela_grafico_cobertura(mensal: pd.DataFrame, comparacao: bool = False) -> pd.DataFrame:
    if mensal is None or mensal.empty:
        return pd.DataFrame()
    base = mensal.copy()
    if comparacao:
        data = base["mes_ts"].map(_valor_para_data_grafico) if "mes_ts" in base.columns else base.get("Data", pd.Series(index=base.index)).map(_valor_para_data_grafico)
        out = pd.DataFrame({"Data": data})
        for col in ["Sell-in", "Sell-out 2.0", "Sell-out 3.0", "Cobertura 2.0", "Cobertura 3.0"]:
            out[col] = pd.to_numeric(base.get(col, np.nan), errors="coerce")
    else:
        data_col = "Mês" if "Mês" in base.columns else ("Data" if "Data" in base.columns else None)
        data = base[data_col].map(_valor_para_data_grafico) if data_col else pd.Series([pd.NaT] * len(base), index=base.index)
        out = pd.DataFrame({"Data": data})
        out["Sell-in"] = pd.to_numeric(base.get("Sell-in", np.nan), errors="coerce")
        out["Sell-out"] = pd.to_numeric(base.get("Sell-out", np.nan), errors="coerce")
        cobertura_col = "Cobertura 12M Móvel" if "Cobertura 12M Móvel" in base.columns else "Cobertura"
        out["Cobertura"] = pd.to_numeric(base.get(cobertura_col, np.nan), errors="coerce")
    out = out.dropna(how="all", subset=[c for c in out.columns if c != "Data"])
    out = out[out["Data"].notna()].copy() if out["Data"].notna().any() else out.copy()
    return out.reset_index(drop=True)


def _criar_imagem_grafico_cobertura(df: pd.DataFrame, titulo: str):
    if df is None or df.empty:
        return None
    try:
        import io
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return None

    plot_df = df.copy().reset_index(drop=True)
    x = np.arange(len(plot_df))
    labels = []
    for v in plot_df.get("Data", pd.Series(range(len(plot_df)))):
        if pd.notna(v):
            try:
                labels.append(pd.Timestamp(v).strftime("%m/%y"))
            except Exception:
                labels.append(str(v))
        else:
            labels.append("")

    fig, ax1 = plt.subplots(figsize=(12.6, 3.7), dpi=120)
    fig.patch.set_facecolor("white")
    ax1.set_facecolor("white")

    cores_linha = {
        "Sell-in": "#1f77b4",
        "Sell-out": "#ff7f0e",
        "Sell-out 2.0": "#ff7f0e",
        "Sell-out 3.0": "#2ca02c",
    }
    for col in ["Sell-in", "Sell-out", "Sell-out 2.0", "Sell-out 3.0"]:
        if col in plot_df.columns:
            y = pd.to_numeric(plot_df[col], errors="coerce")
            if y.notna().any():
                ax1.plot(x, y, marker="o", linewidth=2.4, markersize=6.8, label=col, color=cores_linha.get(col))

    ax1.set_title(titulo, fontsize=13, fontweight="bold", pad=26)
    ax1.set_ylabel("Volume", fontsize=10)
    ax1.grid(axis="y", color="#D9D9D9", linewidth=0.6)
    ax1.set_axisbelow(True)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha="right")
    ax1.margins(x=0.03)

    ax2 = ax1.twinx()
    cores_barra = {
        "Cobertura": "#9ECAE1",
        "Cobertura 2.0": "#74A9CF",
        "Cobertura 3.0": "#FDAE6B",
    }
    cov_cols = [c for c in ["Cobertura", "Cobertura 2.0", "Cobertura 3.0"] if c in plot_df.columns and pd.to_numeric(plot_df[c], errors="coerce").notna().any()]
    cov_max = 1.0
    if cov_cols and len(plot_df):
        idxs_validos = []
        for idx in range(len(plot_df)):
            if any(pd.notna(pd.to_numeric(pd.Series([plot_df.loc[idx, c]]), errors="coerce").iloc[0]) for c in cov_cols):
                idxs_validos.append(idx)
        ultimo = idxs_validos[-1] if idxs_validos else len(plot_df) - 1
        n = len(cov_cols)
        offsets = np.linspace(-0.12, 0.12, n) if n > 1 else [0]
        for pos, col in zip(offsets, cov_cols):
            val = pd.to_numeric(pd.Series([plot_df.loc[ultimo, col]]), errors="coerce").iloc[0]
            if pd.isna(val):
                continue
            cov_max = max(cov_max, float(val))
            ax2.bar([ultimo + float(pos)], [float(val)], width=0.12, label=col, color=cores_barra.get(col), alpha=0.9)
        ax2.set_ylim(0, max(1.0, cov_max * 1.38))
        ylim_top = ax2.get_ylim()[1]
        for k, (pos, col) in enumerate(zip(offsets, cov_cols)):
            val = pd.to_numeric(pd.Series([plot_df.loc[ultimo, col]]), errors="coerce").iloc[0]
            if pd.isna(val):
                continue
            ax2.text(
                ultimo + float(pos),
                float(val) + ylim_top * (0.045 + 0.035 * k),
                f"{float(val):.1%}",
                ha="center", va="bottom", fontsize=9, fontweight="bold", color="black"
            )
    else:
        ax2.set_ylim(0, 1)
    ax2.set_ylabel("Cobertura", fontsize=10)

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    handles = h1 + h2
    labels_leg = l1 + l2
    if handles:
        ax1.legend(handles, labels_leg, loc="upper center", bbox_to_anchor=(0.5, 1.25), ncol=min(len(handles), 5), frameon=False, fontsize=10)

    fig.tight_layout(rect=[0, 0, 1, 0.93])
    bio = io.BytesIO()
    fig.savefig(bio, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    bio.seek(0)
    return bio


def escrever_graficos_cobertura(writer, itens: List[Dict[str, object]], tipo: str = "Categoria", comparacao: bool = False):
    workbook = writer.book
    sheet_name = "Gráficos Cobertura"
    if sheet_name in writer.sheets:
        return
    ws = workbook.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = ws

    fmt_titulo = workbook.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white", "align": "center", "valign": "vcenter", "border": 1})
    fmt_header = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "align": "center", "valign": "vcenter"})
    fmt_data = workbook.add_format({"num_format": "mmm/yy", "border": 1, "align": "center"})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
    fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
    fmt_text = workbook.add_format({"border": 1})

    ws.set_column("A:A", 12)
    ws.set_column("B:F", 15)
    ws.set_column("G:H", 15)
    ws.set_default_row(18)
    tipo_legivel = _tipo_grafico_legivel(tipo)
    row_cursor = 0
    itens_validos = 0

    for item in itens or []:
        categoria = str(item.get("categoria") or item.get("Categoria") or item.get("categoria_key") or "Categoria")
        mensal = item.get("mensal", pd.DataFrame())
        tabela = _preparar_tabela_grafico_cobertura(mensal, comparacao=comparacao)
        if tabela.empty:
            continue
        itens_validos += 1
        titulo = f"{tipo_legivel}: {categoria}"
        ultima_col = max(5, min(len(tabela.columns) - 1, 7))
        ws.merge_range(row_cursor, 0, row_cursor, ultima_col, titulo, fmt_titulo)
        imagem = _criar_imagem_grafico_cobertura(tabela, titulo)
        if imagem is not None:
            ws.insert_image(row_cursor + 1, 0, "grafico_cobertura.png", {"image_data": imagem, "x_scale": 1.0, "y_scale": 1.0})
        else:
            ws.write(row_cursor + 1, 0, "Matplotlib não disponível para gerar a imagem do gráfico.", fmt_text)

        table_row = row_cursor + 24
        for j, col in enumerate(tabela.columns):
            ws.write(table_row, j, col, fmt_header)
        for i, (_, r) in enumerate(tabela.reset_index(drop=True).iterrows(), start=1):
            excel_row = table_row + i
            for j, col in enumerate(tabela.columns):
                val = r.get(col)
                if col == "Data" and pd.notna(val):
                    ws.write_datetime(excel_row, j, pd.Timestamp(val).to_pydatetime(), fmt_data)
                elif "Cobertura" in col:
                    write_number_ou_branco(ws, excel_row, j, val, fmt_pct)
                elif isinstance(val, (int, float, np.integer, np.floating)) and not pd.isna(val):
                    write_number_ou_branco(ws, excel_row, j, val, fmt_num)
                else:
                    ws.write(excel_row, j, "" if pd.isna(val) else str(val), fmt_text)
        row_cursor = table_row + max(len(tabela), 1) + 4

    if itens_validos == 0:
        ws.write(0, 0, "Nenhum dado mensal encontrado para gerar gráficos de cobertura.", fmt_text)


def gerar_excel(
    saida: Path,
    resultados: List[Dict[str, object]],
    detalhe_skus: pd.DataFrame,
    resumo_skus: pd.DataFrame,
    parametros: Dict[str, str],
    avisos: List[str],
    crosschecks: Optional[Dict[str, pd.DataFrame]] = None,
    base_contribuicao_sellout: Optional[pd.DataFrame] = None,
    output_options: Optional[Dict[str, object]] = None,
):
    saida = Path(saida)
    output_options = normalizar_opcoes_saida(output_options)
    parametros = dict(parametros or {})
    parametros["Opções de saída geradas"] = opcoes_saida_para_parametros(output_options)
    detalhe_skus = limpar_dataframe_excel(detalhe_skus)
    resumo_skus = limpar_dataframe_excel(resumo_skus)
    with pd.ExcelWriter(saida, engine="xlsxwriter", engine_kwargs={"options": {"nan_inf_to_errors": True}}) as writer:
        workbook = writer.book

        fmt_titulo = workbook.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white"})
        fmt_link = workbook.add_format({"font_color": "#0563C1", "underline": 1, "border": 1})

        # Define antecipadamente o nome real de cada aba para criar hyperlinks no Resumo Categorias.
        usados = {"Resumo Categorias", "Gráficos Cobertura", "Base SKUs", "SKUs por Categoria", "Base Contribuição Sell-out", "Crosschecks", "Parâmetros", "Avisos", "Descrição Cálculos"}
        for r in resultados:
            r["_nome_aba_excel"] = nome_aba_seguro(str(r.get("categoria", "Categoria")), usados)

        # Resumo categorias
        # Mantém as mesmas métricas do código antigo: MAT/YTD, variações, cobertura e 6M.
        resumo_linhas = [montar_linha_resumo_categoria(r) for r in resultados]
        resumo_cat = pd.DataFrame(resumo_linhas)
        if not resumo_cat.empty:
            resumo_cat.insert(0, "Abrir aba", [r.get("_nome_aba_excel", "") for r in resultados])

        fabricante_por_categoria = {}
        if not resumo_skus.empty:
            resumo_skus_tmp = resumo_skus.copy()
            if "Categoria" in resumo_skus_tmp.columns and "Fabricante_referencia_Sellin" in resumo_skus_tmp.columns:
                fabricante_por_categoria = (
                    resumo_skus_tmp[["Categoria", "Fabricante_referencia_Sellin"]]
                    .dropna(subset=["Categoria"])
                    .drop_duplicates("Categoria")
                    .set_index("Categoria")["Fabricante_referencia_Sellin"]
                    .to_dict()
                )
            resumo_cat = resumo_cat.merge(
                resumo_skus_tmp.drop(columns=["categoria_key"], errors="ignore"),
                on="Categoria",
                how="left",
            )

        if not resumo_cat.empty:
            # Colunas amigáveis de quantidade de SKUs no resumo principal.
            # Mantém as colunas técnicas originais, mas adiciona nomes diretos para leitura rápida.
            if "SKUs_Sellin" in resumo_cat.columns and "Qtd SKUs Sell-in" not in resumo_cat.columns:
                resumo_cat["Qtd SKUs Sell-in"] = resumo_cat["SKUs_Sellin"]
            if "SKUs_Sellout" in resumo_cat.columns and "Qtd SKUs Sell-out" not in resumo_cat.columns:
                resumo_cat["Qtd SKUs Sell-out"] = resumo_cat["SKUs_Sellout"]
            if "SKUs_Em_Comum" in resumo_cat.columns and "Qtd SKUs em comum" not in resumo_cat.columns:
                resumo_cat["Qtd SKUs em comum"] = resumo_cat["SKUs_Em_Comum"]

            # Ordenação solicitada: categoria/PROD mais importante no Sell-in primeiro.
            # A importância usa o Sell-in do período atual usado no cálculo principal (FY ou YTD).
            sellin_atual_num = pd.to_numeric(resumo_cat.get("Sell-in atual", 0), errors="coerce").fillna(0)
            total_sellin_atual = sellin_atual_num.sum()
            resumo_cat["Importância Sell-in"] = np.where(
                total_sellin_atual != 0,
                sellin_atual_num / total_sellin_atual,
                np.nan,
            )
            resumo_cat = resumo_cat.sort_values(
                ["Importância Sell-in", "Sell-in atual", "Categoria"],
                ascending=[False, False, True],
                na_position="last",
            ).reset_index(drop=True)

            # Mantém as colunas novas visíveis no início do resumo.
            colunas_inicio = [
                "Abrir aba", "Categoria", "Status Sell-in", "Importância Sell-in",
                "Tendência Sell-in", "Tendência Sell-out", "Diferença Tendência",
                "Meses consecutivos Sell-in", "Qtd SKUs Sell-in", "Qtd SKUs Sell-out", "Qtd SKUs em comum",
            ]
            colunas_inicio = [c for c in colunas_inicio if c in resumo_cat.columns]
            demais_colunas = [c for c in resumo_cat.columns if c not in colunas_inicio]
            resumo_cat = resumo_cat[colunas_inicio + demais_colunas]

        # Define o fabricante que será mostrado em G2 nas abas de cobertura.
        fabricante_digitado = str(parametros.get("Fabricante selecionado", "") or "").strip()
        if fabricante_digitado.lower() in {"automático / todos", "automatico / todos", "todos", ""}:
            fabricante_digitado = ""
        for r in resultados:
            cat = str(r.get("categoria", ""))
            r["fabricante_exibido"] = fabricante_digitado or str(fabricante_por_categoria.get(cat, "") or "Não informado")

        if output_options.get("resumo_categorias", True):
            resumo_cat = limpar_colunas_resumo_categorias(resumo_cat)
            resumo_cat = remover_colunas_duplicadas(resumo_cat)
            resumo_cat = limpar_dataframe_excel(resumo_cat)
            resumo_cat.to_excel(writer, sheet_name="Resumo Categorias", index=False, startrow=1)
            ws = writer.sheets["Resumo Categorias"]
            ws.merge_range("A1:L1", "Resumo Categorias", fmt_titulo)
            aplicar_formatos_basicos(
                writer, "Resumo Categorias", resumo_cat, startrow=1, startcol=0,
                percent_cols={
                    "Cobertura atual", "Cobertura MAT", "Cobertura 6M Atual",
                    "Tendência Sell-in", "Tendência Sell-out", "Tendência Sell-in 6M", "Tendência Sell-out 6M",
                    "Tendência Sell-in", "Tendência Sell-out", "Diferença Tendência",
                    "Importância Sell-in",
                    "Cobertura importância Sell-in", "Cobertura importância Sell-out",
                    "Cobertura importância Sell-out Fab. Sell-in", "% Sell-in usado para detectar fabricante"
                },
                number_cols={
                    "Sell-in atual", "Sell-out atual", "Meses consecutivos Sell-in",
                    "SKUs_Sellin", "SKUs_Sellout", "SKUs_Em_Comum",
                    "Qtd SKUs Sell-in", "Qtd SKUs Sell-out", "Qtd SKUs em comum",
                    "SKUs_Em_Comum_Fabricante_Sellin", "Volume_Sellin", "Volume_Sellout",
                    "Volume_Sellout_Fabricante_Sellin", "Volume_Sellin_Em_Comum", "Volume_Sellout_Em_Comum",
                    "Volume_Sellout_Em_Comum_Fabricante_Sellin", "Total_Sellout_Fabricante_Sellin"
                },
                date_cols={"Período comum início", "Período comum fim"},
            )
            if not resumo_cat.empty and "Abrir aba" in resumo_cat.columns:
                ws.set_column(0, 0, 24, fmt_link)
                for i, nome_aba in enumerate(resumo_cat["Abrir aba"].fillna("")):
                    nome_aba = str(nome_aba).strip()
                    if nome_aba:
                        ws.write_url(2 + i, 0, f"internal:'{nome_aba}'!A1", fmt_link, string=nome_aba)
        else:
            resumo_cat = limpar_colunas_resumo_categorias(resumo_cat)
            resumo_cat = remover_colunas_duplicadas(resumo_cat)
            resumo_cat = limpar_dataframe_excel(resumo_cat)

        # Abas por categoria.
        if output_options.get("abas_categorias", True):
            for r in resultados:
                nome_aba = r.get("_nome_aba_excel") or nome_aba_seguro(r["categoria"], usados)
                escrever_categoria(writer, nome_aba, r)

        # Gráficos Cobertura com imagem estática Matplotlib + dados editáveis.
        if output_options.get("graficos_cobertura", True):
            escrever_graficos_cobertura(writer, resultados, tipo=parametros.get("Regra categoria/PROD", "Categoria"), comparacao=False)

        # Base SKUs
        if output_options.get("base_skus", True):
            detalhe_skus = limpar_dataframe_excel(detalhe_skus)
            detalhe_skus.to_excel(writer, sheet_name="Base SKUs", index=False, startrow=1)
            ws = writer.sheets["Base SKUs"]
            ws.merge_range("A1:Q1", "Base SKUs - Sell-in x Sell-out", fmt_titulo)
            aplicar_formatos_basicos(
                writer, "Base SKUs", detalhe_skus, startrow=1, startcol=0,
                percent_cols={"Importância Sell-in SKU", "Importância Sell-out SKU", "Importância Sell-out SKU Fab. Sell-in"},
                number_cols={"Volume Sell-in SKU", "Volume Sell-out SKU", "Volume Sell-out Fab. Sell-in SKU"},
            )

        # Base auxiliar para contribuição SKU 2.0 x 3.0.
        # Essa aba fica oculta e é usada pelo modo "Comparação de Estudo de Cobertura".
        base_contrib = aplicar_abas_categoria_base_contribuicao(base_contribuicao_sellout, resultados) if output_options.get("base_contribuicao_sellout", True) else pd.DataFrame()
        base_contrib = limpar_dataframe_excel(base_contrib)
        if output_options.get("base_contribuicao_sellout", True) and not base_contrib.empty:
            limite_excel = 1_048_000
            if len(base_contrib) > limite_excel:
                avisos.append(
                    "Base Contribuição Sell-out excedeu o limite de linhas do Excel e foi limitada. "
                    "Para manter a comparação por SKU, rode o estudo com filtro de fabricante/categoria mais restrito."
                )
                base_contrib = base_contrib.head(limite_excel).copy()
            base_contrib.to_excel(writer, sheet_name="Base Contribuição Sell-out", index=False, startrow=1)
            ws = writer.sheets["Base Contribuição Sell-out"]
            ws.merge_range("A1:I1", "Base Contribuição Sell-out - uso interno da comparação", fmt_titulo)
            aplicar_formatos_basicos(
                writer, "Base Contribuição Sell-out", base_contrib, startrow=1, startcol=0,
                number_cols={"Volume Sell-out", "Ano"},
                date_cols={"Data"},
            )
            try:
                ws.hide()
            except Exception:
                pass

        # SKUs por Categoria
        if output_options.get("skus_por_categoria", True):
            resumo_skus_saida = resumo_skus.drop(columns=["categoria_key"], errors="ignore").copy()
            resumo_skus_saida = limpar_dataframe_excel(resumo_skus_saida)
            resumo_skus_saida.to_excel(writer, sheet_name="SKUs por Categoria", index=False, startrow=1)
            ws = writer.sheets["SKUs por Categoria"]
            ws.merge_range("A1:R1", "SKUs por Categoria", fmt_titulo)
            aplicar_formatos_basicos(
                writer, "SKUs por Categoria", resumo_skus_saida, startrow=1, startcol=0,
                percent_cols={"Cobertura importância Sell-in", "Cobertura importância Sell-out", "Cobertura importância Sell-out Fab. Sell-in", "% Sell-in usado para detectar fabricante"},
                number_cols=set(resumo_skus_saida.columns) - {"Categoria", "Fabricante_referencia_Sellin", "Cobertura importância Sell-in", "Cobertura importância Sell-out", "Cobertura importância Sell-out Fab. Sell-in", "% Sell-in usado para detectar fabricante"},
            )

        # Crosschecks por UF e mês
        if output_options.get("crosschecks", True) and crosschecks:
            escrever_crosschecks(writer, crosschecks)

        # Parâmetros
        if output_options.get("parametros", True):
            detalhe_divisor = parametros.get("__detalhe_divisor_categoria__", [])
            parametros_visiveis = {k: v for k, v in parametros.items() if not str(k).startswith("__")}
            parametros_df = pd.DataFrame([{"Parâmetro": k, "Valor": v} for k, v in parametros_visiveis.items()])
            parametros_df.to_excel(writer, sheet_name="Parâmetros", index=False, startrow=1)
            ws = writer.sheets["Parâmetros"]
            ws.merge_range("A1:B1", "Parâmetros usados", fmt_titulo)
            aplicar_formatos_basicos(writer, "Parâmetros", parametros_df, startrow=1, startcol=0)

            # Divisor por categoria: uma linha por categoria/PROD, para não ficar acumulado em uma única célula.
            if isinstance(detalhe_divisor, list) and detalhe_divisor:
                divisor_df = pd.DataFrame(detalhe_divisor).copy()
                divisor_df = divisor_df.rename(columns={"Divisor": "Divisor aplicado"})
                colunas_divisor = [c for c in ["Categoria", "Coluna ajustada", "Divisor aplicado", "Ratio antes", "Ratio depois", "Status"] if c in divisor_df.columns]
                divisor_df = divisor_df[colunas_divisor]
                start_divisor = len(parametros_df) + 4
                ws.merge_range(start_divisor, 0, start_divisor, max(len(colunas_divisor) - 1, 1), "Divisor aplicado na volumetria por categoria", fmt_titulo)
                divisor_df.to_excel(writer, sheet_name="Parâmetros", index=False, startrow=start_divisor + 1, startcol=0)
                aplicar_formatos_basicos(
                    writer, "Parâmetros", divisor_df, startrow=start_divisor + 1, startcol=0,
                    number_cols={"Divisor aplicado", "Ratio antes", "Ratio depois"},
                )
                ws.set_column(0, 0, 34)
                ws.set_column(1, max(len(colunas_divisor) - 1, 1), 18)

        # Descrição dos cálculos
        if output_options.get("descricao_calculos", True):
            descricao_df = gerar_descricao_calculos()
            descricao_df.to_excel(writer, sheet_name="Descrição Cálculos", index=False, startrow=1)
            ws = writer.sheets["Descrição Cálculos"]
            ws.merge_range("A1:D1", "Descrição dos cálculos e colunas calculadas", fmt_titulo)
            aplicar_formatos_basicos(writer, "Descrição Cálculos", descricao_df, startrow=1, startcol=0)
            ws.set_column(0, 0, 24)
            ws.set_column(1, 1, 34)
            ws.set_column(2, 3, 70)

        # Avisos
        if output_options.get("avisos", True):
            avisos_df = pd.DataFrame({"Aviso": avisos if avisos else ["Nenhum aviso."]})
            avisos_df.to_excel(writer, sheet_name="Avisos", index=False, startrow=1)
            ws = writer.sheets["Avisos"]
            ws.write("A1", "Avisos", fmt_titulo)
            aplicar_formatos_basicos(writer, "Avisos", avisos_df, startrow=1, startcol=0)

        garantir_aba_info_se_vazio(writer, workbook, "Estudo de Cobertura", output_options)


# ============================================================
# Interface / execução
# ============================================================


def nome_arquivo_exibicao(caminho: Optional[str]) -> str:
    if not caminho:
        return "Nenhum arquivo selecionado"
    return nome_arquivo_curto(caminho) or "Nenhum arquivo selecionado"


def centralizar_janela(root, largura: int, altura: int):
    try:
        root.update_idletasks()
        x = int((root.winfo_screenwidth() - largura) / 2)
        y = int((root.winfo_screenheight() - altura) / 2)
        root.geometry(f"{largura}x{altura}+{x}+{y}")
    except Exception:
        root.geometry(f"{largura}x{altura}")




def listar_fabricantes_sellout(arquivo: str) -> List[str]:
    """Lê a coluna Fabricante do Sell-out sem carregar CSV grande inteiro."""
    try:
        caminho = Path(arquivo)
        if not caminho.exists():
            return []

        if eh_csv(caminho):
            enc, sep, header_row_idx, headers, header_start, header_end = preparar_leitura_csv_com_cabecalho(
                arquivo=caminho,
                obrigatorias=["Fabricante"],
                max_linhas_scan=200,
            )
            c_fab = localizar_coluna(pd.DataFrame(columns=headers), ["Fabricante", "Fabricante SKU", "FABRICANTE"], obrigatoria=False)
            if not c_fab:
                return []
            header_pos = {nome: header_start + i for i, nome in enumerate(headers)}
            usecols = [header_pos[c_fab]]
            engine = "python" if sep is None else "c"
            kwargs = dict(
                filepath_or_buffer=caminho,
                sep=sep,
                engine=engine,
                header=None,
                skiprows=header_row_idx,
                usecols=usecols,
                names=[c_fab],
                dtype=str,
                keep_default_na=False,
                skip_blank_lines=False,
                on_bad_lines="skip",
                encoding=enc,
                chunksize=300_000,
            )
            if engine == "c":
                kwargs["low_memory"] = False

            vistos = set()
            valores = []
            for chunk in pd.read_csv(**kwargs):
                vals = chunk[c_fab].dropna().astype(str).str.strip()
                for v in vals.unique().tolist():
                    key = normalizar_texto(v)
                    if v and key not in {"", "nan", "none", "total"} and key not in vistos:
                        vistos.add(key)
                        valores.append(v)
            return sorted(valores, key=lambda x: normalizar_texto(x))

        aba = escolher_aba(caminho, ["Sell-out - SM", "Sell-out", "Sellout", "VTA", "Dados", "Publicar"])
        raw, _ = ler_aba_com_cabecalho(caminho, aba, ["Fabricante"])
        c_fab = localizar_coluna(raw, ["Fabricante", "Fabricante SKU", "FABRICANTE"], obrigatoria=False)
        if not c_fab:
            return []
        vals = raw[c_fab].dropna().astype(str).str.strip()
        vals = [v for v in vals.unique().tolist() if v and normalizar_texto(v) not in {"nan", "none", "total"}]
        return sorted(vals, key=lambda x: normalizar_texto(x))
    except Exception:
        return []

def obter_configuracao_gui(args=None) -> Optional[Dict[str, str]]:
    """Interface única para selecionar Sell-in, Sell-out, saída e parâmetros."""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
    except Exception:
        return None

    try:
        root = tk.Tk()
    except Exception:
        return None

    root.title("Estudo de Cobertura - Sell-in x Sell-out")
    root.resizable(True, True)
    centralizar_janela(root, 900, 680)
    try:
        root.minsize(760, 520)
    except Exception:
        pass

    try:
        root.option_add("*Font", "{Segoe UI} 10")
    except Exception:
        pass

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"), foreground="#17365D")
    style.configure("Sub.TLabel", font=("Segoe UI", 9), foreground="#555555")
    style.configure("Section.TLabelframe.Label", font=("Segoe UI", 10, "bold"), foreground="#17365D")
    style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"))

    resultado: Dict[str, str] = {}

    sellin_var = tk.StringVar(value=getattr(args, "sellin", None) or "")
    sellout_var = tk.StringVar(value=getattr(args, "sellout", None) or "")
    sellout2_var = tk.StringVar(value=getattr(args, "sellout2", None) or "")
    sku_dash_var = tk.StringVar(value=getattr(args, "sku", None) or "")
    fabricante_dash_arquivo_var = tk.StringVar(value=(getattr(args, "congelado_estudo", None) or getattr(args, "arquivo_fabricante", None) or ""))
    sugestao = f"estudo_cobertura_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    saida_var = tk.StringVar(value=getattr(args, "saida", None) or "")
    metrica_inicial = (getattr(args, "metrica", None) or "volume").lower()
    nivel_inicial = (getattr(args, "nivel", None) or "CATEGORIA").upper()
    if nivel_inicial not in {"CATEGORIA", "NIVEL1", "NIVEL2", "ESTMER7"}:
        nivel_inicial = "CATEGORIA"
    if metrica_eh_volume_variavel(metrica_inicial):
        metrica_inicial = "volume_variavel"
    if metrica_inicial not in {"volume", "quantia", "volume_variavel"}:
        metrica_inicial = "volume"
    metrica_var = tk.StringVar(value=metrica_inicial)
    nivel_var = tk.StringVar(value=nivel_inicial)
    fabricante_var = tk.StringVar(value=getattr(args, "fabricante", None) or "")
    # Deixa o campo de fabricante disponível desde o início.
    # Se ficar vazio, o código não aplica filtro.
    usar_fabricante_var = tk.BooleanVar(value=True)
    resumo_categorias_var = tk.BooleanVar(value=bool(getattr(args, "gerar_resumo_categorias", True)))
    abas_categorias_var = tk.BooleanVar(value=bool(getattr(args, "gerar_abas_categorias", True)))
    base_skus_var = tk.BooleanVar(value=bool(getattr(args, "gerar_base_skus", True)))
    base_contribuicao_var = tk.BooleanVar(value=bool(getattr(args, "gerar_base_contribuicao_sellout", False)))
    skus_por_categoria_var = tk.BooleanVar(value=bool(getattr(args, "gerar_skus_por_categoria", True)))
    crosschecks_var = tk.BooleanVar(value=bool(getattr(args, "gerar_crosschecks", True)))
    parametros_var = tk.BooleanVar(value=bool(getattr(args, "gerar_parametros", True)))
    descricao_calculos_var = tk.BooleanVar(value=bool(getattr(args, "gerar_descricao_calculos", True)))
    avisos_var = tk.BooleanVar(value=bool(getattr(args, "gerar_avisos", True)))
    abas_auxiliares_comparacao_var = tk.BooleanVar(value=bool(getattr(args, "gerar_abas_auxiliares_comparacao", True)))
    gerar_top20_sku_canal_uf_var = tk.BooleanVar(value=bool(getattr(args, "gerar_top20_sku_canal_uf", False)))
    fabricantes_disponiveis: List[str] = []
    modo_inicial = (getattr(args, "modo", None) or "estudo").lower()
    usar_segundo_sellout_inicial = bool(
        (getattr(args, "sellout2", None) or "")
        or modo_inicial in {"estudo2out", "estudo_2_sellouts", "dois_sellouts"}
    )
    if modo_inicial == "comparacao":
        modo_label_inicial = "Comparação de Estudo de Cobertura"
    elif modo_inicial in {"dash", "cobertura_dash", "coberturadash"}:
        modo_label_inicial = "Cobertura Dash"
    else:
        modo_label_inicial = "Estudo de Cobertura"
    modo_var = tk.StringVar(value=modo_label_inicial)
    usar_segundo_sellout_var = tk.BooleanVar(value=usar_segundo_sellout_inicial and modo_label_inicial == "Estudo de Cobertura")

    # Área rolável: evita que opções/botões fiquem escondidos em telas menores
    outer = ttk.Frame(root)
    outer.pack(fill="both", expand=True)

    canvas = tk.Canvas(outer, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    main = ttk.Frame(canvas, padding=18)
    canvas_window = canvas.create_window((0, 0), window=main, anchor="nw")

    def ajustar_scrollregion(_event=None):
        try:
            canvas.configure(scrollregion=canvas.bbox("all"))
        except Exception:
            pass

    def ajustar_largura_canvas(event):
        try:
            canvas.itemconfigure(canvas_window, width=event.width)
        except Exception:
            pass

    def rolar_mouse(event):
        try:
            if event.num == 4:
                canvas.yview_scroll(-3, "units")
            elif event.num == 5:
                canvas.yview_scroll(3, "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass

    main.bind("<Configure>", ajustar_scrollregion)
    canvas.bind("<Configure>", ajustar_largura_canvas)
    canvas.bind_all("<MouseWheel>", rolar_mouse)
    canvas.bind_all("<Button-4>", rolar_mouse)
    canvas.bind_all("<Button-5>", rolar_mouse)

    modo_combo = ttk.Combobox(
        main,
        textvariable=modo_var,
        values=["Estudo de Cobertura", "Comparação de Estudo de Cobertura", "Cobertura Dash"],
        state="readonly",
        width=42,
    )
    modo_combo.grid(row=0, column=0, sticky="w")
    subtitulo_var = tk.StringVar(value="Selecione os arquivos, defina a métrica e escolha onde o Excel final será salvo.")
    ttk.Label(
        main,
        textvariable=subtitulo_var,
        style="Sub.TLabel",
    ).grid(row=1, column=0, sticky="w", pady=(2, 14))

    files_frame = ttk.LabelFrame(main, text="Arquivos", padding=14, style="Section.TLabelframe")
    files_frame.grid(row=2, column=0, sticky="ew")
    files_frame.columnconfigure(1, weight=1)

    def modo_eh_comparacao() -> bool:
        return modo_var.get().strip() == "Comparação de Estudo de Cobertura"

    def modo_eh_estudo_2out() -> bool:
        return modo_var.get().strip() == "Estudo de Cobertura" and bool(usar_segundo_sellout_var.get())

    def modo_eh_dash() -> bool:
        return modo_var.get().strip() == "Cobertura Dash"

    def escolher_sellin():
        if modo_eh_comparacao():
            titulo = "Selecione o estudo de cobertura da versão 2.0"
            tipos = [("Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")]
        else:
            titulo = "Selecione a planilha de Sell-in"
            tipos = [("Excel ou CSV", "*.xlsx *.xlsm *.csv *.txt *.tsv"), ("Excel", "*.xlsx *.xlsm"), ("CSV/TXT", "*.csv *.txt *.tsv"), ("Todos os arquivos", "*.*")]
        caminho = filedialog.askopenfilename(title=titulo, filetypes=tipos)
        if caminho:
            sellin_var.set(caminho)
            if not saida_var.get().strip():
                nome_saida = f"comparacao_cobertura_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx" if (modo_eh_comparacao() or modo_eh_estudo_2out()) else (f"cobertura_dash_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx" if modo_eh_dash() else sugestao)
                saida_var.set(str(Path(caminho).with_name(nome_saida)))

    def criar_popup_loading(titulo: str, mensagem: str):
        popup = tk.Toplevel(root)
        popup.title(titulo)
        popup.resizable(False, False)
        popup.transient(root)
        popup.grab_set()
        try:
            popup.option_add("*Font", "{Segoe UI} 10")
        except Exception:
            pass

        frame = ttk.Frame(popup, padding=20)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text=mensagem, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 10))
        barra = ttk.Progressbar(frame, mode="indeterminate", length=360)
        barra.pack(fill="x")
        ttk.Label(frame, text="Aguarde até terminar a leitura do arquivo.", style="Sub.TLabel").pack(anchor="w", pady=(10, 0))
        barra.start(12)

        popup.update_idletasks()
        try:
            x = root.winfo_rootx() + int((root.winfo_width() - popup.winfo_width()) / 2)
            y = root.winfo_rooty() + int((root.winfo_height() - popup.winfo_height()) / 2)
            popup.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        except Exception:
            pass
        popup.protocol("WM_DELETE_WINDOW", lambda: None)
        return popup, barra

    def aplicar_lista_fabricantes(lista: List[str]):
        nonlocal fabricantes_disponiveis
        fabricantes_disponiveis = lista or []
        valores = [""] + fabricantes_disponiveis
        try:
            fabricante_combo["values"] = valores
        except Exception:
            pass
        if fabricante_var.get().strip() and fabricantes_disponiveis and fabricante_var.get().strip() not in fabricantes_disponiveis:
            fabricante_var.set("")
        atualizar_estado_fabricante()
        atualizar_status()

    def atualizar_estado_fabricante(*_):
        try:
            if usar_fabricante_var.get():
                if fabricantes_disponiveis:
                    fabricante_combo.configure(state="readonly")
                else:
                    # Permite digitar manualmente se a lista ainda não foi carregada.
                    fabricante_combo.configure(state="normal")
                carregar_fab_btn.configure(state="normal")
            else:
                fabricante_var.set("")
                fabricante_combo.configure(state="disabled")
                carregar_fab_btn.configure(state="disabled")
        except Exception:
            pass

    def carregar_fabricantes():
        usar_fabricante_var.set(True)
        caminho = sellout_var.get().strip()
        if not caminho:
            aplicar_lista_fabricantes([])
            return []

        try:
            fabricante_combo.configure(state="disabled")
            fabricante_combo["values"] = ["Carregando..."]
        except Exception:
            pass

        popup, barra = criar_popup_loading(
            "Carregando Sell-out",
            "Lendo fabricantes do Sell-out...",
        )
        fila = queue.Queue()

        def worker():
            try:
                fila.put(("ok", listar_fabricantes_sellout(caminho)))
            except Exception as exc:
                fila.put(("erro", exc))

        threading.Thread(target=worker, daemon=True).start()

        def verificar():
            try:
                status, payload = fila.get_nowait()
            except queue.Empty:
                root.after(120, verificar)
                return

            try:
                barra.stop()
                popup.grab_release()
                popup.destroy()
            except Exception:
                pass

            if status == "ok":
                aplicar_lista_fabricantes(payload)
            else:
                aplicar_lista_fabricantes([])
                messagebox.showwarning(
                    "Fabricantes não carregados",
                    "Não consegui carregar a lista de fabricantes automaticamente.\n"
                    "Você ainda pode gerar o estudo sem selecionar fabricante, ou revisar o arquivo Sell-out.\n\n"
                    f"Detalhe: {payload}",
                )

        root.after(120, verificar)
        return []

    def escolher_sellout():
        if modo_eh_comparacao():
            titulo = "Selecione o estudo de cobertura da versão 3.0"
            tipos = [("Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")]
        else:
            if modo_eh_dash():
                titulo = "Selecione o arquivo Vendas UF"
            elif modo_eh_estudo_2out():
                titulo = "Selecione a planilha de Sell-out 2.0 / Publicar"
            else:
                titulo = "Selecione a planilha de Sell-out / Publicar"
            tipos = [("Excel ou CSV", "*.xlsx *.xlsm *.csv *.txt *.tsv"), ("Excel", "*.xlsx *.xlsm"), ("CSV/TXT", "*.csv *.txt *.tsv"), ("Todos os arquivos", "*.*")]
        caminho = filedialog.askopenfilename(title=titulo, filetypes=tipos)
        if caminho:
            sellout_var.set(caminho)
            if modo_eh_dash() and not saida_var.get().strip():
                saida_var.set(str(Path(caminho).with_name(f"cobertura_dash_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")))
            # Não carrega fabricantes automaticamente, para evitar travar em Sell-out/CSV grande.
            aplicar_lista_fabricantes([])

    def escolher_sellout2():
        titulo = "Selecione o arquivo Vendas SKU (opcional)" if modo_eh_dash() else "Selecione a planilha de Sell-out 3.0 / Publicar"
        tipos = [("Excel ou CSV", "*.xlsx *.xlsm *.csv *.txt *.tsv"), ("Excel", "*.xlsx *.xlsm"), ("CSV/TXT", "*.csv *.txt *.tsv"), ("Todos os arquivos", "*.*")]
        caminho = filedialog.askopenfilename(title=titulo, filetypes=tipos)
        if caminho:
            sellout2_var.set(caminho)

    def escolher_sku_dash():
        titulo = "Selecione o arquivo SKU"
        tipos = [("Excel ou CSV", "*.xlsx *.xlsm *.csv *.txt *.tsv"), ("Excel", "*.xlsx *.xlsm"), ("CSV/TXT", "*.csv *.txt *.tsv"), ("Todos os arquivos", "*.*")]
        caminho = filedialog.askopenfilename(title=titulo, filetypes=tipos)
        if caminho:
            sku_dash_var.set(caminho)

    def escolher_fabricante_dash_arquivo():
        titulo = "Selecione o arquivo Congelado"
        tipos = [("Excel ou CSV", "*.xlsx *.xlsm *.csv *.txt *.tsv"), ("Excel", "*.xlsx *.xlsm"), ("CSV/TXT", "*.csv *.txt *.tsv"), ("Todos os arquivos", "*.*")]
        caminho = filedialog.askopenfilename(title=titulo, filetypes=tipos)
        if caminho:
            fabricante_dash_arquivo_var.set(caminho)

    def escolher_saida():
        inicial = saida_var.get().strip()
        initialdir = str(Path(inicial).parent) if inicial else None
        initialfile = Path(inicial).name if inicial else sugestao
        caminho = filedialog.asksaveasfilename(
            title="Salvar estudo de cobertura como",
            initialdir=initialdir,
            initialfile=initialfile,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if caminho:
            saida_var.set(caminho)

    label_entrada_1_var = tk.StringVar(value="Sell-in")
    label_entrada_2_var = tk.StringVar(value="Sell-out")
    label_entrada_3_var = tk.StringVar(value="Sell-out 3.0")
    label_entrada_4_var = tk.StringVar(value="SKU")
    label_entrada_5_var = tk.StringVar(value="Congelado")
    label_saida_var = tk.StringVar(value="Salvar em")

    def linha_arquivo(row: int, titulo_var, var, comando, texto_botao: str):
        lbl = ttk.Label(files_frame, textvariable=titulo_var)
        lbl.grid(row=row, column=0, sticky="w", padx=(0, 10), pady=7)
        entrada = ttk.Entry(files_frame, textvariable=var, width=76, state="readonly")
        entrada.grid(row=row, column=1, sticky="ew", pady=7)
        btn = ttk.Button(files_frame, text=texto_botao, command=comando, width=14)
        btn.grid(row=row, column=2, sticky="e", padx=(10, 0), pady=7)
        return (lbl, entrada, btn)

    row_entrada_1 = linha_arquivo(0, label_entrada_1_var, sellin_var, escolher_sellin, "Selecionar")
    row_entrada_2 = linha_arquivo(1, label_entrada_2_var, sellout_var, escolher_sellout, "Selecionar")
    row_entrada_3 = linha_arquivo(2, label_entrada_3_var, sellout2_var, escolher_sellout2, "Selecionar")
    row_entrada_4 = linha_arquivo(3, label_entrada_4_var, sku_dash_var, escolher_sku_dash, "Selecionar")
    row_entrada_5 = linha_arquivo(4, label_entrada_5_var, fabricante_dash_arquivo_var, escolher_fabricante_dash_arquivo, "Selecionar")
    row_saida = linha_arquivo(5, label_saida_var, saida_var, escolher_saida, "Salvar como")

    opts_frame = ttk.LabelFrame(main, text="Parâmetros", padding=14, style="Section.TLabelframe")
    opts_frame.grid(row=3, column=0, sticky="ew", pady=(14, 0))
    opts_frame.columnconfigure(3, weight=1)

    ttk.Label(opts_frame, text="Cálculo da cobertura:").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=5)
    metrica_combo = ttk.Combobox(opts_frame, textvariable=metrica_var, values=["volume", "quantia", "volume_variavel"], state="readonly", width=22)
    metrica_combo.grid(row=0, column=1, sticky="w", padx=(0, 28), pady=5)

    ttk.Label(opts_frame, text="Agrupar por:").grid(row=0, column=2, sticky="w", padx=(0, 8), pady=5)
    nivel_combo = ttk.Combobox(opts_frame, textvariable=nivel_var, values=["CATEGORIA", "NIVEL1", "NIVEL2", "ESTMER7"], state="readonly", width=18)
    nivel_combo.grid(row=0, column=3, sticky="w", pady=5)

    usar_segundo_sellout_check = ttk.Checkbutton(
        opts_frame,
        text="Usar segundo Sell-out (comparar Sell-out 2.0 x Sell-out 3.0)",
        variable=usar_segundo_sellout_var,
    )
    usar_segundo_sellout_check.grid(row=1, column=0, columnspan=4, sticky="w", pady=(10, 3))

    ttk.Checkbutton(
        opts_frame,
        text="Aplicar filtro opcional por fabricante do Sell-out",
        variable=usar_fabricante_var,
        command=atualizar_estado_fabricante,
    ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 3))

    ttk.Label(opts_frame, text="Fabricante do Sell-out:").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=5)
    fabricante_combo = ttk.Combobox(opts_frame, textvariable=fabricante_var, values=[""], state="normal", width=44)
    fabricante_combo.grid(row=3, column=1, columnspan=2, sticky="ew", padx=(0, 8), pady=5)
    carregar_fab_btn = ttk.Button(opts_frame, text="Carregar lista", command=carregar_fabricantes, width=15, state="disabled")
    carregar_fab_btn.grid(row=3, column=3, sticky="w", pady=5)

    ajuda = (
        "Fabricante é opcional. Você pode digitar exatamente como está no Sell-out para filtrar antes da análise.\n"
        "A lista de fabricantes não é carregada automaticamente para evitar travamento em arquivos grandes; carregar lista é opcional.\n"
        "Sell-in: usa a aba do template e prioriza cabeçalho na linha 14, coluna B.\n"
        "Categoria: usa primeiro a categoria existente no Sell-out; quando não encontrar, tenta usar a Categoria do Sell-in por SKU/EAN. Se houver Congelado opcional, usa a Categoria do Congelado por SKU/EAN, sem fallback para Total Fabricante.\n"
        "NIVEL1/NIVEL2/ESTMER7: usa o PROD do Sell-out; se houver Congelado opcional, ele é ignorado nesses modos.\n"
        "Segundo Sell-out: quando marcado, o primeiro arquivo vira Sell-out 2.0 e o segundo vira Sell-out 3.0; quando desmarcado, o estudo usa apenas um Sell-out."
    )
    ttk.Label(opts_frame, text=ajuda, style="Sub.TLabel", wraplength=790).grid(
        row=5, column=0, columnspan=4, sticky="w", pady=(10, 0)
    )
    atualizar_estado_fabricante()

    outputs_frame = ttk.LabelFrame(main, text="O que gerar no Excel", padding=14, style="Section.TLabelframe")
    outputs_frame.grid(row=4, column=0, sticky="ew", pady=(14, 0))
    outputs_frame.columnconfigure(0, weight=1)
    outputs_frame.columnconfigure(1, weight=1)

    opcoes_checkboxes = [
        ("Gerar Resumo Categorias", resumo_categorias_var),
        ("Gerar abas por categoria/PROD", abas_categorias_var),
        ("Gerar Base SKUs", base_skus_var),
        ("Gerar SKUs por Categoria", skus_por_categoria_var),
        ("Gerar Crosschecks", crosschecks_var),
        ("Gerar Parâmetros", parametros_var),
        ("Gerar Descrição Cálculos", descricao_calculos_var),
        ("Gerar Avisos", avisos_var),
        ("Gerar Base Contribuição Sell-out (oculta)", base_contribuicao_var),
        ("Gerar abas auxiliares da comparação", abas_auxiliares_comparacao_var),
        ("Gerar TOP 20 SKU por UF/Canal", gerar_top20_sku_canal_uf_var),
    ]
    for idx, (texto, var_opcao) in enumerate(opcoes_checkboxes):
        ttk.Checkbutton(outputs_frame, text=texto, variable=var_opcao).grid(
            row=idx // 2, column=idx % 2, sticky="w", padx=(0, 18), pady=3
        )
    ttk.Label(
        outputs_frame,
        text="Padrão: principais abas do estudo ficam marcadas. A Base Contribuição Sell-out e o TOP 20 começam desmarcados.",
        style="Sub.TLabel",
        wraplength=790,
    ).grid(row=6, column=0, columnspan=2, sticky="w", pady=(8, 0))

    def coletar_opcoes_saida_gui() -> Dict[str, bool]:
        return normalizar_opcoes_saida({
            "resumo_categorias": resumo_categorias_var.get(),
            "abas_categorias": abas_categorias_var.get(),
            "base_skus": base_skus_var.get(),
            "base_contribuicao_sellout": base_contribuicao_var.get(),
            "skus_por_categoria": skus_por_categoria_var.get(),
            "crosschecks": crosschecks_var.get(),
            "parametros": parametros_var.get(),
            "descricao_calculos": descricao_calculos_var.get(),
            "avisos": avisos_var.get(),
            "abas_auxiliares_comparacao": abas_auxiliares_comparacao_var.get(),
            "top20_sku_canal_uf": gerar_top20_sku_canal_uf_var.get(),
        })

    status_frame = ttk.LabelFrame(main, text="Arquivos carregados", padding=14, style="Section.TLabelframe")
    status_frame.grid(row=5, column=0, sticky="ew", pady=(14, 0))
    status_frame.columnconfigure(1, weight=1)

    sellin_nome = tk.StringVar()
    sellout_nome = tk.StringVar()
    sellout2_nome = tk.StringVar()
    sku_dash_nome = tk.StringVar()
    fabricante_dash_arquivo_nome = tk.StringVar()
    saida_nome = tk.StringVar()
    fabricante_nome = tk.StringVar()
    status_label_1_var = tk.StringVar(value="Sell-in:")
    status_label_2_var = tk.StringVar(value="Sell-out:")
    status_label_3_var = tk.StringVar(value="Sell-out 3.0:")
    status_label_4_var = tk.StringVar(value="Fabricante:")

    def atualizar_status(*_):
        sellin_nome.set(nome_arquivo_exibicao(sellin_var.get()))
        sellout_nome.set(nome_arquivo_exibicao(sellout_var.get()))
        sellout2_nome.set(nome_arquivo_exibicao(sellout2_var.get()))
        sku_dash_nome.set(nome_arquivo_exibicao(sku_dash_var.get()))
        fabricante_dash_arquivo_nome.set(nome_arquivo_exibicao(fabricante_dash_arquivo_var.get()))
        saida_nome.set(nome_arquivo_exibicao(saida_var.get()))
        if modo_eh_comparacao():
            fabricante_nome.set("Comparação de estudos gerados")
        elif modo_eh_estudo_2out():
            fabricante_nome.set(fabricante_var.get().strip() or "Automático / todos")
        elif usar_fabricante_var.get():
            fabricante_nome.set(fabricante_var.get().strip() or "Selecionado, mas ainda vazio")
        else:
            fabricante_nome.set("Automático / todos")

    for var in (sellin_var, sellout_var, sellout2_var, sku_dash_var, fabricante_dash_arquivo_var, saida_var, fabricante_var, usar_segundo_sellout_var):
        var.trace_add("write", atualizar_status)
    try:
        usar_fabricante_var.trace_add("write", atualizar_status)
    except Exception:
        pass
    atualizar_status()

    ttk.Label(status_frame, textvariable=status_label_1_var, foreground="#555555").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
    ttk.Label(status_frame, textvariable=sellin_nome, font=("Segoe UI", 9, "bold")).grid(row=0, column=1, sticky="w", pady=2)
    ttk.Label(status_frame, textvariable=status_label_2_var, foreground="#555555").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=2)
    ttk.Label(status_frame, textvariable=sellout_nome, font=("Segoe UI", 9, "bold")).grid(row=1, column=1, sticky="w", pady=2)
    status_terceiro_label = ttk.Label(status_frame, textvariable=status_label_3_var, foreground="#555555")
    status_terceiro_valor = ttk.Label(status_frame, textvariable=sellout2_nome, font=("Segoe UI", 9, "bold"))
    status_terceiro_label.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=2)
    status_terceiro_valor.grid(row=2, column=1, sticky="w", pady=2)
    status_sku_dash_label = ttk.Label(status_frame, text="SKU:", foreground="#555555")
    status_sku_dash_valor = ttk.Label(status_frame, textvariable=sku_dash_nome, font=("Segoe UI", 9, "bold"))
    status_sku_dash_label.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=2)
    status_sku_dash_valor.grid(row=3, column=1, sticky="w", pady=2)
    status_fabricante_dash_label = ttk.Label(status_frame, text="Arquivo Congelado:", foreground="#555555")
    status_fabricante_dash_valor = ttk.Label(status_frame, textvariable=fabricante_dash_arquivo_nome, font=("Segoe UI", 9, "bold"))
    status_fabricante_dash_label.grid(row=4, column=0, sticky="w", padx=(0, 8), pady=2)
    status_fabricante_dash_valor.grid(row=4, column=1, sticky="w", pady=2)
    ttk.Label(status_frame, text="Saída:", foreground="#555555").grid(row=5, column=0, sticky="w", padx=(0, 8), pady=2)
    ttk.Label(status_frame, textvariable=saida_nome, font=("Segoe UI", 9, "bold")).grid(row=5, column=1, sticky="w", pady=2)
    ttk.Label(status_frame, textvariable=status_label_4_var, foreground="#555555").grid(row=6, column=0, sticky="w", padx=(0, 8), pady=2)
    ttk.Label(status_frame, textvariable=fabricante_nome, font=("Segoe UI", 9, "bold")).grid(row=6, column=1, sticky="w", pady=2)

    buttons = ttk.Frame(main)
    buttons.grid(row=6, column=0, sticky="e", pady=(18, 0))

    def cancelar():
        root.destroy()

    def confirmar():
        entrada_1 = sellin_var.get().strip()
        entrada_2 = sellout_var.get().strip()
        entrada_3 = sellout2_var.get().strip()
        saida = saida_var.get().strip()
        comparacao = modo_eh_comparacao()
        estudo_2out = modo_eh_estudo_2out()
        dash = modo_eh_dash()

        if not entrada_1:
            msg = "Selecione o estudo de cobertura da versão 2.0." if comparacao else "Selecione a planilha de Sell-in."
            messagebox.showerror("Arquivo obrigatório", msg)
            return
        if not Path(entrada_1).exists():
            nome = "Estudo 2.0" if comparacao else "Sell-in"
            messagebox.showerror("Arquivo não encontrado", f"{nome} não encontrado:\n{entrada_1}")
            return
        if not entrada_2:
            if dash:
                msg = "Selecione o arquivo Vendas UF."
            elif estudo_2out:
                msg = "Selecione a planilha de Sell-out 2.0."
            else:
                msg = "Selecione o estudo de cobertura da versão 3.0." if comparacao else "Selecione a planilha de Sell-out."
            messagebox.showerror("Arquivo obrigatório", msg)
            return
        if not Path(entrada_2).exists():
            if dash:
                nome = "Vendas UF"
            elif estudo_2out:
                nome = "Sell-out 2.0"
            else:
                nome = "Estudo 3.0" if comparacao else "Sell-out"
            messagebox.showerror("Arquivo não encontrado", f"{nome} não encontrado:\n{entrada_2}")
            return
        if estudo_2out and not entrada_3:
            messagebox.showerror("Arquivo obrigatório", "Selecione a planilha de Sell-out 3.0.")
            return
        if estudo_2out and not Path(entrada_3).exists():
            messagebox.showerror("Arquivo não encontrado", f"Sell-out 3.0 não encontrado:\n{entrada_3}")
            return
        if dash:
            if not sku_dash_var.get().strip():
                messagebox.showerror("Arquivo obrigatório", "Selecione o arquivo SKU.")
                return
            if not Path(sku_dash_var.get().strip()).exists():
                messagebox.showerror("Arquivo não encontrado", f"SKU não encontrado:\n{sku_dash_var.get().strip()}")
                return
            if not fabricante_dash_arquivo_var.get().strip():
                messagebox.showerror("Arquivo obrigatório", "Selecione o arquivo Congelado.")
                return
            if not Path(fabricante_dash_arquivo_var.get().strip()).exists():
                messagebox.showerror("Arquivo não encontrado", f"Arquivo Congelado não encontrado:\n{fabricante_dash_arquivo_var.get().strip()}")
                return
        elif fabricante_dash_arquivo_var.get().strip() and not Path(fabricante_dash_arquivo_var.get().strip()).exists():
            messagebox.showerror("Arquivo não encontrado", f"Arquivo Congelado opcional não encontrado:\n{fabricante_dash_arquivo_var.get().strip()}")
            return
        if not saida:
            messagebox.showerror("Saída obrigatória", "Escolha onde o arquivo final será salvo.")
            return
        if not saida.lower().endswith(".xlsx"):
            saida += ".xlsx"

        if comparacao:
            resultado.update({
                "modo": "comparacao",
                "estudo20": entrada_1,
                "estudo30": entrada_2,
                "saida": saida,
                "gerar_top20_sku_canal_uf": bool(gerar_top20_sku_canal_uf_var.get()),
                "output_options": coletar_opcoes_saida_gui(),
            })
        elif estudo_2out:
            resultado.update({
                "modo": "estudo2out",
                "sellin": entrada_1,
                "sellout20": entrada_2,
                "sellout30": entrada_3,
                "saida": saida,
                "metrica": metrica_var.get().strip().lower(),
                "nivel": nivel_var.get().strip().upper(),
                "fabricante": fabricante_var.get().strip() if usar_fabricante_var.get() else "",
                "congelado": fabricante_dash_arquivo_var.get().strip(),
                "gerar_top20_sku_canal_uf": bool(gerar_top20_sku_canal_uf_var.get()),
                "output_options": coletar_opcoes_saida_gui(),
            })
        elif dash:
            resultado.update({
                "modo": "dash",
                "sellin": entrada_1,
                "sellout_dash": entrada_2,
                "vendas_sku": entrada_3,
                "sku": sku_dash_var.get().strip(),
                "arquivo_fabricante": fabricante_dash_arquivo_var.get().strip(),
                "saida": saida,
                "metrica": metrica_var.get().strip().lower(),
                "nivel": nivel_var.get().strip().upper(),
                "fabricante": fabricante_var.get().strip() if usar_fabricante_var.get() else "",
                "output_options": coletar_opcoes_saida_gui(),
            })
        else:
            resultado.update({
                "modo": "estudo",
                "sellin": entrada_1,
                "sellout": entrada_2,
                "congelado": fabricante_dash_arquivo_var.get().strip(),
                "saida": saida,
                "metrica": metrica_var.get().strip().lower(),
                "nivel": nivel_var.get().strip().upper(),
                "fabricante": fabricante_var.get().strip() if usar_fabricante_var.get() else "",
                "output_options": coletar_opcoes_saida_gui(),
            })
        root.destroy()

    ttk.Button(buttons, text="Cancelar", command=cancelar, width=14).pack(side="right", padx=(10, 0))
    gerar_btn = ttk.Button(buttons, text="Gerar estudo", command=confirmar, style="Primary.TButton", width=24)
    gerar_btn.pack(side="right")

    def atualizar_modo_interface(*_):
        try:
            if modo_var.get().strip() == "Estudo de Cobertura":
                usar_segundo_sellout_check.grid(row=1, column=0, columnspan=4, sticky="w", pady=(10, 3))
            else:
                usar_segundo_sellout_check.grid_remove()
        except Exception:
            pass

        if modo_eh_comparacao():
            label_entrada_1_var.set("Estudo 2.0")
            for w in row_entrada_1:
                w.grid()
            try:
                status_frame.grid_slaves(row=0, column=0)[0].grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
                status_frame.grid_slaves(row=0, column=1)[0].grid(row=0, column=1, sticky="w", pady=2)
            except Exception:
                pass
            label_entrada_2_var.set("Estudo 3.0")
            label_entrada_3_var.set("Sell-out 3.0")
            subtitulo_var.set("Selecione dois arquivos gerados por este código para comparar as coberturas.")
            status_label_1_var.set("Estudo 2.0:")
            status_label_2_var.set("Estudo 3.0:")
            status_label_3_var.set("Sell-out 3.0:")
            status_label_4_var.set("Modo:")
            gerar_btn.configure(text="Gerar comparação")
            for w in row_entrada_3:
                w.grid_remove()
            for w in row_entrada_4:
                w.grid_remove()
            for w in row_entrada_5:
                w.grid_remove()
            status_terceiro_label.grid_remove()
            status_terceiro_valor.grid_remove()
            status_sku_dash_label.grid_remove()
            status_sku_dash_valor.grid_remove()
            status_fabricante_dash_label.grid_remove()
            status_fabricante_dash_valor.grid_remove()
            try:
                opts_frame.grid_remove()
            except Exception:
                pass
        elif modo_eh_estudo_2out():
            label_entrada_1_var.set("Sell-in")
            for w in row_entrada_1:
                w.grid()
            try:
                status_frame.grid_slaves(row=0, column=0)[0].grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
                status_frame.grid_slaves(row=0, column=1)[0].grid(row=0, column=1, sticky="w", pady=2)
            except Exception:
                pass
            label_entrada_2_var.set("Sell-out 2.0")
            label_entrada_3_var.set("Sell-out 3.0")
            label_entrada_5_var.set("Congelado (opcional)")
            subtitulo_var.set("Selecione um Sell-in, dois Sell-outs e, se quiser, o Congelado para definir a categoria dos SKUs pela base congelada.")
            status_label_1_var.set("Sell-in:")
            status_label_2_var.set("Sell-out 2.0:")
            status_label_3_var.set("Sell-out 3.0:")
            status_label_4_var.set("Fabricante:")
            gerar_btn.configure(text="Gerar comparação")
            for w in row_entrada_3:
                w.grid()
            for w in row_entrada_4:
                w.grid_remove()
            for w in row_entrada_5:
                w.grid()
            status_terceiro_label.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=2)
            status_terceiro_valor.grid(row=2, column=1, sticky="w", pady=2)
            status_sku_dash_label.grid_remove()
            status_sku_dash_valor.grid_remove()
            status_fabricante_dash_label.configure(text="Congelado opcional:")
            status_fabricante_dash_label.grid(row=4, column=0, sticky="w", padx=(0, 8), pady=2)
            status_fabricante_dash_valor.grid(row=4, column=1, sticky="w", pady=2)
            try:
                opts_frame.grid(row=3, column=0, sticky="ew", pady=(14, 0))
            except Exception:
                pass
        elif modo_eh_dash():
            label_entrada_1_var.set("Sell-in")
            label_entrada_2_var.set("Vendas UF")
            label_entrada_3_var.set("Vendas SKU (opcional)")
            label_entrada_4_var.set("SKU")
            label_entrada_5_var.set("Congelado")
            subtitulo_var.set("Selecione Sell-in, Vendas UF, SKU, Congelado e, se tiver, Vendas SKU para gerar a Cobertura Dash.")
            status_label_1_var.set("Sell-in:")
            status_label_2_var.set("Vendas UF:")
            status_label_3_var.set("Vendas SKU:")
            status_label_4_var.set("Fabricante filtro:")
            gerar_btn.configure(text="Gerar Cobertura Dash")
            for w in row_entrada_1:
                w.grid()
            for w in row_entrada_3:
                w.grid()
            for w in row_entrada_4:
                w.grid()
            for w in row_entrada_5:
                w.grid()
            # No Dash com Sell-in, mantém a primeira linha de status visível.
            try:
                status_frame.grid_slaves(row=0, column=0)[0].grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
                status_frame.grid_slaves(row=0, column=1)[0].grid(row=0, column=1, sticky="w", pady=2)
            except Exception:
                pass
            status_terceiro_label.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=2)
            status_terceiro_valor.grid(row=2, column=1, sticky="w", pady=2)
            status_sku_dash_label.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=2)
            status_sku_dash_valor.grid(row=3, column=1, sticky="w", pady=2)
            status_fabricante_dash_label.configure(text="Arquivo Congelado:")
            status_fabricante_dash_label.grid(row=4, column=0, sticky="w", padx=(0, 8), pady=2)
            status_fabricante_dash_valor.grid(row=4, column=1, sticky="w", pady=2)
            try:
                opts_frame.grid(row=3, column=0, sticky="ew", pady=(14, 0))
            except Exception:
                pass
        else:
            label_entrada_1_var.set("Sell-in")
            for w in row_entrada_1:
                w.grid()
            try:
                status_frame.grid_slaves(row=0, column=0)[0].grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
                status_frame.grid_slaves(row=0, column=1)[0].grid(row=0, column=1, sticky="w", pady=2)
            except Exception:
                pass
            label_entrada_2_var.set("Sell-out")
            label_entrada_3_var.set("Sell-out 3.0")
            label_entrada_5_var.set("Congelado (opcional)")
            subtitulo_var.set("Selecione Sell-in, Sell-out e, se quiser, o Congelado para definir a categoria dos SKUs pela base congelada.")
            status_label_1_var.set("Sell-in:")
            status_label_2_var.set("Sell-out:")
            status_label_3_var.set("Sell-out 3.0:")
            status_label_4_var.set("Fabricante:")
            gerar_btn.configure(text="Gerar estudo")
            for w in row_entrada_3:
                w.grid_remove()
            for w in row_entrada_4:
                w.grid_remove()
            for w in row_entrada_5:
                w.grid()
            status_terceiro_label.grid_remove()
            status_terceiro_valor.grid_remove()
            status_sku_dash_label.grid_remove()
            status_sku_dash_valor.grid_remove()
            status_fabricante_dash_label.configure(text="Congelado opcional:")
            status_fabricante_dash_label.grid(row=4, column=0, sticky="w", padx=(0, 8), pady=2)
            status_fabricante_dash_valor.grid(row=4, column=1, sticky="w", pady=2)
            try:
                opts_frame.grid(row=3, column=0, sticky="ew", pady=(14, 0))
            except Exception:
                pass
        atualizar_status()
        try:
            canvas.configure(scrollregion=canvas.bbox("all"))
        except Exception:
            pass

    modo_var.trace_add("write", atualizar_modo_interface)
    usar_segundo_sellout_var.trace_add("write", atualizar_modo_interface)
    atualizar_modo_interface()

    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()
    return resultado or None


def obter_args():
    parser = argparse.ArgumentParser(description="Gera estudo de cobertura usando Sell-in e Sell-out.")
    parser.add_argument("--sellin", help="Caminho da planilha de Sell-in. Também usado no modo Cobertura Dash.")
    parser.add_argument("--sellout", help="Caminho da planilha de Sell-out/Publicar. No modo dash, também pode ser usado como Vendas UF.")
    parser.add_argument("--vendas-uf", "--vendas_uf", dest="vendas_uf", help="Arquivo Vendas UF para o modo Cobertura Dash.")
    parser.add_argument("--vendas-sku", "--vendas_sku", dest="vendas_sku", help="Arquivo Vendas SKU opcional para o modo Cobertura Dash.")
    parser.add_argument("--metrica", choices=["volume", "quantia", "volume_variavel"], help="Métrica da cobertura: volume, quantia ou volume_variavel.")
    parser.add_argument("--nivel", choices=["CATEGORIA", "Categoria", "categoria", "NIVEL1", "NIVEL2", "ESTMER7", "nivel1", "nivel2", "estmer7"], help="Regra de categoria: CATEGORIA, NIVEL1, NIVEL2 ou ESTMER7.")
    parser.add_argument("--saida", help="Caminho do arquivo de saída.")
    parser.add_argument("--fabricante", help="Fabricante do Sell-out para filtrar quando o Sell-in não tiver SKU/EAN ou quando desejar limitar a análise.")
    parser.add_argument("--modo", choices=["estudo", "comparacao", "estudo2out", "dash"], default="estudo", help="Modo de execução: estudo, comparacao ou dash. Para comparar 2 Sell-outs no estudo, use --sellout2. O valor estudo2out permanece aceito por compatibilidade.")
    parser.add_argument("--estudo20", help="Arquivo de estudo/cobertura da versão 2.0 para comparação.")
    parser.add_argument("--estudo30", help="Arquivo de estudo/cobertura da versão 3.0 para comparação.")
    parser.add_argument("--sellout2", help="Segundo Sell-out/Publicar opcional no modo Estudo de Cobertura. Quando informado, o primeiro Sell-out vira 2.0 e este vira 3.0.")
    parser.add_argument("--sku", help="Arquivo SKU para o modo Cobertura Dash.")
    parser.add_argument("--arquivo-fabricante", "--arquivo_fabricante", "--congelado", "--arquivo-congelado", "--arquivo_congelado", dest="arquivo_fabricante", help="Arquivo Congelado para o modo Cobertura Dash ou Congelado opcional para o Estudo de Cobertura.")
    parser.add_argument("--congelado-estudo", "--congelado_estudo", dest="congelado_estudo", help="Arquivo Congelado opcional para o modo Estudo de Cobertura.")
    parser.add_argument("--gerar-top20-sku-canal-uf", "--gerar_top20_sku_canal_uf", dest="gerar_top20_sku_canal_uf", action="store_true", help="Gera a tabela TOP 20 SKU por UF/Canal na comparação com 2 Sell-outs. Por padrão, não gera.")

    # Opções para desligar abas/saídas que antes eram geradas automaticamente.
    parser.set_defaults(
        gerar_resumo_categorias=True,
        gerar_abas_categorias=True,
        gerar_base_skus=True,
        gerar_base_contribuicao_sellout=False,
        gerar_skus_por_categoria=True,
        gerar_crosschecks=True,
        gerar_parametros=True,
        gerar_descricao_calculos=True,
        gerar_avisos=True,
        gerar_abas_auxiliares_comparacao=True,
    )
    parser.add_argument("--sem-resumo-categorias", dest="gerar_resumo_categorias", action="store_false", help="Não gera a aba Resumo Categorias.")
    parser.add_argument("--sem-abas-categorias", dest="gerar_abas_categorias", action="store_false", help="Não gera as abas individuais por categoria/PROD.")
    parser.add_argument("--sem-base-skus", dest="gerar_base_skus", action="store_false", help="Não gera a aba Base SKUs.")
    parser.add_argument("--gerar-base-contribuicao-sellout", dest="gerar_base_contribuicao_sellout", action="store_true", help="Gera a aba oculta Base Contribuição Sell-out. Por padrão, não gera.")
    parser.add_argument("--sem-base-contribuicao-sellout", dest="gerar_base_contribuicao_sellout", action="store_false", help="Mantém desativada a aba oculta Base Contribuição Sell-out.")
    parser.add_argument("--sem-skus-por-categoria", dest="gerar_skus_por_categoria", action="store_false", help="Não gera a aba SKUs por Categoria.")
    parser.add_argument("--sem-crosschecks", dest="gerar_crosschecks", action="store_false", help="Não gera a aba Crosschecks.")
    parser.add_argument("--sem-parametros", dest="gerar_parametros", action="store_false", help="Não gera a aba Parâmetros.")
    parser.add_argument("--sem-descricao-calculos", dest="gerar_descricao_calculos", action="store_false", help="Não gera a aba Descrição Cálculos.")
    parser.add_argument("--sem-avisos", dest="gerar_avisos", action="store_false", help="Não gera a aba Avisos.")
    parser.add_argument("--sem-abas-auxiliares-comparacao", dest="gerar_abas_auxiliares_comparacao", action="store_false", help="Não gera as abas auxiliares na comparação 2.0 x 3.0.")
    return parser.parse_args()


def opcoes_saida_de_args(args) -> Dict[str, bool]:
    """Monta as opções de saída a partir dos argumentos de terminal ou dos padrões da GUI."""
    return normalizar_opcoes_saida({
        "resumo_categorias": getattr(args, "gerar_resumo_categorias", True),
        "abas_categorias": getattr(args, "gerar_abas_categorias", True),
        "base_skus": getattr(args, "gerar_base_skus", True),
        "base_contribuicao_sellout": getattr(args, "gerar_base_contribuicao_sellout", False),
        "skus_por_categoria": getattr(args, "gerar_skus_por_categoria", True),
        "crosschecks": getattr(args, "gerar_crosschecks", True),
        "parametros": getattr(args, "gerar_parametros", True),
        "descricao_calculos": getattr(args, "gerar_descricao_calculos", True),
        "avisos": getattr(args, "gerar_avisos", True),
        "abas_auxiliares_comparacao": getattr(args, "gerar_abas_auxiliares_comparacao", True),
        "top20_sku_canal_uf": getattr(args, "gerar_top20_sku_canal_uf", False),
    })




# ============================================================
# Ajuste de volumetria
# ============================================================

ESCALAS_VOLUMETRIA = [
    (1, "Unidade / sem ajuste"),
    (10, "Dezenas"),
    (100, "Centenas"),
    (1_000, "Milhares / kg ↔ g / toneladas ↔ kg"),
    (10_000, "Dezena de milhares"),
    (100_000, "Centena de milhares"),
    (1_000_000, "Milhões / toneladas ↔ g"),
    (10_000_000, "Dezena de milhões"),
    (100_000_000, "Centena de milhões"),
    (1_000_000_000, "Bilhões"),
    (10_000_000_000, "Dezena de bilhões"),
    (100_000_000_000, "Centena de bilhões"),
    (1_000_000_000_000, "Trilhões"),
]

LIMITE_INFERIOR_SEM_AJUSTE_VOLUMETRIA = 0.20
LIMITE_SUPERIOR_SEM_AJUSTE_VOLUMETRIA = 5.0
LIMITE_INFERIOR_APOS_AJUSTE_VOLUMETRIA = 0.04
LIMITE_SUPERIOR_APOS_AJUSTE_VOLUMETRIA = 25.0
MELHORIA_MINIMA_LOG_AJUSTE_VOLUMETRIA = 0.70


def nome_escala_volumetria(fator: float) -> str:
    try:
        fator_int = int(round(float(fator)))
    except Exception:
        return str(fator)
    for f, nome in ESCALAS_VOLUMETRIA:
        if fator_int == f:
            return nome
    return f"{fator_int:,.0f}".replace(",", ".")


def _score_fator_volumetria(valor: float, fator: float) -> float:
    if valor <= 0 or fator <= 0:
        return float("inf")
    return abs(math.log10(valor / fator))


def detectar_e_ajustar_volumetria(
    sellin: pd.DataFrame,
    sellout: pd.DataFrame,
    avisos: List[str],
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, object]]:
    """
    Identifica diferença de escala entre Sell-in e Sell-out por categoria/PROD,
    evitando aplicar um único divisor no arquivo inteiro.
    """
    info = {
        "Ajuste volumetria aplicado": "Não",
        "Coluna ajustada": "Nenhuma",
        "Divisor aplicado": 1,
        "Divisor aplicado por categoria": "",
        "Detalhe divisor por categoria": [],
        "Volumetria final comparativa": "Sem alteração",
        "Mediana Sell-out/Sell-in antes do ajuste": np.nan,
        "Mediana Sell-out/Sell-in após o ajuste": np.nan,
        "Critério": "Não havia dados suficientes para inferir escala por categoria.",
    }

    if sellin.empty or sellout.empty:
        avisos.append("Volumetria: não foi possível comparar escalas porque Sell-in ou Sell-out está vazio.")
        return sellin, sellout, info

    if "categoria_key" not in sellin.columns or "categoria_key" not in sellout.columns:
        avisos.append("Volumetria: categoria_key não encontrada em uma das bases; sem ajuste aplicado.")
        return sellin, sellout, info

    si = sellin.copy()
    so = sellout.copy()

    si_cat = (
        si.groupby("categoria_key", dropna=False, as_index=False)
        .agg(sellin_total=("valor_sellin", "sum"), categoria_si=("categoria", "first"))
    )
    so_cat = (
        so.groupby("categoria_key", dropna=False, as_index=False)
        .agg(sellout_total=("valor_sellout", "sum"), categoria_so=("categoria", "first"))
    )
    comp = si_cat.merge(so_cat, on="categoria_key", how="inner")
    comp = comp[(comp["sellin_total"] > 0) & (comp["sellout_total"] > 0)].copy()

    if comp.empty:
        avisos.append("Volumetria: não havia categorias com Sell-in e Sell-out positivos para comparar escala.")
        return si, so, info

    comp["ratio"] = comp["sellout_total"] / comp["sellin_total"].replace(0, np.nan)
    ratios = comp["ratio"].replace([np.inf, -np.inf], np.nan).dropna()
    if ratios.empty:
        avisos.append("Volumetria: a razão Sell-out/Sell-in por categoria ficou vazia; sem ajuste aplicado.")
        return si, so, info

    info["Mediana Sell-out/Sell-in antes do ajuste"] = float(ratios.median())
    candidatos = [f for f, _ in ESCALAS_VOLUMETRIA]
    registros = []
    ajustes = []

    def nome_categoria_linha(row) -> str:
        nome = row.get("categoria_si")
        if pd.isna(nome) or str(nome).strip() == "":
            nome = row.get("categoria_so")
        if pd.isna(nome) or str(nome).strip() == "":
            nome = row.get("categoria_key", "Categoria")
        return str(nome).strip()

    for _, row in comp.iterrows():
        cat_key = row["categoria_key"]
        cat_nome = nome_categoria_linha(row)
        ratio = float(row["ratio"])
        col_ajustada = "Nenhuma"
        divisor = 1.0
        ratio_pos = ratio
        status = "sem ajuste"

        if math.isfinite(ratio) and ratio > 0 and not (LIMITE_INFERIOR_SEM_AJUSTE_VOLUMETRIA <= ratio <= LIMITE_SUPERIOR_SEM_AJUSTE_VOLUMETRIA):
            if ratio > 1:
                alvo = ratio
                melhor_fator = min(candidatos, key=lambda f: _score_fator_volumetria(alvo, f))
                ratio_pre = ratio / melhor_fator
                melhoria_log = abs(math.log10(ratio)) - abs(math.log10(ratio_pre))
                if (
                    melhor_fator >= 10
                    and LIMITE_INFERIOR_APOS_AJUSTE_VOLUMETRIA <= ratio_pre <= LIMITE_SUPERIOR_APOS_AJUSTE_VOLUMETRIA
                    and melhoria_log >= MELHORIA_MINIMA_LOG_AJUSTE_VOLUMETRIA
                ):
                    mask = so["categoria_key"].astype(str) == str(cat_key)
                    if "valor_sellout_original_volumetria" not in so.columns:
                        so["valor_sellout_original_volumetria"] = so["valor_sellout"]
                    so.loc[mask, "valor_sellout"] = pd.to_numeric(so.loc[mask, "valor_sellout"], errors="coerce").fillna(0) / melhor_fator
                    col_ajustada = "Sell-out"
                    divisor = float(melhor_fator)
                    ratio_pos = ratio_pre
                    status = "ajustado"
            else:
                alvo = 1 / ratio
                melhor_fator = min(candidatos, key=lambda f: _score_fator_volumetria(alvo, f))
                ratio_pre = ratio * melhor_fator
                melhoria_log = abs(math.log10(ratio)) - abs(math.log10(ratio_pre))
                if (
                    melhor_fator >= 10
                    and LIMITE_INFERIOR_APOS_AJUSTE_VOLUMETRIA <= ratio_pre <= LIMITE_SUPERIOR_APOS_AJUSTE_VOLUMETRIA
                    and melhoria_log >= MELHORIA_MINIMA_LOG_AJUSTE_VOLUMETRIA
                ):
                    mask = si["categoria_key"].astype(str) == str(cat_key)
                    if "valor_sellin_original_volumetria" not in si.columns:
                        si["valor_sellin_original_volumetria"] = si["valor_sellin"]
                    si.loc[mask, "valor_sellin"] = pd.to_numeric(si.loc[mask, "valor_sellin"], errors="coerce").fillna(0) / melhor_fator
                    col_ajustada = "Sell-in"
                    divisor = float(melhor_fator)
                    ratio_pos = ratio_pre
                    status = "ajustado"

        registros.append({
            "Categoria": cat_nome,
            "Coluna ajustada": col_ajustada,
            "Divisor": divisor,
            "Ratio antes": ratio,
            "Ratio depois": ratio_pos,
            "Status": status,
        })
        if status == "ajustado":
            ajustes.append(registros[-1])

    # Recalcula a mediana após os ajustes aplicados por categoria.
    si_pos = si.groupby("categoria_key", dropna=False, as_index=False)["valor_sellin"].sum().rename(columns={"valor_sellin": "si"})
    so_pos = so.groupby("categoria_key", dropna=False, as_index=False)["valor_sellout"].sum().rename(columns={"valor_sellout": "so"})
    comp_pos = si_pos.merge(so_pos, on="categoria_key", how="inner")
    comp_pos = comp_pos[(comp_pos["si"] > 0) & (comp_pos["so"] > 0)].copy()
    if not comp_pos.empty:
        ratios_pos = (comp_pos["so"] / comp_pos["si"].replace(0, np.nan)).replace([np.inf, -np.inf], np.nan).dropna()
        if not ratios_pos.empty:
            info["Mediana Sell-out/Sell-in após o ajuste"] = float(ratios_pos.median())

    def fmt_registro(r: Dict[str, object]) -> str:
        divisor = float(r.get("Divisor", 1) or 1)
        if str(r.get("Coluna ajustada", "Nenhuma")) == "Nenhuma":
            return f"{r.get('Categoria', '')}: 1 (sem ajuste)"
        return f"{r.get('Categoria', '')}: {r.get('Coluna ajustada')} ÷ {divisor:g}"

    info["Divisor aplicado por categoria"] = "; ".join(fmt_registro(r) for r in registros)
    info["Detalhe divisor por categoria"] = registros
    info["Critério"] = f"Inferência por categoria/PROD. Categorias avaliadas: {len(registros)}. Ajustes aplicados: {len(ajustes)}."

    if ajustes:
        colunas = sorted(set(str(r["Coluna ajustada"]) for r in ajustes))
        divisores = sorted(set(float(r["Divisor"]) for r in ajustes))
        info["Ajuste volumetria aplicado"] = "Sim"
        info["Coluna ajustada"] = " / ".join(colunas) if len(colunas) <= 2 else "Múltiplas"
        info["Divisor aplicado"] = divisores[0] if len(divisores) == 1 else "Por categoria"
        info["Volumetria final comparativa"] = "Ajuste por categoria/PROD"
        avisos.append(
            f"Volumetria: ajuste por categoria aplicado em {len(ajustes)} de {len(registros)} categoria(s). "
            "Consulte a aba Parâmetros para ver o divisor aplicado em cada categoria."
        )
    else:
        info["Coluna ajustada"] = "Nenhuma"
        info["Divisor aplicado"] = 1
        info["Volumetria final comparativa"] = "Sem alteração por categoria"
        avisos.append(
            f"Volumetria: sem ajuste aplicado por categoria. Mediana Sell-out/Sell-in = {info['Mediana Sell-out/Sell-in antes do ajuste']:.4f}."
        )

    return si, so, info


def executar_estudo(
    sellin_path: str,
    sellout_path: str,
    saida_path: str,
    metrica: str,
    nivel: str,
    fabricante_filtro: str = "",
    congelado_path: str = "",
    log_callback=None,
    output_options: Optional[Dict[str, object]] = None,
) -> Path:
    """Executa a geração do estudo. Pode ser chamada pelo terminal ou pela interface."""
    def log(msg: str, pct: Optional[float] = None):
        if log_callback:
            try:
                log_callback(msg, pct)
            except TypeError:
                log_callback(msg)
        else:
            if pct is None:
                print(msg)
            else:
                print(f"{int(round(pct))}% - {msg}")

    output_options = normalizar_opcoes_saida(output_options)
    metrica = "volume_variavel" if metrica_eh_volume_variavel(metrica) else str(metrica or "").lower()
    if metrica not in {"volume", "quantia", "volume_variavel"}:
        raise ValueError("Métrica inválida. Use 'volume', 'quantia' ou 'volume_variavel'.")

    nivel = str(nivel or "").upper()
    if nivel not in {"CATEGORIA", "NIVEL1", "NIVEL2", "ESTMER7"}:
        raise ValueError("Regra de categoria inválida. Use 'CATEGORIA', 'NIVEL1', 'NIVEL2' ou 'ESTMER7'.")

    sellin_file = Path(sellin_path)
    sellout_file = Path(sellout_path)
    saida = Path(saida_path)

    if not sellin_file.exists():
        raise FileNotFoundError(f"Arquivo Sell-in não encontrado: {sellin_file}")
    if not sellout_file.exists():
        raise FileNotFoundError(f"Arquivo Sell-out não encontrado: {sellout_file}")
    if saida.suffix.lower() != ".xlsx":
        saida = saida.with_suffix(".xlsx")

    avisos: List[str] = []

    log("Lendo Sell-in...", 5)
    sellin_raw, av = ler_sellin(sellin_file, metrica)
    avisos.extend(av)

    mapa_congelado = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
    congelado_file = Path(str(congelado_path or "").strip()) if str(congelado_path or "").strip() else None
    usar_congelado_categoria = bool(congelado_file) and nivel == "CATEGORIA"

    if congelado_file and nivel != "CATEGORIA":
        avisos.append(
            f"Congelado opcional informado, mas ignorado porque a regra escolhida foi {nivel}. "
            "O Congelado só define categoria quando a opção escolhida é CATEGORIA; para NIVEL1/NIVEL2/ESTMER7, o estudo usa as PRODs do Sell-out."
        )

    if usar_congelado_categoria:
        log("Lendo Congelado opcional para Categoria...", 12)
        mapa_congelado, av = ler_mapa_congelado_categoria(
            congelado_file,
            "CATEGORIA",
            fabricante_filtro=fabricante_filtro,
        )
        avisos.extend(av)
        if mapa_congelado.empty:
            raise ValueError(
                "O Congelado foi selecionado e a regra escolhida foi CATEGORIA, mas nenhum SKU/EAN ficou mapeado.\n\n"
                "Neste fluxo, o código não pode cair para Total Fabricante, porque a Categoria deve vir obrigatoriamente do Congelado.\n"
                "Verifique se o Congelado possui a coluna de código 'Código Barras SKU' e a coluna de categoria 'Categoría congelada ScannMarket'. "
                "Também são aceitas variações como SKU, EAN, CODIGO_BARRAS_CONTENIDO, Categoria congelada ScannMarket, Categoria atual Data Excellence ou Est Mer 6 (Categoria)."
            )

    log("Lendo Sell-out...", 18)
    sellout, mapa_prod, av = ler_sellout(
        sellout_file,
        metrica,
        nivel,
        sellin_raw,
        fabricante_filtro=fabricante_filtro,
        progress_callback=log,
        mapa_categoria_forcado=mapa_congelado if usar_congelado_categoria else None,
    )
    avisos.extend(av)

    log("Mapeando categoria/PROD no Sell-in...", 36)
    mapa_para_sellin = mapa_congelado if usar_congelado_categoria else mapa_prod
    sellin, av = aplicar_mapeamento_prod_no_sellin(
        sellin_raw,
        mapa_para_sellin,
        fabricante_filtro=fabricante_filtro,
        forcar_mapa_sem_fallback=usar_congelado_categoria,
        origem_mapa="Congelado opcional" if usar_congelado_categoria else "Sell-out",
    )
    avisos.extend(av)
    if usar_congelado_categoria:
        avisos.append("Categoria do Estudo de Cobertura definida pelo Congelado opcional para Sell-in e Sell-out, via SKU/EAN, sem fallback para Total Fabricante.")

    info_volume_variavel = {
        "Volume variável aplicado": "Não",
        "SKUs Sell-out com peso localizado": 0,
        "Gramatura média global": "",
        "Linhas Sell-in convertidas": 0,
        "Critério volume variável": "Não aplicado",
    }

    if usar_congelado_categoria:
        total_fab_key = normalizar_categoria("Total Fabricante")
        bases_com_total = []
        for nome_base, base_check in [("Sell-in", sellin), ("Sell-out", sellout)]:
            if "categoria_key" in base_check.columns:
                tem_total = base_check["categoria_key"].astype(str).str.contains(total_fab_key, na=False).any()
                if tem_total:
                    bases_com_total.append(nome_base)
        if bases_com_total:
            raise ValueError(
                "Falha de segurança: mesmo com Congelado + CATEGORIA, apareceu categoria Total Fabricante em "
                + ", ".join(bases_com_total)
                + ".\nO processamento foi interrompido para evitar gerar estudo com categoria incorreta."
            )

    if nivel in {"NIVEL1", "NIVEL2"}:
        log("Separando análise por valor do PROD escolhido no Sell-out...", 42)
        sellin, sellout, av = reagrupar_por_categoria_original_e_nivel(
            sellin, sellout, nivel, fabricante_filtro=fabricante_filtro
        )
        avisos.extend(av)

    if metrica_eh_volume_variavel(metrica):
        log("Convertendo Sell-in por volume variável/gramatura média...", 44)
        sellin, sellout, info_volume_variavel = aplicar_volume_variavel_por_gramatura(sellin, sellout, avisos)

    log("Verificando volumetria Sell-in x Sell-out...", 45)
    sellin, sellout, info_volumetria = detectar_e_ajustar_volumetria(sellin, sellout, avisos)

    avisos.append(
        "UF comparação: Sell-in e Sell-out foram consolidados para a mesma abertura de UF/grupo: "
        "AL - SE, MA - PI, RR - AM - RO - AC, TO - PA - AP e SP único somando INT/MET/Interior/RegMet. "
        "A tabela por UF, em modo mensal, usa os últimos 12 meses móveis, alinhada à cobertura."
    )

    categorias_sellin = set(sellin["categoria_key"].dropna().unique())
    categorias_sellout = set(sellout["categoria_key"].dropna().unique())
    categorias_comuns = sorted(categorias_sellin & categorias_sellout)

    if not categorias_comuns:
        if usar_congelado_categoria:
            raise ValueError(
                "O Congelado foi informado e a regra escolhida foi CATEGORIA, mas não houve Categoria em comum após o mapeamento por SKU/EAN.\n\n"
                "Neste fluxo, o código não cria Total Fabricante, porque a Categoria deve vir obrigatoriamente do Congelado.\n"
                "Verifique se os SKUs/EANs do Sell-in e do Sell-out existem no Congelado e se a coluna de Categoria do Congelado foi reconhecida.\n\n"
                + montar_mensagem_diagnostico(
                    sellin=sellin,
                    sellout=sellout,
                    avisos=avisos,
                    metrica=metrica,
                    nivel=nivel,
                    fabricante_filtro=fabricante_filtro,
                )
            )
        if str(fabricante_filtro or "").strip():
            categoria_total = f"Total Fabricante - {fabricante_filtro.strip()}"
            aviso_total = (
                "Não havia categoria/PROD em comum. Como havia fabricante selecionado, "
                "foi criada uma visão total pelo fabricante para permitir o cálculo."
            )
        else:
            categoria_total = "Total disponível"
            aviso_total = (
                "Não havia categoria/PROD em comum. Para não impedir a geração, "
                "foi criada uma visão de Total disponível usando os dados que existem em Sell-in e Sell-out."
            )

        sellin = sellin.copy()
        sellout = sellout.copy()
        sellin["categoria"] = categoria_total
        sellout["categoria"] = categoria_total
        sellin["categoria_key"] = normalizar_categoria(categoria_total)
        sellout["categoria_key"] = normalizar_categoria(categoria_total)
        avisos.append(aviso_total)
        categorias_sellin = set(sellin["categoria_key"].dropna().unique())
        categorias_sellout = set(sellout["categoria_key"].dropna().unique())
        categorias_comuns = sorted(categorias_sellin & categorias_sellout)

    if not categorias_comuns:
        raise ValueError(
            montar_mensagem_diagnostico(
                sellin=sellin,
                sellout=sellout,
                avisos=avisos,
                metrica=metrica,
                nivel=nivel,
                fabricante_filtro=fabricante_filtro,
            )
        )

    nome_categoria = (
        pd.concat([
            sellin[["categoria_key", "categoria"]],
            sellout[["categoria_key", "categoria"]],
        ], ignore_index=True)
        .dropna()
        .drop_duplicates("categoria_key")
        .set_index("categoria_key")["categoria"]
        .to_dict()
    )

    log("Calculando Base SKUs...", 48)
    detalhe_skus, resumo_skus = preparar_skus(sellin, sellout)

    log("Calculando coberturas por categoria...", 58)
    resultados = []
    total_categorias = len(categorias_comuns)
    for i, cat_key in enumerate(categorias_comuns, start=1):
        cat_nome = nome_categoria.get(cat_key, cat_key)
        if total_categorias > 1:
            progresso_cat = 58 + (i / max(total_categorias, 1)) * 28
            log(f"Calculando categoria {i}/{total_categorias}: {cat_nome}", progresso_cat)
        r = calcular_cobertura_categoria(cat_key, cat_nome, sellin, sellout)
        if r:
            resultados.append(r)

    if not resultados:
        raise ValueError(
            "As categorias em comum foram encontradas, mas não houve período compatível para calcular.\n\n"
            + montar_mensagem_diagnostico(
                sellin=sellin,
                sellout=sellout,
                avisos=avisos,
                metrica=metrica,
                nivel=nivel,
                fabricante_filtro=fabricante_filtro,
            )
        )

    log("Montando crosschecks por categoria/PROD...", 88)
    crosschecks = montar_crosschecks_fabricante(sellout, resultados)

    if usar_congelado_categoria:
        regra_categoria = (
            "Congelado opcional: como a regra escolhida foi CATEGORIA, o SKU/EAN do Sell-in e do Sell-out é procurado no Congelado; "
            "a Categoria usada no estudo vem do Congelado, sem fallback para Total Fabricante."
        )
    elif nivel == "CATEGORIA":
        regra_categoria = "Categoria: usa a coluna Categoria do Sell-out; se ela não existir, usa a Categoria do Sell-in mapeada por SKU."
    else:
        regra_categoria = (
            f"{nivel}: o SKU do Sell-in é usado para localizar o PROD correspondente no Sell-out; "
            f"cada valor do {nivel} vira uma aba própria e o cálculo usa apenas os SKUs mapeados para aquele valor."
        )

    parametros = {
        "Arquivo Sell-in": nome_arquivo_curto(sellin_file),
        "Arquivo Sell-out": nome_arquivo_curto(sellout_file),
        "Arquivo Congelado opcional": nome_arquivo_curto(congelado_file) if congelado_file else "Não informado",
        "Uso do Congelado opcional": "Aplicado para Categoria" if usar_congelado_categoria else ("Ignorado: regra diferente de CATEGORIA" if congelado_file and nivel != "CATEGORIA" else "Não aplicado"),
        "SKUs mapeados pelo Congelado opcional": str(len(mapa_congelado)) if usar_congelado_categoria else "0",
        "Métrica cobertura": metrica,
        "Volume variável aplicado": str(info_volume_variavel.get("Volume variável aplicado", "Não")),
        "Gramatura média global volume variável": str(info_volume_variavel.get("Gramatura média global", "")),
        "SKUs com peso localizado no Dash/Vendas SKU": str(info_volume_variavel.get("SKUs Sell-out com peso localizado", 0)),
        "Linhas Sell-in convertidas por volume variável": str(info_volume_variavel.get("Linhas Sell-in convertidas", 0)),
        "Critério volume variável": str(info_volume_variavel.get("Critério volume variável", "Não aplicado")),
        "Regra categoria/PROD": nivel,
        "Fabricante selecionado": fabricante_filtro or "Automático / todos",
        "Regra de categoria": regra_categoria,
        "Regra leitura Sell-in": "Aba do template Sell-in; cabeçalho procurado primeiro de baixo para cima entre as linhas 15 e 1; se não encontrar, continua até o limite atual de leitura.",
        "Regra cobertura mensal": "12 meses móveis: SOMA(Sell-out últimos 12 meses) / SOMA(Sell-in últimos 12 meses)",
        "Regra MAT": "Prioriza anos fechados consecutivos. Se não houver, usa YTD contra mesmo período do ano anterior.",
        "Regra fabricante": "Fabricante referência = fabricante predominante dos SKUs do Sell-in encontrados no Sell-out, ponderado pelo Sell-in.",
        "Regra UF comparação": "Sell-in e Sell-out são comparados com UF padronizada: AL - SE, MA - PI, RR - AM - RO - AC, TO - PA - AP e SP consolidado incluindo INT/MET/Interior/RegMet. Em modo mensal, a tabela por UF usa os últimos 12 meses móveis, alinhada à cobertura.",
        "Ajuste volumetria aplicado": str(info_volumetria.get("Ajuste volumetria aplicado", "Não")),
        "Coluna ajustada por volumetria": str(info_volumetria.get("Coluna ajustada", "Nenhuma")),
        "Divisor aplicado na volumetria": str(info_volumetria.get("Divisor aplicado", 1)),
        "Divisor aplicado na volumetria por categoria": "Ver tabela abaixo na aba Parâmetros",
        "__detalhe_divisor_categoria__": info_volumetria.get("Detalhe divisor por categoria", []),
        "Escala/volumetria comparativa final": str(info_volumetria.get("Volumetria final comparativa", "Sem alteração")),
        "Mediana Sell-out/Sell-in antes do ajuste": str(info_volumetria.get("Mediana Sell-out/Sell-in antes do ajuste", "")),
        "Mediana Sell-out/Sell-in após ajuste": str(info_volumetria.get("Mediana Sell-out/Sell-in após o ajuste", "")),
        "Critério ajuste volumetria": str(info_volumetria.get("Critério", "")),
        "Data de geração": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    if output_options.get("base_contribuicao_sellout", False):
        log("Preparando base auxiliar de contribuição por SKU...", 89)
        base_contribuicao_sellout = preparar_base_contribuicao_sellout(sellout)
    else:
        log("Pulando Base Contribuição Sell-out porque a opção está desativada.", 89)
        base_contribuicao_sellout = pd.DataFrame()

    log("Gerando Excel...", 90)
    gerar_excel(
        saida, resultados, detalhe_skus, resumo_skus, parametros, avisos,
        crosschecks=crosschecks,
        base_contribuicao_sellout=base_contribuicao_sellout,
        output_options=output_options,
    )

    log(f"Arquivo gerado com sucesso: {saida}", 100)
    return saida


def executar_cobertura_dash(
    sellin_path: str,
    vendas_uf_path: str,
    sku_path: str,
    congelado_path: str,
    saida_path: str,
    metrica: str,
    nivel: str,
    fabricante_filtro: str = "",
    vendas_sku_path: str | Path | None = None,
    log_callback=None,
    output_options: Optional[Dict[str, object]] = None,
) -> Path:
    """
    Modo Cobertura Dash com Sell-in.

    Usa:
    - Sell-in do template para preencher o lado de referência da cobertura;
    - Vendas UF do Dash com Dia de Selector Data, UF e Vendas Medida;
    - Vendas SKU opcional com Dia de Selector Data, EAN + Nome Sku e Vendas Medida;
    - arquivo SKU com Categoria Scanntech/NIVEL1/NIVEL2;
    - arquivo Congelado opcional com Marca/Fabricante e categoria congelada.

    Quando Vendas UF não traz SKU/categoria, Vendas SKU é usado para distribuir
    Vendas UF por categoria/PROD. O resultado distribuído vira o Sell-out do Dash.
    O Sell-in externo permanece como referência do cálculo de cobertura.
    """
    def log(msg: str, pct: Optional[float] = None):
        if log_callback:
            try:
                log_callback(msg, pct)
            except TypeError:
                log_callback(msg)
        else:
            if pct is None:
                print(msg)
            else:
                print(f"{int(round(pct))}% - {msg}")

    output_options = normalizar_opcoes_saida(output_options)
    metrica = "volume_variavel" if metrica_eh_volume_variavel(metrica) else str(metrica or "").lower()
    if metrica not in {"volume", "quantia", "volume_variavel"}:
        raise ValueError("Métrica inválida. Use 'volume', 'quantia' ou 'volume_variavel'.")

    nivel = str(nivel or "").upper()
    if nivel not in {"CATEGORIA", "NIVEL1", "NIVEL2", "ESTMER7"}:
        raise ValueError("Regra de categoria inválida. Use 'CATEGORIA', 'NIVEL1', 'NIVEL2' ou 'ESTMER7'.")

    sellin_file = Path(sellin_path)
    vendas_uf_file = Path(vendas_uf_path)
    vendas_sku_file = Path(vendas_sku_path) if vendas_sku_path else None
    sku_file = Path(sku_path)
    congelado_file = Path(congelado_path) if congelado_path else None
    saida = Path(saida_path)

    for rotulo, caminho in [
        ("Sell-in", sellin_file),
        ("Vendas UF", vendas_uf_file),
        ("SKU", sku_file),
    ]:
        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo {rotulo} não encontrado: {caminho}")
    if congelado_file is not None and not congelado_file.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {congelado_file}")
    if vendas_sku_file is not None and not vendas_sku_file.exists():
        raise FileNotFoundError(f"Arquivo Vendas SKU não encontrado: {vendas_sku_file}")
    if saida.suffix.lower() != ".xlsx":
        saida = saida.with_suffix(".xlsx")

    avisos: List[str] = []

    log("Lendo Sell-in...", 6)
    sellin_raw, av = ler_sellin(sellin_file, metrica)
    avisos.extend(av)

    log("Lendo arquivos auxiliares do Dash...", 14)
    aux_dash = preparar_auxiliares_dash(
        sku_file,
        congelado_file,
        avisos,
        fabricante_filtro=fabricante_filtro,
    )
    mapa_aux_dash = montar_mapa_categoria_aux_dash(aux_dash, nivel, avisos)

    sellout_sku_base = None
    if vendas_sku_file is not None:
        log("Lendo Vendas SKU...", 24)
        sellout_sku_base, mapa_prod_sku, av = ler_sellout_dash(
            vendas_sku_file,
            metrica,
            nivel,
            aux_dash,
            sellin_para_categoria=sellin_raw,
            fabricante_filtro=fabricante_filtro,
        )
        avisos.extend(av)

        log("Lendo Vendas UF...", 34)
        vendas_uf, mapa_prod_uf, av = ler_sellout_dash(
            vendas_uf_file,
            metrica,
            nivel,
            aux_dash,
            sellin_para_categoria=sellin_raw,
            fabricante_filtro=fabricante_filtro,
        )
        avisos.extend(av)

        log("Distribuindo Vendas UF por participação de Vendas SKU...", 40)
        sellout_distribuido = distribuir_vendas_uf_por_vendas_sku(vendas_uf, sellout_sku_base, avisos)
        if sellout_distribuido is None or sellout_distribuido.empty:
            avisos.append(
                "Cobertura Dash: Vendas UF não pôde ser distribuído por Vendas SKU. "
                "Usando Vendas UF como Sell-out Total Dash para não bloquear a geração."
            )
            sellout_distribuido = vendas_uf.copy()
        sellout = sellout_distribuido.copy()
        mapa_prod = mapa_prod_sku if mapa_prod_sku is not None and not mapa_prod_sku.empty else mapa_prod_uf
    else:
        log("Lendo Vendas UF...", 30)
        vendas_uf, mapa_prod, av = ler_sellout_dash(
            vendas_uf_file,
            metrica,
            nivel,
            aux_dash,
            sellin_para_categoria=sellin_raw,
            fabricante_filtro=fabricante_filtro,
        )
        avisos.extend(av)
        sellout = vendas_uf.copy()

    log("Mapeando categoria/PROD no Sell-in...", 42)
    # No modo Dash, o Sell-in deve ser classificado pelo mesmo de/para de SKU/Congelado
    # usado no Dash. O mapa vindo de Vendas SKU pode conter só os EANs vendidos no Dash;
    # por isso ele é complementado com o mapa auxiliar completo do arquivo SKU/Congelado.
    mapas_para_sellin = []
    if mapa_prod is not None and not mapa_prod.empty:
        mapas_para_sellin.append(mapa_prod[["ean", "categoria_map_prod", "categoria_key_map_prod"]].copy())
    if mapa_aux_dash is not None and not mapa_aux_dash.empty:
        mapas_para_sellin.append(mapa_aux_dash[["ean", "categoria_map_prod", "categoria_key_map_prod"]].copy())
    if mapas_para_sellin:
        mapa_prod_sellin = pd.concat(mapas_para_sellin, ignore_index=True)
        mapa_prod_sellin["ean"] = mapa_prod_sellin["ean"].map(ean_texto)
        mapa_prod_sellin = mapa_prod_sellin[mapa_prod_sellin["ean"] != ""].drop_duplicates("ean", keep="first")
    else:
        mapa_prod_sellin = pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"])
    sellin, av = aplicar_mapeamento_prod_no_sellin(sellin_raw, mapa_prod_sellin, fabricante_filtro=fabricante_filtro)
    avisos.extend(av)

    info_volume_variavel = {
        "Volume variável aplicado": "Não",
        "SKUs Sell-out com peso localizado": 0,
        "Gramatura média global": "",
        "Linhas Sell-in convertidas": 0,
        "Critério volume variável": "Não aplicado",
    }

    if nivel in {"NIVEL1", "NIVEL2", "ESTMER7"}:
        log("Separando análise por valor do PROD escolhido no Dash...", 44)
        sellin, sellout, av = reagrupar_por_categoria_original_e_nivel(
            sellin, sellout, nivel, fabricante_filtro=fabricante_filtro
        )
        avisos.extend(av)

    if metrica_eh_volume_variavel(metrica):
        log("Convertendo Sell-in por volume variável/gramatura média...", 44)
        sellin, sellout, info_volume_variavel = aplicar_volume_variavel_por_gramatura(sellin, sellout, avisos)

    log("Verificando volumetria Sell-in x Dash...", 48)
    sellin, sellout, info_volumetria = detectar_e_ajustar_volumetria(sellin, sellout, avisos)

    if (sellout is None or sellout.empty) and sellin is not None and not sellin.empty:
        avisos.append(
            "Cobertura Dash: a base Dash/Sell-out ficou vazia após os cruzamentos. "
            "Para não bloquear a geração, foi criada uma base comparada a partir do Sell-in disponível."
        )
        sellout = sellin.rename(columns={"valor_sellin": "valor_sellout"}).copy()
        if "valor_sellout" not in sellout.columns:
            sellout["valor_sellout"] = pd.to_numeric(sellin.get("valor_sellin", 0), errors="coerce").fillna(0)

    avisos.append(
        "Modo Cobertura Dash com Sell-in: o arquivo Sell-in preenche o lado Sell-in da cobertura; "
        "Vendas UF/Vendas SKU preenchem o lado Dash/Sell-out."
    )
    avisos.append(
        "UF comparação: as bases foram consolidadas para a mesma abertura de UF/grupo: "
        "AL - SE, MA - PI, RR - AM - RO - AC, TO - PA - AP e SP único somando INT/MET/Interior/RegMet. "
        "A tabela por UF, em modo mensal, usa os últimos 12 meses móveis, alinhada à cobertura."
    )

    categorias_sellin = set(sellin["categoria_key"].dropna().unique()) if not sellin.empty else set()
    categorias_sellout = set(sellout["categoria_key"].dropna().unique()) if not sellout.empty else set()
    categorias_comuns = sorted(categorias_sellin & categorias_sellout)

    if not categorias_comuns:
        if str(fabricante_filtro or "").strip():
            categoria_total = f"Total Fabricante - {fabricante_filtro.strip()}"
            aviso_total = (
                "Não havia categoria/PROD em comum no modo Dash. Como havia fabricante selecionado, "
                "foi criada uma visão total pelo fabricante para permitir o cálculo."
            )
        else:
            categoria_total = "Total Dash"
            aviso_total = (
                "Não havia categoria/PROD em comum no modo Dash. Para não impedir a geração, "
                "foi criada uma visão Total Dash usando os dados disponíveis."
            )
        sellin = sellin.copy()
        sellout = sellout.copy()
        sellin["categoria"] = categoria_total
        sellout["categoria"] = categoria_total
        sellin["categoria_key"] = normalizar_categoria(categoria_total)
        sellout["categoria_key"] = normalizar_categoria(categoria_total)
        avisos.append(aviso_total)
        categorias_comuns = sorted(set(sellin["categoria_key"].dropna().unique()) & set(sellout["categoria_key"].dropna().unique()))

    if not categorias_comuns:
        raise ValueError(
            montar_mensagem_diagnostico(
                sellin=sellin,
                sellout=sellout,
                avisos=avisos,
                metrica=metrica,
                nivel=nivel,
                fabricante_filtro=fabricante_filtro,
            )
        )

    nome_categoria = (
        pd.concat([
            sellin[["categoria_key", "categoria"]],
            sellout[["categoria_key", "categoria"]],
        ], ignore_index=True)
        .dropna()
        .drop_duplicates("categoria_key")
        .set_index("categoria_key")["categoria"]
        .to_dict()
    )

    log("Calculando Base SKUs...", 54)
    sellout_para_skus = sellout_sku_base if sellout_sku_base is not None and not sellout_sku_base.empty else sellout
    detalhe_skus, resumo_skus = preparar_skus(sellin, sellout_para_skus)

    log("Calculando coberturas por categoria...", 62)
    resultados = []
    total_categorias = len(categorias_comuns)
    for i, cat_key in enumerate(categorias_comuns, start=1):
        cat_nome = nome_categoria.get(cat_key, cat_key)
        if total_categorias > 1:
            progresso_cat = 62 + (i / max(total_categorias, 1)) * 24
            log(f"Calculando categoria {i}/{total_categorias}: {cat_nome}", progresso_cat)
        r = calcular_cobertura_categoria(cat_key, cat_nome, sellin, sellout)
        if r:
            resultados.append(r)

    if not resultados:
        raise ValueError(
            "As categorias em comum foram encontradas, mas não houve período compatível para calcular.\n\n"
            + montar_mensagem_diagnostico(
                sellin=sellin,
                sellout=sellout,
                avisos=avisos,
                metrica=metrica,
                nivel=nivel,
                fabricante_filtro=fabricante_filtro,
            )
        )

    log("Montando crosschecks por categoria/PROD...", 88)
    crosschecks = montar_crosschecks_fabricante(sellout, resultados)

    if nivel == "CATEGORIA":
        regra_categoria = "Categoria: usa Categoria Scanntech do arquivo SKU por EAN; se faltar, usa categoria do Congelado e depois categoria do próprio Dash."
    elif nivel == "ESTMER7":
        regra_categoria = "ESTMER7: usa o SKU/EAN do Vendas SKU/Dash para localizar Est Mer 7 (Subcategoria) no Congelado; cada valor vira uma aba própria."
    else:
        regra_categoria = (
            f"{nivel}: usa o SKU/EAN do Vendas SKU/Dash para localizar o {nivel} no arquivo SKU; "
            f"cada valor do {nivel} vira uma aba própria."
        )

    parametros = {
        "Modo": "Cobertura Dash",
        "Arquivo Sell-in": nome_arquivo_curto(sellin_file),
        "Arquivo Vendas UF": nome_arquivo_curto(vendas_uf_file),
        "Arquivo Vendas SKU": nome_arquivo_curto(vendas_sku_file) if vendas_sku_file is not None else "Não informado",
        "Arquivo SKU": nome_arquivo_curto(sku_file),
        "Arquivo Congelado": nome_arquivo_curto(congelado_file),
        "Métrica cobertura": metrica,
        "Volume variável aplicado": str(info_volume_variavel.get("Volume variável aplicado", "Não")),
        "Gramatura média global volume variável": str(info_volume_variavel.get("Gramatura média global", "")),
        "SKUs com peso localizado no Dash/Vendas SKU": str(info_volume_variavel.get("SKUs Sell-out com peso localizado", 0)),
        "Linhas Sell-in convertidas por volume variável": str(info_volume_variavel.get("Linhas Sell-in convertidas", 0)),
        "Critério volume variável": str(info_volume_variavel.get("Critério volume variável", "Não aplicado")),
        "Regra categoria/PROD": nivel,
        "Fabricante selecionado": fabricante_filtro or "Automático / todos",
        "Regra de categoria": regra_categoria,
        "Regra leitura Sell-in": "Aba do template Sell-in; cabeçalho procurado primeiro de baixo para cima entre as linhas 15 e 1; se não encontrar, continua até o limite atual de leitura.",
        "Regra leitura Vendas UF": "Cabeçalho na primeira linha ou cabeçalho flexível; espera Dia de Selector Data, UF e Vendas Medida.",
        "Regra leitura Vendas SKU": "Opcional; cabeçalho na primeira linha; espera Dia de Selector Data, EAN + Nome Sku e Vendas Medida.",
        "Regra leitura SKU/Congelado": "Cabeçalho na primeira linha.",
        "Regra cobertura mensal": "12 meses móveis: SOMA(Dash/Sell-out últimos 12 meses) / SOMA(Sell-in últimos 12 meses)",
        "Regra MAT": "Prioriza anos fechados consecutivos. Se não houver, usa YTD contra mesmo período do ano anterior.",
        "Regra Congelado": "Marca/Fabricante são preenchidos pelo arquivo Congelado usando Código Barras SKU/CODIGO_BARRAS_CONTENIDO. A categoria do Congelado é usada como double check/fallback da Categoria Scanntech do SKU. Est Mer 7 pode ser usado como abertura adicional.",
        "Regra UF comparação": "Sell-in e Dash/Sell-out são comparados com UF padronizada: AL - SE, MA - PI, RR - AM - RO - AC, TO - PA - AP e SP consolidado incluindo INT/MET/Interior/RegMet. Em modo mensal, a tabela por UF usa os últimos 12 meses móveis, alinhada à cobertura.",
        "Ajuste volumetria aplicado": str(info_volumetria.get("Ajuste volumetria aplicado", "Não")),
        "Coluna ajustada por volumetria": str(info_volumetria.get("Coluna ajustada", "Nenhuma")),
        "Divisor aplicado na volumetria": str(info_volumetria.get("Divisor aplicado", 1)),
        "Divisor aplicado na volumetria por categoria": "Ver tabela abaixo na aba Parâmetros",
        "__detalhe_divisor_categoria__": info_volumetria.get("Detalhe divisor por categoria", []),
        "Escala/volumetria comparativa final": str(info_volumetria.get("Volumetria final comparativa", "Sem alteração")),
        "Mediana comparativa antes do ajuste": str(info_volumetria.get("Mediana Sell-out/Sell-in antes do ajuste", "")),
        "Mediana comparativa após ajuste": str(info_volumetria.get("Mediana Sell-out/Sell-in após o ajuste", "")),
        "Critério ajuste volumetria": str(info_volumetria.get("Critério", "")),
        "Data de geração": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    if output_options.get("base_contribuicao_sellout", False):
        log("Preparando base auxiliar de contribuição por SKU...", 89)
        base_contribuicao_sellout = preparar_base_contribuicao_sellout(sellout)
    else:
        log("Pulando Base Contribuição Sell-out porque a opção está desativada.", 89)
        base_contribuicao_sellout = pd.DataFrame()

    log("Gerando Excel...", 90)
    gerar_excel(
        saida, resultados, detalhe_skus, resumo_skus, parametros, avisos,
        crosschecks=crosschecks,
        base_contribuicao_sellout=base_contribuicao_sellout,
        output_options=output_options,
    )

    log(f"Arquivo gerado com sucesso: {saida}", 100)
    return saida




# ============================================================
# Comparação entre dois estudos de cobertura já gerados
# ============================================================

ABAS_SISTEMA_COMPARACAO = {
    normalizar_texto(x)
    for x in [
        "Resumo Categorias",
        "Base SKUs",
        "SKUs por Categoria",
        "Gráficos Cobertura",
        "Graficos Cobertura",
        "Base Contribuição Sell-out",
        "Base Contribuicao Sell-out",
        "Crosschecks",
        "Parâmetros",
        "Parametros",
        "Avisos",
        "Descrição Cálculos",
        "Descricao Calculos",
    ]
}


def _eh_aba_sistema(nome: str) -> bool:
    return normalizar_texto(nome) in ABAS_SISTEMA_COMPARACAO


def _valor_numerico_planilha(valor) -> float:
    try:
        if pd.isna(valor):
            return np.nan
    except Exception:
        pass
    if isinstance(valor, str) and valor.strip() == "":
        return np.nan
    num = numero_brasil(valor)
    return np.nan if pd.isna(num) else float(num)


def _chave_periodo_comparacao(valor):
    mes = converter_mes(valor)
    if pd.notna(mes):
        return ("MES", pd.Timestamp(mes).strftime("%Y-%m"))
    texto = str(valor).strip()
    if texto == "" or normalizar_texto(texto) in {"nan", "none"}:
        return None
    return ("TXT", normalizar_texto(texto))


def _rotulo_periodo_comparacao(valor, chave) -> str:
    if chave and chave[0] == "MES":
        try:
            return label_mes_pt(pd.Timestamp(chave[1] + "-01"))
        except Exception:
            pass
    return str(valor).strip() if str(valor).strip() else str(chave[1] if chave else "")


def _linha_anterior_tem_titulo_sku_comum(raw: pd.DataFrame, header_row: int, janela: int = 4) -> bool:
    """Indica se a linha de cabeçalho pertence a uma tabela separada de SKU em comum.

    A comparação precisa ler a tabela mensal principal. Em alguns layouts existe
    uma segunda tabela com cabeçalho parecido, como Data / Sell-in / Sell-out /
    Cobertura, mas precedida por um título contendo "SKU em Comum". Essa tabela
    não pode ser usada como base da comparação principal.
    """
    inicio = max(0, header_row - janela)
    for rr in range(inicio, header_row):
        try:
            texto = " ".join(normalizar_texto(x) for x in raw.iloc[rr].tolist())
        except Exception:
            continue
        if "sku em comum" in texto or "skus em comum" in texto:
            return True
    return False


def _localizar_tabela_mensal_em_aba(raw: pd.DataFrame) -> Optional[Tuple[int, Dict[str, int]]]:
    """Localiza a tabela mensal principal das abas geradas pelo estudo.

    Correção importante:
    no layout atual, a mesma linha pode ter DUAS tabelas lado a lado:

        Data | Sell-in | Sell-out | Cobertura | ... | UF | Sell-in 12M | Sell-out 12M | ...

    A rotina antiga varria a linha inteira e acabava sobrescrevendo as colunas
    de Sell-in/Sell-out mensal com as colunas da tabela de UF. Por isso, na
    comparação, os valores mensais vinham iguais aos valores por UF.

    Agora, depois de achar a coluna Data/Mês, a função só aceita as colunas
    imediatamente à direita dela e antes do próximo cabeçalho de outra tabela
    (UF, Categoria, SKU etc.).
    """
    max_linhas = min(len(raw), 220)
    candidatos = []

    data_headers = {"data", "mes", "mês"}
    sellin_headers = {"sell in", "sellin"}
    sellout_headers = {"sell out", "sellout"}
    cobertura_headers = {"cobertura", "cobertura 12m movel", "cobertura 12 meses moveis"}
    sellin_comum_headers = {"sell in sku em comum", "sellin sku em comum", "sell in skus em comum", "sellin skus em comum"}
    sellout_comum_headers = {"sell out sku em comum", "sellout sku em comum", "sell out skus em comum", "sellout skus em comum"}
    cobertura_comum_headers = {"cobertura sku em comum", "cobertura skus em comum"}

    def proximo_inicio_outra_tabela(valores: List[str], idx_data: int) -> int:
        """Identifica onde começa a próxima tabela lateral, normalmente em UF."""
        marcadores_outra_tabela = {
            "uf", "estado", "categoria", "categoria prod", "prod", "sku", "ean",
            "marca", "fabricante", "status sku",
        }
        for c in range(idx_data + 1, len(valores)):
            v = valores[c]
            if v in marcadores_outra_tabela:
                return c
            # Proteção extra: se aparecer Sell-in 12M/Sell-out 12M após alguns espaços,
            # normalmente já é a tabela por UF e não a mensal.
            if c > idx_data + 4 and v in {"sell in 12m", "sellin 12m", "sell out 12m", "sellout 12m"}:
                return c
        return len(valores)

    for r in range(max_linhas):
        valores = [normalizar_texto(x) for x in raw.iloc[r].tolist()]

        # Pode haver mais de uma ocorrência de Data/Mês em layouts futuros.
        idx_datas = [c for c, v in enumerate(valores) if v in data_headers]
        for idx_data in idx_datas:
            limite = proximo_inicio_outra_tabela(valores, idx_data)
            janela = list(range(idx_data + 1, min(limite, idx_data + 10)))

            idx_si = idx_so = idx_cov = None
            idx_si_comum = idx_so_comum = idx_cov_comum = None

            for c in janela:
                v = valores[c]
                # Match estrito: a tabela mensal principal não deve usar Sell-in 12M da UF.
                if v in sellin_comum_headers:
                    idx_si_comum = c
                elif v in sellout_comum_headers:
                    idx_so_comum = c
                elif v in cobertura_comum_headers:
                    idx_cov_comum = c
                elif v in sellin_headers and idx_si is None:
                    idx_si = c
                elif v in sellout_headers and idx_so is None:
                    idx_so = c
                elif v in cobertura_headers and idx_cov is None:
                    idx_cov = c

            if idx_si is not None and idx_so is not None:
                contexto_sku = _linha_anterior_tem_titulo_sku_comum(raw, r)
                candidatos.append((contexto_sku, r, {
                    "data": idx_data,
                    "sellin": idx_si,
                    "sellout": idx_so,
                    "cobertura": idx_cov if idx_cov is not None else -1,
                    "sellin_comum": idx_si_comum if idx_si_comum is not None else -1,
                    "sellout_comum": idx_so_comum if idx_so_comum is not None else -1,
                    "cobertura_comum": idx_cov_comum if idx_cov_comum is not None else -1,
                }))

    if not candidatos:
        return None

    # Prioriza a tabela principal, ou seja, a primeira que não esteja sob um
    # título de SKU em comum. Se só existir a tabela SKU em comum, usa fallback.
    for contexto_sku, r, cols in candidatos:
        if not contexto_sku:
            return r, cols
    _, r, cols = candidatos[0]
    return r, cols

def _extrair_mensal_estudo(caminho: str | Path, aba: str) -> pd.DataFrame:
    raw = pd.read_excel(caminho, sheet_name=aba, header=None, engine="openpyxl")
    loc = _localizar_tabela_mensal_em_aba(raw)
    base_cols = ["chave", "data", "Sell-in", "Sell-out", "Cobertura", "Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"]
    if not loc:
        return pd.DataFrame(columns=base_cols)

    header_row, cols = loc
    linhas = []
    blanks_consecutivos = 0
    for r in range(header_row + 1, len(raw)):
        data_val = raw.iat[r, cols["data"]] if cols["data"] < raw.shape[1] else None
        si_val = raw.iat[r, cols["sellin"]] if cols["sellin"] < raw.shape[1] else None
        so_val = raw.iat[r, cols["sellout"]] if cols["sellout"] < raw.shape[1] else None
        cov_val = raw.iat[r, cols["cobertura"]] if cols.get("cobertura", -1) >= 0 and cols["cobertura"] < raw.shape[1] else None
        si_comum_val = raw.iat[r, cols["sellin_comum"]] if cols.get("sellin_comum", -1) >= 0 and cols["sellin_comum"] < raw.shape[1] else None
        so_comum_val = raw.iat[r, cols["sellout_comum"]] if cols.get("sellout_comum", -1) >= 0 and cols["sellout_comum"] < raw.shape[1] else None
        cov_comum_val = raw.iat[r, cols["cobertura_comum"]] if cols.get("cobertura_comum", -1) >= 0 and cols["cobertura_comum"] < raw.shape[1] else None

        chave = _chave_periodo_comparacao(data_val)
        si = _valor_numerico_planilha(si_val)
        so = _valor_numerico_planilha(so_val)
        cov = _valor_numerico_planilha(cov_val)
        si_comum = _valor_numerico_planilha(si_comum_val)
        so_comum = _valor_numerico_planilha(so_comum_val)
        cov_comum = _valor_numerico_planilha(cov_comum_val)

        linha_vazia = chave is None and pd.isna(si) and pd.isna(so)
        if linha_vazia:
            blanks_consecutivos += 1
            if blanks_consecutivos >= 2 and linhas:
                break
            continue
        blanks_consecutivos = 0

        if chave is None:
            continue
        linhas.append({
            "chave": chave,
            "data": _rotulo_periodo_comparacao(data_val, chave),
            "Sell-in": si,
            "Sell-out": so,
            "Cobertura": cov,
            "Sell-in SKU em Comum": si_comum,
            "Sell-out SKU em Comum": so_comum,
            "Cobertura SKU em Comum": cov_comum,
        })

    df = pd.DataFrame(linhas)
    if df.empty:
        return pd.DataFrame(columns=base_cols)
    df = df.drop_duplicates("chave", keep="first")

    # Layout novo: os dados de SKU em comum ficam em tabela separada.
    # Quando a tabela mensal principal não trouxer essas colunas, tenta localizar e mesclar pelo mês.
    try:
        comum_df = _extrair_sku_comum_separado_estudo(raw)
        if not comum_df.empty:
            df = df.drop(columns=["Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"], errors="ignore")
            df = df.merge(comum_df, on="chave", how="left")
    except Exception:
        pass

    for col in ["Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"]:
        if col not in df.columns:
            df[col] = np.nan
    return df



def _extrair_sku_comum_separado_estudo(raw: pd.DataFrame) -> pd.DataFrame:
    """Extrai a tabela separada de SKU em comum do layout novo do estudo individual."""
    linhas = []
    max_linhas = min(len(raw), 220)
    header_row = None
    cols = {}
    for r in range(max_linhas):
        valores = [normalizar_texto(x) for x in raw.iloc[r].tolist()]
        idx_data = idx_si = idx_so = idx_cov = None
        tem_sku_comum = False
        for c, v in enumerate(valores):
            if "sku em comum" in v:
                tem_sku_comum = True
            if v in {"data", "mes", "mês"}:
                idx_data = c
            elif v in {"sell in sku em comum", "sellin sku em comum", "sell in", "sellin"}:
                idx_si = c
            elif v in {"sell out sku em comum", "sellout sku em comum", "sell out", "sellout"}:
                idx_so = c
            elif v in {"cobertura sku em comum", "cobertura"}:
                idx_cov = c
        # Precisa estar abaixo de um título contendo SKU em comum ou ter headers explícitos.
        titulo_proximo = False
        for rr in range(max(0, r - 2), r + 1):
            titulo_vals = [normalizar_texto(x) for x in raw.iloc[rr].tolist()]
            if any("sku em comum" in x for x in titulo_vals):
                titulo_proximo = True
                break
        if idx_data is not None and idx_si is not None and idx_so is not None and (tem_sku_comum or titulo_proximo):
            header_row = r
            cols = {"data": idx_data, "sellin": idx_si, "sellout": idx_so, "cobertura": idx_cov if idx_cov is not None else -1}
            break
    if header_row is None:
        return pd.DataFrame(columns=["chave", "Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"])

    blanks = 0
    for r in range(header_row + 1, len(raw)):
        data_val = raw.iat[r, cols["data"]] if cols["data"] < raw.shape[1] else None
        si_val = raw.iat[r, cols["sellin"]] if cols["sellin"] < raw.shape[1] else None
        so_val = raw.iat[r, cols["sellout"]] if cols["sellout"] < raw.shape[1] else None
        cov_val = raw.iat[r, cols["cobertura"]] if cols.get("cobertura", -1) >= 0 and cols["cobertura"] < raw.shape[1] else None
        chave = _chave_periodo_comparacao(data_val)
        si = _valor_numerico_planilha(si_val)
        so = _valor_numerico_planilha(so_val)
        if chave is None and pd.isna(si) and pd.isna(so):
            blanks += 1
            if blanks >= 2 and linhas:
                break
            continue
        blanks = 0
        if chave is None:
            continue
        linhas.append({
            "chave": chave,
            "Sell-in SKU em Comum": si,
            "Sell-out SKU em Comum": so,
            "Cobertura SKU em Comum": _valor_numerico_planilha(cov_val),
        })
    df = pd.DataFrame(linhas)
    if df.empty:
        return pd.DataFrame(columns=["chave", "Sell-in SKU em Comum", "Sell-out SKU em Comum", "Cobertura SKU em Comum"])
    return df.drop_duplicates("chave", keep="first")

def _ler_estudo_gerado_para_comparacao(caminho: str | Path) -> Dict[str, pd.DataFrame]:
    caminho = Path(caminho)
    wb = load_workbook(caminho, read_only=True, data_only=True)
    abas = list(wb.sheetnames)
    wb.close()

    dados: Dict[str, pd.DataFrame] = {}
    for aba in abas:
        if _eh_aba_sistema(aba):
            continue
        try:
            df = _extrair_mensal_estudo(caminho, aba)
        except Exception:
            continue
        if not df.empty:
            dados[aba] = df
    return dados


def _montar_comparacao_mensal(df20: pd.DataFrame, df30: pd.DataFrame) -> pd.DataFrame:
    a = df20.rename(columns={
        "data": "Data 2.0",
        "Sell-in": "Sell-in 2.0",
        "Sell-out": "Sell-out 2.0",
        "Cobertura": "Cobertura 2.0",
    })
    b = df30.rename(columns={
        "data": "Data 3.0",
        "Sell-in": "Sell-in 3.0",
        "Sell-out": "Sell-out 3.0",
        "Cobertura": "Cobertura 3.0",
    })
    comp = a.merge(b, on="chave", how="outer")

    def sort_key(chave):
        if isinstance(chave, tuple) and len(chave) == 2 and chave[0] == "MES":
            return (0, chave[1])
        return (1, str(chave))

    comp["_ordem"] = comp["chave"].map(sort_key)
    comp = comp.sort_values("_ordem").drop(columns=["_ordem"])
    comp["Data"] = comp["Data 3.0"].where(comp["Data 3.0"].notna(), comp["Data 2.0"])
    comp["Sell-in"] = comp["Sell-in 3.0"].where(comp["Sell-in 3.0"].notna(), comp["Sell-in 2.0"])
    comp["Diferença Sell-out"] = comp["Sell-out 3.0"].fillna(0) - comp["Sell-out 2.0"].fillna(0)
    comp["Diferença % Sell-out"] = comp["Diferença Sell-out"] / comp["Sell-out 2.0"].replace(0, np.nan)
    comp["Diferença Cobertura"] = comp["Cobertura 3.0"] - comp["Cobertura 2.0"]
    return comp[[
        "Data", "Sell-in", "Sell-out 2.0", "Sell-out 3.0", "Diferença Sell-out",
        "Diferença % Sell-out", "Cobertura 2.0", "Cobertura 3.0", "Diferença Cobertura",
    ]]


def _escrever_aba_comparacao_categoria(writer, nome_aba: str, categoria: str, mensal: pd.DataFrame):
    workbook = writer.book
    ws = workbook.add_worksheet(nome_aba)
    writer.sheets[nome_aba] = ws

    azul_escuro = "#1F4E78"
    azul_claro = "#D9EAF7"
    azul_linha = "#156082"
    laranja_linha = "#E97132"
    verde_linha = "#70AD47"
    cinza = "#595959"

    fmt_titulo = workbook.add_format({"bold": True, "font_size": 15, "font_color": "white", "bg_color": azul_escuro, "align": "center", "valign": "vcenter", "border": 1})
    fmt_secao = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
    fmt_header = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
    fmt_text = workbook.add_format({"border": 1})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
    fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
    fmt_link = workbook.add_format({"font_color": "blue", "underline": 1})

    ws.set_column("A:A", 2.7)
    ws.set_column("B:B", 12)
    ws.set_column("C:F", 14)
    ws.set_column("G:J", 14)
    ws.set_column("L:U", 12)

    ws.merge_range("B1:J1", f"Comparação de Estudo de Cobertura - {categoria}", fmt_titulo)
    ws.write_url("B2", "internal:'Resumo Categorias'!A1", fmt_link, string="Voltar ao Resumo Categorias")
    ws.write("B4", "Categoria/PROD", fmt_secao)
    ws.merge_range("C4:J4", categoria, fmt_text)

    start = 6
    headers = list(mensal.columns)
    for j, h in enumerate(headers, start=1):
        ws.write(start, j, h, fmt_header)

    for i, row in mensal.reset_index(drop=True).iterrows():
        r = start + 1 + i
        for j, h in enumerate(headers, start=1):
            val = row[h]
            if h == "Data":
                ws.write(r, j, "" if pd.isna(val) else str(val), fmt_text)
            elif "Cobertura" in h or "%" in h:
                if pd.isna(val):
                    ws.write_blank(r, j, None, fmt_pct)
                else:
                    ws.write_number(r, j, float(val), fmt_pct)
            else:
                if pd.isna(val):
                    ws.write_blank(r, j, None, fmt_num)
                else:
                    ws.write_number(r, j, float(val), fmt_num)

    if len(mensal) > 0:
        first = start + 1
        last = start + len(mensal)
        chart_row = start + len(mensal) + 3

        chart = workbook.add_chart({"type": "line"})
        chart.add_series({"name": "Sell-in", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 2, last, 2], "line": {"color": azul_linha, "width": 2.25}})
        chart.add_series({"name": "Sell-out 2.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 3, last, 3], "line": {"color": laranja_linha, "width": 2.25}})
        chart.add_series({"name": "Sell-out 3.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 4, last, 4], "line": {"color": verde_linha, "width": 2.25}})
        chart.set_title({"name": "Sell-in x Sell-out 2.0 x Sell-out 3.0 por mês", "name_font": {"color": "#000000", "size": 14, "bold": True}})
        chart.set_x_axis({"name": "Mês", "num_font": {"rotation": -90}})
        chart.set_y_axis({"name": "Volume", "num_format": "#,##0"})
        chart.set_legend({"position": "bottom"})
        chart.set_size({"width": 720, "height": 340})
        chart.set_plotarea({"border": {"none": True}})
        chart.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_row, 1, chart)

        chart_cov = workbook.add_chart({"type": "line"})
        chart_cov.add_series({"name": "Cobertura 2.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 7, last, 7], "line": {"color": laranja_linha, "width": 2.25}})
        chart_cov.add_series({"name": "Cobertura 3.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 8, last, 8], "line": {"color": verde_linha, "width": 2.25}})
        chart_cov.set_title({"name": "Cobertura 2.0 x Cobertura 3.0", "name_font": {"color": cinza, "size": 13, "bold": True}})
        chart_cov.set_x_axis({"name": "Mês", "num_font": {"rotation": -90}})
        chart_cov.set_y_axis({"name": "Cobertura", "num_format": "0%"})
        chart_cov.set_legend({"position": "bottom"})
        chart_cov.set_size({"width": 720, "height": 300})
        chart_cov.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_row + 18, 1, chart_cov)


def gerar_comparacao_estudos(estudo20: str | Path, estudo30: str | Path, saida: str | Path, log_callback=None, gerar_top20_sku_canal_uf: bool = False, output_options: Optional[Dict[str, object]] = None) -> Path:
    def log(msg, pct=None):
        if log_callback:
            log_callback(msg, pct)

    output_options = normalizar_opcoes_saida(output_options)
    if gerar_top20_sku_canal_uf:
        output_options["top20_sku_canal_uf"] = True
    gerar_top20_sku_canal_uf = bool(output_options.get("top20_sku_canal_uf", False))
    estudo20 = Path(estudo20)
    estudo30 = Path(estudo30)
    saida = Path(saida)

    log("Lendo estudo 2.0...", 10)
    dados20 = _ler_estudo_gerado_para_comparacao(estudo20)
    log("Lendo estudo 3.0...", 25)
    dados30 = _ler_estudo_gerado_para_comparacao(estudo30)

    if not dados20:
        raise ValueError(f"Não encontrei abas de cobertura mensal no estudo 2.0: {estudo20}")
    if not dados30:
        raise ValueError(f"Não encontrei abas de cobertura mensal no estudo 3.0: {estudo30}")

    todas = sorted(set(dados20) | set(dados30), key=normalizar_texto)
    usados = {"Resumo Categorias", "Gráficos Cobertura"}
    mapas_abas = {cat: nome_aba_seguro(cat, usados) for cat in todas}

    resumos = []
    comparacoes = {}
    total = len(todas)
    for i, cat in enumerate(todas, start=1):
        if total:
            log(f"Comparando categoria/PROD {i}/{total}: {cat}", 30 + (i / total) * 45)
        df20 = dados20.get(cat, pd.DataFrame(columns=["chave", "data", "Sell-in", "Sell-out", "Cobertura"]))
        df30 = dados30.get(cat, pd.DataFrame(columns=["chave", "data", "Sell-in", "Sell-out", "Cobertura"]))
        mensal = _montar_comparacao_mensal(df20, df30)
        comparacoes[cat] = mensal

        so20 = mensal["Sell-out 2.0"].sum(skipna=True)
        so30 = mensal["Sell-out 3.0"].sum(skipna=True)
        si = mensal["Sell-in"].sum(skipna=True)
        cov20 = divisao_segura(so20, si)
        cov30 = divisao_segura(so30, si)
        meses_comuns = int(((mensal["Sell-out 2.0"].fillna(0) != 0) & (mensal["Sell-out 3.0"].fillna(0) != 0)).sum())
        status = "Em ambos" if cat in dados20 and cat in dados30 else ("Só 2.0" if cat in dados20 else "Só 3.0")
        resumos.append({
            "Abrir aba": "Abrir",
            "Categoria/PROD": cat,
            "Status Comparação": status,
            "Meses em comum": meses_comuns,
            "Sell-in": si,
            "Sell-out 2.0": so20,
            "Sell-out 3.0": so30,
            "Diferença Sell-out": so30 - so20,
            "Diferença % Sell-out": divisao_segura(so30 - so20, so20),
            "Cobertura 2.0": cov20,
            "Cobertura 3.0": cov30,
            "Diferença Cobertura": cov30 - cov20,
            "sheet": mapas_abas[cat],
        })

    resumo = pd.DataFrame(resumos)
    if not resumo.empty:
        resumo = resumo.sort_values(["Status Comparação", "Categoria/PROD"], ascending=[True, True])

    log("Gerando Excel da comparação...", 82)
    with pd.ExcelWriter(saida, engine="xlsxwriter", datetime_format="dd/mm/yyyy") as writer:
        workbook = writer.book
        azul_escuro = "#1F4E78"
        azul_claro = "#D9EAF7"
        fmt_titulo = workbook.add_format({"bold": True, "font_size": 15, "font_color": "white", "bg_color": azul_escuro, "align": "center", "valign": "vcenter", "border": 1})
        fmt_header = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
        fmt_text = workbook.add_format({"border": 1})
        fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
        fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
        fmt_link = workbook.add_format({"font_color": "blue", "underline": 1, "border": 1})

        ws = workbook.add_worksheet("Resumo Categorias")
        writer.sheets["Resumo Categorias"] = ws
        ws.merge_range("A1:L1", "Comparação de Estudo de Cobertura", fmt_titulo)
        ws.write("A2", "Estudo 2.0", fmt_header)
        ws.write("B2", nome_arquivo_curto(estudo20), fmt_text)
        ws.write("A3", "Estudo 3.0", fmt_header)
        ws.write("B3", nome_arquivo_curto(estudo30), fmt_text)

        resumo = remover_colunas_duplicadas(resumo)
        export_cols = [c for c in resumo.columns if c != "sheet"]
        start = 5
        for j, col in enumerate(export_cols):
            ws.write(start, j, col, fmt_header)
        for i, row in resumo.reset_index(drop=True).iterrows():
            r = start + 1 + i
            for j, col in enumerate(export_cols):
                val = row[col]
                if col == "Abrir aba":
                    ws.write_url(r, j, f"internal:'{row['sheet']}'!A1", fmt_link, string="Abrir")
                elif col in {"Diferença % Sell-out", "Cobertura 2.0", "Cobertura 3.0", "Diferença Cobertura"}:
                    if pd.isna(val):
                        ws.write_blank(r, j, None, fmt_pct)
                    else:
                        ws.write_number(r, j, float(val), fmt_pct)
                elif isinstance(val, (int, float, np.integer, np.floating)) and not pd.isna(val):
                    ws.write_number(r, j, float(val), fmt_num)
                else:
                    ws.write(r, j, "" if pd.isna(val) else str(val), fmt_text)
        ws.set_column("A:A", 12)
        ws.set_column("B:C", 26)
        ws.set_column("D:D", 13)
        ws.set_column("E:L", 15)

        for cat, mensal in comparacoes.items():
            _escrever_aba_comparacao_categoria(writer, mapas_abas[cat], cat, mensal)

        avisos = pd.DataFrame({"Avisos": [
            "Comparação gerada a partir de dois arquivos de estudo de cobertura já gerados pelo código.",
            "A tabela mensal compara Sell-out 2.0 x Sell-out 3.0, Cobertura 2.0 x Cobertura 3.0 e diferença mensal.",
            "A comparação preserva os valores/cálculos dos estudos originais; o Sell-in usa 2.0 como referência visual e só usa 3.0 quando 2.0 estiver vazio.",
        ]})
        avisos.to_excel(writer, sheet_name="Avisos", index=False)
        aplicar_formatos_basicos(writer, "Avisos", avisos)

    log(f"Comparação gerada com sucesso: {saida}", 100)
    return saida

# ============================================================
# Comparação avançada entre estudos / geração direta com 2 Sell-outs
# ============================================================

def _comparacao_chave_para_mes(chave):
    if isinstance(chave, tuple) and len(chave) == 2 and chave[0] == "MES":
        try:
            return pd.Timestamp(str(chave[1]) + "-01")
        except Exception:
            return pd.NaT
    return pd.NaT


def _localizar_tabela_uf_em_aba(raw: pd.DataFrame) -> Optional[Tuple[int, Dict[str, int]]]:
    max_linhas = min(len(raw), 160)
    for r in range(max_linhas):
        valores = [normalizar_texto(x) for x in raw.iloc[r].tolist()]
        idx = {"uf": None, "sellin": None, "sellout": None, "cobertura": None, "imp_si": None, "imp_so": None}
        for c, v in enumerate(valores):
            if v == "uf":
                idx["uf"] = c
            elif v.startswith("sell in") or v in {"sellin", "sell in 12m", "sell in mat"}:
                idx["sellin"] = c
            elif v.startswith("sell out") or v in {"sellout", "sell out 12m", "sell out mat"}:
                idx["sellout"] = c
            elif v == "cobertura" or v.startswith("cobertura"):
                idx["cobertura"] = c
            elif "importancia" in v and "sell in" in v:
                idx["imp_si"] = c
            elif "importancia" in v and "sell out" in v:
                idx["imp_so"] = c
        if idx["uf"] is not None and idx["sellin"] is not None and idx["sellout"] is not None:
            return r, idx
    return None


def _extrair_uf_estudo(caminho: str | Path, aba: str) -> pd.DataFrame:
    raw = pd.read_excel(caminho, sheet_name=aba, header=None, engine="openpyxl")
    loc = _localizar_tabela_uf_em_aba(raw)
    if not loc:
        return pd.DataFrame(columns=["UF", "Sell-in", "Sell-out", "Cobertura", "Importância Sell-in", "Importância Sell-out"])
    header_row, cols = loc
    linhas = []
    blanks = 0
    for r in range(header_row + 1, len(raw)):
        uf_val = raw.iat[r, cols["uf"]] if cols["uf"] is not None and cols["uf"] < raw.shape[1] else None
        uf_txt = "" if pd.isna(uf_val) else str(uf_val).strip()
        si_val = raw.iat[r, cols["sellin"]] if cols["sellin"] is not None and cols["sellin"] < raw.shape[1] else None
        so_val = raw.iat[r, cols["sellout"]] if cols["sellout"] is not None and cols["sellout"] < raw.shape[1] else None
        if not uf_txt and pd.isna(_valor_numerico_planilha(si_val)) and pd.isna(_valor_numerico_planilha(so_val)):
            blanks += 1
            if blanks >= 2 and linhas:
                break
            continue
        blanks = 0
        if not uf_txt:
            continue
        cov_val = raw.iat[r, cols["cobertura"]] if cols.get("cobertura") is not None and cols["cobertura"] < raw.shape[1] else None
        imp_si_val = raw.iat[r, cols["imp_si"]] if cols.get("imp_si") is not None and cols["imp_si"] < raw.shape[1] else None
        imp_so_val = raw.iat[r, cols["imp_so"]] if cols.get("imp_so") is not None and cols["imp_so"] < raw.shape[1] else None
        linhas.append({
            "UF": uf_txt,
            "Sell-in": _valor_numerico_planilha(si_val),
            "Sell-out": _valor_numerico_planilha(so_val),
            "Cobertura": _valor_numerico_planilha(cov_val),
            "Importância Sell-in": _valor_numerico_planilha(imp_si_val),
            "Importância Sell-out": _valor_numerico_planilha(imp_so_val),
        })
    return pd.DataFrame(linhas)


def _extrair_fabricante_estudo(caminho: str | Path, aba: str) -> str:
    try:
        raw = pd.read_excel(caminho, sheet_name=aba, header=None, engine="openpyxl", nrows=4)
        # G2 no layout atual.
        if raw.shape[0] >= 2 and raw.shape[1] >= 7:
            val = raw.iat[1, 6]
            if not pd.isna(val) and str(val).strip():
                return str(val).strip()
    except Exception:
        pass
    return ""


def _localizar_tabela_skus_excluidos_estudo(raw: pd.DataFrame) -> Optional[Tuple[int, Dict[str, int]]]:
    """Localiza a tabela "SKUs excluídos do cálculo de SKU em Comum" em uma aba de estudo."""
    if raw is None or raw.empty:
        return None

    max_linhas = min(len(raw), 1200)
    title_row = None
    for r in range(max_linhas):
        valores = [normalizar_texto(x) for x in raw.iloc[r].tolist()]
        texto = " ".join(v for v in valores if v)
        if ("skus excluidos" in texto or "sku excluidos" in texto or "sku excluido" in texto) and "sku em comum" in texto:
            title_row = r
            break

    if title_row is None:
        return None

    for r in range(title_row + 1, min(title_row + 8, len(raw))):
        valores = [normalizar_texto(x) for x in raw.iloc[r].tolist()]
        idx = {"sku": None, "status": None, "sellin": None, "sellout": None, "nome_sku": None, "marca": None, "fabricante": None}
        for c, v in enumerate(valores):
            if v in {"sku", "ean"}:
                idx["sku"] = c
            elif v == "status" or v.startswith("status"):
                idx["status"] = c
            elif v in {"sell in", "sellin"}:
                idx["sellin"] = c
            elif v in {"sell out", "sellout"}:
                idx["sellout"] = c
            elif v in {"nome sku", "nome do sku", "descricao sku", "descricao do sku", "producto", "produto"}:
                idx["nome_sku"] = c
            elif v == "marca":
                idx["marca"] = c
            elif v == "fabricante":
                idx["fabricante"] = c

        if idx["sku"] is not None and idx["status"] is not None:
            return r, idx

    return None


def _extrair_skus_excluidos_estudo(caminho: str | Path, aba: str) -> pd.DataFrame:
    """Extrai a lista de SKUs excluídos do bloco de SKU em comum de uma aba de estudo já gerada."""
    colunas = ["SKU", "Status", "Sell-in", "Sell-out", "Nome SKU", "Marca", "Fabricante"]
    try:
        raw = pd.read_excel(caminho, sheet_name=aba, header=None, engine="openpyxl")
    except Exception:
        return pd.DataFrame(columns=colunas)

    loc = _localizar_tabela_skus_excluidos_estudo(raw)
    if not loc:
        return pd.DataFrame(columns=colunas)

    header_row, cols = loc
    linhas = []
    blanks = 0

    def celula(row_idx: int, chave: str):
        col_idx = cols.get(chave)
        if col_idx is None or col_idx < 0 or col_idx >= raw.shape[1]:
            return None
        return raw.iat[row_idx, col_idx]

    for r in range(header_row + 1, len(raw)):
        sku_val = celula(r, "sku")
        status_val = celula(r, "status")
        sku_txt = "" if pd.isna(sku_val) else str(sku_val).strip()
        status_txt = "" if pd.isna(status_val) else str(status_val).strip()
        linha_norm = normalizar_texto(" ".join("" if pd.isna(x) else str(x) for x in raw.iloc[r].tolist()))

        # Linha de mensagem quando não houve exclusão.
        if "nenhum sku excluido" in linha_norm or "todos os skus relevantes" in linha_norm:
            continue

        if not sku_txt and not status_txt:
            blanks += 1
            if blanks >= 2 and linhas:
                break
            continue
        blanks = 0

        if not sku_txt:
            continue

        status_norm = normalizar_texto(status_txt)
        if status_norm not in {"somente no sell in", "somente no sell out"}:
            # Evita capturar início de outro bloco ou textos soltos.
            if linhas:
                break
            continue

        linhas.append({
            "SKU": sku_txt,
            "Status": status_txt,
            "Sell-in": _valor_numerico_planilha(celula(r, "sellin")),
            "Sell-out": _valor_numerico_planilha(celula(r, "sellout")),
            "Nome SKU": "" if pd.isna(celula(r, "nome_sku")) else str(celula(r, "nome_sku")).strip(),
            "Marca": "" if pd.isna(celula(r, "marca")) else str(celula(r, "marca")).strip(),
            "Fabricante": "" if pd.isna(celula(r, "fabricante")) else str(celula(r, "fabricante")).strip(),
        })

    if not linhas:
        return pd.DataFrame(columns=colunas)
    out = pd.DataFrame(linhas)
    for c in colunas:
        if c not in out.columns:
            out[c] = np.nan if c in {"Sell-in", "Sell-out"} else ""
    return out[colunas].reset_index(drop=True)


def _ler_estudo_gerado_para_comparacao_completo(caminho: str | Path) -> Dict[str, Dict[str, object]]:
    caminho = Path(caminho)
    wb = load_workbook(caminho, read_only=True, data_only=True)
    abas = list(wb.sheetnames)
    wb.close()
    dados: Dict[str, Dict[str, object]] = {}
    for aba in abas:
        if _eh_aba_sistema(aba):
            continue
        try:
            mensal = _extrair_mensal_estudo(caminho, aba)
            uf = _extrair_uf_estudo(caminho, aba)
            fabricante = _extrair_fabricante_estudo(caminho, aba)
            skus_excluidos = _extrair_skus_excluidos_estudo(caminho, aba)
        except Exception:
            continue
        if not mensal.empty or not uf.empty or not skus_excluidos.empty:
            dados[aba] = {"mensal": mensal, "uf": uf, "fabricante": fabricante, "skus_excluidos": skus_excluidos}
    return dados


def _montar_comparacao_mensal_avancada(df20: pd.DataFrame, df30: pd.DataFrame) -> pd.DataFrame:
    a = df20.rename(columns={
        "data": "Data 2.0",
        "Sell-in": "Sell-in 2.0",
        "Sell-out": "Sell-out 2.0",
        "Cobertura": "Cobertura 2.0",
        "Sell-in SKU em Comum": "Sell-in SKU em Comum 2.0",
        "Sell-out SKU em Comum": "Sell-out SKU em Comum 2.0",
        "Cobertura SKU em Comum": "Cobertura SKU em Comum 2.0",
    })
    b = df30.rename(columns={
        "data": "Data 3.0",
        "Sell-in": "Sell-in 3.0",
        "Sell-out": "Sell-out 3.0",
        "Cobertura": "Cobertura 3.0",
        "Sell-in SKU em Comum": "Sell-in SKU em Comum 3.0",
        "Sell-out SKU em Comum": "Sell-out SKU em Comum 3.0",
        "Cobertura SKU em Comum": "Cobertura SKU em Comum 3.0",
    })
    comp = a.merge(b, on="chave", how="outer")
    def sort_key(chave):
        if isinstance(chave, tuple) and len(chave) == 2 and chave[0] == "MES":
            return (0, chave[1])
        return (1, str(chave))
    comp["_ordem"] = comp["chave"].map(sort_key)
    comp = comp.sort_values("_ordem").drop(columns=["_ordem"])
    comp["Data"] = comp.get("Data 2.0", np.nan).where(comp.get("Data 2.0", np.nan).notna(), comp.get("Data 3.0", np.nan))
    # A comparação não pode alterar o Sell-in calculado nos estudos originais.
    # Por padrão, usa o Sell-in do estudo 2.0 como referência visual; só usa o 3.0 se o 2.0 estiver vazio.
    comp["Sell-in"] = comp.get("Sell-in 2.0", np.nan).where(comp.get("Sell-in 2.0", np.nan).notna(), comp.get("Sell-in 3.0", np.nan))
    for col in [
        "Sell-out 2.0", "Sell-out 3.0", "Cobertura 2.0", "Cobertura 3.0",
        "Sell-in SKU em Comum 2.0", "Sell-out SKU em Comum 2.0", "Cobertura SKU em Comum 2.0",
        "Sell-in SKU em Comum 3.0", "Sell-out SKU em Comum 3.0", "Cobertura SKU em Comum 3.0",
    ]:
        if col not in comp.columns:
            comp[col] = np.nan
    comp["Diferença Sell-out"] = comp["Sell-out 3.0"].fillna(0) - comp["Sell-out 2.0"].fillna(0)
    comp["Diferença % Sell-out"] = comp["Diferença Sell-out"] / comp["Sell-out 2.0"].replace(0, np.nan)
    comp["Diferença Cobertura"] = comp["Cobertura 3.0"] - comp["Cobertura 2.0"]
    comp["mes_ts"] = comp["chave"].map(_comparacao_chave_para_mes)
    return comp[[
        "chave", "mes_ts", "Data", "Sell-in", "Sell-out 2.0", "Sell-out 3.0",
        "Cobertura 2.0", "Cobertura 3.0",
        "Sell-in SKU em Comum 2.0", "Sell-out SKU em Comum 2.0", "Cobertura SKU em Comum 2.0",
        "Sell-in SKU em Comum 3.0", "Sell-out SKU em Comum 3.0", "Cobertura SKU em Comum 3.0",
        "Diferença Sell-out", "Diferença % Sell-out", "Diferença Cobertura",
    ]]


def _normalizar_tabela_uf_para_comparacao(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garante que a tabela de UF tenha as colunas esperadas antes do merge da comparação.
    Isso evita erro 'UF' quando uma categoria existe em apenas um estudo ou quando a tabela
    por UF não foi localizada em uma das abas.
    """
    cols_base = ["UF", "Sell-in", "Sell-out", "Cobertura", "Importância Sell-in", "Importância Sell-out"]

    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(columns=cols_base)

    out = df.copy()

    # Renomeia variações comuns para o padrão usado na comparação.
    renomear = {}
    for col in out.columns:
        n = normalizar_texto(col)
        if n == "uf":
            renomear[col] = "UF"
        elif n in {"sell in", "sellin", "sell in 12m", "sell in mat"}:
            renomear[col] = "Sell-in"
        elif n in {"sell out", "sellout", "sell out 12m", "sell out mat"}:
            renomear[col] = "Sell-out"
        elif n == "cobertura" or n.startswith("cobertura"):
            renomear[col] = "Cobertura"
        elif "importancia" in n and "sell in" in n:
            renomear[col] = "Importância Sell-in"
        elif "importancia" in n and "sell out" in n:
            renomear[col] = "Importância Sell-out"

    if renomear:
        out = out.rename(columns=renomear)

    # Sem UF não há como fazer a comparação por UF; devolve tabela vazia padronizada.
    if "UF" not in out.columns:
        return pd.DataFrame(columns=cols_base)

    out["UF"] = out["UF"].fillna("SEM UF").astype(str).str.strip().replace("", "SEM UF")

    for col in cols_base:
        if col not in out.columns:
            out[col] = np.nan

    return out[cols_base]


def _montar_comparacao_uf_avancada(uf20: pd.DataFrame, uf30: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "UF", "Sell-in 12M", "Sell-out 12M 2.0", "Sell-out 12M 3.0",
        "Cobertura 2.0", "Cobertura 3.0", "Importância Sell-in",
        "Importância Sell-out 2.0", "Importância Sell-out 3.0"
    ]

    uf20 = _normalizar_tabela_uf_para_comparacao(uf20)
    uf30 = _normalizar_tabela_uf_para_comparacao(uf30)

    if uf20.empty and uf30.empty:
        return pd.DataFrame(columns=cols)

    a = uf20.rename(columns={
        "Sell-in": "Sell-in 2.0",
        "Sell-out": "Sell-out 2.0",
        "Cobertura": "Cobertura 2.0",
        "Importância Sell-in": "Importância Sell-in 2.0",
        "Importância Sell-out": "Importância Sell-out 2.0",
    })
    b = uf30.rename(columns={
        "Sell-in": "Sell-in 3.0",
        "Sell-out": "Sell-out 3.0",
        "Cobertura": "Cobertura 3.0",
        "Importância Sell-in": "Importância Sell-in 3.0",
        "Importância Sell-out": "Importância Sell-out 3.0",
    })

    comp = a.merge(b, on="UF", how="outer")

    sellin_30 = comp["Sell-in 3.0"] if "Sell-in 3.0" in comp.columns else pd.Series(np.nan, index=comp.index)
    sellin_20 = comp["Sell-in 2.0"] if "Sell-in 2.0" in comp.columns else pd.Series(np.nan, index=comp.index)
    imp_si_30 = comp["Importância Sell-in 3.0"] if "Importância Sell-in 3.0" in comp.columns else pd.Series(np.nan, index=comp.index)
    imp_si_20 = comp["Importância Sell-in 2.0"] if "Importância Sell-in 2.0" in comp.columns else pd.Series(np.nan, index=comp.index)

    # A comparação deve preservar os valores dos estudos. Para o Sell-in por UF,
    # usa o 2.0 como referência e só cai para 3.0 se o 2.0 estiver vazio.
    comp["Sell-in 12M"] = sellin_20.where(sellin_20.notna(), sellin_30)
    comp["Importância Sell-in"] = imp_si_20.where(imp_si_20.notna(), imp_si_30)
    comp["Sell-out 12M 2.0"] = comp["Sell-out 2.0"] if "Sell-out 2.0" in comp.columns else np.nan
    comp["Sell-out 12M 3.0"] = comp["Sell-out 3.0"] if "Sell-out 3.0" in comp.columns else np.nan
    comp["Cobertura 2.0"] = comp["Cobertura 2.0"] if "Cobertura 2.0" in comp.columns else np.nan
    comp["Cobertura 3.0"] = comp["Cobertura 3.0"] if "Cobertura 3.0" in comp.columns else np.nan
    comp["Importância Sell-out 2.0"] = comp["Importância Sell-out 2.0"] if "Importância Sell-out 2.0" in comp.columns else np.nan
    comp["Importância Sell-out 3.0"] = comp["Importância Sell-out 3.0"] if "Importância Sell-out 3.0" in comp.columns else np.nan

    for c in cols:
        if c not in comp.columns:
            comp[c] = np.nan

    # Garante ordem padrão e total no final quando existirem.
    comp["_ord"] = comp["UF"].map(lambda x: 99_999 if normalizar_texto(x) == "total" else UF_COMPARACAO_RANK.get(str(x), 10_000))
    comp = comp.sort_values(["_ord", "UF"]).drop(columns="_ord")
    return comp[cols]


def _periodos_resumo_comparacao(mensal: pd.DataFrame) -> Dict[str, object]:
    meses = [m for m in mensal.get("mes_ts", pd.Series(dtype="datetime64[ns]")).dropna().tolist()]
    if not meses:
        return {"tipo": "Total disponível", "label_anterior": "Anterior", "label_atual": "Atual", "inicio_anterior": pd.NaT, "fim_anterior": pd.NaT, "inicio_atual": pd.NaT, "fim_atual": pd.NaT}
    return determinar_periodos_mat_ou_ytd(meses, max(meses))


def _somar_periodo_comp(mensal: pd.DataFrame, coluna: str, inicio, fim) -> float:
    if coluna not in mensal.columns or pd.isna(inicio) or pd.isna(fim) or "mes_ts" not in mensal.columns:
        return np.nan
    mask = (mensal["mes_ts"] >= inicio) & (mensal["mes_ts"] <= fim)
    return float(pd.to_numeric(mensal.loc[mask, coluna], errors="coerce").fillna(0).sum())


def _resumos_topo_comparacao(mensal: pd.DataFrame):
    periodos = _periodos_resumo_comparacao(mensal)
    la = periodos["label_anterior"]
    lat = periodos["label_atual"]
    ia, fa = periodos["inicio_anterior"], periodos["fim_anterior"]
    it, ft = periodos["inicio_atual"], periodos["fim_atual"]
    meses = [m for m in mensal.get("mes_ts", pd.Series(dtype="datetime64[ns]")).dropna().tolist()]
    max_mes = max(meses) if meses else pd.NaT
    tendencia_disponivel = tem_janela_movel_completa(meses, max_mes, MESES_MAT * 2)

    linhas = []
    for indicador, coluna in [("Sellin", "Sell-in"), ("Sellout 2.0", "Sell-out 2.0"), ("Sellout 3.0", "Sell-out 3.0")]:
        ant = _somar_periodo_comp(mensal, coluna, ia, fa)
        atu = _somar_periodo_comp(mensal, coluna, it, ft)
        linhas.append({"Indicador": indicador, la: ant, lat: atu, "Tendência": variacao(atu, ant) if tendencia_disponivel else np.nan})
    mat_df = pd.DataFrame(linhas)

    if meses:
        inicio_6_atual = (max_mes - pd.DateOffset(months=MESES_MOVEL - 1)).to_period("M").to_timestamp()
        fim_6_anterior = (inicio_6_atual - pd.DateOffset(months=1)).to_period("M").to_timestamp()
        inicio_6_anterior = (fim_6_anterior - pd.DateOffset(months=MESES_MOVEL - 1)).to_period("M").to_timestamp()
        label_6_ant = f"{formatar_periodo(inicio_6_anterior)} a {formatar_periodo(fim_6_anterior)}"
        label_6_atual = f"{formatar_periodo(inicio_6_atual)} a {formatar_periodo(max_mes)}"
        linhas6 = []
        seis_disponivel = tem_janela_movel_completa(meses, max_mes, MESES_MOVEL * 2)
        for indicador, coluna in [("Sellin", "Sell-in"), ("Sellout 2.0", "Sell-out 2.0"), ("Sellout 3.0", "Sell-out 3.0")]:
            ant = _somar_periodo_comp(mensal, coluna, inicio_6_anterior, fim_6_anterior)
            atu = _somar_periodo_comp(mensal, coluna, inicio_6_atual, max_mes)
            linhas6.append({
                "Indicador": indicador,
                label_6_ant: ant if seis_disponivel else np.nan,
                label_6_atual: atu if seis_disponivel else np.nan,
                "Tendência": variacao(atu, ant) if seis_disponivel else np.nan,
            })
        seis_df = pd.DataFrame(linhas6)
    else:
        label_6_ant, label_6_atual = "6M Ant.", "6M Atual"
        seis_df = pd.DataFrame({"Indicador": ["Sellin", "Sellout 2.0", "Sellout 3.0"], label_6_ant: [np.nan]*3, label_6_atual: [np.nan]*3, "Tendência": [np.nan]*3})
    return periodos, mat_df, seis_df


def _resumo_mat_movel_comparacao(mensal: pd.DataFrame) -> Tuple[Dict[str, object], pd.DataFrame]:
    meses = [m for m in mensal.get("mes_ts", pd.Series(dtype="datetime64[ns]")).dropna().tolist()]
    max_mes = max(meses) if meses else pd.NaT
    periodos_mat = calcular_periodos_mat_movel(meses, max_mes)
    la = periodos_mat["label_anterior"]
    lat = periodos_mat["label_atual"]
    if not tem_janela_movel_completa(meses, max_mes, MESES_MAT * 2):
        periodos_mat["tipo"] = "MAT móvel indisponível - menos de 24 meses"
        linhas = [{"Indicador": indicador, la: np.nan, lat: np.nan, "Tendência": np.nan} for indicador in ["Sellin", "Sellout 2.0", "Sellout 3.0"]]
        return periodos_mat, pd.DataFrame(linhas)
    ia, fa = periodos_mat["inicio_anterior"], periodos_mat["fim_anterior"]
    it, ft = periodos_mat["inicio_atual"], periodos_mat["fim_atual"]
    linhas = []
    for indicador, coluna in [("Sellin", "Sell-in"), ("Sellout 2.0", "Sell-out 2.0"), ("Sellout 3.0", "Sell-out 3.0")]:
        ant = _somar_periodo_comp(mensal, coluna, ia, fa)
        atu = _somar_periodo_comp(mensal, coluna, it, ft)
        linhas.append({"Indicador": indicador, la: ant, lat: atu, "Tendência": variacao(atu, ant)})
    return periodos_mat, pd.DataFrame(linhas)


def _escrever_aba_comparacao_categoria(writer, nome_aba: str, categoria: str, mensal: pd.DataFrame, uf: Optional[pd.DataFrame] = None, fabricante: str = "", skus_excluidos20: Optional[pd.DataFrame] = None, skus_excluidos30: Optional[pd.DataFrame] = None, contribuicao_skus: Optional[pd.DataFrame] = None, info_contribuicao: str = "", gerar_top20_contribuicao: bool = False):
    workbook = writer.book
    ws = workbook.add_worksheet(nome_aba)
    writer.sheets[nome_aba] = ws

    azul_escuro = "#1F4E78"
    azul_claro = "#D9EAF7"
    azul_linha = "#156082"
    laranja_linha = "#E97132"
    verde_linha = "#70AD47"
    fmt_titulo = workbook.add_format({"bold": True, "font_size": 16, "font_color": "white", "bg_color": azul_escuro, "align": "center", "valign": "vcenter", "border": 1})
    fmt_top_label = workbook.add_format({"bold": True, "font_color": "white", "bg_color": azul_escuro, "align": "center", "valign": "vcenter", "border": 1})
    fmt_top_value = workbook.add_format({"bold": True, "font_color": azul_escuro, "bg_color": "#EAF3F8", "align": "center", "valign": "vcenter", "border": 1})
    fmt_secao = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
    fmt_header = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
    fmt_text = workbook.add_format({"border": 1, "align": "left", "valign": "vcenter"})
    fmt_text_center = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
    fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
    fmt_date = workbook.add_format({"num_format": "mmm/yy", "border": 1, "align": "center"})
    fmt_blank = workbook.add_format({"border": 1})
    fmt_link = workbook.add_format({"font_color": "#0563C1", "underline": 1, "border": 1, "align": "center"})

    ws.set_column("A:A", 2.7)
    ws.set_column("B:G", 12)
    ws.set_column("H:P", 12)
    ws.set_column("L:P", 16)
    ws.set_default_row(18)

    periodos, fy_df, seis_df = _resumos_topo_comparacao(mensal)
    periodos_mat_movel, mat_df = _resumo_mat_movel_comparacao(mensal)
    label_ant = periodos.get("label_anterior", "FY/YTD-1")
    label_atual = periodos.get("label_atual", "FY/YTD")
    label_mat_ant = periodos_mat_movel.get("label_anterior", "MAT-1")
    label_mat_atual = periodos_mat_movel.get("label_atual", "MAT")
    periodo_atual_txt = periodo_texto(periodos.get("inicio_atual"), periodos.get("fim_atual"))

    ws.write_url("L1", "internal:'Resumo Categorias'!A1", fmt_link, string="Voltar ao Resumo")
    ws.merge_range("B1:D2", "Estudo de Cobertura", fmt_titulo)
    ws.write("E1", "Categoria", fmt_top_label)
    ws.write("E2", categoria, fmt_top_value)
    ws.write("G1", "Fabricante", fmt_top_label)
    ws.write("G2", fabricante or "Não informado", fmt_top_value)

    # Tabela superior: tendência FY/YTD. Mantida separada da tendência MAT.
    ws.write("B6", "Tipo de Cálculo", fmt_secao)
    ws.merge_range("C6:F6", f"{periodos.get('tipo', '')} - {periodo_atual_txt}".strip(" -"), fmt_secao)
    ws.write("B7", "12 meses", fmt_header)
    ws.write_blank("C7", None, fmt_blank)
    ws.write("D7", f"Volume {label_ant}", fmt_header)
    ws.write("E7", f"Volume {label_atual}", fmt_header)
    ws.write("F7", "Tendência", fmt_header)
    for i, row in fy_df.reset_index(drop=True).iterrows():
        r = 7 + i
        ws.write(r, 2, row.get("Indicador", ""), fmt_text_center)
        write_number_ou_branco(ws, r, 3, row.get(label_ant, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 4, row.get(label_atual, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 5, row.get("Tendência", np.nan), fmt_pct)

    cols_6m = [c for c in seis_df.columns if c not in {"Indicador", "Tendência"}]
    label_6_ant = cols_6m[0] if len(cols_6m) else "6M Ant."
    label_6_atual = cols_6m[1] if len(cols_6m) > 1 else "6M Atual"
    ws.merge_range("H6:K6", f"6 meses móveis - {label_6_atual}", fmt_secao)
    ws.write_blank("H7", None, fmt_blank)
    ws.write("I7", "Volume 6M Ant.", fmt_header)
    ws.write("J7", "Volume 6M Atual", fmt_header)
    ws.write("K7", "Tendência", fmt_header)
    for i, row in seis_df.reset_index(drop=True).iterrows():
        r = 7 + i
        ws.write(r, 7, row.get("Indicador", ""), fmt_text_center)
        write_number_ou_branco(ws, r, 8, row.get(label_6_ant, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 9, row.get(label_6_atual, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 10, row.get("Tendência", np.nan), fmt_pct)

    ws.merge_range("B12:D12", "Período", fmt_secao)
    ws.write("C13", "Dt Inic", fmt_header)
    ws.write("D13", "Dt fim", fmt_header)
    for i, indicador in enumerate(["Sellin", "Sellout 2.0", "Sellout 3.0"], start=14):
        ws.write(i - 1, 1, indicador, fmt_text_center)
        for c, key in [(2, "inicio_atual"), (3, "fim_atual")]:
            dt = periodos.get(key)
            if pd.notna(dt):
                ws.write_datetime(i - 1, c, pd.Timestamp(dt).to_pydatetime(), fmt_date)
            else:
                ws.write_blank(i - 1, c, None, fmt_date)

    # Tabela MAT: últimos 12 meses móveis contra os 12 meses imediatamente anteriores.
    ws.merge_range("H12:K12", "MAT", fmt_secao)
    ws.write("I13", f"Volume {label_mat_ant}", fmt_header)
    ws.write("J13", f"Volume {label_mat_atual}", fmt_header)
    ws.write("K13", "Tendência", fmt_header)
    for i, row in mat_df.reset_index(drop=True).iterrows():
        r = 13 + i
        ws.write(r, 7, row.get("Indicador", ""), fmt_text_center)
        write_number_ou_branco(ws, r, 8, row.get(label_mat_ant, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 9, row.get(label_mat_atual, np.nan), fmt_num)
        write_number_ou_branco(ws, r, 10, row.get("Tendência", np.nan), fmt_pct)

    # Tabela mensal principal em B:G, conforme layout de comparação.
    mensal_cols = ["Data", "Sell-in", "Sell-out 2.0", "Sell-out 3.0", "Cobertura 2.0", "Cobertura 3.0"]
    for col in mensal_cols:
        if col not in mensal.columns:
            mensal[col] = np.nan
    mensal_saida = mensal[mensal_cols].copy()
    start_mensal = 17
    for j, h in enumerate(mensal_saida.columns, start=1):
        ws.write(start_mensal, j, h, fmt_header)
    for i, row in mensal_saida.reset_index(drop=True).iterrows():
        r = start_mensal + 1 + i
        for j, h in enumerate(mensal_saida.columns, start=1):
            val = row[h]
            if h == "Data":
                chave = mensal.iloc[i].get("chave") if i < len(mensal) else None
                ts = _comparacao_chave_para_mes(chave)
                if pd.notna(ts):
                    ws.write_datetime(r, j, ts.to_pydatetime(), fmt_date)
                else:
                    ws.write(r, j, "" if pd.isna(val) else str(val), fmt_text)
            elif "Cobertura" in h or "%" in h:
                write_number_ou_branco(ws, r, j, val, fmt_pct)
            else:
                write_number_ou_branco(ws, r, j, val, fmt_num)

    # Tabela por UF ao lado da tabela mensal, em I:Q.
    if uf is None:
        uf = pd.DataFrame()
    uf = limpar_dataframe_excel(uf)
    if skus_excluidos20 is None:
        skus_excluidos20 = pd.DataFrame()
    if skus_excluidos30 is None:
        skus_excluidos30 = pd.DataFrame()
    skus_excluidos20 = limpar_dataframe_excel(skus_excluidos20)
    skus_excluidos30 = limpar_dataframe_excel(skus_excluidos30)
    start_uf = start_mensal
    uf_cols = ["UF", "Sell-in 12M", "Sell-out 12M 2.0", "Sell-out 12M 3.0", "Cobertura 2.0", "Cobertura 3.0", "Importância Sell-in", "Importância Sell-out 2.0", "Importância Sell-out 3.0"]
    for j, h in enumerate(uf_cols, start=8):
        ws.write(start_uf, j, h, fmt_header)
    for i, row in uf.reset_index(drop=True).iterrows():
        r = start_uf + 1 + i
        uf_nome = str(row.get("UF", ""))
        for j, h in enumerate(uf_cols, start=8):
            val = row.get(h, np.nan)
            if h == "UF":
                ws.write(r, j, uf_nome, fmt_text)
            elif "Cobertura" in h or "Importância" in h:
                write_number_ou_branco(ws, r, j, val, fmt_pct)
            else:
                write_number_ou_branco(ws, r, j, val, fmt_num)

    first = start_mensal + 1
    last = start_mensal + len(mensal_saida)
    chart_row = max(start_mensal + len(mensal_saida) + 1, start_uf + len(uf) + 2)
    if len(mensal_saida) > 0:
        chart = workbook.add_chart({"type": "line"})
        chart.add_series({"name": "Sell-in", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 2, last, 2], "line": {"color": azul_linha, "width": 2.25}})
        chart.add_series({"name": "Sell-out 2.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 3, last, 3], "line": {"color": laranja_linha, "width": 2.25}})
        chart.add_series({"name": "Sell-out 3.0", "categories": [nome_aba, first, 1, last, 1], "values": [nome_aba, first, 4, last, 4], "line": {"color": verde_linha, "width": 2.25}})
        chart.set_title({"name": "Sell-in x Sell-out 2.0 x Sell-out 3.0 por mês", "name_font": {"color": "#000000", "size": 14, "bold": True}})
        chart.set_x_axis({"name": "Mês", "num_font": {"rotation": -90}, "num_format": "mmm/yy"})
        chart.set_y_axis({"name": "Volume", "num_format": "#,##0"})
        chart.set_legend({"position": "bottom"})
        chart.set_size({"width": 760, "height": 340})
        chart.set_plotarea({"border": {"none": True}})
        chart.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_row, 1, chart)

    # Tabela adicional opcional: contribuição por UF, Canal e EAN usando apenas Sell-out 2.0 x 3.0.
    fim_contrib = chart_row + 17
    if gerar_top20_contribuicao:
        contrib_title_row = chart_row + 18
        if contribuicao_skus is None:
            contribuicao_skus = pd.DataFrame()
        contribuicao_skus = limpar_dataframe_excel(contribuicao_skus)
        contrib_cols = list(contribuicao_skus.columns) if not contribuicao_skus.empty else ["Aviso"]
        max_col_contrib = max(1 + len(contrib_cols) - 1, 1)
        try:
            ws.merge_range(contrib_title_row, 1, contrib_title_row, max_col_contrib, "Top 20 contribuição por cada UF / Canal / EAN - Sell-out 2.0 x 3.0", fmt_secao)
        except Exception:
            ws.write(contrib_title_row, 1, "Top 20 contribuição por cada UF / Canal / EAN - Sell-out 2.0 x 3.0", fmt_secao)
        info_row = contrib_title_row + 1
        info_txt = info_contribuicao or "Tabela calculada a partir da base auxiliar de Sell-out dos estudos comparados."
        try:
            ws.merge_range(info_row, 1, info_row, max_col_contrib, info_txt, fmt_text)
        except Exception:
            ws.write(info_row, 1, info_txt, fmt_text)
        contrib_header_row = contrib_title_row + 2

        if contribuicao_skus.empty:
            ws.write(contrib_header_row, 1, "Aviso", fmt_header)
            ws.write(contrib_header_row + 1, 1, info_txt, fmt_text)
            fim_contrib = contrib_header_row + 1
        else:
            for jj, h in enumerate(contrib_cols, start=1):
                ws.write(contrib_header_row, jj, h, fmt_header)
            pct_cols_norm = {normalizar_texto(c) for c in contrib_cols if "contrib" in normalizar_texto(c)}
            for ii, (_, contrib_row) in enumerate(contribuicao_skus.reset_index(drop=True).iterrows()):
                rr = contrib_header_row + 1 + ii
                for jj, h in enumerate(contrib_cols, start=1):
                    val = contrib_row.get(h, np.nan)
                    h_norm = normalizar_texto(h)
                    if h in {"UF", "Canal", "EAN"}:
                        ws.write(rr, jj, "" if pd.isna(val) else str(val), fmt_text)
                    elif h_norm in pct_cols_norm:
                        write_number_ou_branco(ws, rr, jj, val, fmt_pct)
                    elif isinstance(val, (int, float, np.integer, np.floating)) and not pd.isna(val):
                        write_number_ou_branco(ws, rr, jj, val, fmt_num)
                    else:
                        ws.write(rr, jj, "" if pd.isna(val) else str(val), fmt_text)
            fim_contrib = contrib_header_row + len(contribuicao_skus)
            ws.set_column(1, min(max_col_contrib, 13), 15)

    # Tabelas separadas de SKU em Comum, conforme o layout enviado.
    sku_title_row = fim_contrib + 3
    ws.merge_range(sku_title_row, 1, sku_title_row, 4, "SKU em Comum 2.0", fmt_secao)
    ws.merge_range(sku_title_row, 6, sku_title_row, 9, "SKU em Comum 3.0", fmt_secao)
    sku_header_row = sku_title_row + 1
    for j, h in enumerate(["Data", "Sell-in", "Sell-out 2.0", "Cobertura 2.0"], start=1):
        ws.write(sku_header_row, j, h, fmt_header)
    for j, h in enumerate(["Data", "Sell-in", "Sell-out 3.0", "Cobertura 3.0"], start=6):
        ws.write(sku_header_row, j, h, fmt_header)

    primeira_sku_row = sku_header_row + 1
    for i in range(len(mensal)):
        r = primeira_sku_row + i
        chave = mensal.iloc[i].get("chave") if i < len(mensal) else None
        ts = _comparacao_chave_para_mes(chave)
        if pd.notna(ts):
            ws.write_datetime(r, 1, ts.to_pydatetime(), fmt_date)
            ws.write_datetime(r, 6, ts.to_pydatetime(), fmt_date)
        else:
            data_txt = mensal.iloc[i].get("Data", "")
            ws.write(r, 1, "" if pd.isna(data_txt) else str(data_txt), fmt_text)
            ws.write(r, 6, "" if pd.isna(data_txt) else str(data_txt), fmt_text)

        write_number_ou_branco(ws, r, 2, mensal.iloc[i].get("Sell-in SKU em Comum 2.0", np.nan), fmt_num)
        write_number_ou_branco(ws, r, 3, mensal.iloc[i].get("Sell-out SKU em Comum 2.0", np.nan), fmt_num)
        write_number_ou_branco(ws, r, 7, mensal.iloc[i].get("Sell-in SKU em Comum 3.0", np.nan), fmt_num)
        write_number_ou_branco(ws, r, 8, mensal.iloc[i].get("Sell-out SKU em Comum 3.0", np.nan), fmt_num)
        # Não reescreve fórmulas na comparação: preserva a cobertura já calculada no estudo original.
        write_number_ou_branco(ws, r, 4, mensal.iloc[i].get("Cobertura SKU em Comum 2.0", np.nan), fmt_pct)
        write_number_ou_branco(ws, r, 9, mensal.iloc[i].get("Cobertura SKU em Comum 3.0", np.nan), fmt_pct)

    # Tabelas dos SKUs excluídos do cálculo de SKU em comum, uma para cada versão.
    # A comparação apenas lê e reproduz a lista salva em cada estudo original.
    def escrever_skus_excluidos_comp(start_row: int, start_col: int, titulo: str, df_excluidos: pd.DataFrame) -> int:
        cols_excluidos = ["SKU", "Status", "Sell-in", "Sell-out", "Nome SKU", "Marca", "Fabricante"]
        ws.merge_range(start_row, start_col, start_row, start_col + len(cols_excluidos) - 1, titulo, fmt_secao)
        header_row = start_row + 1
        for jj, h in enumerate(cols_excluidos):
            ws.write(header_row, start_col + jj, h, fmt_header)

        if df_excluidos is None or df_excluidos.empty:
            ws.merge_range(header_row + 1, start_col, header_row + 1, start_col + len(cols_excluidos) - 1, "Nenhum SKU excluído ou tabela não localizada no estudo original.", fmt_text_center)
            return header_row + 1

        df_excluidos = df_excluidos.copy()
        for c in cols_excluidos:
            if c not in df_excluidos.columns:
                df_excluidos[c] = np.nan if c in {"Sell-in", "Sell-out"} else ""

        for ii, (_, sku_row) in enumerate(df_excluidos[cols_excluidos].reset_index(drop=True).iterrows()):
            rr = header_row + 1 + ii
            ws.write(rr, start_col + 0, str(sku_row.get("SKU", "")), fmt_text)
            ws.write(rr, start_col + 1, str(sku_row.get("Status", "")), fmt_text_center)
            write_number_ou_branco(ws, rr, start_col + 2, sku_row.get("Sell-in", np.nan), fmt_num)
            write_number_ou_branco(ws, rr, start_col + 3, sku_row.get("Sell-out", np.nan), fmt_num)
            ws.write(rr, start_col + 4, str(sku_row.get("Nome SKU", "")), fmt_text)
            ws.write(rr, start_col + 5, str(sku_row.get("Marca", "")), fmt_text)
            ws.write(rr, start_col + 6, str(sku_row.get("Fabricante", "")), fmt_text)
        return header_row + len(df_excluidos)

    excluidos_title_row = primeira_sku_row + max(len(mensal), 1) + 2
    fim_excluidos20 = escrever_skus_excluidos_comp(excluidos_title_row, 1, "SKUs excluídos do SKU em Comum - 2.0", skus_excluidos20)
    fim_excluidos30 = escrever_skus_excluidos_comp(excluidos_title_row, 9, "SKUs excluídos do SKU em Comum - 3.0", skus_excluidos30)
    fim_excluidos = max(fim_excluidos20, fim_excluidos30)

    if len(mensal) > 0:
        last_sku = primeira_sku_row + len(mensal) - 1
        chart_comum_row = fim_excluidos + 2
        chart_comum = workbook.add_chart({"type": "line"})
        chart_comum.add_series({"name": "Sell-in SKU comum 2.0", "categories": [nome_aba, primeira_sku_row, 1, last_sku, 1], "values": [nome_aba, primeira_sku_row, 2, last_sku, 2], "line": {"color": azul_linha, "width": 2.25, "dash_type": "dash"}})
        chart_comum.add_series({"name": "Sell-out SKU comum 2.0", "categories": [nome_aba, primeira_sku_row, 1, last_sku, 1], "values": [nome_aba, primeira_sku_row, 3, last_sku, 3], "line": {"color": laranja_linha, "width": 2.25}})
        chart_comum.add_series({"name": "Sell-in SKU comum 3.0", "categories": [nome_aba, primeira_sku_row, 6, last_sku, 6], "values": [nome_aba, primeira_sku_row, 7, last_sku, 7], "line": {"color": azul_linha, "width": 2.25}})
        chart_comum.add_series({"name": "Sell-out SKU comum 3.0", "categories": [nome_aba, primeira_sku_row, 6, last_sku, 6], "values": [nome_aba, primeira_sku_row, 8, last_sku, 8], "line": {"color": verde_linha, "width": 2.25}})
        chart_comum.set_title({"name": "SKU em comum - Sell-in x Sell-out", "name_font": {"color": "#000000", "size": 14, "bold": True}})
        chart_comum.set_x_axis({"name": "Mês", "num_font": {"rotation": -90}, "num_format": "mmm/yy"})
        chart_comum.set_y_axis({"name": "Volume", "num_format": "#,##0"})
        chart_comum.set_legend({"position": "bottom"})
        chart_comum.set_size({"width": 760, "height": 340})
        chart_comum.set_plotarea({"border": {"none": True}})
        chart_comum.set_chartarea({"border": {"color": "#BFBFBF"}})
        ws.insert_chart(chart_comum_row, 1, chart_comum)




# ============================================================
# Abas auxiliares da comparação 2.0 x 3.0
# ============================================================

def _ler_tabela_auxiliar_estudo(caminho: str | Path, sheet_name: str, header: int = 1) -> pd.DataFrame:
    """Lê uma aba tabular dos estudos gerados, ignorando a linha de título visual."""
    try:
        df = pd.read_excel(caminho, sheet_name=sheet_name, header=header)
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    # Remove colunas automáticas vazias do Excel.
    remover = []
    for c in df.columns:
        if str(c).startswith("Unnamed") and df[c].isna().all():
            remover.append(c)
    if remover:
        df = df.drop(columns=remover)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _ler_matriz_auxiliar_estudo(caminho: str | Path, sheet_name: str) -> pd.DataFrame:
    """Lê uma aba visual, como Crosschecks, como matriz crua para exibir lado a lado."""
    try:
        df = pd.read_excel(caminho, sheet_name=sheet_name, header=None)
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    return df.dropna(axis=0, how="all").dropna(axis=1, how="all")


def _achar_coluna_aux(df: pd.DataFrame, alternativas: List[str]) -> Optional[str]:
    if df is None or df.empty:
        return None
    mapa = {normalizar_texto(c): c for c in df.columns}
    for alt in alternativas:
        n = normalizar_texto(alt)
        if n in mapa:
            return mapa[n]
    for alt in alternativas:
        n = normalizar_texto(alt)
        for cn, cr in mapa.items():
            if n and (n == cn or n in cn):
                return cr
    return None


def _preparar_chaves_aux(df: pd.DataFrame, chaves: List[Tuple[str, List[str]]]) -> Tuple[pd.DataFrame, List[str]]:
    """Renomeia as chaves encontradas para nomes canônicos e devolve as chaves utilizáveis."""
    if df is None or df.empty:
        return pd.DataFrame(), []
    out = df.copy()
    usadas = []
    ren = {}
    for canon, alternativas in chaves:
        col = _achar_coluna_aux(out, alternativas)
        if col:
            ren[col] = canon
            usadas.append(canon)
    if ren:
        out = out.rename(columns=ren)
    for c in usadas:
        if c in out.columns:
            out[c] = out[c].fillna("SEM INFORMAÇÃO").astype(str).str.strip().replace("", "SEM INFORMAÇÃO")
    return out, usadas


def _limpar_para_merge_aux(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    # Evita colunas duplicadas no merge.
    novas = []
    cont = {}
    for c in out.columns:
        nome = str(c).strip()
        if nome in cont:
            cont[nome] += 1
            nome = f"{nome}_{cont[nome]}"
        else:
            cont[nome] = 1
        novas.append(nome)
    out.columns = novas
    return out


def _combinar_tabelas_2versoes(
    df20: pd.DataFrame,
    df30: pd.DataFrame,
    chaves: List[Tuple[str, List[str]]],
    label20: str = "2.0",
    label30: str = "3.0",
) -> pd.DataFrame:
    """
    Junta duas abas tabulares dos estudos, mantendo as mesmas colunas lado a lado:
    Coluna X 2.0, Coluna X 3.0.
    """
    df20 = _limpar_para_merge_aux(df20)
    df30 = _limpar_para_merge_aux(df30)
    df20, keys20 = _preparar_chaves_aux(df20, chaves)
    df30, keys30 = _preparar_chaves_aux(df30, chaves)
    keys = [k for k, _ in chaves if k in keys20 or k in keys30]

    if df20.empty and df30.empty:
        return pd.DataFrame()

    # Quando não existe chave confiável, replica as linhas lado a lado por posição.
    if not keys:
        a = df20.reset_index(drop=True).add_suffix(f" {label20}") if not df20.empty else pd.DataFrame()
        b = df30.reset_index(drop=True).add_suffix(f" {label30}") if not df30.empty else pd.DataFrame()
        max_len = max(len(a), len(b))
        a = a.reindex(range(max_len))
        b = b.reindex(range(max_len))
        return pd.concat([a, b], axis=1)

    for k in keys:
        if k not in df20.columns:
            df20[k] = "SEM INFORMAÇÃO"
        if k not in df30.columns:
            df30[k] = "SEM INFORMAÇÃO"

    cols20_orig = [c for c in df20.columns if c not in keys]
    cols30_orig = [c for c in df30.columns if c not in keys]
    ordem_metricas = []
    for c in cols20_orig + cols30_orig:
        if c not in ordem_metricas:
            ordem_metricas.append(c)

    a = df20.rename(columns={c: f"{c} {label20}" for c in cols20_orig})
    b = df30.rename(columns={c: f"{c} {label30}" for c in cols30_orig})

    try:
        comp = a.merge(b, on=keys, how="outer")
    except Exception:
        # Fallback por posição para evitar travar caso existam duplicidades inesperadas de chave.
        a = df20.reset_index(drop=True).add_suffix(f" {label20}")
        b = df30.reset_index(drop=True).add_suffix(f" {label30}")
        max_len = max(len(a), len(b))
        return pd.concat([a.reindex(range(max_len)), b.reindex(range(max_len))], axis=1)

    colunas = list(keys)
    for c in ordem_metricas:
        c20 = f"{c} {label20}"
        c30 = f"{c} {label30}"
        # Garante o mesmo conjunto de colunas em 2.0 e 3.0.
        # Se uma coluna existir só em uma versão, a outra é criada em branco,
        # mantendo sempre o par lado a lado: Coluna 2.0 | Coluna 3.0.
        if c20 not in comp.columns:
            comp[c20] = np.nan
        if c30 not in comp.columns:
            comp[c30] = np.nan
        colunas.append(c20)
        colunas.append(c30)
    extras = [c for c in comp.columns if c not in colunas]
    comp = comp[colunas + extras]
    return remover_colunas_duplicadas(comp)


def _montar_resumo_comparacao_completo(resumo_base: pd.DataFrame, estudo20: Path, estudo30: Path) -> pd.DataFrame:
    """Acrescenta ao Resumo da comparação todas as colunas do Resumo individual, lado a lado 2.0/3.0."""
    df20 = _ler_tabela_auxiliar_estudo(estudo20, "Resumo Categorias", header=1)
    df30 = _ler_tabela_auxiliar_estudo(estudo30, "Resumo Categorias", header=1)
    df20 = remover_colunas_duplicadas(df20)
    df30 = remover_colunas_duplicadas(df30)
    for df in (df20, df30):
        if not df.empty:
            # Remove o hyperlink original para criar um novo hyperlink da comparação.
            for c in list(df.columns):
                if normalizar_texto(c) == "abrir aba":
                    df.drop(columns=[c], inplace=True)
            cat_col = _achar_coluna_aux(df, ["Categoria", "Categoria/PROD"])
            if cat_col and cat_col != "Categoria/PROD":
                df.rename(columns={cat_col: "Categoria/PROD"}, inplace=True)

    detalhe = _combinar_tabelas_2versoes(
        df20, df30,
        [("Categoria/PROD", ["Categoria/PROD", "Categoria"])],
        "2.0", "3.0",
    )
    if resumo_base is None or resumo_base.empty:
        return remover_colunas_duplicadas(detalhe)
    out = remover_colunas_duplicadas(resumo_base.copy())
    if not detalhe.empty and "Categoria/PROD" in detalhe.columns:
        out = out.merge(remover_colunas_duplicadas(detalhe), on="Categoria/PROD", how="left")
    return remover_colunas_duplicadas(out)


def _escrever_dataframe_comparacao_aux(writer, sheet_name: str, titulo: str, df: pd.DataFrame):
    workbook = writer.book
    fmt_titulo = workbook.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white"})
    if df is None or df.empty:
        df = pd.DataFrame({"Aviso": ["Não foram encontrados dados dessa aba em nenhum dos estudos comparados."]})
    df = remover_colunas_duplicadas(df)
    df = limpar_dataframe_excel(df)
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    ws = writer.sheets[sheet_name]
    last_col = max(len(df.columns) - 1, 0)
    try:
        ws.merge_range(0, 0, 0, last_col, titulo, fmt_titulo)
    except Exception:
        ws.write(0, 0, titulo, fmt_titulo)
    percent_cols = {c for c in df.columns if any(t in normalizar_texto(c) for t in ["cobertura", "importancia", "variacao", "tendencia", "diferenca tendencia", "%"])}
    number_cols = {c for c in df.columns if c not in percent_cols}
    aplicar_formatos_basicos(writer, sheet_name, df, startrow=1, startcol=0, percent_cols=percent_cols, number_cols=number_cols)


def _escrever_crosschecks_comparacao_lado_a_lado(writer, estudo20: Path, estudo30: Path):
    """Replica Crosschecks dos dois estudos lado a lado, preservando a estrutura visual de cada um."""
    df20 = _ler_matriz_auxiliar_estudo(estudo20, "Crosschecks")
    df30 = _ler_matriz_auxiliar_estudo(estudo30, "Crosschecks")
    if df20.empty and df30.empty:
        _escrever_dataframe_comparacao_aux(writer, "Crosschecks", "Crosschecks 2.0 x 3.0", pd.DataFrame({"Aviso": ["Crosschecks não encontrados nos estudos comparados."]}))
        return

    workbook = writer.book
    ws = workbook.add_worksheet("Crosschecks")
    writer.sheets["Crosschecks"] = ws
    fmt_titulo = workbook.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E78", "font_color": "white", "align": "center"})
    fmt_header = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1})
    fmt_text = workbook.add_format({"border": 1})
    fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})

    n20 = df20.shape[1] if not df20.empty else 0
    offset30 = max(n20 + 3, 4)
    ws.merge_range(0, 0, 0, max(n20 - 1, 0), "Crosschecks - Sell-out 2.0", fmt_titulo)
    ws.merge_range(0, offset30, 0, offset30 + max((df30.shape[1] if not df30.empty else 1) - 1, 0), "Crosschecks - Sell-out 3.0", fmt_titulo)

    def escrever_matriz(df: pd.DataFrame, start_col: int):
        if df.empty:
            ws.write(1, start_col, "Sem dados", fmt_header)
            return
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                val = df.iat[i, j]
                if pd.isna(val):
                    ws.write_blank(i + 1, start_col + j, None, fmt_text)
                elif isinstance(val, (int, float, np.integer, np.floating)):
                    ws.write_number(i + 1, start_col + j, float(val), fmt_num)
                else:
                    ws.write(i + 1, start_col + j, str(val), fmt_text)
        for j in range(df.shape[1]):
            ws.set_column(start_col + j, start_col + j, 14)

    escrever_matriz(df20, 0)
    escrever_matriz(df30, offset30)


def _escrever_abas_auxiliares_comparacao(writer, estudo20: Path, estudo30: Path, output_options: Optional[Dict[str, object]] = None):
    """
    Cria na comparação as mesmas abas auxiliares do Estudo de Cobertura individual.

    Abas replicadas:
    - Base SKUs
    - SKUs por Categoria
    - Crosschecks
    - Parâmetros
    - Avisos
    - Descrição Cálculos

    As colunas de métricas são gravadas em pares: campo 2.0 ao lado do campo 3.0.
    """
    output_options = normalizar_opcoes_saida(output_options)
    specs = []
    if output_options.get("base_skus", True):
        specs.append(("Base SKUs", "Base SKUs - comparação 2.0 x 3.0", [("Categoria", ["Categoria"]), ("ean", ["ean", "EAN", "SKU"])], 1))
    if output_options.get("skus_por_categoria", True):
        specs.append(("SKUs por Categoria", "SKUs por Categoria - comparação 2.0 x 3.0", [("Categoria", ["Categoria", "Categoria/PROD"])], 1))
    if output_options.get("parametros", True):
        specs.append(("Parâmetros", "Parâmetros usados - comparação 2.0 x 3.0", [("Parâmetro", ["Parâmetro", "Parametro"])], 1))
    if output_options.get("descricao_calculos", True):
        specs.append(("Descrição Cálculos", "Descrição dos cálculos - comparação 2.0 x 3.0", [("Nome da aba", ["Nome da aba"]), ("Nome da coluna calculada", ["Nome da coluna calculada"])], 1))
    if output_options.get("avisos", True):
        specs.append(("Avisos", "Avisos - comparação 2.0 x 3.0", [], 1))
    for sheet, titulo, chaves, header in specs:
        df20 = _ler_tabela_auxiliar_estudo(estudo20, sheet, header=header)
        df30 = _ler_tabela_auxiliar_estudo(estudo30, sheet, header=header)
        combinado = _combinar_tabelas_2versoes(df20, df30, chaves, "2.0", "3.0")
        _escrever_dataframe_comparacao_aux(writer, sheet, titulo, combinado)
    if output_options.get("crosschecks", True):
        _escrever_crosschecks_comparacao_lado_a_lado(writer, estudo20, estudo30)

def _categoria_key_resumo(valor) -> str:
    return normalizar_texto(valor)


def _linha_resumo_original_por_categoria(estudo: str | Path, categoria: str) -> Dict[str, object]:
    """Lê a linha original do Resumo Categorias do estudo gerado, sem recalcular nada."""
    df = _ler_tabela_auxiliar_estudo(estudo, "Resumo Categorias", header=1)
    if df is None or df.empty:
        return {}
    df = remover_colunas_duplicadas(df)
    cat_col = _achar_coluna_aux(df, ["Categoria", "Categoria/PROD"])
    if not cat_col:
        return {}
    keys = df[cat_col].map(_categoria_key_resumo)
    alvo = _categoria_key_resumo(categoria)
    hit = df.loc[keys == alvo]
    if hit.empty:
        return {}
    return hit.iloc[0].to_dict()


def _obter_valor_resumo_original(linha: Dict[str, object], alternativas: List[str]):
    if not linha:
        return np.nan
    mapa = {normalizar_texto(k): k for k in linha.keys()}
    for alt in alternativas:
        k = mapa.get(normalizar_texto(alt))
        if k is not None:
            return linha.get(k)
    for alt in alternativas:
        n = normalizar_texto(alt)
        for kn, kr in mapa.items():
            if n and n in kn:
                return linha.get(kr)
    return np.nan


def _valor_resumo_original_num(linha: Dict[str, object], alternativas: List[str]):
    return _valor_numerico_planilha(_obter_valor_resumo_original(linha, alternativas))


def _montar_resumo_base_preservado(categoria: str, linha20: Dict[str, object], linha30: Dict[str, object], sheet: str, status: str) -> Dict[str, object]:
    """Monta as colunas principais da comparação usando valores já calculados nos estudos."""
    si = _valor_resumo_original_num(linha20, ["Sell-in MAT", "Sell in MAT", "Sell-in atual", "Sell-in"])
    if pd.isna(si):
        si = _valor_resumo_original_num(linha30, ["Sell-in MAT", "Sell in MAT", "Sell-in atual", "Sell-in"])
    so20 = _valor_resumo_original_num(linha20, ["Sell-out MAT", "Sell out MAT", "Sell-out atual", "Sell-out"])
    so30 = _valor_resumo_original_num(linha30, ["Sell-out MAT", "Sell out MAT", "Sell-out atual", "Sell-out"])
    cov20 = _valor_resumo_original_num(linha20, ["Cobertura MAT", "Cobertura atual", "Cobertura"])
    cov30 = _valor_resumo_original_num(linha30, ["Cobertura MAT", "Cobertura atual", "Cobertura"])
    diff_so = so30 - so20 if pd.notna(so30) and pd.notna(so20) else np.nan
    diff_so_pct = divisao_segura(diff_so, so20) if pd.notna(diff_so) else np.nan
    diff_cov = cov30 - cov20 if pd.notna(cov30) and pd.notna(cov20) else np.nan
    return {
        "Abrir aba": "Abrir",
        "Categoria/PROD": categoria,
        "Status Comparação": status,
        "Sell-in": si,
        "Sell-out 2.0": so20,
        "Sell-out 3.0": so30,
        "Diferença Sell-out": diff_so,
        "Diferença % Sell-out": diff_so_pct,
        "Cobertura 2.0": cov20,
        "Cobertura 3.0": cov30,
        "Diferença Cobertura": diff_cov,
        "sheet": sheet,
    }



def _extrair_base_contribuicao_estudo(caminho: str | Path) -> pd.DataFrame:
    """Lê a aba oculta Base Contribuição Sell-out de um estudo individual."""
    df = _ler_tabela_auxiliar_estudo(caminho, "Base Contribuição Sell-out", header=1)
    colunas = ["Aba Categoria", "Categoria", "categoria_key", "UF", "Canal", "EAN", "Data", "Ano", "Volume Sell-out"]
    if df is None or df.empty:
        return pd.DataFrame(columns=colunas)

    ren = {}
    alternativas = {
        "Aba Categoria": ["Aba Categoria", "Nome aba", "Aba"],
        "Categoria": ["Categoria", "Categoria/PROD"],
        "categoria_key": ["categoria_key", "Categoria Key"],
        "UF": ["UF", "uf"],
        "Canal": ["Canal", "PDV_CANAL", "PDV Canal"],
        "EAN": ["EAN", "SKU", "ean"],
        "Data": ["Data", "Mês", "Mes", "mes"],
        "Ano": ["Ano", "ano"],
        "Volume Sell-out": ["Volume Sell-out", "valor_sellout", "Sell-out", "Volume"],
    }
    for canon, alts in alternativas.items():
        col = _achar_coluna_aux(df, alts)
        if col:
            ren[col] = canon
    if ren:
        df = df.rename(columns=ren)

    for c in colunas:
        if c not in df.columns:
            df[c] = np.nan if c in {"Data", "Ano", "Volume Sell-out"} else ""

    out = df[colunas].copy()
    out["Aba Categoria"] = out["Aba Categoria"].fillna("").astype(str).str.strip()
    out["Categoria"] = out["Categoria"].fillna("").astype(str).str.strip()
    out["categoria_key"] = out["categoria_key"].fillna("").astype(str).map(normalizar_categoria)
    out.loc[out["categoria_key"].eq(""), "categoria_key"] = out.loc[out["categoria_key"].eq(""), "Categoria"].map(normalizar_categoria)
    out["UF"] = out["UF"].fillna("TOTAL").astype(str).str.strip().replace("", "TOTAL")
    out["Canal"] = out["Canal"].fillna("TOTAL").astype(str).str.strip().replace("", "TOTAL")
    out["EAN"] = out["EAN"].map(ean_texto)
    out["Data"] = out["Data"].map(converter_mes)
    out["Ano"] = pd.to_numeric(out["Ano"], errors="coerce")
    out.loc[out["Data"].notna(), "Ano"] = out.loc[out["Data"].notna(), "Data"].dt.year
    out["Volume Sell-out"] = limpar_coluna_numerica_vetorizada(out["Volume Sell-out"])
    out = out[(out["EAN"] != "") & (out["Volume Sell-out"].fillna(0) != 0)].copy()
    return out.reset_index(drop=True)


def _filtrar_base_contribuicao_categoria(base: pd.DataFrame, categoria_aba: str) -> pd.DataFrame:
    if base is None or base.empty:
        return pd.DataFrame()
    alvo_aba = str(categoria_aba or "").strip()
    alvo_norm = normalizar_categoria(alvo_aba)
    out = base.copy()
    if "Aba Categoria" in out.columns:
        hit = out[out["Aba Categoria"].fillna("").astype(str).str.strip() == alvo_aba].copy()
        if not hit.empty:
            return hit
    if "Categoria" in out.columns:
        hit = out[out["Categoria"].map(normalizar_categoria) == alvo_norm].copy()
        if not hit.empty:
            return hit
    if "categoria_key" in out.columns:
        hit = out[out["categoria_key"].map(normalizar_categoria) == alvo_norm].copy()
        if not hit.empty:
            return hit
    return pd.DataFrame()


def _meses_comuns_contribuicao(base20: pd.DataFrame, base30: pd.DataFrame) -> List[pd.Timestamp]:
    meses20 = set(pd.to_datetime(base20.get("Data", pd.Series(dtype="datetime64[ns]")), errors="coerce").dropna().map(lambda x: pd.Timestamp(x).to_period("M").to_timestamp()))
    meses30 = set(pd.to_datetime(base30.get("Data", pd.Series(dtype="datetime64[ns]")), errors="coerce").dropna().map(lambda x: pd.Timestamp(x).to_period("M").to_timestamp()))
    return sorted(meses20 & meses30)


def _selecionar_periodos_contribuicao(base20: pd.DataFrame, base30: pd.DataFrame) -> Optional[Dict[str, object]]:
    """
    Escolhe os dois períodos da análise de contribuição.

    Prioridade:
    1) dois últimos anos completos disponíveis nos dois Sell-outs;
    2) YTD do último ano contra o mesmo período do ano anterior;
    3) MAT atual contra MAT anterior, usando meses comuns;
    4) dois últimos anos disponíveis, quando só houver ano e não houver mês.
    """
    meses = _meses_comuns_contribuicao(base20, base30)
    meses_set = set(meses)

    if meses:
        anos = sorted({m.year for m in meses})
        anos_completos = []
        for ano in anos:
            esperado = {pd.Timestamp(year=ano, month=m, day=1) for m in range(1, 13)}
            if esperado <= meses_set:
                anos_completos.append(ano)
        if len(anos_completos) >= 2:
            ano1, ano2 = anos_completos[-2], anos_completos[-1]
            return {
                "tipo": "ANOS_COMPLETOS",
                "label1": str(ano1),
                "label2": str(ano2),
                "meses1": [pd.Timestamp(year=ano1, month=m, day=1) for m in range(1, 13)],
                "meses2": [pd.Timestamp(year=ano2, month=m, day=1) for m in range(1, 13)],
                "criterio": f"Período usado: anos completos {ano1} e {ano2}.",
            }

        ultimo = meses[-1]
        ano2 = ultimo.year
        ano1 = ano2 - 1
        mes_limite = ultimo.month
        meses1 = [pd.Timestamp(year=ano1, month=m, day=1) for m in range(1, mes_limite + 1)]
        meses2 = [pd.Timestamp(year=ano2, month=m, day=1) for m in range(1, mes_limite + 1)]
        if set(meses1) <= meses_set and set(meses2) <= meses_set:
            mes_label = MESES_PT_ABREV.get(mes_limite, str(mes_limite)).replace(".", "")
            return {
                "tipo": "YTD",
                "label1": f"YTD {ano1}",
                "label2": f"YTD {ano2}",
                "meses1": meses1,
                "meses2": meses2,
                "criterio": f"Período usado: YTD Jan-{mes_label}/{ano1} contra Jan-{mes_label}/{ano2}, com exatamente os mesmos meses.",
            }

        fim2 = ultimo
        inicio2 = (fim2 - pd.DateOffset(months=11)).to_period("M").to_timestamp()
        meses2 = list(pd.date_range(inicio2, fim2, freq="MS"))
        fim1 = inicio2 - pd.DateOffset(months=1)
        inicio1 = (fim1 - pd.DateOffset(months=11)).to_period("M").to_timestamp()
        meses1 = list(pd.date_range(inicio1, fim1, freq="MS"))
        if set(meses1) <= meses_set and set(meses2) <= meses_set:
            return {
                "tipo": "MAT",
                "label1": "MAT ant.",
                "label2": "MAT atual",
                "meses1": meses1,
                "meses2": meses2,
                "criterio": f"Período usado: MAT anterior ({formatar_periodo(inicio1)} a {formatar_periodo(fim1)}) contra MAT atual ({formatar_periodo(inicio2)} a {formatar_periodo(fim2)}).",
            }

    anos20 = set(pd.to_numeric(base20.get("Ano", pd.Series(dtype=float)), errors="coerce").dropna().astype(int).tolist())
    anos30 = set(pd.to_numeric(base30.get("Ano", pd.Series(dtype=float)), errors="coerce").dropna().astype(int).tolist())
    anos = sorted(anos20 & anos30)
    if len(anos) >= 2:
        ano1, ano2 = anos[-2], anos[-1]
        return {
            "tipo": "ANO_DISPONIVEL",
            "label1": str(ano1),
            "label2": str(ano2),
            "ano1": ano1,
            "ano2": ano2,
            "criterio": f"Período usado: anos disponíveis {ano1} e {ano2}. O arquivo não trouxe mês suficiente para validar YTD/MAT.",
        }
    return None


def _filtrar_periodo_contribuicao(base: pd.DataFrame, periodo: Dict[str, object], lado: int) -> pd.DataFrame:
    if base is None or base.empty or periodo is None:
        return pd.DataFrame()
    out = base.copy()
    if periodo.get("meses1") is not None and periodo.get("meses2") is not None:
        meses = periodo["meses1"] if lado == 1 else periodo["meses2"]
        meses_set = {pd.Timestamp(m).to_period("M").to_timestamp() for m in meses}
        datas = pd.to_datetime(out.get("Data", pd.Series(dtype="datetime64[ns]")), errors="coerce").map(lambda x: pd.Timestamp(x).to_period("M").to_timestamp() if pd.notna(x) else pd.NaT)
        return out[datas.isin(meses_set)].copy()
    ano = periodo.get("ano1") if lado == 1 else periodo.get("ano2")
    if ano is None:
        return pd.DataFrame()
    anos = pd.to_numeric(out.get("Ano", pd.Series(dtype=float)), errors="coerce")
    return out[anos == int(ano)].copy()


def _agregar_contribuicao_periodo(base: pd.DataFrame, periodo: Dict[str, object], lado: int, nome_valor: str) -> pd.DataFrame:
    filtrado = _filtrar_periodo_contribuicao(base, periodo, lado)
    if filtrado.empty:
        return pd.DataFrame(columns=["UF", "Canal", "EAN", nome_valor])
    return (
        filtrado.groupby(["UF", "Canal", "EAN"], dropna=False, as_index=False)["Volume Sell-out"]
        .sum()
        .rename(columns={"Volume Sell-out": nome_valor})
    )


def _denominador_contribuicao_periodo(base: pd.DataFrame, periodo: Dict[str, object], lado: int, nome_valor: str) -> pd.DataFrame:
    filtrado = _filtrar_periodo_contribuicao(base, periodo, lado)
    if filtrado.empty:
        return pd.DataFrame(columns=["UF", "Canal", nome_valor])
    return (
        filtrado.groupby(["UF", "Canal"], dropna=False, as_index=False)["Volume Sell-out"]
        .sum()
        .rename(columns={"Volume Sell-out": nome_valor})
    )


def _montar_contribuicao_sku_categoria(base20: pd.DataFrame, base30: pd.DataFrame, categoria_aba: str) -> Tuple[pd.DataFrame, str]:
    """Monta Top 20 SKUs por UF + Canal com maior diferença de contribuição 2.0 x 3.0."""
    cols_min = ["UF", "Canal", "EAN", "Aviso"]
    if base20 is None or base20.empty or base30 is None or base30.empty:
        return pd.DataFrame(columns=cols_min), "Base auxiliar de contribuição não encontrada nos estudos. Gere os estudos 2.0 e 3.0 novamente com esta versão do código."

    periodo = _selecionar_periodos_contribuicao(base20, base30)
    if periodo is None:
        return pd.DataFrame(columns=cols_min), "Não foi possível definir dois períodos comparáveis para contribuição."

    b20 = _filtrar_base_contribuicao_categoria(base20, categoria_aba)
    b30 = _filtrar_base_contribuicao_categoria(base30, categoria_aba)
    if b20.empty and b30.empty:
        return pd.DataFrame(columns=cols_min), "Categoria não localizada na base auxiliar de contribuição."

    label1 = str(periodo.get("label1", "Período 1"))
    label2 = str(periodo.get("label2", "Período 2"))
    c20_p1 = f"2.0 {label1}"
    c20_p2 = f"2.0 {label2}"
    c30_p1 = f"3.0 {label1}"
    c30_p2 = f"3.0 {label2}"

    partes = [
        _agregar_contribuicao_periodo(b20, periodo, 1, c20_p1),
        _agregar_contribuicao_periodo(b20, periodo, 2, c20_p2),
        _agregar_contribuicao_periodo(b30, periodo, 1, c30_p1),
        _agregar_contribuicao_periodo(b30, periodo, 2, c30_p2),
    ]
    comp = partes[0]
    for parte in partes[1:]:
        comp = comp.merge(parte, on=["UF", "Canal", "EAN"], how="outer")
    if comp.empty:
        return pd.DataFrame(columns=cols_min), periodo.get("criterio", "") + " Sem volume para a categoria nos períodos selecionados."
    for c in [c20_p1, c20_p2, c30_p1, c30_p2]:
        comp[c] = pd.to_numeric(comp.get(c, 0), errors="coerce").fillna(0)

    den20_p1 = _denominador_contribuicao_periodo(b20, periodo, 1, "den20_p1")
    den20_p2 = _denominador_contribuicao_periodo(b20, periodo, 2, "den20_p2")
    den30_p1 = _denominador_contribuicao_periodo(b30, periodo, 1, "den30_p1")
    for den in [den20_p1, den20_p2, den30_p1]:
        comp = comp.merge(den, on=["UF", "Canal"], how="left")
    for c in ["den20_p1", "den20_p2", "den30_p1"]:
        comp[c] = pd.to_numeric(comp.get(c, 0), errors="coerce").fillna(0)

    contrib1 = f"Contribuição {label1}"
    contrib2 = f"Contribuição {label2}"
    comp[contrib1] = (comp[c30_p1] - comp[c20_p1]) / comp["den20_p1"].replace(0, np.nan)
    comp[contrib2] = (comp[c30_p2] - comp[c20_p2]) / comp["den20_p2"].replace(0, np.nan)
    comp["Contribuição do 2.0"] = (comp[c20_p2] - comp[c20_p1]) / comp["den20_p1"].replace(0, np.nan)
    comp["Contribuição do 3.0"] = (comp[c30_p2] - comp[c30_p1]) / comp["den30_p1"].replace(0, np.nan)
    comp["Dif Contribuição"] = comp["Contribuição do 3.0"] - comp["Contribuição do 2.0"]
    comp["Total Geral"] = comp[[c20_p1, c20_p2, c30_p1, c30_p2]].sum(axis=1)

    # Ranking correto: Top 20 dentro de CADA combinação UF + Canal.
    # Exemplo: SP / Atacado tem seus 20 SKUs, SP / 10+ tem outros 20 SKUs, e assim por diante.
    comp["_rank_abs"] = comp["Dif Contribuição"].abs().fillna(0)
    comp["_ordem_uf"] = comp["UF"].map(lambda x: UF_COMPARACAO_RANK.get(str(x), 10_000))
    comp = comp.sort_values(
        ["_ordem_uf", "UF", "Canal", "_rank_abs", "Total Geral", "EAN"],
        ascending=[True, True, True, False, False, True],
    )
    comp = comp.groupby(["UF", "Canal"], sort=False, group_keys=False).head(20).copy()

    cols = [
        "UF", "Canal", "EAN",
        c20_p1, c20_p2, c30_p1, c30_p2,
        "Total Geral",
        contrib1, contrib2, "Contribuição do 2.0", "Contribuição do 3.0", "Dif Contribuição",
    ]
    comp = comp[cols].reset_index(drop=True)
    info = str(periodo.get("criterio", "")) + " Ranking: Top 20 SKUs por cada UF + Canal, usando o maior valor absoluto de Dif Contribuição dentro de cada grupo."
    return comp, info

def gerar_comparacao_estudos(estudo20: str | Path, estudo30: str | Path, saida: str | Path, log_callback=None, gerar_top20_sku_canal_uf: bool = False, output_options: Optional[Dict[str, object]] = None) -> Path:
    def log(msg, pct=None):
        if log_callback:
            log_callback(msg, pct)
    output_options = normalizar_opcoes_saida(output_options)
    if gerar_top20_sku_canal_uf:
        output_options["top20_sku_canal_uf"] = True
    gerar_top20_sku_canal_uf = bool(output_options.get("top20_sku_canal_uf", False))
    estudo20 = Path(estudo20)
    estudo30 = Path(estudo30)
    saida = Path(saida)
    log("Lendo estudo 2.0...", 10)
    dados20 = _ler_estudo_gerado_para_comparacao_completo(estudo20)
    log("Lendo estudo 3.0...", 25)
    dados30 = _ler_estudo_gerado_para_comparacao_completo(estudo30)
    if gerar_top20_sku_canal_uf:
        log("Lendo bases auxiliares de contribuição SKU...", 28)
        base_contrib20 = _extrair_base_contribuicao_estudo(estudo20)
        base_contrib30 = _extrair_base_contribuicao_estudo(estudo30)
    else:
        base_contrib20 = pd.DataFrame()
        base_contrib30 = pd.DataFrame()
    if not dados20:
        raise ValueError(f"Não encontrei abas de cobertura no estudo 2.0: {estudo20}")
    if not dados30:
        raise ValueError(f"Não encontrei abas de cobertura no estudo 3.0: {estudo30}")
    todas = sorted(set(dados20) | set(dados30), key=normalizar_texto)
    usados = {"Resumo Categorias"}
    mapas_abas = {cat: nome_aba_seguro(cat, usados) for cat in todas}
    comparacoes = {}
    resumos = []
    total = len(todas)
    for i, cat in enumerate(todas, start=1):
        log(f"Comparando categoria/PROD {i}/{total}: {cat}", 30 + (i / max(total, 1)) * 45)
        d20 = dados20.get(cat, {})
        d30 = dados30.get(cat, {})
        df20 = d20.get("mensal", pd.DataFrame(columns=["chave", "data", "Sell-in", "Sell-out", "Cobertura"]))
        df30 = d30.get("mensal", pd.DataFrame(columns=["chave", "data", "Sell-in", "Sell-out", "Cobertura"]))
        mensal = _montar_comparacao_mensal_avancada(df20, df30)
        uf = _montar_comparacao_uf_avancada(d20.get("uf", pd.DataFrame()), d30.get("uf", pd.DataFrame()))
        fabricante = d30.get("fabricante") or d20.get("fabricante") or ""
        if gerar_top20_sku_canal_uf:
            contribuicao_skus, info_contribuicao = _montar_contribuicao_sku_categoria(base_contrib20, base_contrib30, cat)
        else:
            contribuicao_skus = pd.DataFrame()
            info_contribuicao = ""
        comparacoes[cat] = {
            "mensal": mensal,
            "uf": uf,
            "fabricante": fabricante,
            "skus_excluidos20": d20.get("skus_excluidos", pd.DataFrame()),
            "skus_excluidos30": d30.get("skus_excluidos", pd.DataFrame()),
            "contribuicao_skus": contribuicao_skus,
            "info_contribuicao": info_contribuicao,
        }
        status = "Em ambos" if cat in dados20 and cat in dados30 else ("Só 2.0" if cat in dados20 else "Só 3.0")
        linha20 = _linha_resumo_original_por_categoria(estudo20, cat)
        linha30 = _linha_resumo_original_por_categoria(estudo30, cat)
        resumos.append(_montar_resumo_base_preservado(cat, linha20, linha30, mapas_abas[cat], status))
    resumo = pd.DataFrame(resumos)
    if not resumo.empty:
        resumo = resumo.sort_values(["Status Comparação", "Categoria/PROD"], ascending=[True, True])
    # Acrescenta no Resumo Categorias as mesmas colunas do estudo individual,
    # com os valores 2.0 e 3.0 lado a lado.
    resumo = _montar_resumo_comparacao_completo(resumo, estudo20, estudo30)
    log("Gerando Excel da comparação...", 82)
    with pd.ExcelWriter(saida, engine="xlsxwriter", engine_kwargs={"options": {"nan_inf_to_errors": True}}) as writer:
        workbook = writer.book
        azul_escuro = "#1F4E78"
        azul_claro = "#D9EAF7"
        fmt_titulo = workbook.add_format({"bold": True, "font_size": 15, "font_color": "white", "bg_color": azul_escuro, "align": "center", "valign": "vcenter", "border": 1})
        fmt_header = workbook.add_format({"bold": True, "bg_color": azul_claro, "border": 1, "align": "center", "valign": "vcenter"})
        fmt_text = workbook.add_format({"border": 1})
        fmt_num = workbook.add_format({"num_format": "#,##0.0", "border": 1})
        fmt_pct = workbook.add_format({"num_format": "0.0%", "border": 1})
        fmt_link = workbook.add_format({"font_color": "blue", "underline": 1, "border": 1})
        if output_options.get("resumo_categorias", True):
            ws = workbook.add_worksheet("Resumo Categorias")
            writer.sheets["Resumo Categorias"] = ws
            ws.merge_range("A1:L1", "Comparação de Estudo de Cobertura", fmt_titulo)
            ws.write("A2", "Estudo/Sell-out 2.0", fmt_header)
            ws.write("B2", nome_arquivo_curto(estudo20), fmt_text)
            ws.write("A3", "Estudo/Sell-out 3.0", fmt_header)
            ws.write("B3", nome_arquivo_curto(estudo30), fmt_text)
            ws.write("A4", "Opções de saída", fmt_header)
            ws.write("B4", opcoes_saida_para_parametros(output_options), fmt_text)
            resumo = limpar_colunas_resumo_categorias(resumo)
            resumo = remover_colunas_duplicadas(resumo)
            export_cols = [c for c in resumo.columns if c != "sheet"]
            start = 5
            for j, col in enumerate(export_cols):
                ws.write(start, j, col, fmt_header)
            for i, row in resumo.reset_index(drop=True).iterrows():
                r = start + 1 + i
                for j, col in enumerate(export_cols):
                    val = row[col]
                    if col == "Abrir aba":
                        ws.write_url(r, j, f"internal:'{row['sheet']}'!A1", fmt_link, string="Abrir")
                    elif col in {"Diferença % Sell-out", "Cobertura 2.0", "Cobertura 3.0", "Diferença Cobertura"}:
                        write_number_ou_branco(ws, r, j, val, fmt_pct)
                    elif isinstance(val, (int, float, np.integer, np.floating)) and not pd.isna(val):
                        ws.write_number(r, j, float(val), fmt_num)
                    else:
                        ws.write(r, j, "" if pd.isna(val) else str(val), fmt_text)
            ws.set_column("A:A", 12)
            ws.set_column("B:C", 28)
            ws.set_column("D:L", 15)

        if output_options.get("abas_categorias", True):
            for cat, obj in comparacoes.items():
                _escrever_aba_comparacao_categoria(
                    writer, mapas_abas[cat], cat, obj["mensal"],
                    uf=obj["uf"],
                    fabricante=obj.get("fabricante", ""),
                    skus_excluidos20=obj.get("skus_excluidos20", pd.DataFrame()),
                    skus_excluidos30=obj.get("skus_excluidos30", pd.DataFrame()),
                    contribuicao_skus=obj.get("contribuicao_skus", pd.DataFrame()),
                    info_contribuicao=obj.get("info_contribuicao", ""),
                    gerar_top20_contribuicao=gerar_top20_sku_canal_uf,
                )

        if output_options.get("graficos_cobertura", True):
            itens_grafico = [{"categoria": cat, "mensal": obj.get("mensal", pd.DataFrame())} for cat, obj in comparacoes.items()]
            escrever_graficos_cobertura(writer, itens_grafico, tipo="Categoria/PROD", comparacao=True)

        # Abas auxiliares equivalentes ao estudo individual, sempre com colunas 2.0 e 3.0 lado a lado.
        # Isso também vale para o modo "Estudo de Cobertura com 2 Sell-out", pois ele gera
        # os dois estudos temporários e depois passa por esta mesma rotina de comparação.
        if output_options.get("abas_auxiliares_comparacao", True):
            _escrever_abas_auxiliares_comparacao(writer, estudo20, estudo30, output_options=output_options)

        garantir_aba_info_se_vazio(writer, workbook, "Comparação de Estudo de Cobertura", output_options)
    log(f"Comparação gerada com sucesso: {saida}", 100)
    return saida


def executar_estudo_dois_sellouts(
    sellin_path: str,
    sellout20_path: str,
    sellout30_path: str,
    saida_path: str,
    metrica: str,
    nivel: str,
    fabricante_filtro: str = "",
    congelado_path: str = "",
    log_callback=None,
    gerar_top20_sku_canal_uf: bool = False,
    output_options: Optional[Dict[str, object]] = None,
) -> Path:
    """Gera dois estudos temporários usando o mesmo Sell-in e compara os resultados em um único arquivo final."""
    def log(msg: str, pct: Optional[float] = None):
        if log_callback:
            try:
                log_callback(msg, pct)
            except TypeError:
                log_callback(msg)
        else:
            print(msg if pct is None else f"{int(round(pct))}% - {msg}")

    output_options = normalizar_opcoes_saida(output_options)
    if gerar_top20_sku_canal_uf:
        output_options["top20_sku_canal_uf"] = True
    gerar_top20_sku_canal_uf = bool(output_options.get("top20_sku_canal_uf", False))

    # Os estudos temporários precisam manter todas as abas técnicas para que a comparação consiga ler a base,
    # mesmo quando algumas saídas estão desativadas no arquivo final do usuário.
    opcoes_temporarias = normalizar_opcoes_saida()
    opcoes_temporarias.update({
        "resumo_categorias": True,
        "abas_categorias": True,
        "base_skus": True,
        "base_contribuicao_sellout": True,
        "skus_por_categoria": True,
        "crosschecks": True,
        "parametros": True,
        "descricao_calculos": True,
        "avisos": True,
        "abas_auxiliares_comparacao": True,
        "top20_sku_canal_uf": gerar_top20_sku_canal_uf,
    })

    nivel_norm = str(nivel or "").upper()
    congelado_path = str(congelado_path or "").strip()
    if congelado_path and nivel_norm != "CATEGORIA":
        log(
            f"Congelado informado, mas será ignorado porque a regra escolhida foi {nivel_norm}. "
            "O Congelado só define categoria quando a opção é CATEGORIA.",
            3,
        )
    elif congelado_path and nivel_norm == "CATEGORIA":
        log(
            "Congelado opcional informado para o modo com 2 Sell-outs. "
            "A Categoria dos dois estudos temporários será definida por SKU/EAN no Congelado, sem fallback para Total Fabricante.",
            3,
        )

    saida = Path(saida_path)
    with tempfile.TemporaryDirectory(prefix="estudo_cobertura_2out_") as tmpdir:
        tmp = Path(tmpdir)
        estudo20 = tmp / "estudo_sellout_2_0.xlsx"
        estudo30 = tmp / "estudo_sellout_3_0.xlsx"

        log("Gerando estudo temporário do Sell-out 2.0...", 5)
        executar_estudo(
            sellin_path, sellout20_path, estudo20, metrica, nivel, fabricante_filtro,
            congelado_path=congelado_path if nivel_norm == "CATEGORIA" else "",
            log_callback=lambda msg, pct=None: log(f"2.0 - {msg}", 5 + (float(pct or 0) * 0.38)),
            output_options=opcoes_temporarias,
        )
        log("Gerando estudo temporário do Sell-out 3.0...", 44)
        executar_estudo(
            sellin_path, sellout30_path, estudo30, metrica, nivel, fabricante_filtro,
            congelado_path=congelado_path if nivel_norm == "CATEGORIA" else "",
            log_callback=lambda msg, pct=None: log(f"3.0 - {msg}", 44 + (float(pct or 0) * 0.38)),
            output_options=opcoes_temporarias,
        )
        log("Comparando os estudos temporários...", 84)
        gerar_comparacao_estudos(
            estudo20, estudo30, saida,
            log_callback=lambda msg, pct=None: log(msg, 84 + (float(pct or 0) * 0.16)),
            gerar_top20_sku_canal_uf=gerar_top20_sku_canal_uf,
            output_options=output_options,
        )
    log(f"Arquivo final gerado com sucesso: {saida}", 100)
    return saida


def executar_estudo_com_loading(cfg: Dict[str, str]) -> bool:
    """Executa o estudo com janela de loading, sem travar a interface."""
    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except Exception:
        if cfg.get("modo") == "comparacao":
            gerar_comparacao_estudos(
                cfg["estudo20"], cfg["estudo30"], cfg["saida"],
                gerar_top20_sku_canal_uf=bool(cfg.get("gerar_top20_sku_canal_uf", False)),
                output_options=cfg.get("output_options"),
            )
        elif cfg.get("modo") == "estudo2out":
            executar_estudo_dois_sellouts(
                cfg["sellin"], cfg["sellout20"], cfg["sellout30"], cfg["saida"],
                cfg["metrica"], cfg["nivel"], cfg.get("fabricante", ""), cfg.get("congelado", ""),
                gerar_top20_sku_canal_uf=bool(cfg.get("gerar_top20_sku_canal_uf", False)),
                output_options=cfg.get("output_options"),
            )
        elif cfg.get("modo") == "dash":
            executar_cobertura_dash(
                cfg["sellin"], cfg["sellout_dash"], cfg["sku"], cfg["arquivo_fabricante"], cfg["saida"],
                cfg["metrica"], cfg["nivel"], cfg.get("fabricante", ""), vendas_sku_path=cfg.get("vendas_sku") or None,
                output_options=cfg.get("output_options"),
            )
        else:
            executar_estudo(
                cfg["sellin"], cfg["sellout"], cfg["saida"], cfg["metrica"], cfg["nivel"], cfg.get("fabricante", ""), cfg.get("congelado", ""),
                output_options=cfg.get("output_options"),
            )
        return True

    root = tk.Tk()
    comparacao = cfg.get("modo") == "comparacao"
    estudo2out = cfg.get("modo") == "estudo2out"
    dash = cfg.get("modo") == "dash"
    root.title("Gerando comparação de cobertura" if (comparacao or estudo2out) else ("Gerando Cobertura Dash" if dash else "Gerando estudo de cobertura"))
    root.resizable(False, False)
    try:
        root.option_add("*Font", "{Segoe UI} 10")
    except Exception:
        pass

    frame = ttk.Frame(root, padding=22)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Gerando comparação de cobertura" if (comparacao or estudo2out) else ("Gerando Cobertura Dash" if dash else "Gerando estudo de cobertura"), font=("Segoe UI", 13, "bold")).pack(anchor="w")
    status_var = tk.StringVar(value="Preparando comparação..." if (comparacao or estudo2out) else ("Preparando Cobertura Dash..." if dash else "Preparando leitura dos arquivos..."))
    ttk.Label(frame, textvariable=status_var, wraplength=470).pack(anchor="w", pady=(8, 12))

    progresso_var = tk.DoubleVar(value=0)
    barra = ttk.Progressbar(frame, mode="determinate", maximum=100, variable=progresso_var, length=470)
    barra.pack(fill="x")
    pct_var = tk.StringVar(value="0%")
    ttk.Label(frame, textvariable=pct_var, foreground="#555555").pack(anchor="e", pady=(4, 0))
    ttk.Label(
        frame,
        text="O percentual é atualizado por etapas principais, sem varrer linha por linha para não deixar o processo mais lento.",
        foreground="#555555",
        wraplength=470,
    ).pack(anchor="w", pady=(8, 0))
    try:
        root.update_idletasks()
        largura = 540
        altura = 170
        x = int((root.winfo_screenwidth() - largura) / 2)
        y = int((root.winfo_screenheight() - altura) / 2)
        root.geometry(f"{largura}x{altura}+{x}+{y}")
    except Exception:
        pass

    fila_logs = queue.Queue()
    fila_resultado = queue.Queue()

    def log(msg: str, pct: Optional[float] = None):
        fila_logs.put((str(msg), pct))

    def worker():
        try:
            if comparacao:
                saida = gerar_comparacao_estudos(
                    cfg["estudo20"],
                    cfg["estudo30"],
                    cfg["saida"],
                    log_callback=log,
                    gerar_top20_sku_canal_uf=bool(cfg.get("gerar_top20_sku_canal_uf", False)),
                    output_options=cfg.get("output_options"),
                )
            elif estudo2out:
                saida = executar_estudo_dois_sellouts(
                    cfg["sellin"],
                    cfg["sellout20"],
                    cfg["sellout30"],
                    cfg["saida"],
                    cfg["metrica"],
                    cfg["nivel"],
                    cfg.get("fabricante", ""),
                    cfg.get("congelado", ""),
                    log_callback=log,
                    gerar_top20_sku_canal_uf=bool(cfg.get("gerar_top20_sku_canal_uf", False)),
                    output_options=cfg.get("output_options"),
                )
            elif dash:
                saida = executar_cobertura_dash(
                    cfg["sellin"],
                    cfg["sellout_dash"],
                    cfg["sku"],
                    cfg["arquivo_fabricante"],
                    cfg["saida"],
                    cfg["metrica"],
                    cfg["nivel"],
                    cfg.get("fabricante", ""),
                    vendas_sku_path=cfg.get("vendas_sku") or None,
                    log_callback=log,
                    output_options=cfg.get("output_options"),
                )
            else:
                saida = executar_estudo(
                    cfg["sellin"],
                    cfg["sellout"],
                    cfg["saida"],
                    cfg["metrica"],
                    cfg["nivel"],
                    cfg.get("fabricante", ""),
                    cfg.get("congelado", ""),
                    log_callback=log,
                    output_options=cfg.get("output_options"),
                )
            fila_resultado.put(("ok", saida))
        except Exception as exc:
            fila_resultado.put(("erro", exc))

    threading.Thread(target=worker, daemon=True).start()

    finalizado = {"valor": False}

    def bloquear_fechamento():
        if not finalizado["valor"]:
            messagebox.showinfo("Processamento em andamento", "Aguarde a geração do estudo terminar.")
        else:
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", bloquear_fechamento)

    def verificar():
        while True:
            try:
                item = fila_logs.get_nowait()
                if isinstance(item, tuple):
                    msg, pct = item
                else:
                    msg, pct = str(item), None
                status_var.set(msg)
                if pct is not None:
                    pct_num = max(0, min(100, float(pct)))
                    progresso_var.set(pct_num)
                    pct_var.set(f"{int(round(pct_num))}%")
            except queue.Empty:
                break

        try:
            status, payload = fila_resultado.get_nowait()
        except queue.Empty:
            root.after(150, verificar)
            return

        finalizado["valor"] = True
        if status == "ok":
            progresso_var.set(100)
            pct_var.set("100%")
            status_var.set("Concluído.")
            messagebox.showinfo("Estudo gerado", f"Arquivo gerado com sucesso:\n{payload}")
            root.destroy()
        else:
            status_var.set("Erro ao gerar estudo.")
            messagebox.showerror("Erro", str(payload))
            root.destroy()

    root.after(150, verificar)
    root.mainloop()
    return finalizado["valor"]


def main():
    args = obter_args()
    output_options = opcoes_saida_de_args(args)

    if getattr(args, "modo", "estudo") == "comparacao":
        precisa_gui = not (args.estudo20 and args.estudo30 and args.saida)
        if precisa_gui:
            cfg = obter_configuracao_gui(args)
            if not cfg:
                raise ValueError("Operação cancelada ou interface gráfica indisponível.")
            executar_estudo_com_loading(cfg)
            return
        gerar_comparacao_estudos(
            args.estudo20, args.estudo30, args.saida,
            gerar_top20_sku_canal_uf=bool(getattr(args, "gerar_top20_sku_canal_uf", False)),
            output_options=output_options,
        )
        return

    if getattr(args, "modo", "estudo") == "estudo2out":
        precisa_gui = not (args.sellin and args.sellout and args.sellout2 and args.saida and args.metrica and args.nivel)
        if precisa_gui:
            cfg = obter_configuracao_gui(args)
            if not cfg:
                raise ValueError("Operação cancelada ou interface gráfica indisponível.")
            executar_estudo_com_loading(cfg)
            return
        executar_estudo_dois_sellouts(
            args.sellin,
            args.sellout,
            args.sellout2,
            args.saida,
            args.metrica.lower(),
            args.nivel.upper(),
            args.fabricante or "",
            getattr(args, "congelado_estudo", None) or getattr(args, "arquivo_fabricante", None) or "",
            gerar_top20_sku_canal_uf=bool(getattr(args, "gerar_top20_sku_canal_uf", False)),
            output_options=output_options,
        )
        return

    if getattr(args, "modo", "estudo") == "dash":
        vendas_uf_arg = args.vendas_uf or args.sellout
        precisa_gui = not (args.sellin and vendas_uf_arg and args.sku and args.saida and args.metrica and args.nivel)
        if precisa_gui:
            cfg = obter_configuracao_gui(args)
            if not cfg:
                raise ValueError("Operação cancelada ou interface gráfica indisponível.")
            executar_estudo_com_loading(cfg)
            return
        executar_cobertura_dash(
            args.sellin,
            vendas_uf_arg,
            args.sku,
            args.arquivo_fabricante or "",
            args.saida,
            args.metrica.lower(),
            args.nivel.upper(),
            args.fabricante or "",
            vendas_sku_path=args.vendas_sku or None,
            output_options=output_options,
        )
        return

    if getattr(args, "sellout2", None):
        # Modo unificado: Estudo de Cobertura + segundo Sell-out opcional.
        # Quando --sellout2 é informado, o primeiro Sell-out é tratado como 2.0
        # e o segundo como 3.0, reaproveitando o fluxo otimizado de comparação.
        precisa_gui = not (args.sellin and args.sellout and args.sellout2 and args.saida and args.metrica and args.nivel)
        if precisa_gui:
            cfg = obter_configuracao_gui(args)
            if not cfg:
                raise ValueError("Operação cancelada ou interface gráfica indisponível.")
            executar_estudo_com_loading(cfg)
            return
        executar_estudo_dois_sellouts(
            args.sellin,
            args.sellout,
            args.sellout2,
            args.saida,
            args.metrica.lower(),
            args.nivel.upper(),
            args.fabricante or "",
            getattr(args, "congelado_estudo", None) or getattr(args, "arquivo_fabricante", None) or "",
            gerar_top20_sku_canal_uf=bool(getattr(args, "gerar_top20_sku_canal_uf", False)),
            output_options=output_options,
        )
        return

    precisa_gui = not (args.sellin and args.sellout and args.saida and args.metrica and args.nivel)
    if precisa_gui:
        cfg = obter_configuracao_gui(args)
        if not cfg:
            raise ValueError("Operação cancelada ou interface gráfica indisponível.")
        executar_estudo_com_loading(cfg)
        return

    executar_estudo(
        args.sellin,
        args.sellout,
        args.saida,
        args.metrica.lower(),
        args.nivel.upper(),
        args.fabricante or "",
        getattr(args, "congelado_estudo", None) or getattr(args, "arquivo_fabricante", None) or "",
        output_options=output_options,
    )


# ============================================================
# Override v9 - Leitura rápida do Congelado opcional
# ============================================================
# Motivo: a versão anterior fazia uma detecção flexível pesada no Congelado:
# lia abas inteiras, depois relia sem cabeçalho e testava até 150 possíveis linhas
# de cabeçalho. Em arquivos grandes, isso deixava a barra parada em
# "Lendo Congelado opcional para Categoria...".
#
# Esta versão localiza o cabeçalho por streaming com openpyxl e depois lê somente
# as colunas úteis: Código Barras SKU/EAN, Categoria congelada, Fabricante e
# colunas auxiliares de volume/valor, já podendo filtrar pelo Fabricante durante
# a leitura.

_ler_congelado_flexivel_lento_v8 = ler_congelado_flexivel
_ler_mapa_congelado_categoria_v8 = ler_mapa_congelado_categoria


def _score_nome_codigo_congelado_v9(nome_coluna) -> int:
    col_norm = normalizar_texto(nome_coluna)
    if not col_norm:
        return 0
    prioridade_exata = {
        normalizar_texto("Código Barras SKU"): 1000,
        normalizar_texto("Codigo Barras SKU"): 1000,
        normalizar_texto("Código_Barras_SKU"): 1000,
        normalizar_texto("Codigo_Barras_SKU"): 1000,
        normalizar_texto("Código de Barras SKU"): 980,
        normalizar_texto("Codigo de Barras SKU"): 980,
        normalizar_texto("CODIGO_BARRAS_SKU"): 980,
        normalizar_texto("CODIGO BARRAS SKU"): 980,
        normalizar_texto("COD_BARRAS_SKU"): 970,
        normalizar_texto("COD BARRAS SKU"): 970,
        normalizar_texto("CODIGO_BARRAS_CONTENIDO"): 950,
        normalizar_texto("CODIGO BARRAS CONTENIDO"): 950,
        normalizar_texto("Código Barras Contenido"): 950,
        normalizar_texto("Codigo Barras Contenido"): 950,
        normalizar_texto("Código Barras Conteúdo"): 950,
        normalizar_texto("Codigo Barras Conteudo"): 950,
        normalizar_texto("EAN"): 900,
        normalizar_texto("SKU/EAN"): 890,
        normalizar_texto("EAN/SKU"): 890,
        normalizar_texto("SKU EAN"): 880,
        normalizar_texto("EAN SKU"): 880,
        normalizar_texto("SKU"): 600,
    }
    if col_norm in prioridade_exata:
        return prioridade_exata[col_norm]
    tokens = set(col_norm.split())
    termos_descritivos = {
        "nome", "nombre", "descricao", "descrição", "descripcion", "desc",
        "categoria", "fabricante", "marca", "fornecedor", "proveedor",
        "preco", "precio", "price", "valor", "vendas", "volume",
        "conteudo", "contenido", "qtd", "cant", "medida", "analise", "análise",
    }
    if ("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens) and "sku" in tokens:
        return 920
    if ("codigo" in tokens or "cod" in tokens) and ("barras" in tokens or "barra" in tokens):
        return 850
    if "ean" in tokens and not (tokens & termos_descritivos):
        return 800
    if col_norm == "sku":
        return 600
    if "sku" in tokens and not (tokens & termos_descritivos):
        return 520
    return 0


def _score_nome_categoria_congelado_v9(nome_coluna, alternativas_cat: List[str]) -> int:
    col_norm = normalizar_texto(nome_coluna)
    if not col_norm:
        return 0
    prioridade_exata = [
        "Categoría congelada ScannMarket",
        "Categoria congelada ScannMarket",
        "Categoria atual Data Excellence",
        "Est Mer 6 (Categoria)",
        "Est Mer 6 Categoria",
        "Categoria Scanntech",
        "Categoria SM",
        "CATEGORIA SCANN",
        "Categoria",
        "CATEGORIA",
        "PROD",
    ]
    alts = prioridade_exata + list(alternativas_cat or [])
    mapa_scores = {normalizar_texto(a): 1000 - i for i, a in enumerate(alts) if normalizar_texto(a)}
    if col_norm in mapa_scores:
        return mapa_scores[col_norm]
    score = 0
    for alt in alts:
        alt_norm = normalizar_texto(alt)
        if alt_norm and (alt_norm in col_norm or col_norm in alt_norm):
            score = max(score, 500 + len(alt_norm))
    if "categoria" in col_norm:
        score += 200
    if "congelada" in col_norm or "congelado" in col_norm:
        score += 150
    if "scannmarket" in col_norm or "scann" in col_norm:
        score += 100
    if "data excellence" in col_norm:
        score += 100
    return score


def _score_nome_fabricante_congelado_v9(nome_coluna) -> int:
    col_norm = normalizar_texto(nome_coluna)
    if not col_norm:
        return 0
    prioridade = [
        "Fabricante SKU", "Fabricante", "Fabricante do SKU", "Fabricante del SKU",
        "Proveedor SKU", "Proveedor", "Fornecedor SKU", "Fornecedor",
        "Fabricante Marca", "Nome Fabricante", "Nombre Fabricante",
    ]
    for i, alt in enumerate(prioridade):
        if col_norm == normalizar_texto(alt):
            return 1000 - i
    tokens = set(col_norm.split())
    termos_ruins = {
        "categoria", "congelada", "scannmarket", "ean", "codigo", "barras",
        "marca", "nome", "nombre", "descricao", "descripcion", "volume", "valor",
        "venda", "vendas", "qtd", "quantidade", "canal", "uf", "data", "mes",
    }
    score = 0
    if "fabricante" in tokens or "fabricante" in col_norm:
        score += 500
    if "proveedor" in tokens or "fornecedor" in tokens:
        score += 450
    if "sku" in tokens:
        score += 50
    if tokens & termos_ruins:
        score -= 200
    return max(score, 0)


def _score_nome_auxiliar_congelado_v9(nome_coluna) -> int:
    """Colunas usadas só para desempatar SKU duplicado no mapa do Congelado."""
    col_norm = normalizar_texto(nome_coluna)
    if not col_norm:
        return 0
    alts = [
        "Cant Vta", "Quantidade Venda", "Volume", "Vendas em volume", "Qtd_de_Vendas", "Qtd de Vendas",
        "Imp Vta (Ult.24 Meses)", "Imp Vta Ult 24 Meses", "Imp Vta", "Valor", "Vendas em valor",
    ]
    for i, alt in enumerate(alts):
        alt_norm = normalizar_texto(alt)
        if col_norm == alt_norm:
            return 500 - i
        if alt_norm and alt_norm in col_norm:
            return 300
    return 0


def _headers_congelado_v9(valores_linha) -> List[str]:
    valores = list(valores_linha or [])
    while valores and str(valores[-1]).strip() in {"", "nan", "None"}:
        valores.pop()
    if not valores:
        return []
    return nomes_unicos(valores)


def _detectar_layout_congelado_excel_v9(caminho: Path, alternativas_cat: List[str], max_linhas_scan: int = 80):
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        nomes_abas = wb.sheetnames
        preferidas_norm = [normalizar_texto(x) for x in ["Base Congelada", "Congelado", "Fabricante", "Dados", "Planilha1", "Sheet1"]]
        abas_ordenadas = sorted(
            nomes_abas,
            key=lambda n: 0 if any(p and p in normalizar_texto(n) for p in preferidas_norm) else 1,
        )
        melhor = None
        for aba in abas_ordenadas:
            ws = wb[aba]
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_linhas_scan, values_only=True), start=1):
                headers = _headers_congelado_v9(row)
                if not headers:
                    continue
                scores_codigo = [_score_nome_codigo_congelado_v9(h) for h in headers]
                scores_cat = [_score_nome_categoria_congelado_v9(h, alternativas_cat) for h in headers]
                scores_fab = [_score_nome_fabricante_congelado_v9(h) for h in headers]
                max_codigo = max(scores_codigo) if scores_codigo else 0
                max_cat = max(scores_cat) if scores_cat else 0
                max_fab = max(scores_fab) if scores_fab else 0
                if max_codigo <= 0 or max_cat <= 0:
                    continue
                # Penaliza linhas muito vazias/curtas e premia o layout exato do usuário.
                score = max_codigo * 10 + max_cat * 8 + max_fab + min(len(headers), 80)
                if any(normalizar_texto(h) == normalizar_texto("Código Barras SKU") for h in headers):
                    score += 5000
                if any(normalizar_texto(h) == normalizar_texto("Categoría congelada ScannMarket") for h in headers):
                    score += 5000
                candidato = (score, aba, idx, headers)
                if melhor is None or candidato[0] > melhor[0]:
                    melhor = candidato
        return melhor
    finally:
        wb.close()


def _ler_congelado_excel_stream_v9(
    caminho: Path,
    alternativas_cat: List[str],
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> pd.DataFrame:
    if avisos is None:
        avisos = []
    layout = _detectar_layout_congelado_excel_v9(caminho, alternativas_cat)
    if not layout:
        return pd.DataFrame()

    score, aba, header_row, headers = layout
    idx_codigo = [i for i, h in enumerate(headers) if _score_nome_codigo_congelado_v9(h) > 0]
    idx_cat = [i for i, h in enumerate(headers) if _score_nome_categoria_congelado_v9(h, alternativas_cat) > 0]
    idx_fab = [i for i, h in enumerate(headers) if _score_nome_fabricante_congelado_v9(h) > 0]
    idx_aux = [i for i, h in enumerate(headers) if _score_nome_auxiliar_congelado_v9(h) > 0]

    if not idx_codigo or not idx_cat:
        return pd.DataFrame()

    # Mantém somente colunas úteis. Isso evita carregar centenas de colunas do Congelado.
    indices = sorted(set(idx_codigo + idx_cat + idx_fab + idx_aux))
    nomes = [headers[i] for i in indices]
    pos_fab_rel = None
    fab_key = normalizar_texto(fabricante_filtro)
    if fabricante_filtro and idx_fab:
        melhor_fab_idx = max(idx_fab, key=lambda i: _score_nome_fabricante_congelado_v9(headers[i]))
        if melhor_fab_idx in indices:
            pos_fab_rel = indices.index(melhor_fab_idx)

    dados = []
    linhas_lidas = 0
    linhas_mantidas = 0
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[aba]
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            linhas_lidas += 1
            valores = []
            for i in indices:
                valores.append(row[i] if i < len(row) else None)

            if not any(v is not None and str(v).strip() not in {"", "nan", "None"} for v in valores):
                continue

            # Filtro por fabricante durante a leitura, quando a coluna foi encontrada.
            if fab_key and pos_fab_rel is not None:
                valor_fab = normalizar_texto(valores[pos_fab_rel])
                if not (valor_fab == fab_key or fab_key in valor_fab):
                    continue

            dados.append(valores)
            linhas_mantidas += 1
    finally:
        wb.close()

    df = pd.DataFrame(dados, columns=nomes)
    df = remover_linhas_colunas_vazias(limpar_bom_dataframe(df))
    msg_filtro = ""
    if fabricante_filtro and pos_fab_rel is not None:
        msg_filtro = f"; filtro de Fabricante aplicado durante a leitura: '{fabricante_filtro}', linhas mantidas: {linhas_mantidas} de {linhas_lidas}"
    elif fabricante_filtro:
        msg_filtro = "; Fabricante informado, mas a coluna de fabricante não foi detectada nesta leitura rápida"

    avisos.append(
        f"Congelado lido em modo rápido: aba '{aba}', cabeçalho na linha {header_row}, "
        f"colunas úteis lidas: {len(nomes)} de {len(headers)}{msg_filtro}.".replace(",", ".")
    )
    return df


def ler_congelado_flexivel(
    caminho: Path,
    alternativas_cat: List[str],
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> pd.DataFrame:
    """Versão rápida da leitura do Congelado opcional."""
    if avisos is None:
        avisos = []
    caminho = Path(caminho)

    # CSV/TXT geralmente já é mais simples. Mantém a leitura antiga para evitar quebrar formatos variados.
    if eh_csv(caminho):
        return _ler_congelado_flexivel_lento_v8(caminho, alternativas_cat, avisos)

    try:
        rapido = _ler_congelado_excel_stream_v9(
            caminho,
            alternativas_cat,
            avisos=avisos,
            fabricante_filtro=fabricante_filtro,
        )
        if rapido is not None and not rapido.empty:
            return rapido
        avisos.append("Leitura rápida do Congelado não encontrou layout válido; tentando fallback compatível da versão anterior.")
    except Exception as exc:
        avisos.append(f"Leitura rápida do Congelado falhou ({exc}); tentando fallback compatível da versão anterior.")

    # Fallback só para arquivos pequenos. Em arquivos grandes, o método antigo é justamente o que travava.
    try:
        tamanho_mb = Path(caminho).stat().st_size / (1024 * 1024)
    except Exception:
        tamanho_mb = 999
    if tamanho_mb <= 30:
        return _ler_congelado_flexivel_lento_v8(caminho, alternativas_cat, avisos)

    avisos.append(
        f"Congelado '{Path(caminho).name}' tem {tamanho_mb:.1f} MB e não será lido pelo fallback lento. "
        "Confira se o cabeçalho possui 'Código Barras SKU' e 'Categoría congelada ScannMarket'."
    )
    return pd.DataFrame()


def ler_mapa_congelado_categoria(
    congelado_path: str | Path,
    nivel_prod: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Versão v9: passa o fabricante para a leitura rápida do Congelado, permitindo
    filtrar durante a própria leitura do arquivo.
    """
    if avisos is None:
        avisos = []

    caminho = Path(str(congelado_path or "").strip())
    if not str(caminho):
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {caminho}")

    alternativas_cat, origem = _alternativas_categoria_congelado(nivel_prod)
    raw = ler_congelado_flexivel(caminho, alternativas_cat, avisos, fabricante_filtro=fabricante_filtro)
    if raw is None or raw.empty:
        avisos.append(f"Congelado opcional '{caminho.name}' foi lido, mas está vazio ou sem cabeçalho útil.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    # Mantém a validação/filtro oficial. Se a leitura rápida já filtrou, aqui apenas confirma.
    raw = filtrar_congelado_por_fabricante(
        raw,
        fabricante_filtro=fabricante_filtro,
        avisos=avisos,
        contexto="Congelado opcional do Estudo de Cobertura",
        obrigatorio=bool(str(fabricante_filtro or "").strip()),
    )

    if raw is None or raw.empty:
        avisos.append(f"Congelado opcional '{caminho.name}' ficou vazio após o filtro de fabricante.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    colunas_codigo = localizar_colunas_codigo_congelado(raw)
    c_cat = localizar_coluna_categoria_congelado(raw, alternativas_cat)
    c_vol = localizar_coluna(raw, ["Cant Vta", "Quantidade Venda", "Volume", "Vendas em volume", "Qtd_de_Vendas", "Qtd de Vendas"], obrigatoria=False)
    c_val = localizar_coluna(raw, ["Imp Vta (Ult.24 Meses)", "Imp Vta Ult 24 Meses", "Imp Vta", "Valor", "Vendas em valor"], obrigatoria=False)

    if not colunas_codigo:
        raise ValueError(
            "No Congelado opcional, não encontrei coluna de SKU/EAN/Código de Barras preenchida.\n"
            "Para o seu layout, a coluna esperada é 'Código Barras SKU'.\n"
            f"Alternativas aceitas: {', '.join(EAN_SKU_ALTERNATIVAS_AMPLAS)}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(raw.columns)}"
        )
    if not c_cat:
        raise ValueError(
            f"No Congelado opcional, não encontrei a coluna de categoria para a regra {normalizar_nivel_prod(nivel_prod)}.\n"
            "Para o seu layout, a coluna esperada é 'Categoría congelada ScannMarket'.\n"
            f"Alternativas testadas: {', '.join(alternativas_cat)}.\n"
            f"Colunas disponíveis: {resumo_colunas_disponiveis(raw.columns)}"
        )

    partes_codigo = []
    for c_cod in colunas_codigo:
        temp = pd.DataFrame({
            "ean": raw[c_cod].map(ean_texto),
            "categoria_map_prod": raw[c_cat].astype(str).str.strip(),
            "_rank_congelado": (
                (limpar_coluna_numerica_vetorizada(raw[c_vol]) if c_vol else 0.0)
                + (limpar_coluna_numerica_vetorizada(raw[c_val]) if c_val else 0.0) / 1_000_000_000
            ),
            "_coluna_codigo_congelado": c_cod,
        })
        partes_codigo.append(temp)

    base = pd.concat(partes_codigo, ignore_index=True) if partes_codigo else pd.DataFrame(columns=["ean", "categoria_map_prod", "_rank_congelado", "_coluna_codigo_congelado"])
    base["categoria_key_map_prod"] = base["categoria_map_prod"].map(normalizar_categoria)
    base = base[(base["ean"] != "") & (base["categoria_key_map_prod"] != "")].copy()

    if base.empty:
        avisos.append(f"Congelado opcional '{caminho.name}' foi lido, mas nenhum SKU ficou com categoria válida em '{c_cat}'.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    mapa = (
        base.sort_values(["ean", "_rank_congelado"], ascending=[True, False])
        .drop_duplicates("ean", keep="first")
        [["ean", "categoria_map_prod", "categoria_key_map_prod"]]
    )

    colunas_codigo_txt = ", ".join(colunas_codigo) if colunas_codigo else "não identificada"
    avisos.append(
        f"Congelado opcional aplicado no Estudo de Cobertura: {len(mapa):,} códigos de SKU/EAN mapeados usando {origem} "
        f"pela coluna de categoria '{c_cat}' e coluna(s) de código {colunas_codigo_txt}.".replace(",", ".")
    )
    return mapa, avisos


# ============================================================
# Override v10 - Congelado direto para mapa, sem carregar DataFrame gigante
# ============================================================
# A v9 já localizava o cabeçalho por streaming, mas ainda montava um DataFrame
# com as linhas úteis do Congelado e depois fazia outro filtro/normalização.
# Em Congelados grandes, isso ainda podia deixar a tela parada em 12%.
#
# Esta versão monta diretamente o mapa EAN -> Categoria durante a leitura:
# - lê somente Código Barras SKU/EAN, Categoría congelada ScannMarket e Fabricante SKU;
# - filtra Fabricante durante o loop;
# - não lê colunas de volume/valor para rank;
# - não guarda o Congelado inteiro em memória;
# - não executa o filtro de fabricante duas vezes.


def _valor_celula_seguro_v10(row, idx):
    try:
        return row[idx] if idx is not None and idx < len(row) else None
    except Exception:
        return None


def _fabricante_match_rapido_v10(valor, filtro_original: str, filtro_norm: str, filtro_upper: str) -> bool:
    """Filtro rápido para fabricante, evitando normalizar Unicode em toda linha quando não precisa."""
    if not filtro_original:
        return True
    if valor is None:
        return False
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return False

    texto_upper = texto.upper()
    if texto_upper == filtro_upper or filtro_upper in texto_upper:
        return True

    texto_norm = normalizar_texto(texto)
    return texto_norm == filtro_norm or (filtro_norm and filtro_norm in texto_norm)


def _indices_por_score_v10(headers: List[str], func_score, limite: Optional[int] = None) -> List[int]:
    candidatos = []
    for i, h in enumerate(headers):
        try:
            score = int(func_score(h))
        except Exception:
            score = 0
        if score > 0:
            candidatos.append((score, i))
    candidatos.sort(reverse=True)
    indices = [i for _, i in candidatos]
    if limite is not None:
        indices = indices[:limite]
    return indices


def _ler_mapa_congelado_excel_stream_v10(
    caminho: Path,
    alternativas_cat: List[str],
    origem: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Lê Excel do Congelado em streaming e retorna diretamente o mapa EAN -> Categoria.
    Não retorna o Congelado bruto.
    """
    if avisos is None:
        avisos = []

    caminho = Path(caminho)
    layout = _detectar_layout_congelado_excel_v9(caminho, alternativas_cat, max_linhas_scan=80)
    if not layout:
        avisos.append(
            "Leitura rápida v10 do Congelado não encontrou cabeçalho com Código Barras SKU/EAN e Categoria."
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    _, aba, header_row, headers = layout

    # Código: usa no máximo as 3 melhores colunas para evitar iterar dezenas de colunas parecidas.
    # Para o layout do Congelado, a prioridade máxima é Código Barras SKU.
    idx_codigos = _indices_por_score_v10(headers, _score_nome_codigo_congelado_v9, limite=3)
    idx_cats = _indices_por_score_v10(
        headers,
        lambda h: _score_nome_categoria_congelado_v9(h, alternativas_cat),
        limite=1,
    )
    idx_fabs = _indices_por_score_v10(headers, _score_nome_fabricante_congelado_v9, limite=1)

    if not idx_codigos:
        avisos.append("Congelado v10: nenhuma coluna de código foi detectada.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not idx_cats:
        avisos.append("Congelado v10: nenhuma coluna de categoria foi detectada.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    c_cat_idx = idx_cats[0]
    c_fab_idx = idx_fabs[0] if idx_fabs else None

    fabricante_filtro = str(fabricante_filtro or "").strip()
    filtro_norm = normalizar_texto(fabricante_filtro)
    filtro_upper = fabricante_filtro.upper()

    if fabricante_filtro and c_fab_idx is None:
        raise ValueError(
            "Foi informado Fabricante para filtrar o Congelado, mas não encontrei coluna de Fabricante no Congelado.\n"
            "Colunas esperadas: Fabricante SKU, Fabricante, Fabricante do SKU, Proveedor ou Fornecedor.\n"
            f"Colunas detectadas no cabeçalho: {resumo_colunas_disponiveis(headers)}"
        )

    mapa = {}
    linhas_lidas = 0
    linhas_fabricante = 0
    linhas_categoria_valida = 0
    linhas_codigo_valido = 0

    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[aba]
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            linhas_lidas += 1

            if fabricante_filtro:
                valor_fab = _valor_celula_seguro_v10(row, c_fab_idx)
                if not _fabricante_match_rapido_v10(valor_fab, fabricante_filtro, filtro_norm, filtro_upper):
                    continue

            linhas_fabricante += 1

            categoria = _valor_celula_seguro_v10(row, c_cat_idx)
            if categoria is None:
                continue
            categoria = str(categoria).strip()
            if not categoria or categoria.lower() in {"nan", "none"}:
                continue

            categoria_key = normalizar_categoria(categoria)
            if not categoria_key:
                continue

            linhas_categoria_valida += 1

            achou_codigo_na_linha = False
            for idx_cod in idx_codigos:
                ean = ean_texto(_valor_celula_seguro_v10(row, idx_cod))
                if not ean:
                    continue
                achou_codigo_na_linha = True
                # Mantém a primeira ocorrência válida. No Congelado filtrado por fabricante,
                # o mesmo EAN deve ter a mesma categoria; não vale a pena ranquear por volume/valor.
                if ean not in mapa:
                    mapa[ean] = (categoria, categoria_key)

            if achou_codigo_na_linha:
                linhas_codigo_valido += 1
    finally:
        wb.close()

    if not mapa:
        avisos.append(
            f"Congelado v10 lido, mas nenhum EAN foi mapeado. Aba '{aba}', cabeçalho linha {header_row}. "
            f"Linhas lidas: {linhas_lidas:,}; linhas após filtro Fabricante: {linhas_fabricante:,}; "
            f"linhas com categoria válida: {linhas_categoria_valida:,}; linhas com código válido: {linhas_codigo_valido:,}.".replace(",", ".")
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    mapa_df = pd.DataFrame(
        [(ean, cat, cat_key) for ean, (cat, cat_key) in mapa.items()],
        columns=["ean", "categoria_map_prod", "categoria_key_map_prod"],
    )

    codigos_usados = ", ".join(headers[i] for i in idx_codigos)
    cat_usada = headers[c_cat_idx]
    fab_usada = headers[c_fab_idx] if c_fab_idx is not None else "não filtrado"

    avisos.append(
        f"Congelado v10 aplicado: {len(mapa_df):,} EANs mapeados diretamente por streaming; "
        f"aba '{aba}', cabeçalho linha {header_row}; código(s): {codigos_usados}; "
        f"categoria: {cat_usada}; fabricante: {fab_usada}; "
        f"linhas lidas: {linhas_lidas:,}; linhas após filtro Fabricante: {linhas_fabricante:,}.".replace(",", ".")
    )

    return mapa_df, avisos


def ler_mapa_congelado_categoria(
    congelado_path: str | Path,
    nivel_prod: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Override v10.
    Para Excel, monta o mapa direto em streaming, sem carregar o Congelado inteiro.
    Para CSV/TXT/TSV, mantém o fluxo v9 por compatibilidade.
    """
    if avisos is None:
        avisos = []

    caminho = Path(str(congelado_path or "").strip())
    if not str(caminho):
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {caminho}")

    alternativas_cat, origem = _alternativas_categoria_congelado(nivel_prod)

    # CSV continua no v9 por compatibilidade. Excel usa o modo mais leve.
    if eh_csv(caminho):
        return _ler_mapa_congelado_categoria_v8(
            caminho,
            nivel_prod,
            avisos=avisos,
            fabricante_filtro=fabricante_filtro,
        )

    try:
        return _ler_mapa_congelado_excel_stream_v10(
            caminho,
            alternativas_cat,
            origem,
            avisos=avisos,
            fabricante_filtro=fabricante_filtro,
        )
    except Exception as exc:
        # Não volta para fallback lento automaticamente, porque esse fallback era o que travava.
        # O erro precisa aparecer claramente para ajuste de cabeçalho/fabricante.
        raise RuntimeError(
            "Falha na leitura rápida v10 do Congelado. O fallback lento foi bloqueado para evitar travamento.\n"
            f"Detalhe: {exc}"
        ) from exc



# ============================================================
# Override v11 - Congelado ainda mais rápido
# ============================================================
# Melhorias em relação à v10:
# - Lê somente o intervalo entre as colunas úteis do Excel, em vez da linha inteira.
# - Para a leitura quando encontra muitas linhas vazias consecutivas nas colunas úteis.
# - Cacheia o mapa do Congelado para o modo com 2 Sell-outs não ler o mesmo arquivo duas vezes.
# - Adiciona leitura rápida também para Congelado em CSV/TXT/TSV, sem voltar ao fallback lento.

_CACHE_MAPA_CONGELADO_V11: Dict[Tuple[str, int, int, str, str], pd.DataFrame] = {}


def _cache_key_congelado_v11(caminho: Path, nivel_prod: str, fabricante_filtro: str) -> Tuple[str, int, int, str, str]:
    caminho = Path(caminho)
    st = caminho.stat()
    return (
        str(caminho.resolve()),
        int(st.st_size),
        int(st.st_mtime_ns),
        normalizar_nivel_prod(nivel_prod),
        normalizar_texto(fabricante_filtro or ""),
    )


def _celula_vazia_v11(valor) -> bool:
    if valor is None:
        return True
    txt = str(valor).strip()
    return txt == "" or txt.lower() in {"nan", "none"}


def _linha_util_vazia_v11(row, indices_relativos: List[int]) -> bool:
    for idx in indices_relativos:
        if idx is None:
            continue
        if idx < len(row) and not _celula_vazia_v11(row[idx]):
            return False
    return True


def _ler_mapa_congelado_excel_stream_v11(
    caminho: Path,
    alternativas_cat: List[str],
    origem: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
    max_linhas_vazias_seguidas: int = 5000,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Lê Excel do Congelado direto para mapa EAN -> Categoria.

    Otimização principal: após localizar o cabeçalho, não percorre a linha inteira.
    O openpyxl só entrega as células entre a menor e a maior coluna útil
    Código/Categoria/Fabricante. Isso reduz muito o custo em planilhas largas.
    """
    if avisos is None:
        avisos = []

    caminho = Path(caminho)
    layout = _detectar_layout_congelado_excel_v9(caminho, alternativas_cat, max_linhas_scan=80)
    if not layout:
        avisos.append(
            "Congelado v11: não encontrei cabeçalho com Código Barras SKU/EAN e Categoria."
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    _, aba, header_row, headers = layout

    idx_codigos = _indices_por_score_v10(headers, _score_nome_codigo_congelado_v9, limite=3)
    idx_cats = _indices_por_score_v10(
        headers,
        lambda h: _score_nome_categoria_congelado_v9(h, alternativas_cat),
        limite=1,
    )
    idx_fabs = _indices_por_score_v10(headers, _score_nome_fabricante_congelado_v9, limite=1)

    if not idx_codigos:
        avisos.append("Congelado v11: nenhuma coluna de código foi detectada.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not idx_cats:
        avisos.append("Congelado v11: nenhuma coluna de categoria foi detectada.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    c_cat_idx = idx_cats[0]
    c_fab_idx = idx_fabs[0] if idx_fabs else None

    fabricante_filtro = str(fabricante_filtro or "").strip()
    filtro_norm = normalizar_texto(fabricante_filtro)
    filtro_upper = fabricante_filtro.upper()

    if fabricante_filtro and c_fab_idx is None:
        raise ValueError(
            "Foi informado Fabricante para filtrar o Congelado, mas não encontrei coluna de Fabricante no Congelado.\n"
            "Colunas esperadas: Fabricante SKU, Fabricante, Fabricante do SKU, Proveedor ou Fornecedor.\n"
            f"Colunas detectadas no cabeçalho: {resumo_colunas_disponiveis(headers)}"
        )

    indices_abs = sorted(set(idx_codigos + [c_cat_idx] + ([c_fab_idx] if c_fab_idx is not None else [])))
    min_idx = min(indices_abs)
    max_idx = max(indices_abs)

    # Como vamos ler somente min_col:max_col, os índices precisam virar relativos.
    idx_codigos_rel = [i - min_idx for i in idx_codigos]
    c_cat_rel = c_cat_idx - min_idx
    c_fab_rel = c_fab_idx - min_idx if c_fab_idx is not None else None
    indices_util_rel = sorted(set(idx_codigos_rel + [c_cat_rel] + ([c_fab_rel] if c_fab_rel is not None else [])))

    mapa: Dict[str, Tuple[str, str]] = {}
    linhas_lidas = 0
    linhas_fabricante = 0
    linhas_categoria_valida = 0
    linhas_codigo_valido = 0
    linhas_vazias_seguidas = 0

    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[aba]
        for row in ws.iter_rows(
            min_row=header_row + 1,
            min_col=min_idx + 1,
            max_col=max_idx + 1,
            values_only=True,
        ):
            linhas_lidas += 1

            if _linha_util_vazia_v11(row, indices_util_rel):
                linhas_vazias_seguidas += 1
                if linhas_vazias_seguidas >= max_linhas_vazias_seguidas:
                    break
                continue
            linhas_vazias_seguidas = 0

            if fabricante_filtro:
                valor_fab = _valor_celula_seguro_v10(row, c_fab_rel)
                if not _fabricante_match_rapido_v10(valor_fab, fabricante_filtro, filtro_norm, filtro_upper):
                    continue

            linhas_fabricante += 1

            categoria = _valor_celula_seguro_v10(row, c_cat_rel)
            if categoria is None:
                continue
            categoria = str(categoria).strip()
            if not categoria or categoria.lower() in {"nan", "none"}:
                continue

            categoria_key = normalizar_categoria(categoria)
            if not categoria_key:
                continue

            linhas_categoria_valida += 1

            achou_codigo_na_linha = False
            for idx_cod_rel in idx_codigos_rel:
                ean = ean_texto(_valor_celula_seguro_v10(row, idx_cod_rel))
                if not ean:
                    continue
                achou_codigo_na_linha = True
                if ean not in mapa:
                    mapa[ean] = (categoria, categoria_key)

            if achou_codigo_na_linha:
                linhas_codigo_valido += 1
    finally:
        wb.close()

    if not mapa:
        avisos.append(
            f"Congelado v11 lido, mas nenhum EAN foi mapeado. Aba '{aba}', cabeçalho linha {header_row}. "
            f"Linhas lidas: {linhas_lidas:,}; linhas após filtro Fabricante: {linhas_fabricante:,}; "
            f"linhas com categoria válida: {linhas_categoria_valida:,}; linhas com código válido: {linhas_codigo_valido:,}.".replace(",", ".")
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    mapa_df = pd.DataFrame(
        [(ean, cat, cat_key) for ean, (cat, cat_key) in mapa.items()],
        columns=["ean", "categoria_map_prod", "categoria_key_map_prod"],
    )

    codigos_usados = ", ".join(headers[i] for i in idx_codigos)
    cat_usada = headers[c_cat_idx]
    fab_usada = headers[c_fab_idx] if c_fab_idx is not None else "não filtrado"

    avisos.append(
        f"Congelado v11 aplicado: {len(mapa_df):,} EANs mapeados; "
        f"aba '{aba}', cabeçalho linha {header_row}; código(s): {codigos_usados}; "
        f"categoria: {cat_usada}; fabricante: {fab_usada}; "
        f"colunas lidas no Excel: {min_idx + 1} até {max_idx + 1}; "
        f"linhas lidas: {linhas_lidas:,}; linhas após filtro Fabricante: {linhas_fabricante:,}.".replace(",", ".")
    )

    return mapa_df, avisos


def _detectar_layout_congelado_csv_v11(caminho: Path, alternativas_cat: List[str], max_linhas_scan: int = 120):
    enc, sep, sample = detectar_csv_formato(caminho, nrows=max_linhas_scan)
    melhor = None
    sample = limpar_bom_dataframe(sample)
    for idx in range(min(max_linhas_scan, len(sample))):
        headers = _headers_congelado_v9(sample.iloc[idx].tolist())
        if not headers:
            continue
        scores_codigo = [_score_nome_codigo_congelado_v9(h) for h in headers]
        scores_cat = [_score_nome_categoria_congelado_v9(h, alternativas_cat) for h in headers]
        scores_fab = [_score_nome_fabricante_congelado_v9(h) for h in headers]
        max_codigo = max(scores_codigo) if scores_codigo else 0
        max_cat = max(scores_cat) if scores_cat else 0
        max_fab = max(scores_fab) if scores_fab else 0
        if max_codigo <= 0 or max_cat <= 0:
            continue
        score = max_codigo * 10 + max_cat * 8 + max_fab + min(len(headers), 80)
        if any(normalizar_texto(h) == normalizar_texto("Código Barras SKU") for h in headers):
            score += 5000
        if any(normalizar_texto(h) == normalizar_texto("Categoría congelada ScannMarket") for h in headers):
            score += 5000
        candidato = (score, enc, sep, idx + 1, headers)
        if melhor is None or candidato[0] > melhor[0]:
            melhor = candidato
    return melhor


def _ler_mapa_congelado_csv_stream_v11(
    caminho: Path,
    alternativas_cat: List[str],
    origem: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
    chunksize: int = 200_000,
) -> Tuple[pd.DataFrame, List[str]]:
    """Lê Congelado CSV/TXT/TSV em blocos, somente com código/categoria/fabricante."""
    if avisos is None:
        avisos = []

    layout = _detectar_layout_congelado_csv_v11(caminho, alternativas_cat)
    if not layout:
        avisos.append("Congelado CSV v11: não encontrei cabeçalho com Código Barras SKU/EAN e Categoria.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    _, enc, sep, header_row, headers = layout
    idx_codigos = _indices_por_score_v10(headers, _score_nome_codigo_congelado_v9, limite=3)
    idx_cats = _indices_por_score_v10(headers, lambda h: _score_nome_categoria_congelado_v9(h, alternativas_cat), limite=1)
    idx_fabs = _indices_por_score_v10(headers, _score_nome_fabricante_congelado_v9, limite=1)

    if not idx_codigos or not idx_cats:
        avisos.append("Congelado CSV v11: código ou categoria não detectados após localizar cabeçalho.")
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    fabricante_filtro = str(fabricante_filtro or "").strip()
    c_fab_idx = idx_fabs[0] if idx_fabs else None
    if fabricante_filtro and c_fab_idx is None:
        raise ValueError(
            "Foi informado Fabricante para filtrar o Congelado CSV, mas não encontrei coluna de Fabricante.\n"
            f"Colunas detectadas no cabeçalho: {resumo_colunas_disponiveis(headers)}"
        )

    usecols = sorted(set(idx_codigos + idx_cats + ([c_fab_idx] if c_fab_idx is not None else [])))
    nomes_usecols = [headers[i] for i in usecols]
    pos_para_nome = {i: headers[i] for i in usecols}
    nomes_codigo = [pos_para_nome[i] for i in idx_codigos if i in pos_para_nome]
    nome_cat = pos_para_nome[idx_cats[0]]
    nome_fab = pos_para_nome[c_fab_idx] if c_fab_idx is not None else None

    engine = "python" if sep is None else "c"
    read_kwargs = dict(
        filepath_or_buffer=caminho,
        sep=sep,
        engine=engine,
        header=None,
        skiprows=header_row,
        usecols=usecols,
        names=nomes_usecols,
        dtype=str,
        keep_default_na=False,
        skip_blank_lines=False,
        on_bad_lines="skip",
        encoding=enc,
        chunksize=chunksize,
    )
    if engine == "c":
        read_kwargs["low_memory"] = False

    mapa: Dict[str, Tuple[str, str]] = {}
    linhas_lidas = 0
    linhas_fabricante = 0

    filtro_norm = normalizar_texto(fabricante_filtro)
    filtro_upper = fabricante_filtro.upper()

    for chunk in pd.read_csv(**read_kwargs):
        chunk = limpar_bom_dataframe(chunk)
        linhas_lidas += len(chunk)

        if fabricante_filtro and nome_fab:
            mask_fab = chunk[nome_fab].map(lambda x: _fabricante_match_rapido_v10(x, fabricante_filtro, filtro_norm, filtro_upper))
            chunk = chunk[mask_fab].copy()

        linhas_fabricante += len(chunk)
        if chunk.empty:
            continue

        chunk[nome_cat] = chunk[nome_cat].fillna("").astype(str).str.strip()
        chunk = chunk[chunk[nome_cat].map(normalizar_categoria).ne("")].copy()
        if chunk.empty:
            continue

        for nome_codigo in nomes_codigo:
            if nome_codigo not in chunk.columns:
                continue
            temp = chunk[[nome_codigo, nome_cat]].copy()
            temp["ean"] = temp[nome_codigo].map(ean_texto)
            temp = temp[temp["ean"] != ""]
            if temp.empty:
                continue
            for ean, categoria in zip(temp["ean"].tolist(), temp[nome_cat].tolist()):
                if ean not in mapa:
                    cat = str(categoria).strip()
                    mapa[ean] = (cat, normalizar_categoria(cat))

    if not mapa:
        avisos.append(
            f"Congelado CSV v11 lido, mas nenhum EAN foi mapeado. Linhas lidas: {linhas_lidas:,}; "
            f"linhas após filtro Fabricante: {linhas_fabricante:,}.".replace(",", ".")
        )
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos

    mapa_df = pd.DataFrame(
        [(ean, cat, cat_key) for ean, (cat, cat_key) in mapa.items()],
        columns=["ean", "categoria_map_prod", "categoria_key_map_prod"],
    )
    avisos.append(
        f"Congelado CSV v11 aplicado: {len(mapa_df):,} EANs mapeados; "
        f"código(s): {', '.join(nomes_codigo)}; categoria: {nome_cat}; fabricante: {nome_fab or 'não filtrado'}; "
        f"linhas lidas: {linhas_lidas:,}; linhas após filtro Fabricante: {linhas_fabricante:,}.".replace(",", ".")
    )
    return mapa_df, avisos


def ler_mapa_congelado_categoria(
    congelado_path: str | Path,
    nivel_prod: str,
    avisos: Optional[List[str]] = None,
    fabricante_filtro: str = "",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Override v11.
    Usa cache + leitura enxuta do Congelado para evitar travamento, principalmente
    no modo Estudo de Cobertura com 2 Sell-outs.
    """
    if avisos is None:
        avisos = []

    caminho = Path(str(congelado_path or "").strip())
    if not str(caminho):
        return pd.DataFrame(columns=["ean", "categoria_map_prod", "categoria_key_map_prod"]), avisos
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo Congelado não encontrado: {caminho}")

    alternativas_cat, origem = _alternativas_categoria_congelado(nivel_prod)
    cache_key = _cache_key_congelado_v11(caminho, nivel_prod, fabricante_filtro)

    if cache_key in _CACHE_MAPA_CONGELADO_V11:
        mapa_cache = _CACHE_MAPA_CONGELADO_V11[cache_key].copy()
        avisos.append(
            f"Congelado v11 reutilizado do cache: {len(mapa_cache):,} EANs mapeados. ".replace(",", ".")
            + "Isso evita reler o Congelado no segundo Sell-out."
        )
        return mapa_cache, avisos

    try:
        if eh_csv(caminho):
            mapa, avisos = _ler_mapa_congelado_csv_stream_v11(
                caminho,
                alternativas_cat,
                origem,
                avisos=avisos,
                fabricante_filtro=fabricante_filtro,
            )
        else:
            mapa, avisos = _ler_mapa_congelado_excel_stream_v11(
                caminho,
                alternativas_cat,
                origem,
                avisos=avisos,
                fabricante_filtro=fabricante_filtro,
            )
    except Exception as exc:
        raise RuntimeError(
            "Falha na leitura rápida v11 do Congelado. O fallback lento continua bloqueado para evitar travamento.\n"
            f"Detalhe: {exc}"
        ) from exc

    if mapa is not None and not mapa.empty:
        _CACHE_MAPA_CONGELADO_V11[cache_key] = mapa.copy()

    return mapa, avisos

if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nERRO:")
        print(exc)
        sys.exit(1)
